#!/usr/bin/env python
# -*- coding: utf-8 -*-
#-----------------------------------------------------------------------------
# Name:        HeavyBase.py
# Purpose:     Hybrid Online-Offline multiplatform P2P data entry engine 
#              for electronic Case Report Forms and 'Omics' data sharing
#              based on a historiographical "Push-based" Peer-to-Peer DB
#
# Author:      Luca Clivio <luca.clivio@heavybase.org>
#              Contacts: 2nd mail luca@clivio.net, mobile +39-347-2538040
#
# Created:     2006/06/04
# RCS-ID:      $Id: HeavyBase.py $
#
# Copyright:   2006-2014 Luca Clivio
# License:     GNU-GPL v3
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
# Version: 5.8.5.0 - Released: 2014-02-04 12:00:00 [yyyy-mm-dd hh:mm:ss]
#-----------------------------------------------------------------------------

RELEASE="5.8.5"

#PROGNAME="HEAVyBASE"
import sys
PROGNAME=sys.argv[0][:-3]
SUPPORT="luca.clivio@heavybase.org"
if PROGNAME!="HeavyBase": SUPPORT="luca.clivio@p2pdb.net"
DISCLAIMER=PROGNAME+" "+RELEASE+", Open source P2P data management system for eCRF and 'Omics' data sharing\nCopyright (C) 2006-2014 Luca Clivio <"+SUPPORT+">"
print "\n"+DISCLAIMER

DEBUG_MODE=False
LOGFILE=""
import os
if os.path.isfile("debug.txt"): DEBUG_MODE=True
# updates - begin
updates=False
try:
    infile = open("update.py")
    contents=infile.read()
    infile.close()
    updates=True
except: pass
if updates:
    print "updating..."
    if DEBUG_MODE:
        exec(contents)
        print "update ok."
        os.unlink("update.py")
    else:
        try:
            exec(contents)
            print "update ok."
            os.unlink("update.py")
        except:
            print "update failed."
# updates - end

# HeavyBaseService - Begin
#The server object
class HeavyBaseService:
    def __init__(self, peermode=False):
        self.peermode=peermode
        self.queue=[]
        #self.log("-","init")
        self.peers={}
        # autoelezione primario -inzio
        import random
        random.seed()
        self.id_instance=str(int(random.random()*10000000000))[0:10]
        # autoelezione primario - fine
        
    def getServiceInstance(self, project_name, username, password):
        return self.id_instance

    def log(self,project_name,string):
        #print ts+"\t"+string
        #return self.client_ip
        return
        import datetime
        now = datetime.datetime.utcnow()
        ts=now.strftime("%Y%m%d%H%M%S")
        try:
            logfile = open("HeavyBaseService_"+project_name+".log", "a")
            try:
                logfile.write(ts+"\t"+string+"\n")
            finally:
                logfile.close()
        except IOError:
            pass

    def sendMail(self, project_name, username, password, sender, comma_sep_receivers, subject, body):
        self.log(project_name,"sendMail - username="+username+"; password="+password)
        if self.peermode:
            ret = "Sending mail not allowed by peers."
        else:
            import smtplib,datetime
            receivers = comma_sep_receivers.split(",")
            message = "From: " + sender + "\n" + "To: " + comma_sep_receivers + "\n" + "Subject: " + subject + "\n\n" + body
            now = datetime.datetime.utcnow()
            ts=now.strftime("%Y%m%d%H%M%S")
            ret=""
            try:
               smtpObj = smtplib.SMTP('localhost')
               smtpObj.sendmail(sender, receivers, message)
               ret = "Mail sent successfully. Timestamp:" + ts
            except SMTPException:
               ret = "Error(2) sending mail. Timestamp:" + ts
        self.log(project_name,"sendMail - Output: "+ret)
        return ret

    def getRelease(self, project_name, username, password):
        self.log(project_name,"getRelease - username="+username+"; password="+password)
        if self.peermode:
            release="0.0.0"
        else:
            release=""
            if len(project_name)>0:
                if project_name=="sinpe_domus":
                    release="2.0.0"
                else:
                    release=RELEASE
        self.log(project_name,"getRelease - Output: "+release)
        return release

    def checkUpdates(self, project_name, username, password):
        import datetime
        self.log(project_name,"checkUpdates - username="+username+"; password="+password)
        if self.peermode:
            updates=""
        else:
            now = datetime.datetime.utcnow()
            ts=now.strftime("%Y%m%d%H%M%S")
            
            from HeavyBaseService_updates import HeavyBaseService_updates
            up=HeavyBaseService_updates()
            updates=up.GetFilenames(project_name)

        self.log(project_name,"checkUpdates - Output: "+updates)
        return updates

    def downloadFile(self, project_name, username, password, filename):
        import base64, zlib, datetime
        self.log(project_name,"downloadFile - username="+username+"; password="+password+"; filename="+filename)
        if self.peermode:
            encoded=""
        else:
            now = datetime.datetime.utcnow()
            ts=now.strftime("%Y%m%d%H%M%S")
            encoded=""
            try:
                infile = open(project_name+"/"+filename,"rb")
                contents=infile.read()
                infile.close()
                encoded = base64.b64encode(zlib.compress(contents))
            except:
                pass
            if encoded:
                self.log(project_name,"downloadFile - Output: <file encoded>")
            else:  
                self.log(project_name,"downloadFile - Output: <nothing>")
        return encoded

    def uploadFile(self, project_name, username, password, filename, encoded):
        import base64, zlib, datetime, os
        self.log(project_name,"uploadFile - username="+username+"; password="+password+"; filename="+filename)
        if self.peermode:
            return "File upload denied."
        else:
            try: os.makedirs("upload/"+project_name)
            except: pass
            data = zlib.decompress(base64.b64decode(encoded))
            outfile = open("upload/"+project_name+"/"+filename, 'wb')
            outfile.write(data)
            outfile.close()
            return "File "+filename+" successfully uploaded."

    def findPeers(self,projectname,username,password,id_instance, internal_ip, use_proxy, client_ip):
        import datetime
        if client_ip!="": self.peers[id_instance]=client_ip
        return str(self.peers)
        
        now = datetime.datetime.utcnow()
        ts=now.strftime("%Y%m%d%H%M%S")

        self.peers={}
        try: self.peers=pickle.load(file("peers_"+project_name+".pkl", "r"))
        except: pass

        #cleaning
        isclean=False
        while not isclean:
            idx=-1
            ct=0
            for q in self.peers:
                if (now-q["timestamp"]).seconds>180:
                    idx=ct
                    break
                ct=ct+1
            if idx!=-1:
                del self.peers[idx]
            else:
                isclean=True

        #adding current
        idx=-1
        ct=0
        #for q in self.peers:
        #    if     

    
    def formatExceptionInfo(maxTBlevel=5):
        import sys
        import traceback
        cla, exc, trbk = sys.exc_info()
        excName = cla.__name__
        try:
            excArgs = exc.__dict__["args"]
        except KeyError:
            excArgs = "<no args>"
        excTb = traceback.format_tb(trbk, maxTBlevel)
        return (excName, excArgs, excTb)

    def p2p(self, project_name, username, password, id_instance, state, partner, encoded):
        self.log(project_name,"p2p - username="+username+"; password="+password+"; id_instance="+id_instance+"; state="+state)
        import base64, zlib, datetime, time
        now = datetime.datetime.utcnow()
        ts=now.strftime("%Y%m%d%H%M%S")

        compatibilityMatrix={}
        compatibilityMatrix["42x"]=["4.2.5","4.2.6","4.2.7","4.2.8","4.2.9","4.3.0","4.3.1","4.3.2","4.3.3","4.3.4","4.3.5","4.3.6","4.3.7","4.3.8","4.3.9","4.4.0","4.4.1","4.4.2","4.4.3","4.4.4","4.4.5","4.4.6","4.4.7"]    #the keys are the compatibility groups logical code
        compatibilityMatrix["45x"]=["4.5.0","4.5.1"]    #the keys are the compatibility groups logical code
        compatibilityMatrix["452x"]=["4.5.2","4.5.3","4.5.4","4.5.5","4.5.6","4.5.7","4.5.8","4.5.9","4.6.0","4.6.1","4.6.2","4.6.3","4.6.4","4.6.5","4.6.6","4.6.7","4.6.8","4.6.9","4.7.0","4.7.1"
                                    ,"4.7.2","4.7.3","4.7.4","4.7.5","4.7.6","4.7.7","4.7.8","4.7.9","4.8.0","4.8.1","4.8.2","4.8.3","4.8.4","4.8.5","4.8.6","4.8.7","4.8.8","4.8.9","4.9.0","4.9.1"
                                    ,"4.9.2","4.9.3","4.9.4","4.9.5","4.9.6","4.9.7","4.9.8","4.9.9","5.0.0","5.0.1","5.0.2","5.0.3","5.0.4","5.0.5","5.0.6","5.0.7","5.0.8","5.0.9","5.1.0","5.1.1"
                                    ,"5.1.2","5.1.3","5.1.4","5.1.5","5.1.6","5.1.7","5.1.8","5.1.9","5.2.0","5.2.1","5.2.2","5.2.3","5.2.4","5.2.5","5.2.6","5.2.7","5.2.8","5.2.9","5.3.0","5.3.1"
                                    ,"5.3.2","5.3.3","5.3.4","5.3.5","5.3.6","5.3.7","5.3.8","5.3.9","5.4.0","5.4.1","5.4.2","5.4.3","5.4.4","5.4.5","5.4.6","5.4.7","5.4.8","5.4.9","5.5.0","5.5.1"
                                    ,"5.5.2","5.5.3","5.5.4","5.5.5","5.5.6","5.5.7","5.5.8","5.5.9","5.6.0","5.6.1","5.6.2","5.6.3","5.6.4","5.6.5","5.6.6","5.6.7","5.6.8","5.6.9","5.7.0","5.7.1"
                                    ,"5.7.2","5.7.3","5.7.4","5.7.5","5.7.6","5.7.7","5.7.8","5.7.9","5.8.0","5.8.1","5.8.2","5.8.3","5.8.4","5.8.5"]
        if project_name.find("--")<0: project_name=project_name+"--old"  #case teorically impossible
        for key in compatibilityMatrix:
            if project_name.split("--")[1] in compatibilityMatrix[key]:
                project_name=project_name.split("--")[0]+"--comp"+key
        
        returnValue=",,"+"NOP"
        try:
            #locking
            lock_ok=False
            reason_lock=""
            if len(project_name)>0 and len(username)>0 and len(password)>0:
                #while not lock_ok:
                import pickle
                try:
                    sem_datetime=pickle.load(file("sem_"+project_name+".pkl", "r"))
                    if (now-sem_datetime).seconds>90: 
                        lock_ok=True
                    else:
                        reason_lock="["+`(now-sem_datetime).seconds`+" sec]"
                except:
                    lock_ok=True
                #print "lock status: "+str(lock_ok)
                if lock_ok:
                    pickle.dump(now,file("sem_"+project_name+".pkl", "w"))
                    self.log(project_name,id_instance+" - ("+project_name+" locked)")
                    self.queue=[]
                    try:
                        self.queue=pickle.load(file("queue_"+project_name+".pkl", "r"))
                    except: pass
                else:
                    self.log(project_name,id_instance+" locked "+reason_lock+". waiting.")
                    returnValue=",,"+"NOP"
            #lock_ok=True
            #encoded=""
            reset=False
            reason_reset=""
            if lock_ok:
                self.log(project_name,id_instance+" - step 1: cleaning")
                #cleaning
                isclean=False
                while not isclean:
                    idxClean=-1
                    ct=0
                    for q in self.queue:
                        self.log(project_name,q["instance"]+" not seen in the last "+`(now-q["timestamp"]).seconds`+" seconds")
                        if (now-q["timestamp"]).seconds>180:
                            idxClean=ct
                            if q["instance"]==id_instance:
                                reset=True
                                reason_reset="[cleaning]"
                            break
                        ct=ct+1
                    if idxClean!=-1:
                        self.log(project_name,"cleaning "+self.queue[idxClean]["instance"])
                        self.queue.remove(self.queue[idxClean])
                    else:
                        isclean=True

                if not reset:
                    newInstance=False
                    #common
                    idx=-1
                    ct=0
                    for q in self.queue:
                        if q["instance"]==id_instance:
                            idx=ct
                            break
                        ct=ct+1
                    if idx==-1:
                        q={}
                        q["instance"]=id_instance
                        self.queue.append(q)
                        idx=ct
                        newInstance=True
                    self.queue[idx]["state"]=state
                    self.queue[idx]["timestamp"]=now
                    self.queue[idx]["project_name"]=project_name
                    self.queue[idx]["username"]=username
                    self.queue[idx]["password"]=password

                    if self.queue[idx].has_key("partner"):
                        self.log(project_name,id_instance+" has partner. - step 1.9: partner validation")
                        found=False
                        for q in self.queue:
                            if q["instance"]==id_instance:
                                found=True
                                break
                        if not found: 
                            self.log(project_name,id_instance+" - partner "+self.queue[idx]["partner"]+" removed")
                            del self.queue[idx]["partner"]
                                    
                    #state machine
                    returnValue=",,NOP"
                    if newInstance and state!="A": 
                        reset=True
                        reason_reset="[new instance and state is not A]"
                    if not reset: 
                        self.log(project_name,id_instance+" - step 2: state machine begin")
                        idxPartner=-1
                        if state=="A":
                            if self.queue[idx].has_key("standby"): del self.queue[idx]["standby"]
                            data=zlib.decompress(base64.b64decode(encoded))
                            self.queue[idx]["md5Rows"]=data
                            #partner searching
                            ct=0
                            for q in self.queue:
                                self.log(project_name,q["instance"] + " " + `q.has_key("partner")`)
                                if not self.queue[idx].has_key("partner"):
                                    if not q.has_key("partner"):
                                        #if q["instance"]!=id_instance and q["md5Rows"]!=data and q["project_name"]==project_name:
                                        if q["instance"]!=id_instance and q["md5Rows"]!=data and q["project_name"]==project_name and q["state"] in ["A","B","B1"]:
                                            idxPartner=ct
                                            self.queue[idxPartner]["partner"]=id_instance  #self.queue[idx]["instance"]
                                            #q["partner"]=id_instance  #self.queue[idx]["instance"]
                                            self.queue[idx]["partner"]=q["instance"]    #self.queue[idxPartner]["instance"]
                                            self.log(project_name,id_instance + " <-1-> " + q["instance"])
                                    elif q["partner"]==id_instance:
                                        idxPartner=ct
                                        self.queue[idx]["partner"]=q["instance"]
                                        self.log(project_name,id_instance + " <-2-> " + q["instance"])
                                #elif q["partner"]==id_instance:
                                else:
                                    if not q.has_key("partner"):
                                        pass
                                    elif q["partner"]==id_instance:
                                        idxPartner=ct
                                        self.log(project_name,id_instance + " <-3-> " + q["instance"])
                                ct=ct+1
                                # if idxPartner!=-1: break
                            self.log(project_name,id_instance + " - 2.01: end cycle")
                            if idxPartner!=-1:
                                encoded = base64.b64encode(zlib.compress(self.queue[idxPartner]["md5Rows"]))
                                returnValue=self.queue[idxPartner]["instance"]+","+self.queue[idxPartner]["state"]+","+encoded
                            else:
                                returnValue=",,NOP"
                        else:
                            idxPartner=-1
                            ct=0
                            for q in self.queue:
                                if q.has_key("partner"):
                                    if q["partner"]==id_instance:
                                        idxPartner=ct
                                        self.log(project_name,id_instance + " <-4-> " + q["instance"])
                                ct=ct+1
                                if idxPartner!=-1: break
                            if idxPartner==-1 and state!="E": 
                                reset=True
                                reason_reset="[partner disappeared]"

                        self.log(project_name,id_instance+" - step 2.1: ready for data input")
                        if self.queue[idx].has_key("standby"):
                            self.log(project_name,self.queue[idx]["instance"]+" has been idle for "+`int((now-self.queue[idx]["standby"]).seconds)`+" seconds.")
                            if (now-self.queue[idx]["standby"]).seconds>180: 
                                reset=True
                                reason_reset="[idle for >180 sec]"
                    #print str(reset)
                    self.log(project_name,id_instance+" - step 2.5: reset="+`reset`+" "+reason_reset)
                    if not reset: 
                        if idxPartner!=-1:
                            self.log(project_name,id_instance+" - step 3: data input")
                            if state=="B":
                                if self.queue[idx].has_key("standby"): del self.queue[idx]["standby"]
                                self.queue[idx]["lstRows"]=zlib.decompress(base64.b64decode(encoded))
                            elif state=="B1":
                                if not self.queue[idx].has_key("standby"): 
                                    self.queue[idx]["standby"]=now
                                    self.log(project_name,self.queue[idx]["instance"]+" from now in standby (B1)")
                            elif state=="C":
                                if self.queue[idx].has_key("standby"): del self.queue[idx]["standby"]
                                self.queue[idx]["deltaRows"]=zlib.decompress(base64.b64decode(encoded.split(",")[0]))
                                if len(encoded.split(","))>=2: self.queue[idx]["deltaContents_index"]=zlib.decompress(base64.b64decode(encoded.split(",")[1]))
                                else: self.queue[idx]["deltaContents_index"]=""
                                if len(encoded.split(","))>=3: self.queue[idx]["deltaHeaders"]=zlib.decompress(base64.b64decode(encoded.split(",")[2]))
                                else: self.queue[idx]["deltaHeaders"]=""
                            elif state=="C1":
                                if not self.queue[idx].has_key("standby"): 
                                    self.queue[idx]["standby"]=now
                                    self.log(project_name,self.queue[idx]["instance"]+" from now in standby (C1)")
                            elif state=="D":
                                if self.queue[idx].has_key("standby"): del self.queue[idx]["standby"]
                                self.queue[idx]["deltaDict"]=zlib.decompress(base64.b64decode(encoded))
                            elif state=="D1":
                                if not self.queue[idx].has_key("standby"): 
                                    self.queue[idx]["standby"]=now
                                    self.log(project_name,self.queue[idx]["instance"]+" from now in standby (D1)")


                            self.log(project_name,id_instance+" - step 4: data output")
                            # if self.queue[idxPartner]["state"] in ["A"]:
                            if state in ["A"]:
                                pass
                            # elif self.queue[idxPartner]["state"] in ["B","B1"]:
                            elif state in ["B","B1"] and self.queue[idxPartner]["state"] in ["B","B1","C","C1"]:
                                encoded="NOP"
                                if not self.queue[idx].has_key("sent_B"):
                                    self.queue[idx]["sent_B"]=True
                                    encoded = base64.b64encode(zlib.compress(self.queue[idxPartner]["lstRows"]))
                                    self.log(project_name,id_instance+" - step 4.5: sending md5Rows and lstRows")
                                    self.queue[idxPartner]["md5Rows"]=""
                                    self.queue[idxPartner]["lstRows"]=""
                                returnValue=self.queue[idxPartner]["instance"]+","+self.queue[idxPartner]["state"]+","+encoded
                            # elif self.queue[idxPartner]["state"] in ["C","C1"]:
                            elif state in ["C","C1"] and self.queue[idxPartner]["state"] in ["C","C1","D","D1"]:
                                encoded="NOP"
                                if not self.queue[idx].has_key("sent_C"):
                                    self.queue[idx]["sent_C"]=True
                                    encoded = base64.b64encode(zlib.compress(self.queue[idxPartner]["deltaRows"]))+","+base64.b64encode(zlib.compress(self.queue[idxPartner]["deltaContents_index"]))+","+base64.b64encode(zlib.compress(self.queue[idxPartner]["deltaHeaders"]))
                                    self.log(project_name,id_instance+" - step 4.5: sending deltaRows, deltaContents_index, deltaHeaders")
                                    self.queue[idxPartner]["deltaRows"]=""
                                    self.queue[idxPartner]["deltaContents_index"]=""
                                    self.queue[idxPartner]["deltaHeaders"]=""
                                    #in questo momento si e' perso il legame tra record (che vengono spediti ora) e dati (che devono ancora arrivare)
                                returnValue=self.queue[idxPartner]["instance"]+","+self.queue[idxPartner]["state"]+","+encoded
                            # elif self.queue[idxPartner]["state"] in ["D","D1"]:
                            elif state in ["D","D1"] and self.queue[idxPartner]["state"] in ["D","D1","E"]:
                                encoded="NOP"
                                if not self.queue[idx].has_key("sent_D"):
                                    self.queue[idx]["sent_D"]=True
                                    encoded = base64.b64encode(zlib.compress(self.queue[idxPartner]["deltaDict"]))
                                    self.log(project_name,id_instance+" - step 4.5: sending deltaDict")
                                    self.queue[idxPartner]["deltaDict"]=""
                                returnValue=self.queue[idxPartner]["instance"]+","+self.queue[idxPartner]["state"]+","+encoded
                            else:
                                returnValue=self.queue[idxPartner]["instance"]+","+self.queue[idxPartner]["state"]+","+"NOP"
                            self.log(project_name,id_instance+" - step 5: end data output")

                        if state=="E":
                            if self.queue[idx].has_key("standby"): del self.queue[idx]["standby"]
                            if idxPartner!=-1:
                                if self.queue[idxPartner]["state"] in ["E"]:
                                    self.log(project_name,id_instance+" has completed the syncronization, partner "+self.queue[idxPartner]["instance"]+" has state="+self.queue[idxPartner]["state"])
                                    self.queue.remove(self.queue[idx])
                                    # del self.queue[idxPartner]["partner"]
                                else:
                                    self.log(project_name,id_instance+" waiting the other end for completing the syncronization")
                            else:
                                self.log(project_name,id_instance+" has completed the syncronization, partner "+self.queue[idx]["partner"]+" has already completed.")
                                self.queue.remove(self.queue[idx])
                                returnValue=",,RES"
                    else:
                        self.queue.remove(self.queue[idx])
                        # self.queue.remove(self.queue[idxPartner])
                        returnValue=",,RES"
                        self.log(project_name,id_instance+" sending reset(1) because: "+reason_reset)

                else:
                    #qui reset e' True
                    returnValue=",,RES"
                    self.log(project_name,id_instance+" sending reset(2) because: "+reason_reset)
                    
                #salvataggio dati - inizio
                self.log(project_name,id_instance+" - step 6: pickle.dump")
                pickle.dump(self.queue,file("queue_"+project_name+".pkl", "w"))
                try:
                    import os
                    os.unlink("sem_"+project_name+".pkl")
                    self.log(project_name,id_instance+" - ("+project_name+" unlocked)")
                except: pass
                for key in self.queue:
                    del self.queue[key]
                del self.queue
                #salvataggio dati - fine
            else:
                #qui lock e' False - devo aspettare senza cancellare il semaforo
                pass

            ##salvataggio dati - inizio
            #self.log(project_name,id_instance+" - step 6: pickle.dump")
            #pickle.dump(self.queue,file("queue_"+project_name+".pkl", "w"))
            #if lock_ok:
                #try:
                    #import os
                    #os.unlink("sem_"+project_name+".pkl")
                    #self.log(project_name,id_instance+" - ("+project_name+" unlocked)")
                #except: pass
            #del self.queue
            ##salvataggio dati - fine

            ct=0
            tolog=""
            for elm in returnValue.split(","):
                if ct>0: tolog=tolog+","
                if len(elm)<10: 
                    tolog=tolog+elm
                else:
                    tolog=tolog+"<encoded("+str(len(elm))+")>"
                ct=ct+1
            tolog=tolog+" - ["+str(len(returnValue.split(",")))+"]"
            self.log(project_name,"p2p - Output: "+tolog)
        except:
            self.log(project_name,id_instance+" ERROR:")
            import sys
            self.log(project_name,id_instance+" - "+`sys.exc_info()`)
            self.log(project_name,id_instance+" - "+`self.formatExceptionInfo()`)
        return returnValue

import threading
class HeavyBaseServiceStarter(threading.Thread):
    def __init__(self):
        threading.Thread.__init__(self)
        self.timeToQuit = threading.Event()
        self.timeToQuit.clear()      

    def stop(self):    
        try:
            print "Shutting down "+PROGNAME+"Service"
            self.server.server_close()
            #self.server.shutdown()
        except: pass
        self.timeToQuit.set()
        self.join()

    def run(self):
        import SimpleXMLRPCServer, base64, zlib, datetime, time, socket

        socket.setdefaulttimeout(10)
        
        try:
            self.server = SimpleXMLRPCServer.SimpleXMLRPCServer(("", 60002))
            self.heavybaseservice_object = HeavyBaseService(True)
            self.server.register_instance(HeavyBaseService())
            print "Starting "+PROGNAME+"Service on port 60002"
            while not self.timeToQuit.isSet():
                self.server.handle_request()
        except:
            print PROGNAME+"Service already started on port 60002"
# HeavyBaseService - End

import shutil
import datetime, time
import math, string
# import wxversion
# wxversion.select('2.8')
import wx
import wx.xrc
import wx.aui
import wx.grid as gridlib
import wx.lib.masked as masked

if os.name=='nt':
    import win32api
    import win32con
    import win32gui_struct
    try:
        import winxpgui as win32gui
    except ImportError:
        import win32gui

    
#SCREEN_X=900
SCREEN_X=1024
#SCREEN_Y=640
SCREEN_Y=520

DATABASE_NAME = PROGNAME.lower()+'.db'
if len(sys.argv)>=2:
    DATABASE_NAME = sys.argv[1]
DATABASE = os.path.abspath(DATABASE_NAME)
DATABASE_PATH=DATABASE[:DATABASE.rfind(os.path.sep)+1]
if DEBUG_MODE: print "Database path: "+DATABASE_PATH
PROFILE_NAME = PROGNAME.lower()+'.xrc'
if len(sys.argv)>=3:
    PROFILE = sys.argv[2]
PROFILE = os.path.abspath(PROFILE_NAME)

CRF_DESC="crf"
REPORTS_PATH=".."+os.path.sep+"reports"+os.path.sep
DOC_PATH=".."+os.path.sep+"doc"+os.path.sep
try: os.makedirs(REPORTS_PATH)
except: pass

def ScreenCapture( captureStartPos, captureBmapSize, debug=False ):
    """
    General Desktop screen portion capture - partial or entire Desktop.
    
    My particular screen hardware configuration: 
        wx.Display( 0 ) refers to the extended Desktop display monitor screen.
        wx.Display( 1 ) refers to the primary  Desktop display monitor screen.
    
    Any particular Desktop screen size is :
        screenRect = wx.Display( n ).GetGeometry()
        
    Different wx.Display's in a single system may have different dimensions.        
    """

    # A wx.ScreenDC provides access to the entire Desktop.
    # This includes any extended Desktop monitor screens that are enabled in the OS.
    scrDC = wx.ScreenDC()
    scrDcSize = scrDC.Size
    scrDcSizeX, scrDcSizeY = scrDcSize
    # Cross-platform adaptations :
    scrDcBmap     = scrDC.GetAsBitmap()
    scrDcBmapSize = scrDcBmap.GetSize()
    if debug :
        print 'DEBUG:  Size of scrDC.GetAsBitmap() ', scrDcBmapSize
    # Check if scrDC.GetAsBitmap() method has been implemented on this platform.
    if   not scrDcBmap.IsOk() :   # Not implemented :  Get the screen bitmap the long way.
        if debug :
            print 'DEBUG:  Using memDC.Blit() since scrDC.GetAsBitmap() is nonfunctional.'
        # Create a new empty (black) destination bitmap the size of the scrDC. 
        # Overwrire the invalid original "scrDcBmap".
        scrDcBmap = wx.EmptyBitmap( *scrDcSize )
        scrDcBmapSizeX, scrDcBmapSizeY = scrDcSize
        # Create a DC tool that is associated with scrDcBmap.
        memDC = wx.MemoryDC( scrDcBmap )
        # Copy (blit, "Block Level Transfer") a portion of the screen bitmap 
        #   into the returned capture bitmap.
        # The bitmap associated with memDC (scrDcBmap) is the blit destination.
        memDC.Blit( 0, 0,                           # Copy to this start coordinate.
                    scrDcBmapSizeX, scrDcBmapSizeY, # Copy an area this size.
                    scrDC,                          # Copy from this DC's bitmap.
                    0, 0,                    )      # Copy from this start coordinate.

        memDC.SelectObject( wx.NullBitmap )     # Finish using this wx.MemoryDC.
                                                # Release scrDcBmap for other uses.        
    else :
        if debug :
            print 'DEBUG:  Using scrDC.GetAsBitmap()'
        # This platform has scrDC.GetAsBitmap() implemented.
        scrDcBmap = scrDC.GetAsBitmap()     # So easy !  Copy the entire Desktop bitmap.
        if debug :
            print 'DEBUG:  scrDcBmap.GetSize() ', scrDcBmap.GetSize()
    #end if
    return scrDcBmap.GetSubBitmap( wx.RectPS( captureStartPos, captureBmapSize ) )

def SaveScreenshot(filename,filetype):
    # Capture the entire primary Desktop screen.
    captureBmapSize = (wx.SystemSettings.GetMetric( wx.SYS_SCREEN_X ), 
                       wx.SystemSettings.GetMetric( wx.SYS_SCREEN_Y ) )
    print '>>>>  Screen Size ', captureBmapSize
    captureStartPos = (0, 0)    # Arbitrary U-L position anywhere within the screen
    bitmap = ScreenCapture( captureStartPos, captureBmapSize )
    #bitmap.SaveFile( filename, wx.BITMAP_TYPE_PNG )
    bitmap.SaveFile( filename, filetype )
    
def udisks_info(device):
    from subprocess import check_output as qx
    # get udisks output
    error=False
    try:
        out = qx(['udisks', '--show-info', device]).decode()
    except:
        #maybe not a block device
        error=True
    if error:
        return ""
    else:
        # strip header & footer
        out = out[out.index('\n')+1:]
        i = out.find('=====')
        if i != -1: out = out[:i]
        return out

def IsRemovable(path):
    #drivebits=win32file.GetLogicalDrives()
    #for d in range(1,26):
            #mask=1 << d
            #if drivebits & mask:
                    ## here if the drive is at least there
                    #drname='%c:\\' % chr(ord('A')+d)
                    #t=win32file.GetDriveType(drname)
                    #if t == win32file.DRIVE_REMOVABLE:

    #Or whatever DRIVE_* defined values there are in win32file
    #DRIVE_CDROM 
    #DRIVE_FIXED 
    #DRIVE_NO_ROOT_DIR 
    #DRIVE_RAMDISK 
    #DRIVE_REMOTE 
    #DRIVE_REMOVABLE 
    #DRIVE_UNKNOWN 
    isremovable=False
    if sys.platform[:3] == 'win':
        import win32file
        abspath=os.path.abspath(path)
        drname=abspath.split("\\")[0]+"\\"
        if win32file.GetDriveType(drname)==win32file.DRIVE_REMOVABLE: isremovable=True
    elif sys.platform == 'darwin':
        from plistlib import readPlistFromString as rPFS
        abspath=os.path.abspath(path)
        retry=True
        while retry:
            print "testing "+abspath
            listdrives=subprocess.Popen('df -P "'+abspath+'"', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            listdrivesout, err=listdrives.communicate()
            if err=="": retry=False
            else: abspath=abspath[:abspath.rfind(os.path.sep)]
        isremovable=True
        devicename=listdrivesout.split("\n")[-1].split(" ")[0]
        if devicename=="": devicename=listdrivesout.split("\n")[-2].split(" ")[0]
        ret=subprocess.Popen('diskutil list -plist', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        retout, err=ret.communicate()
        for disk in rPFS(retout)['AllDisks']:
            print disk,devicename
            if "/dev/"+disk==devicename:
                print "finding info on "+disk
                info=subprocess.Popen('diskutil info -plist '+disk, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                infoout, err=info.communicate()
                isremovable=not rPFS(infoout)['Internal']
                return isremovable
    else: #Linux
        abspath=os.path.abspath(path)
        retry=True
        while retry:
            print "testing "+abspath
            listdrives=subprocess.Popen('df -P "'+abspath+'"', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            listdrivesout, err=listdrives.communicate()
            if err=="": retry=False
            else: abspath=abspath[:abspath.rfind(os.path.sep)]
        isremovable=True
        devicename=listdrivesout.split("\n")[-1].split(" ")[0]
        if devicename=="": devicename=listdrivesout.split("\n")[-2].split(" ")[0]
        for elm in udisks_info(devicename).split("\n"):
            if "native-path" in elm.split(":")[0]:
                if "/usb" not in elm[len(elm.split(":")[0]):]:
                    isremovable=False
                    return isremovable
    return isremovable

def IsRunningFromRemovable():
    return IsRemovable(sys.argv[0])

def GetDeviceName(path):
    abspath=os.path.abspath(path)
    ret=""
    if sys.platform[:3] == 'win':
        import win32file
        ret=abspath.split(":")[0]
    else: #Darwin, Linux
        retry=True
        while retry:
            #print "testing "+abspath
            listdrives=subprocess.Popen('df -P "'+abspath+'"', shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            listdrivesout, err=listdrives.communicate()
            if err=="": retry=False
            else: abspath=abspath[:abspath.rfind(os.path.sep)]
        devicename=listdrivesout.split("\n")[-1].split(" ")[0]
        if devicename=="": devicename=listdrivesout.split("\n")[-2].split(" ")[0]
        ret=devicename
    print abspath,ret
    return ret
    
def IsSameDevice(path1,path2):
    samedevice=False
    if sys.platform != 'darwin':
        if GetDeviceName(path1)==GetDeviceName(path2):
            samedevice=True
    return samedevice
        
# Directory walker - Begin
def walker(arg, dirname, names):
#    print arg, dirname, names
    arg.append([dirname,names])

def getFileList(curpath):
    objs=[]
    os.path.walk(curpath, walker, objs)
    strList=[]
    for i1 in range(len(objs)):
        for i2 in range(len(objs[i1][1])):
            elm=objs[i1][0]+os.path.sep+objs[i1][1][i2]
            if not os.path.isdir(elm):
                strList.append(elm)
    strList.sort()
    return strList
# Usage:
# lst=getFileList(".")
# print lst.join('\n')
# Directory walker - End
filelist=getFileList(".")
for elm in filelist:
    basename=os.path.basename(elm)
    if basename.lower()[-4:]==".rtf" or basename.lower()[-4:]==".abw" or basename.lower()[-4:]==".xls" or basename.lower()[-4:]==".doc" or basename.lower()[-4:]==".pdf":
        if basename.lower()[-4:]==".rtf" or basename.lower()[-4:]==".abw":
            destination=REPORTS_PATH+basename
        elif basename.lower()[-4:]==".xls" or basename.lower()[-4:]==".doc" or basename.lower()[-4:]==".pdf":
            destination=DOC_PATH+basename
        try:
            if os.path.exists(destination): os.unlink(destination)
            shutil.move(elm,destination)
        except: pass

#import locale
#locale.setlocale(locale.LC_ALL, '')

# Resource handler class
class MaskedCtrlXmlHandler(wx.xrc.XmlResourceHandler):
    def __init__(self):
        wx.xrc.XmlResourceHandler.__init__(self)
        # Standard styles
        self.AddWindowStyles()
    def CanHandle(self,node):
        return self.IsOfClass(node, 'MaskedCtrl')
    # Process XML parameters and create the object
    def DoCreateResource(self):
        assert self.GetInstance() is None
        fldname=self.GetName()
        if fldname==-1: fldname=self.GetText('name')
        w = masked.Ctrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), autoformat=self.GetText('autoformat'), defaultEncoding='ascii', pos=self.GetPosition(), size=self.GetSize(), style=0, value='')
        self.SetupWindow(w)
        return w

datectrls={}
class DateCtrlXmlHandler(wx.xrc.XmlResourceHandler):
    def __init__(self):
        wx.xrc.XmlResourceHandler.__init__(self)
        # Standard styles
        self.AddWindowStyles()
    def CanHandle(self,node):
        return self.IsOfClass(node, 'wxTextCtrl')
    # Process XML parameters and create the object
    def DoCreateResource(self):
        assert self.GetInstance() is None
        fldname=self.GetName()
        if fldname==-1: fldname=self.GetText('name')
        maxlength=self.GetText('maxlength')
        altered=False
        if datectrls.has_key(fldname):
            str_style=self.GetText('style')
            isReadOnly=False
            if "READONLY" in str_style: isReadOnly=True
            readonlystyle=wx.TE_READONLY if isReadOnly else 0
            if datectrls[fldname]=='R': isReadOnly=True
            if datectrls[fldname]=='date':
                #if sys.platform != 'darwin':
                altered=True
                autoformat="EUDATEDDMMYYYY/"
                w = masked.Ctrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), autoformat=autoformat, defaultEncoding='ascii', pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, value='', invalidBackgroundColour='White', emptyBackgroundColour='White')
                #w = masked.Ctrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), mask="#{2}/#{2}/#{4}", defaultEncoding='ascii', pos=self.GetPosition(), size=self.GetSize(), style=0, value='', emptyBackgroundColour='White')
            elif datectrls[fldname]=='integer':
                if sys.platform != 'darwin':    #il numctl sembra non implementato nel mac
                    altered=True
                    #w = masked.numctrl.NumCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, value=None, decimalChar = '.', allowNegative = True, integerWidth = 10, fractionWidth = 0, allowNone = True, selectOnEntry = False, groupDigits = False, useParensForNegatives = False, signedForegroundColour='Black', invalidBackgroundColour='White')
                    w = masked.numctrl.NumCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, decimalChar = '.', allowNegative = True, integerWidth = 10, fractionWidth = 0, allowNone = True, selectOnEntry = False, groupDigits = False, useParensForNegatives = False, signedForegroundColour='Black', invalidBackgroundColour='White')
            elif datectrls[fldname]=='float':
                if sys.platform != 'darwin':    #il numctl sembra non implementato nel mac
                    altered=True
                    #w = masked.numctrl.NumCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, value=None, decimalChar = '.', allowNegative = True, integerWidth = 10, fractionWidth = 3, allowNone = True, selectOnEntry = False, groupDigits = False, useParensForNegatives = False, signedForegroundColour='Black', invalidBackgroundColour='White')
                    w = masked.numctrl.NumCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, decimalChar = '.', allowNegative = True, integerWidth = 10, fractionWidth = 3, allowNone = True, selectOnEntry = False, groupDigits = False, useParensForNegatives = False, signedForegroundColour='Black', invalidBackgroundColour='White')
            elif datectrls[fldname][:6]=='number':
                if sys.platform != 'darwin':    #il numctl sembra non implementato nel mac
                    altered=True
                    mask=datectrls[fldname][datectrls[fldname].find("=")+1:]
                    #intpart=int(mask[:mask.find(".")])
                    allowNegative=True
                    intpart=0
                    tmp=mask[:mask.find(".")]
                    if len(tmp)>0:
                        if tmp[0]=="+": 
                            allowNegative=False
                            intpart=int(tmp[1:])
                        elif tmp[0]=="-": 
                            intpart=int(tmp[1:])
                        else:
                            intpart=int(tmp)
                    decpart=int(mask[mask.find(".")+1:])
                    if DEBUG_MODE:
                        if intpart==0:
                            print "Error on Field: "+fldname+": "+mask
                    #w = masked.numctrl.NumCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, value=None, decimalChar = '.', allowNegative = allowNegative, integerWidth = intpart, fractionWidth = decpart, allowNone = True, selectOnEntry = False, groupDigits = False, useParensForNegatives = False, signedForegroundColour='Black', invalidBackgroundColour='White')
                    w = masked.numctrl.NumCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, decimalChar = '.', allowNegative = allowNegative, integerWidth = intpart, fractionWidth = decpart, allowNone = True, selectOnEntry = False, groupDigits = False, useParensForNegatives = False, signedForegroundColour='Black', invalidBackgroundColour='White')
            elif datectrls[fldname]=='R':
                altered=True
                if self.GetParentAsWindow().GetName()!="frameSearch":
                    w = wx.TextCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=wx.TE_READONLY, value='')
                else:
                    w = wx.TextCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=0, value='')
            elif str(datectrls[fldname]).split("=")[0]=='mask':
                altered=True
                w = masked.Ctrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), mask=datectrls[fldname].split("=")[1], formatcodes='_.-S', decimalChar='.', defaultEncoding='ascii', signedForegroundColour='Black', useParensForNegatives=False, pos=self.GetPosition(), size=self.GetSize(), style=readonlystyle, value='')
        if not altered:
            str_style=self.GetText('style')
            lng_style=0
            if "MULTILINE" in str_style: lng_style=lng_style+wx.TE_MULTILINE
            if "READONLY" in str_style: lng_style=lng_style+wx.TE_READONLY
            #if fldname=="cod_paz": print "page="+self.GetParentAsWindow().GetName()+" - cod_paz style="+str(lng_style)
            w = wx.TextCtrl(name=fldname, id=self.GetID(), parent=self.GetParentAsWindow(), pos=self.GetPosition(), size=self.GetSize(), style=lng_style, value='')
        self.SetupWindow(w)
        if not altered:
            try: 
                if int(maxlength)!=0: 
                    w.SetMaxLength(int(maxlength))
                    w.maxLength=int(maxlength)
            except: pass
        #w.style=style
        return w

class ProgressStatusBar:
    def __init__ (self, parent, statusbar, sbarfields=1, sbarfield=0, maxcount=100):

        rect = statusbar.GetFieldRect (sbarfield)
        barposn = (rect [0], rect [1])

        # On MSW the X dimension returned from GetFieldRect for the last field is too small.
        #   This hack fills the field but covers over the lower right frame resize handle,
        #    but that's OK since the field size should be unchangable.
        if (sbarfield+1 == sbarfields) and (wx.Platform == '__WXMSW__'):
            barsize = (rect [2]+35, rect [3])   # completely fill the last field
        else:
            barsize = (rect [2],    rect [3])
        #end if

        self.progbar = wx.Gauge (statusbar, -1, maxcount, barposn, barsize)
        self.value=0
    #end def

    def SetValue (self, value):
        #if DEBUG_MODE:
            #self.progbar.SetValue (value)
        #else:
        if value!=self.value:
            self.value=value
            self.progbar.Pulse()
    #end def

#end class ProgressStatusBar

#Custom Settings facilities - Begin
def LoadCustomSetting(key):
    ret=""
    #cur.execute("SELECT setting_value FROM settings WHERE setting_key='"+key+"'")
    
    sqlValue=""
    for i in range(len(key)):
        if key[i]=="'":
            sqlValue+="''"
        else:
            sqlValue+=key[i]
            
    cur.execute("SELECT setting_value FROM settings WHERE setting_key like '"+sqlValue+"'")
    row = cur.fetchone()
    if row!=None:
        ret=row[0]
    return ret

def SaveCustomSetting(key,value):
    cur.execute("SELECT setting_value FROM settings WHERE setting_key='"+key+"'")
    print "Saving setting: "+key+" = "+value
    if cur.fetchone()==None:
        cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('"+key+"','"+GetSqlValue(value)+"')")
    else:
        cur.execute("UPDATE settings SET setting_value='"+GetSqlValue(value)+"' WHERE setting_key='"+key+"'")

def GetSetting(parentForm,key):
    ret=""
    if parentForm.arrSettings.has_key(key): ret=parentForm.arrSettings[key]
    return ret
    
def TT(key):
    interface_profile=LoadCustomSetting("interface_profile")
    TTDICT={}
    if "trial_italiano" in interface_profile:
        TTDICT["Are you sure you wish to clone this record (all the pages) and open the clone?"]="Si vuole creare uno schedario per nuovo Paziente copiato da questo ed aprirlo?"
        TTDICT["Record Clone"]="Duplica Paziente"
        TTDICT["Are you sure you wish to save this record (all the pages)?"]="Si vogliono salvare i dati di questo Paziente (tutte le pagine)?"
        TTDICT["Save record"]="Salvataggio dati Paziente"
        TTDICT["Are you sure you wish to save this record (all the pages) and exit?"]="Si vogliono salvare i dati di questo Paziente (tutte le pagine) e tornare alla pagina di Ricerca?"
        TTDICT["Save record and exit"]="Salvataggio dati Paziente e uscita"
        TTDICT["Are you sure you wish to quit without saving this record (all the pages)?"]="Si vuole chiudere lo schedario di questo Paziente (tutte le pagine) senza salvare eventuali modifiche appena fatte?"
        TTDICT["Quit record"]="Uscita senza salvare"
        TTDICT["Are you sure you wish to delete this record (all the pages)?"]="Si vogliono eliminare i dati di questo Paziente (tutte le pagine)?"
        TTDICT["Delete record"]="Eliminazione dati Paziente"
        TTDICT["Do you wish to list all of the records?"]="Si desidera elencare tutti i Pazienti?"
        
    if "italiano" in interface_profile:
        TTDICT["Authorization failed"]="Autorizzazione negata"
        TTDICT["Delete denied"]="Eliminazione non possibile"
        TTDICT["Heavybase has been updated and needs to be restarted. Can it be closed now?"]="Heavybase e' stato aggiornato e deve essere riavviato. E' possibile chiuderlo adesso?"
        TTDICT["System updated"]="Sistema aggiornato"
        TTDICT["Update in progress, this application should not be closed now. Do you wish to close it anyway?"]="Aggiornamento in corso, l'applicazione non dovrebbe essere chiusa adesso. Si vuole chiuderla comunque?"
        TTDICT["Confirm Exit"]="Conferma chiusura"
        TTDICT["Do you really want to close this application?"]="Si vuole davvero chiudere questa applicazione?"
        TTDICT["Cleaning in progress"]="Chiusura in corso"
        TTDICT["Do you want to close any other HeavyBase process?"]="Si desidera chiudere ogni altro HeavyBase aperto?"
        TTDICT["Do you want to close the application window?"]="Vuoi chiudere l'interfaccia utente?"
        TTDICT["Search"]="Ricerca"
        TTDICT["File extract"]="Estrazione file"
        TTDICT["Do you want to open the extracted file?"]="Si vuole aprire il file estratto?"
        TTDICT["File open cancelled"]="Apertura file annullata"
        TTDICT["not available"]="non disponibile"
        TTDICT["File extraction cancelled"]="Estrazione file annullata"
        TTDICT["File delete cancelled"]="Eliminazione file annullata"
        TTDICT["Cancel"]="Annulla"
        TTDICT["New password"]="Nuova password"
        TTDICT["PASSWORD EXPIRED"]="PASSWORD SCADUTA"
        TTDICT["Repeat new password"]="Ripeti la nuova password"
        TTDICT["User"]="Utente"
        TTDICT["Change user password"]="Cambio password utente"
        TTDICT["Change user password tool must be invoked from Search page"]="La funzione di cambio password utente va richiamata dalla pagina di Ricerca"
        TTDICT["The password cannot be empty"]="La password non puo' essere vuota"
        TTDICT["The two passwords are different"]="Le due password sono diverse"
        TTDICT["The password must be at least 8 characters long"]="La password deve essere lunga almeno 8 caratteri"
        TTDICT["Module disabled"]="Modulo disabilitato"
        TTDICT["Change cypher key"]="Cambio di chiave di cifratura"
        TTDICT["Change cypher key tool must be invoked from Search page"]="La funzione di cambio chiave di cifratura va richiamata dalla pagina di Ricerca"
        TTDICT["Export format"]="Formato di esportazione"
        TTDICT["Authentication failed"]="Autenticazione fallita"
        TTDICT["Bad Username, Password or Cypher key"]="Errore nel campo username o password o chiave di cifratura"
        TTDICT["Initializing"]="Inizializzazione"
        TTDICT["Please wait"]="Attendere prego"
        TTDICT["Loading HEAVyBASE"]="Caricamento HEAVyBASE"
        TTDICT["Code"]="Codice"
        TTDICT["Description"]="Descrizione"
        TTDICT["Activity"]="Attivita'"
        TTDICT["Exit"]="Esci"
        TTDICT["Network settings"]="Configurazione di rete"
        TTDICT["New Data management"]="Nuova Gestione dati"
        TTDICT["Data management"]="Gestione dati"
        TTDICT["Data entry mode"]="Modalita' data entry"
        TTDICT["Data extraction mode"]="Modalita' estrazione dati"
        TTDICT["Data Analysis"]="Analisi dati"
        TTDICT["Audit Trail"]="Audit trail"
        TTDICT["Modify Access Rights"]="Modifica diritti di accesso"
        TTDICT["Network settings"]="Configurazione di rete"
        TTDICT["Username"]="Nome utente"
        TTDICT["Password"]="Password"
        TTDICT["Cypher key"]="Chiave di cifratura"
        TTDICT["Data analysis tool must be invoked from Search page"]="L'Analisi dati va richiamata dalla pagina di Ricerca"
        TTDICT["Already in Data extraction mode"]="La modalita' estrazione dati e' gia' attiva"
        TTDICT["Already in Data entry mode"]="La modalita' data-entry e' gia' attiva"
        TTDICT["Data entry tool must be invoked from Search page"]="La modalita' data-entry va richiamata dalla pagina di Ricerca"
        TTDICT["Data extraction tool must be invoked from Search page"]="La modalita' estrazione dati va richiamata dalla pagina di Ricerca"
        TTDICT["can read"]="possono aprire"
        TTDICT["can modify"]="possono modificare"
        TTDICT["can delete"]="possono eliminare"
        TTDICT["Current Record"]="Record attuale"
        TTDICT["Current Variable"]="Variabile attuale"
        TTDICT["Metadata"]="Metadati"
        TTDICT["Load options"]="Carica opzioni"
        TTDICT["Not an user input field"]="Non e' un campo di input per l'utente"
        TTDICT["Import external dataset"]="Importa dataset esterno"
        TTDICT["Export DB"]="Esporta DB"
        TTDICT["OTP Credentials Keyring"]="Portachiavi Credenziali OTP"
        TTDICT["File saved successfully"]="File salvato correttamente"
        TTDICT["must be invoked from Login or Search page"]="va richiamata dalla pagina di Login o di Ricerca"
        TTDICT["Enter your PIN (8 chars minimum)"]="Inserisci il tuo PIN (minimo 8 caratteri)"
        TTDICT["Keyring PIN"]="PIN Portachiavi"
        TTDICT["The file must be on a removable device (eg. a USB pendrive)"]="Il file deve essere su un dispositivo rimovibile (es. un pendrive USB)"
        TTDICT["Activate"]="Attiva"
        TTDICT["Hide"]="Nascondi"
        TTDICT["Help Desk: Send Screenshot"]="Help Desk: Invia Schermata"
        TTDICT["Timeout. Please retry later"]="Tempo Scaduto. Riprovare piu' tardi"
        TTDICT["Offline. Please retry later"]="Non connesso. Riprovare piu' tardi"
        TTDICT["Wait"]="Attendere"
        TTDICT["Screenshot sent successfully"]="Schermata inviata correttamente"
        TTDICT["Can the application be moved on the Desktop?"]="Posso spostare l'applicazione sul Desktop?"
        TTDICT["Application relocation"]="Spostamento applicazione"
        TTDICT["The application cannot be moved.\nPlease don't click 'EXECUTE' directly from the Browser, click on 'SAVE' instead."]="L'applicazione non puo' essere spostata.\nDon cliccare 'ESEGUI' direttamente dal Browser, utilizzare invece il pulsante 'SALVA'"
        TTDICT["No Records found"]="Non ci sono schede compilate"
        TTDICT["No validation defined"]="Nessuna validazione definita"
        TTDICT["created"]="creato"
        TTDICT["Confidence"]="Confidenza"
        TTDICT["Free text memo / Reason for change"]="Note a testo libero / Motivo cambio dato"
        TTDICT["Valid data"]="Dato valido"
        TTDICT["To be confirmed"]="Da confermare"
        TTDICT["Not available, type 1"]="Non reperibile, tipo 1"
        TTDICT["Not available, type 2"]="Non reperibile, tipo 2"
        TTDICT["All the items must be checked"]="Tutte le voci devono essere cliccate per accettazione"
        TTDICT["Too many columns for a '.xls' file"]="Troppe colonne per un file '.xls'"
        TTDICT["Print preview preparation in progress"]="Preparazione anteprima di stampa in corso"
        TTDICT["Choose a file name"]="Scegliere un nome file"
        TTDICT["Print eCRF"]="Stampa schede"
        TTDICT["Print canceled"]="Stampa annullata"
        TTDICT["alphabetic, non empty variables only"]="alfab., solo variabili non vuote"
        TTDICT["by position, non empty variables only"]="per posiz., solo variabili non vuote"
        TTDICT["alphabetic, all representative vars."]="alfab., tutte le var. rappresentative"
        TTDICT["by position, all representative vars."]="per posiz., tutte le var. rappresentative"
        TTDICT["alphabetic, all variables"]="alfab., tutte le variabili"
        TTDICT["by position, all variables"]="per posiz., tutte le variabili"
        TTDICT["AVAILABLE FIELDS"]="CAMPI DISPONIBILI"
        TTDICT["DATA EXTRACTION"]="ESTRAZIONE DATI"
        TTDICT["QUERY BUILDER"]="COSTRUTTORE DI QUERY"
        TTDICT["field"]="campo"
        TTDICT["condition"]="condizione"
        TTDICT["type"]="tipo"
        TTDICT["value or field"]="valore o campo"
        TTDICT["FREE TEXT"]="TESTO LIBERO"
        TTDICT["mark all"]="marca tutti"
        TTDICT["unmark all"]="smarca tutti"
        TTDICT["use DISTINCT clause"]="usa clausola DISTINCT"
        TTDICT["Run Query"]="Avvia Estrazione"
        TTDICT["Exporter"]="Esportatore"
        TTDICT["no fields selected."]="non sono stati selezionati campi."
        TTDICT["too many fields selected (max. 2000)."]="troppi campi selezionati (max. 2000)."
        #Controlli formali cablati
        TTDICT["possible duplicate"]="possibile duplicato"
        TTDICT["incorrect date"]="data non corretta"
        TTDICT["File not writable"]="File non scrivibile"
        TTDICT["Fields and labels list"]="Lista di campi ed etichette"
        TTDICT["Data structure document"]="Documento di struttura dati"
        TTDICT["Database Administration"]="Amministrazione Database"
        
    key2=key
    if key in TTDICT: key2=TTDICT[key]
    ret=LoadCustomSetting("{"+key+"}")
    if ret=="": ret=key2
    
    import re
    pattern = re.compile("heavy"+"base", re.IGNORECASE)
    ret=pattern.sub(PROGNAME, ret)
    return ret

#Custom Settings facilities - End

def GetSystemProxy():
    proxyEnable = 0
    proxyIp = ""

    import os
    if os.name=='nt':
        import win32api
        import win32con
        import exceptions
        chiave = win32api.RegOpenKey(win32con.HKEY_CURRENT_USER , "Software")
        chiave = win32api.RegOpenKey(chiave , "Microsoft")
        chiave = win32api.RegOpenKey(chiave , "Windows")
        chiave = win32api.RegOpenKey(chiave , "CurrentVersion")
        chiave = win32api.RegOpenKey(chiave , "Internet Settings")
        # search wpad
        i = 0
        wpad_found = False
        wpad_url = ""
        try:
            while(wpad_found == False):
                setting1 = win32api.RegEnumValue(chiave , i)
                valore = win32api.RegQueryValueEx(chiave, setting1[0])
                if(setting1[0] == "AutoConfigURL"):
                    wpad_found = True
                    wpad_url = str(valore[0])                    
                i = i+1
        except Exception:
            pass
        if wpad_found:
            try:
                # wpad based proxy configuration
                wpad_filename="~wpad.dat"
                try: os.unlink(wpad_filename)
                except: pass

                import urllib
                opener = urllib.FancyURLopener({})
                f = opener.open(wpad_url)
                wpad_contents = f.read()
                fout = open(wpad_filename, "w")
                fout.write(wpad_contents)
                fout.close()
                ##sysinfo
                #sysinfofile = open('sysinfo.txt','a')
                #sysinfofile.write("--- WPAD begin ---")
                #sysinfofile.write(wpad_contents)
                #sysinfofile.write("--- WPAD end ---")
                # pacparser
                import pacparser
                pacparser.init()
                pacparser.parse_pac(wpad_filename)
                serverurl = 'xmlrpc.heavybase.org'
                proxylist = pacparser.find_proxy('http://'+serverurl, serverurl)
                firstproxy = proxylist.split(";")[0].strip()
                firstproxymode = firstproxy.split(" ")[0]
                if firstproxymode == "DIRECT": 
                    # no proxy
                    pass
                elif firstproxymode == "PROXY":
                    proxyEnable = 1
                    proxyIp = "http://"+firstproxy.split(" ")[1]
                #print proxyIp
            except:
                print "wpad_mode failed. switching to normal proxy conf"
                wpad_found=False
        if not wpad_found:
            # normal proxy configuration
            flag = 0
            i = 0
            try:
                while(flag != 2):
                    setting1 = win32api.RegEnumValue(chiave , i)
                    valore = win32api.RegQueryValueEx(chiave, setting1[0])
                    if(setting1[0] == "ProxyEnable"):
                        flag = flag+1
                        proxyEnable = valore[0]
                    if(setting1[0] == "ProxyServer"):
                        flag = flag+1
                        proxyIp = str(valore[0])
                        proxyIp = "http://"+proxyIp
                        #print "Tipo chiave proxy : "+str(valore[1])
                    i = i+1
            except Exception:
                pass
        ##sysinfo
        #sysinfofile = open('sysinfo.txt','a')
        #sysinfofile.write("--- PROXY SETTING begin ---")
        #sysinfofile.write("proxyEnable="+str(proxyEnable))
        #sysinfofile.write("proxyIp="+proxyIp)
        #sysinfofile.write("--- PROXY SETTING end ---")
    elif os.name=='posix':
        try:
            proxyIp=os.environ['http_proxy']
            if len(proxyIp)>7:
                if proxyIp[0:7]=="http://":
                    proxyEnable = 1
        except:
            pass
    return proxyEnable,proxyIp

def SetNTLMAPSconfig():
    filename="ntlmaps.cfg"
    fin = open(filename, 'r')
    data = fin.read()
    fin.close()
    
    proxy_ip=""
    proxy_port=""
    Network_Proxy_Setting=LoadCustomSetting("Network_Proxy_Setting")
    if Network_Proxy_Setting=="" or Network_Proxy_Setting=="automatic":
        autoProxySetting=GetSystemProxy()
        if autoProxySetting[0]==1:
            proxy=autoProxySetting[1]
            proxy_ip=proxy.split(":")[-2]
            proxy_port=proxy.split(":")[-1]
    elif Network_Proxy_Setting!="no proxy":
        proxy_ip=LoadCustomSetting("Network_Proxy_Host")
        proxy_port=LoadCustomSetting("Network_Proxy_Port")
    if "//" in proxy_ip: proxy_ip=proxy_ip.split("//")[-1]

    data=data.replace("LISTEN_PORT:5865", "LISTEN_PORT:60001")
    if proxy_ip!="" and proxy_port!="":
        data=data.replace("PARENT_PROXY:your_parentproxy", "PARENT_PROXY:"+proxy_ip)
        data=data.replace("PARENT_PROXY_PORT:8080", "PARENT_PROXY_PORT:"+proxy_port)

        username=LoadCustomSetting("Network_Proxy_Username")
        password=LoadCustomSetting("Network_Proxy_Password")
        if "\\" in username:
            domain=username.split("\\")[0]
            username=username.split("\\")[1]
            data=data.replace("NT_DOMAIN:your_domain", "NT_DOMAIN:"+domain)
        if username!="": data=data.replace("USER:username_to_use", "USER:"+username)
        if password!="": data=data.replace("PASSWORD:your_nt_password", "PASSWORD:"+password)
        data=data.replace("NT_PART:0","NT_PART:1")
        data=data.replace("NTLM_FLAGS: 06820000", "NTLM_FLAGS: 07820000")
        
    fout = open(filename, "w")
    fout.write(data)
    fout.close()

class RtfBuilder:
    def __init__(self, filename=""):
        self.NuovoTestoBold=False
        self.NuovoTestoItalic=False
        self.NuovoTestoSize=12
        self.NuovoTestoName="Arial"
        self.NuovoTestoAlign="Justify"  #Alternativa: "Center"
    
    def IntestazioneRTF(self,Int1pSx, Int1pMx, Int1pDx, Foo1pSx, Foo1pMx, Foo1pDx, IntSx, IntMx, IntDx, FooSx, FooMx):
        buf=[]
        buf.append("{\\rtf1\\ansi\\ansicpg1252\\uc1\\deff0\\stshfdbch0\\stshfloch0\\stshfhich0\\stshfbi0\\deflang1040\\deflangfe1040{\\fonttbl{\\f0\\froman\\fcharset0\\fprq2{\\*\\panose 02020603050405020304}Arial;}{\\f1\\fswiss\\fcharset0\\fprq2{\\*\\panose 020b0604020202020204}Arial;}")
        buf.append("{\\f65\\froman\\fcharset238\\fprq2 Arial CE;}{\\f66\\froman\\fcharset204\\fprq2 Arial Cyr;}{\\f68\\froman\\fcharset161\\fprq2 Arial Greek;}{\\f69\\froman\\fcharset162\\fprq2 Arial Tur;}")
        buf.append("{\\f70\\froman\\fcharset177\\fprq2 Arial (Hebrew);}{\\f71\\froman\\fcharset178\\fprq2 Arial (Arabic);}{\\f72\\froman\\fcharset186\\fprq2 Arial Baltic;}{\\f75\\fswiss\\fcharset238\\fprq2 Arial CE;}{\\f76\\fswiss\\fcharset204\\fprq2 Arial Cyr;}")
        buf.append("{\\f78\\fswiss\\fcharset161\\fprq2 Arial Greek;}{\\f79\\fswiss\\fcharset162\\fprq2 Arial Tur;}{\\f80\\fswiss\\fcharset177\\fprq2 Arial (Hebrew);}{\\f81\\fswiss\\fcharset178\\fprq2 Arial (Arabic);}{\\f82\\fswiss\\fcharset186\\fprq2 Arial Baltic;}}")
        buf.append("{\\colortbl;\\red0\\green0\\blue0;\\red220\\green220\\blue255;\\red0\\green255\\blue255;\\red0\\green255\\blue0;\\red255\\green0\\blue255;\\red255\\green0\\blue0;\\red255\\green255\\blue0;\\red255\\green255\\blue255;\\red0\\green0\\blue128;\\red0\\green128\\blue128;\\red0\\green128\\blue0;")
        buf.append("\\red128\\green0\\blue128;\\red128\\green0\\blue0;\\red128\\green128\\blue0;\\red128\\green128\\blue128;\\red192\\green192\\blue192;}{\\stylesheet{\\ql \\li0\\ri0\\widctlpar\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0")
        buf.append("\\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 \\snext0 Normal;}{\\*\\cs10 \\additive \\ssemihidden Default Paragraph Font;}{\\*")
        buf.append("\\ts11\\tsrowd\\trftsWidthB3\\trpaddl108\\trpaddr108\\trpaddfl3\\trpaddft3\\trpaddfb3\\trpaddfr3\\trcbpat1\\trcfpat1\\tscellwidthfts0\\tsvertalt\\tsbrdrt\\tsbrdrl\\tsbrdrb\\tsbrdrr\\tsbrdrdgl\\tsbrdrdgr\\tsbrdrh\\tsbrdrv")
        buf.append("\\ql \\li0\\ri0\\widctlpar\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs20\\lang1024\\langfe1024\\cgrid\\langnp1024\\langfenp1024 \\snext11 \\ssemihidden Normal Table;}{\\s15\\ql \\li0\\ri0\\widctlpar")
        buf.append("\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 \\sbasedon0 \\snext15 \\styrsid7822823 header;}{\\s16\\ql \\li0\\ri0\\widctlpar")
        buf.append("\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 \\sbasedon0 \\snext16 \\styrsid7822823 footer;}{\\*\\cs17 \\additive \\sbasedon10 \\styrsid7822823 page number;}}")
        buf.append("{\\*\\rsidtbl \\rsid7822823\\rsid13713086}{\\*\\generator Microsoft Word 10.0.5815;}{\\info{\\author luca Clivio}{\\operator luca Clivio}{\\creatim\\yr2004\\mo6\\dy16\\hr6\\min3}{\\revtim\\yr2004\\mo6\\dy16\\hr6\\min3}{\\version2}{\\edmins0}")
        buf.append("{\\nofpages1}{\\nofwords0}{\\nofchars5}{\\nofcharsws5}{\\vern16501}}\\paperw11906\\paperh16838\\margl1134\\margr1134\\margt1418\\margb1134")
        buf.append("\\deftab708\\widowctrl\\ftnbj\\aenddoc\\hyphhotz283\\noxlattoyen\\expshrtn\\noultrlspc\\dntblnsbdb\\nospaceforul\\hyphcaps0\\formshade\\horzdoc\\dgmargin\\dghspace180\\dgvspace180\\dghorigin1134\\dgvorigin1418\\dghshow1\\dgvshow1")
        buf.append("\\jexpand\\viewkind1\\viewscale100\\pgbrdrhead\\pgbrdrfoot\\splytwnine\\ftnlytwnine\\htmautsp\\nolnhtadjtbl\\useltbaln\\alntblind\\lytcalctblwd\\lyttblrtgr\\lnbrkrule\\nobrkwrptbl\\snaptogridincell\\allowfieldendsel\\wrppunct\\asianbrkrule\\rsidroot7822823 \\fet0\\sectd")
        buf.append("\\psz9\\linex0\\headery709\\footery709\\colsx708\\endnhere\\titlepg\\sectlinegrid360\\sectdefaultcl\\sectrsid7822823\\sftnbj {\\header \\pard\\plain \\s15\\ql \\li0\\ri0\\widctlpar\\brdrb\\brdrs\\brdrw15\\brsp20")
        buf.append("\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 {\\f1\\fs20\\insrsid7822823\\charrsid13713086 " + self.CreaTestoRtf(IntSx) + "\\tab " + self.CreaTestoRtf(IntMx) + "\\tab " + self.CreaTestoRtf(IntDx))
        buf.append("\\par }\\pard \\s15\\ql \\li0\\ri0\\widctlpar\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 {\\f1\\fs20\\insrsid7822823\\charrsid13713086")
        buf.append("\\par }}{\\footer \\pard\\plain \\s16\\ql \\li0\\ri0\\widctlpar\\brdrb\\brdrs\\brdrw15\\brsp20 \\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 {\\f1\\fs20\\insrsid7822823\\charrsid13713086")
        buf.append("")
        buf.append("\\par }\\pard \\s16\\ql \\li0\\ri0\\widctlpar\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 {\\f1\\fs20\\insrsid7822823\\charrsid13713086 " + self.CreaTestoRtf(FooSx) + "\\tab " + self.CreaTestoRtf(FooMx) + "\\tab Pag. }{\\field{\\*\\fldinst {")
        buf.append("\\cs17\\f1\\fs20\\insrsid7822823\\charrsid13713086  PAGE }}{\\fldrslt {\\cs17\\f1\\fs20\\lang1024\\langfe1024\\noproof 2}}}{\\cs17\\f1\\fs20\\insrsid7822823\\charrsid13713086 /}{\\field{\\*\\fldinst {\\cs17\\f1\\fs20\\insrsid7822823\\charrsid13713086  NUMPAGES }")
        buf.append("}{\\fldrslt {\\cs17\\f1\\fs20\\lang1024\\langfe1024\\noproof 2}}}{\\f1\\fs20\\insrsid7822823\\charrsid13713086")
        buf.append("\\par }}{\\headerf \\pard\\plain \\s15\\ql \\li0\\ri0\\widctlpar\\brdrb\\brdrs\\brdrw15\\brsp20 \\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 {\\f1\\fs20\\insrsid7822823\\charrsid13713086")
        buf.append(self.CreaTestoRtf(Int1pSx) + "\\tab " + self.CreaTestoRtf(Int1pMx) + "\\tab " + self.CreaTestoRtf(Int1pDx))
        buf.append("\\par }\\pard \\s15\\ql \\li0\\ri0\\widctlpar\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 {\\f1\\fs20\\insrsid7822823\\charrsid13713086")
        buf.append("\\par }}{\\footerf \\pard\\plain \\s16\\ql \\li0\\ri0\\widctlpar\\brdrb\\brdrs\\brdrw15\\brsp20 \\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 \\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 {\\f1\\fs20\\insrsid7822823\\charrsid13713086")
        buf.append("")
        buf.append("\\par }\\pard \\s16\\ql \\li0\\ri0\\widctlpar\\tqc\\tx4819\\tqr\\tx9638\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0 {\\f1\\fs20\\insrsid7822823\\charrsid13713086 " + self.CreaTestoRtf(Foo1pSx) + "\\tab " + self.CreaTestoRtf(Foo1pMx) + "\\tab " + self.CreaTestoRtf(Foo1pDx))
        buf.append("\\par }}{\\*\\pnseclvl1\\pnucrm\\pnstart1\\pnindent720\\pnhang {\\pntxta .}}{\\*\\pnseclvl2\\pnucltr\\pnstart1\\pnindent720\\pnhang {\\pntxta .}}{\\*\\pnseclvl3\\pndec\\pnstart1\\pnindent720\\pnhang {\\pntxta .}}{\\*\\pnseclvl4\\pnlcltr\\pnstart1\\pnindent720\\pnhang {\\pntxta )}}")
        buf.append("{\\*\\pnseclvl5\\pndec\\pnstart1\\pnindent720\\pnhang {\\pntxtb (}{\\pntxta )}}{\\*\\pnseclvl6\\pnlcltr\\pnstart1\\pnindent720\\pnhang {\\pntxtb (}{\\pntxta )}}{\\*\\pnseclvl7\\pnlcrm\\pnstart1\\pnindent720\\pnhang {\\pntxtb (}{\\pntxta )}}{\\*\\pnseclvl8")
        buf.append("\\pnlcltr\\pnstart1\\pnindent720\\pnhang {\\pntxtb (}{\\pntxta )}}{\\*\\pnseclvl9\\pnlcrm\\pnstart1\\pnindent720\\pnhang {\\pntxtb (}{\\pntxta )}}\\pard\\plain \\qj \\li0\\ri0\\widctlpar\\aspalpha\\aspnum\\faauto\\adjustright\\rin0\\lin0\\itap0\\pararsid7822823")
        buf.append("\\fs24\\lang1040\\langfe1040\\cgrid\\langnp1040\\langfenp1040 {\\f1\\fs20")
        buf.append("\\par }{\\f1\\fs20\\charrsid7822823")
        return "\n".join(buf)
    
    def CreaTestoRtf(self,testo):
        if testo!="":
            testo=testo.replace(u"\\"  ,"")
            testo=testo.replace(u"\r\n","\\par ")
            testo=testo.replace(u"\r"  ,"\\par ")
            testo=testo.replace(u"\n"  ,"\\par ")
            testo=testo.replace(u"\t"  ,"\\par ")
            testo=testo.replace(u"`"   ,"\\'92")
            testo=testo.replace(u"\'"  ,"\\'92")
            testo=testo.replace(u'"'   ,"\\'94")
            testo=testo.replace(u"à"   ,"\\'e0")
            testo=testo.replace(u"è"   ,"\\'e8")
            testo=testo.replace(u"é"   ,"\\'e9")
            testo=testo.replace(u"ì"   ,"\\'ec")
            testo=testo.replace(u"ò"   ,"\\'f2")
            testo=testo.replace(u"ù"   ,"\\'f9")
        return testo
    
    def NuovoTesto(self,Testo, Bold, Italic, Size, Name, Align):
        NuovoTestoBold = Bold
        NuovoTestoItalic = Italic
        NuovoTestoSize = Size
        NuovoTestoName = Name
        NuovoTestoAlign = Align
        
        StrAlign=""
        if NuovoTestoAlign=="Justify": StrAlign = "j"
        elif NuovoTestoAlign=="Center": StrAlign = "c"
        
        StrBold=""
        if NuovoTestoBold: StrBold = "\\b"
        
        StrItalic=""
        if NuovoTestoItalic: StrItalic = "\\i"
        
        txt="{\\f1\\q" + StrAlign + "\\fs" + str(Size * 2) + StrBold + StrItalic + " " + self.CreaTestoRtf(Testo) + "}" + "{\\f1\\fs" + str(Size * 2) + "\\par }"
        return txt
    
    def ContinuaTesto(self,Testo):
        Bold = self.NuovoTestoBold
        Italic = self.NuovoTestoItalic
        Size = self.NuovoTestoSize
        Name = self.NuovoTestoName
        Align = self.NuovoTestoAlign
    
        StrAlign=""
        if NuovoTestoAlign=="Justify": StrAlign = "j"
        elif NuovoTestoAlign=="Center": StrAlign = "c"
    
        StrBold=""
        if NuovoTestoBold: StrBold = "\\b"
        
        StrItalic=""
        if NuovoTestoItalic: StrItalic = "\\i"
    
        txt="\\pard\\plain \\s1\\cf0\\rtlch\\af1\\lang255\\ltrch\\dbch\\af3\\loch\\lang255\\q" + StrAlign + "\\rtlch\\afs32\\ltrch\\loch\\fs" + str(Size * 2) + StrBold + StrItalic + " {\\ltrch\\loch\\f1 " + self.CreaTestoRtf(Testo) + "}"
        return txt
    
    def NuovoBreak(self,Tipo):
        StrBreak=""
        if Tipo==0: StrBreak = "{\\f1\\fs20 \\par\\page }"
        return StrBreak

    def NuovaTbl_Inizio(self,NumColonne):
        buf=[]
        buf.append("\\par \\trowd\\trql\\trhdr")
        Largh = 9636
        for i in range(NumColonne-1):
            buf.append("\\clbrdrt\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrl\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrb\\brdrs\\brdrw1\\brdrcf1\\brsp0\\cellx" + str(Largh / NumColonne * (i+1)))
        buf.append("\\clbrdrt\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrl\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrb\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrr\\brdrs\\brdrw1\\brdrcf1\\brsp0\\cellx" + str(Largh))
        return "\n".join(buf)
        
    def NuovaTbl_Riga(self,NumColonne):
        buf=[]
        buf.append("\\cell\\row\\pard \\trowd\\trql")
        Largh = 9636
        for i in range(NumColonne-1):
            buf.append("\\clbrdrl\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrb\\brdrs\\brdrw1\\brdrcf1\\brsp0\\cellx" + str(Largh / NumColonne * (i+1)))
        buf.append("\\clbrdrl\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrb\\brdrs\\brdrw1\\brdrcf1\\brsp0\\clbrdrr\\brdrs\\brdrw1\\brdrcf1\\brsp0\\cellx" + str(Largh))
        return "\n".join(buf)
    
    def NuovaTbl_Colonna(self,Prima, Intestazione, Size, Testo):
        buf=[]
        if Prima:
            buf.append("\\pard\\intbl\\pard\\plain ")
        else:
            buf.append("\\cell\\pard\\plain ")
        
        if Testo!="":
            while Testo[:2]=="\r\n":
                Testo = Testo[2:]
            while Testo[0]=="\n":
                Testo = Testo[1:]
        
        if Intestazione:
            buf.append("\\intbl\\s4\\li113\\ri0\\lin113\\rin0\\fi0\\sb28\\sa28\\cf0\\rtlch\\af1\\lang1040\\ltrch\\dbch\\af3\\langfe255\\qc\\ltrch\\loch\\i\\b\\ltrch\\loch\\fs" + str(Size * 2) + " {\\ltrch\\loch\\f1 " + self.CreaTestoRtf(Testo) + "}")
        else:
            buf.append("\\intbl\\s3\\cf0\\tqc\\tx4818\\tqr\\tx9637\\rtlch\\af1\\lang1040\\ltrch\\dbch\\af3\\langfe255\\li113\\ri0\\lin113\\rin0\\fi0\\sb28\\sa28\\ltrch\\loch\\fs" + str(Size * 2) + " {\\ltrch\\loch\\f1\\b0\\i0 " + self.CreaTestoRtf(Testo) + "}")
        return "\n".join(buf)
    
    def NuovaTbl_Fine(self):
        txt="\\cell\\row\\pard"
        return txt

    def NuovoGrafico(self,Desc, Inizio, Fine, Base):
        for i1 in range(len(Desc) - 1):
            for i2 in range(i1+1,len(Desc)):
                if (Inizio[i1] > Inizio[i2]) or (Inizio[i1] == Inizio[i2] and Fine[i1] > Fine[i2]) or (Inizio[i1] == Inizio[i2] and Fine[i1] == Fine[i2] and Desc[i1] > Desc[i2]):
                    sDesc = Desc[i1]
                    iInizio = Inizio[i1]
                    iFine = Fine[i1]
                    Desc[i1] = Desc[i2]
                    Inizio[i1] = Inizio[i2]
                    Fine[i1] = Fine[i2]
                    Desc[i2] = sDesc
                    Inizio[i2] = iInizio
                    Fine[i2] = iFine
        aDesc=[]
        for elm in Desc: aDesc.append("")
        aInizio=[]
        for elm in Inizio: aInizio.append(0)
        aFine=[]
        for elm in Find: aFine.append(0)
        c = 0
        for i in range(len(Desc)):
            if ((Inizio[i] - Base) < 1 and (Fine[i] - Base) >= 1) or ((Inizio[i] - Base) >= 1 and (Inizio[i] - Base) <= 30) or ((Fine[i] - Base) >= 1 and (Fine[i] - Base) <= 30) or ((Inizio[i] - Base) <= 30 and (Fine[i] - Base) > 30):
                c = c + 1
                aDesc[c] = Desc[i]
                aInizio[c] = Inizio[i]
                aFine[c] = Fine[i]
        aDesc = aDesc[:c]
        aInizio = aInizio[:c]
        aFine = aFine[:c]

        self.NuovoTesto("\nPeriodo gg." + str(Base + 1) + ":" + str(Base + 30) + "\n", False, True, 10, "Arial", "Justify")
        
        Largh = 9636
        buf=[]
        for i in range(len(aDesc)):
            if i == 1:
                buf.append("\\par \\trowd\\trql\\cellx" + str(Largh))
            else:
                buf.append("\\cell\\row\\pard \\trowd\\trql\\cellx" + str(Largh))
            buf.append("\n")
            
            buf.append("\\pard\\intbl\\pard\\plain \\intbl\\s5\\li113\\ri0\\lin113\\rin0\\fi0\\sb28\\sa28{\\*\\hyphen2\\hyphlead2\\hyphtrail2\\hyphmax0}\\rtlch\\afs24\\lang255\\ltrch\\dbch\\af3\\afs24\\langfe255\\loch\\f1\\fs24\\lang255\\ltrch\\loch\\fs20 {\\ltrch\\loch\\f1 " + aDesc[i] + "}")
            buf.append("\n")

            buf.append("\\cell\\row\\pard \\trowd\\trql")
            for i1 in range(30):
                if (aInizio[i] - Base) > (i1+1) or (aFine[i] - Base) < (i1+1):
                    buf.append("\\cellx" + str(Largh / 30 * (i1+1)))
                else:
                    buf.append("\\clcbpat2\\cellx" + str(Largh / 30 * (i1+1)))
            buf.append("\n")
            for i1 in range(30):
                if i1 == 0:
                    buf.append("\\pard\\intbl")
                else:
                    buf.append("\\cell")
                buf.append("\\plain \\intbl\\s5\\li60\\ri0\\lin60\\rin0\\fi0\\sb60\\sa60{\\*\\hyphen2\\hyphlead2\\hyphtrail2\\hyphmax0}\\rtlch\\afs24\\lang255\\ltrch\\dbch\\af3\\afs24\\langfe255\\loch\\f1\\fs24\\lang255\\ltrch\\loch\\fs16 ")
                if (aInizio[i] - Base) > (i1+1) or (aFine[i] - Base) < (i1+1):
                    pass
                else:
                    buf.append(" {\\ltrch\\loch\\f1 " + str(Base + (i1+1)) & "}")
                buf.append("\n")
            if i == (len(aDesc)-1):
                buf.append("\\cell\\row\\pard ")

    def TerminaRtf(self):
        return "\\par }}"


#from pysqlite2 import dbapi2 as sqlite
sqlite_version = None
try:
     import sqlite
     sqlite_version = 1
except ImportError:
     try:
         from pysqlite2 import dbapi2 as sqlite
         if sqlite.version_info < (2,1,0):
             raise ValueError('pysqlite2 minimum version is 2.1.0+ - %s found'%sqlite.version)
         sqlite_version = 2
     except ImportError:
         import sqlite3 as sqlite
         sqlite_version = 3

con = sqlite.connect(DATABASE, isolation_level=None, timeout=60.0)
cur = con.cursor()
try: cur.execute('PRAGMA auto_vacuum = 1')
except: pass
# dangerous performance - begin
# cur.execute('PRAGMA synchronous =  0')
# cur.execute('PRAGMA count_changes = 0')

cur.execute('PRAGMA temp_store =  MEMORY')
# cur.execute('PRAGMA journal_mode = MEMORY')
# dangerous performance - end

## pp begin
import zipfile
zfile = zipfile.ZipFile(PROGNAME+"_library1.zip", "r" )
pp_fnames=("pp/pp.py","pp/ppauto.py","pp/ppcommon.py","pp/pptransport.py","pp/ppworker.py","pp/ppserver.py","pyRijndael.py","ntlmaps.py","ntlmaps.cfg","singletonmixin.py")
for fname in pp_fnames:
    data = zfile.read(fname)
    if "/" in fname: 
        fname_nopath=fname.split("/")[1]
    else:
        fname_nopath=fname
    try: os.unlink(fname_nopath)
    except: pass
    fout = open(fname_nopath, "w")
    fout.write(data)
    fout.close()

sys.path.insert(0, '.')
import pyRijndael as rijndael
import pp
# ppservers = ()
ppservers_local = ()
job_server_local = pp.Server(ppservers=ppservers_local)
ppservers = ("*",)
job_server = pp.Server(ppservers=ppservers, secret="heavybase")
    
time.sleep(1)
#print "Starting pp with", job_server_local.get_ncpus(), "cpu cores"
active_nodes=job_server.get_active_nodes()
print "Starting cluster with:"
for key in active_nodes.keys():
    print "node "+`key`+" - "+`active_nodes[key]`+" cpu cores"

SetNTLMAPSconfig()
import subprocess
if os.name=='nt':
    pp_server = subprocess.Popen(["pythonw","ppserver.py","-w","1","-a","-s","heavybase"])
    ntlmaps_server = subprocess.Popen(["pythonw","ntlmaps.py","-c","ntlmaps.cfg"])
else:
    pp_server = subprocess.Popen(["python","-S","ppserver.py","-w","1","-a","-s","heavybase",">/dev/null","2>&1"])
    ntlmaps_server = subprocess.Popen(["python","ntlmaps.py","-c","ntlmaps.cfg",">/dev/null","2>&1"])
## pp end

if LoadCustomSetting("HeavyBaseService")!="0":
    HeavyBaseServiceTask = HeavyBaseServiceStarter()
    HeavyBaseServiceTask.start()
    #job=job_server_local.submit(HeavyBaseServiceCaller,(), (), ())

# fino a questo punto non mi serve includere nel path lo zip "HeavyBase_library2.zip"
sys.path.append(PROGNAME+'_library1.zip')
sys.path.append(PROGNAME+'_library2.zip')
sys.path.append(PROGNAME+'_library3.zip')
sys.path.append(PROGNAME+'_library4.zip')
sys.path.append(PROGNAME+'_library5.zip')
try: import xlwt
except: from pyExcelerator import *
try: import xlrd
except: pass
try: import openpyxl
except: pass

from hashlib import sha256  
import base64

try: import quickreport
except ImportError: pass
try: import analysis
except ImportError: pass

import socket
socket.setdefaulttimeout(10)

# HeavyBaseServer - Begin
from BaseHTTPServer import BaseHTTPRequestHandler
class HeavyBaseServer(BaseHTTPRequestHandler):
    def CustomInit(self):
        self.machineState = 0
        self.con = sqlite.connect(DATABASE, isolation_level=None, timeout=10.0)
        self.cur = self.con.cursor()
        self.cur.execute('PRAGMA temp_store =  MEMORY')
        self.cur.execute('PRAGMA journal_mode = MEMORY')
        
    def TT(self,key):
        tofind="{"+key+"}"
        self.cur.execute("SELECT setting_value FROM settings WHERE setting_key like '"+tofind+"'")
        row = self.cur.fetchone()
        ret=""
        if row!=None: ret=row[0]
        if ret=="": ret=key
        import re
        pattern = re.compile("heavy"+"base", re.IGNORECASE)
        ret=pattern.sub(PROGNAME, ret)
        return ret
    def do_GET(self):
        try: 
            self.machineState
        except: self.CustomInit()
        
        try:
            filename=self.path
            if self.path=="/": filename="/index.html"
            if filename=="/index.html":
                self.send_response(200)
                self.send_header('Content-type',        'text/html')
                self.end_headers()
                if self.machineState == 0:      #Login page
                    txt=[]
                    txt.append("<HTML><BODY style=\"background-color:#E0E0E0; font-size:11px; font-family:Verdana;\">")
                    txt.append("<FORM>")
                    txt.append("<table border=0 cellpadding=0 cellspacing=0 height=100%><tr><td valign='top'>")
                    txt.append("<TABLE BORDER=0>")
                    txt.append("<TR><TD>"+self.TT('Username')+"</TD><TD><INPUT TYPE='text' NAME='txtUsername'></TD></TR>")
                    txt.append("<TR><TD>"+self.TT('Password')+"</TD><TD><INPUT TYPE='password' NAME='txtPassword'></TD></TR>")
                    txt.append("<TR><TD>"+self.TT('Cypher key')+"</TD><TD><INPUT TYPE='password' NAME='txtCryptoKey'></TD></TR>")
                    txt.append("<TR><TD><input type='submit' value='login' style='font-family: Verdana; font-size: 100%; width:100; height:28;'></TD><TD>&nbsp;</TD></TR>")
                    txt.append("</TABLE>")
                    txt.append("</FORM>")
                    txt.append("</td></tr><tr><td valign='middle'>")
                    txt.append("&nbsp;</td></tr><tr><td valign='bottom'>")
                    txt.append("<p style=\"font:xx-small Verdana,Sans-serif;\">"+DISCLAIMER.replace("<","&lt;").replace(">","&gt;").replace("\n","<br>")+"</p>")
                    txt.append("</td></tr>")
                    txt.append("</BODY></HTML>")
                    self.wfile.write(''.join(txt))
        except IOError:
            self.send_error(404,'File Not Found: %s' % self.path)
            
    def do_POST(self):
        global rootnode
        try:
            #ctype, pdict = cgi.parse_header(self.headers.getheader('content-type'))
            #if ctype == 'multipart/form-data':
                #postvars = cgi.parse_multipart(self.rfile, pdict)
            #elif ctype == 'application/x-www-form-urlencoded':
                #length = int(self.headers.getheader('content-length'))
                #postvars = cgi.parse_qs(self.rfile.read(length), keep_blank_values=1)
            #else:
                #postvars = {}
                
            #content_len = int(self.headers.getheader('content-length'))
            #post_body = self.rfile.read(content_len)

            ctype, pdict = cgi.parse_header(self.headers.getheader('content-type'))
            if ctype=='multipart/form-data':
                query=cgi.parse_multipart(self.rfile, pdict)
            #self.send_response(301)
            self.send_response(200)
            self.end_headers()
        except IOError:
            self.send_error(404,'File Not Found: %s' % self.path)

class HeavyBaseServerStarter(threading.Thread):
    def __init__ (self):
        threading.Thread.__init__(self)
        self.finished = threading.Event()
 
    def stop (self):
        try:
            print "Shutting down "+PROGNAME+"Server"
            self.server.shutdown()
        except: pass
        self.finished.set()
        self.join()

    def run (self):
        from BaseHTTPServer import HTTPServer

        socket.setdefaulttimeout(10)

        try:
            self.server = HTTPServer(('', 60003), HeavyBaseServer)
            
            #Go into the main listener loop
            print "Starting "+PROGNAME+"Server on port 60003"
            self.server.serve_forever()
        except:
            print PROGNAME+"Server already started on port 60003"

if LoadCustomSetting("HeavyBaseServer")=="1":
    HeavyBaseServerTask = HeavyBaseServerStarter()
    HeavyBaseServerTask.start()
# HeavyBaseServer - End
    
# SqliteServer - Begin
import os
import sys
import time
import socket
from threading import Thread
import sqlite3 as sqlite
from struct import pack, unpack

# BON #
def dump(value, file):
    buf = dumps(value)
    file.write(buf)

def load(file):
    buf = file.read()
    return loads(buf)

def dumps(obj):
    buf = serialize(obj)
    _buf = serialize(len(buf))
    return _buf + buf

def loads(buf):
    obj_size, s, e = deserialize(buf, 0, len(buf))
    obj, _, _ = deserialize(buf, e, len(buf))
    return obj

def serialize(obj):
    buflist = []

    if obj is None:
        buflist.append('n')

    elif type(obj) == bool:
        buflist.append('1' if obj else '0')

    elif type(obj) in (int, long):
        if obj >= 0:
            if obj < 2 ** 8:
                buflist.append('B')
                buflist.append(pack('!B', obj))
            elif obj < 2 ** 16:
                buflist.append('H')
                buflist.append(pack('!H', obj))
            elif obj < 2 ** 32:
                buflist.append('I')
                buflist.append(pack('!I', obj))
            elif obj < 2 ** 64:
                buflist.append('L')
                buflist.append(pack('!Q', obj))
            else:
                raise Exception, 'requires %d <= signed number <= %d' % (0, 2 ** 64 - 1)
        else:
            if -(2 ** 7) <= obj < 2 ** 7:
                buflist.append('b')
                buflist.append(pack('!b', obj))
            elif -(2 ** 15) <= obj < 2 ** 15:
                buflist.append('h')
                buflist.append(pack('!h', obj))
            elif -(2 ** 31) <= obj < 2 ** 31:
                buflist.append('i')
                buflist.append(pack('!i', obj))
            elif -(2 ** 63) <= obj < 2 ** 63:
                buflist.append('l')
                buflist.append(pack('!q', obj))
            else:
                raise Exception, 'requires %d <= unsigned number <= %d' % (-(2 ** 63), 2 ** 63 - 1)

    elif type(obj) == float:
        buflist.append('d')
        buflist.append(pack('!d', obj))

    elif type(obj) == str:
        if len(obj) == 1:
            buflist.append('C')
            buflist.append(obj)
        else:
            buflist.append('s')
            buflist.append(serialize(len(obj)))
            buflist.append(obj)
            
    elif type(obj) == unicode:
        buflist.append('u')
        buffer_obj = buffer(obj)
        buflist.append(serialize(len(buffer_obj)))
        buflist.append(str(buffer_obj))
    
    elif type(obj) == tuple:
        buflist.append('(')
        buflist.append(serialize(len(obj)))
        for _obj in obj:
            buflist.append(serialize(_obj))

    elif type(obj) == list:
        buflist.append('[')
        buflist.append(serialize(len(obj)))
        for _obj in obj:
            buflist.append(serialize(_obj))

    elif type(obj) == dict:
        buflist.append('{')
        buflist.append(serialize(len(obj)))
        for _obj_k, _obj_v in obj.iteritems():
            buflist.append(serialize(_obj_k))
            buflist.append(serialize(_obj_v))

    return ''.join(buflist)

def deserialize(buf, s, e):
    obj_type = unpack('!c', buf[s])[0]

    if obj_type == 'n':
        obj = None
        return obj, s, s + 1

    elif obj_type == '0':
        obj = False
        return obj, s, s + 1

    elif obj_type == '1':
        obj = True
        return obj, s, s + 1

    elif obj_type == 'c':
        obj = chr(256 + unpack('!b', buf[s + 1:s + 2])[0])
        return obj, s, s + 2

    elif obj_type == 'C':
        obj = chr(unpack('!B', buf[s + 1:s + 2])[0])
        return obj, s, s + 2

    elif obj_type == 'b':
        obj = unpack('!b', buf[s + 1:s + 2])[0]
        return obj, s, s + 2

    elif obj_type == 'B':
        obj = unpack('!B', buf[s + 1:s + 2])[0]
        return obj, s, s + 2

    elif obj_type == 'h':
        obj = unpack('!h', buf[s + 1:s + 3])[0]
        return obj, s, s + 3

    elif obj_type == 'H':
        obj = unpack('!H', buf[s + 1:s + 3])[0]
        return obj, s, s + 3

    elif obj_type == 'i':
        obj = unpack('!i', buf[s + 1:s + 5])[0]
        return obj, s, s + 5

    elif obj_type == 'I':
        obj = unpack('!I', buf[s + 1:s + 5])[0]
        return obj, s, s + 5

    elif obj_type == 'l':
        obj = unpack('!q', buf[s + 1:s + 9])[0]
        return obj, s, s + 9

    elif obj_type == 'L':
        obj = unpack('!Q', buf[s + 1:s + 9])[0]
        return obj, s, s + 9

    elif obj_type == 'f':
        obj = unpack('!f', buf[s + 1:s + 5])[0]
        return obj, s, s + 5

    elif obj_type == 'd':
        obj = unpack('!d', buf[s + 1:s + 9])[0]
        return obj, s, s + 9

    elif obj_type == 's':
        obj_size, _, _s = deserialize(buf, s + 1, e)
        obj = buf[_s:_s + obj_size]
        return obj, s, _s + obj_size
        
    elif obj_type == 'u':
        obj_size, _, _s = deserialize(buf, s + 1, e)
        obj = buf[_s:_s + obj_size].decode('utf-16')
        return obj, s, _s + obj_size

    elif obj_type == '(':
        obj = []
        obj_size, _, _s = deserialize(buf, s + 1, e)
        for i in range(obj_size):
            _obj, _, _s = deserialize(buf, _s, e)
            obj.append(_obj)
        return tuple(obj), s, _s

    elif obj_type == '[':
        obj = []
        obj_size, _, _s = deserialize(buf, s + 1, e)
        for i in range(obj_size):
            _obj, _, _s = deserialize(buf, _s, e)
            obj.append(_obj)
        return obj, s, _s

    elif obj_type == '{':
        obj = {}
        obj_size, _, _s = deserialize(buf, s + 1, e)
        for i in range(obj_size):
            _obj_k, _, _s = deserialize(buf, _s, e)
            _obj_v, _, _s = deserialize(buf, _s, e)
            obj[_obj_k] = _obj_v
        return obj, s, _s

# COMMON #
def recv(soccon):
    buf = ''
    tmp = ''

    while len(buf) < 11:
        try:
            tmp = soccon.recv(11)
        except socket.error, msg:
            print msg
            break
    
        if tmp:
            buf += tmp
        else:
            print soccon, 'empty string received, socket connection closed'
            soccon.close()
            break

    if len(buf) >= 11:
        buf_size = deserialize(buf[:11],0,len(buf[:11]))[0]
    else:
        return buf
    
    while len(buf) < (2 + buf_size):
        try:
            tmp = soccon.recv(4096 if (11 + buf_size) - len(buf) >= 4096 else (11 + buf_size) - len(buf))
        except socket.error, msg:
            print msg
            break
    
        if tmp:
            buf += tmp
        else:
            print soccon, 'empty string received, socket connection closed'
            soccon.close()
            break
    return buf

# SERVER #
class SqliteServer(object):
    def __init__(self, host, port):
        self._host = host
        self._port = port
        self._socket = None
        
        self._socconadr_sqlcon = {} # {(soccon, socadr): sqlcon}
        self._sqlcon_sqlcuridcur = {} # {sqlcon: {sqlcurid: sqlcur}}
        
        for res in socket.getaddrinfo(self._host, self._port, socket.AF_UNSPEC, socket.SOCK_STREAM, 0, socket.AI_PASSIVE):
            af, socktype, proto, canonname, sa = res
            
            try:
                self._socket = socket.socket(af, socktype, proto)
                self._socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            except socket.error, msg:
                print 'SqliteServer', msg
                self._socket = None
                continue
            
            try:
                self._socket.bind(sa)
                self._socket.listen(1)
                self._socket.setblocking(1)
            except socket.error, msg:
                print 'SqliteServer', msg
                self._socket.close()
                self._socket = None
                continue
            
            break
            
        if not self._socket:
            print 'SqliteServer:'
            print '\terror: could not open socket'
            sys.exit(1)
        
        while True:
            try:
                soccon, socadr = self._socket.accept()
                Thread(target=self.process_connection, args=(soccon, socadr)).start()
            except socket.error, msg:
                print 'SqliteServer', msg
    
    def process_connection(self, soccon, socadr):
        while True:
            # first 8 bytes are total size to be received
            buf = recv(soccon)
            if len(buf) >= 11:
                self.process_request(soccon, socadr, buf)
            else:
                soccon.close()
                break
    
    def process_request(self, soccon, socadr, buf):
        request = loads(buf)
        
        request_class = request['__class__']
        request_func = request['__func__']
        request_kw = dict([(k, v) for k, v in request.iteritems() if k not in ('__class__', '__func__', '__id__')])
        
        if request_class == 'Connection':
            if request_func == '__init__':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = sqlite.connect(**request_kw)
                    self._socconadr_sqlcon[(soccon, socadr)] = sqlcon
                    self._sqlcon_sqlcuridcur[sqlcon] = {}
                        
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
                
            elif request_func == 'close':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    del self._socconadr_sqlcon[(soccon, socadr)]
                    del self._sqlcon_sqlcuridcur[sqlcon]
                    sqlcon.close()
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
                soccon.close()
            
            elif request_func == 'commit':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcon.commit()
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'rollback':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcon.rollback()
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
        
        elif request_class == 'Cursor':
            if request_func == '__init__':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = sqlcon.cursor()
                    self._sqlcon_sqlcuridcur[sqlcon][request['__id__']] = sqlcur
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'close':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    del self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    sqlcur.close()
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'callproc':
                pass
            
            elif request_func == 'description':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    cols = [i[0] for i in sqlcur.description]
                    response['row'] = cols
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'execute':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    sqlcur.execute(request['operation'].encode('utf-8'), request['parameters'])
                except sqlite.OperationalError, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'executemany':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    sqlcur.executemany(request['operation'].encode('utf-8'), request['seq_of_parameters'])
                except sqlite.OperationalError, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'fetchone':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    row = sqlcur.fetchone()
                    response['compression'] = request['compression']
                    if request['compression'] == None:
                        response['row'] = row
                    elif request['compression'] in ('gz', 'gzip', 'zlib'):
                        response['row'] = dumps(row).encode('zlib')
                    elif request['compression'] in ('bz2', 'bzip2'):
                        response['row'] = dumps(row).encode('bz2')
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'fetchmany':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    rows = sqlcur.fetchmany(request['size'])
                    response['compression'] = request['compression']
                    if request['compression'] == None:
                        response['rows'] = rows
                    elif request['compression'] in ('gz', 'gzip', 'zlib'):
                        response['rows'] = dumps(rows).encode('zlib')
                    elif request['compression'] in ('bz2', 'bzip2'):
                        response['rows'] = dumps(rows).encode('bz2')
                except Exception, msg:
                    response['error'] = str(msg)
                soccon.send(dumps(response))
            
            elif request_func == 'fetchall':
                response = {'__class__':request_class, '__func__':request_func, '__id__':request['__id__']}
                try:
                    sqlcon = self._socconadr_sqlcon[(soccon, socadr)]
                    sqlcur = self._sqlcon_sqlcuridcur[sqlcon][request['__id__']]
                    rows = sqlcur.fetchall()
                    response['compression'] = request['compression']
                    if request['compression'] == None:
                        response['rows'] = rows
                    elif request['compression'] in ('gz', 'gzip', 'zlib'):
                        response['rows'] = dumps(rows).encode('zlib')
                    elif request['compression'] in ('bz2', 'bzip2'):
                        response['rows'] = dumps(rows).encode('bz2')
                except Exception, msg:
                    response['error'] = str(msg)
                
                soccon.send(dumps(response))

# CLIENT #
def SqliteServerConnect(host, port, username, password, **kw):
    return Connection(host, port, username, password, **kw)

class Connection(object):
    def __init__(self, host, port, username, password, **kw):
        self._host = host
        self._port = port
        self._username = username
        self._password = password
        
        self._socket = None

        for res in socket.getaddrinfo(self._host, self._port, socket.AF_UNSPEC, socket.SOCK_STREAM):
            af, socktype, proto, canonname, sa = res
        
            try:
                self._socket = socket.socket(af, socktype, proto)
                self._socket.setsockopt(socket.SOL_SOCKET, socket.SO_REUSEADDR, 1)
            except socket.error, msg:
                self._socket = None
                continue
        
            try:
                self._socket.connect(sa)
                self._socket.setblocking(1)
            except socket.error, msg:
                self._socket.close()
                self._socket = None
                continue
        
            break
        
        if not self._socket:
            print 'could not open socket'
            sys.exit(1)
        
        request = {'__class__': 'Connection', '__func__':'__init__', '__id__':id(self)}
        request.update(kw)
        #print "request = "+str(request)+"---"+str(len(str(request)))
        self._socket.send(dumps(request))
        
        buffer = recv(self._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
        
    def close(self):
        request = {'__class__': 'Connection', '__func__':'close', '__id__':id(self)}
        self._socket.send(dumps(request))

        buffer = recv(self._socket)

        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']

        self._socket.close()
        self._socket = None

    
    def commit(self):
        request = {'__class__': 'Connection', '__func__':'commit', '__id__':id(self)}
        self._socket.send(dumps(request))
        
        buffer = recv(self._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
    
    def rollback(self):
        request = {'__class__': 'Connection', '__func__':'rollback', '__id__':id(self)}
        self._socket.send(dumps(request))
        
        buffer = recv(self._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
    
    def cursor(self):
        cursor = Cursor(connection=self)
        return cursor
    
class Cursor(object):
    def __init__(self, connection):
        self._connection = connection
        #self.description = (None, None, None, None, None, None, None)
        self.rowcount = None
        self.arraysize = 1
        
        request = {'__class__': 'Cursor', '__func__':'__init__', '__id__':id(self)}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        #print response
        if 'error' in response:
            raise Exception, response['error']
    
    def close(self):
        request = {'__class__': 'Cursor', '__func__':'close', '__id__':id(self)}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
    
    def description(self):
        request = {'__class__': 'Cursor', '__func__':'description', '__id__':id(self)}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
        
        return response['row']
    
    def callproc(self, procname, parameters=None):
        raise NotImplementedError
    
    def execute(self, operation, parameters=()):
        request = {'__class__': 'Cursor', '__func__':'execute', '__id__':id(self), 'operation':operation, 'parameters':parameters}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
    
    def executemany(self, operation, seq_of_parameters=()):
        request = {'__class__': 'Cursor', '__func__':'executemany', '__id__':id(self), 'operation':operation, 'seq_of_parameters':seq_of_parameters}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
    
    def fetchone(self, compression=None):
        request = {'__class__': 'Cursor', '__func__':'fetchone', '__id__':id(self), 'compression':compression}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
        
        if response['compression'] == None:
            return response['row']
        elif response['compression'] in ('gz', 'gzip', 'zlib'):
            return loads(response['row'].decode('zlib'))
        elif response['compression'] in ('bz2', 'bzip2'):
            return loads(response['row'].decode('bz2'))
    
    def fetchmany(self, size=1, compression=None):
        request = {'__class__': 'Cursor', '__func__':'fetchmany', '__id__':id(self), 'size':size, 'compression':compression}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
        
        if response['compression'] == None:
            return response['rows']
        elif response['compression'] in ('gz', 'gzip', 'zlib'):
            return loads(response['rows'].decode('zlib'))
        elif response['compression'] in ('bz2', 'bzip2'):
            return loads(response['rows'].decode('bz2'))
        
    def fetchall(self, compression=None):
        request = {'__class__': 'Cursor', '__func__':'fetchall', '__id__':id(self), 'compression':compression}
        self._connection._socket.send(dumps(request))
        
        buffer = recv(self._connection._socket)
        response = loads(buffer)
        if 'error' in response:
            raise Exception, response['error']
        
        if response['compression'] == None:
            return response['rows']
        elif response['compression'] in ('gz', 'gzip', 'zlib'):
            return loads(response['rows'].decode('zlib'))
        elif response['compression'] in ('bz2', 'bzip2'):
            return loads(response['rows'].decode('bz2'))
    
    def nextset(self):
        raise NotImplementedError
    
    def setinputsizes(self, sizes):
        raise NotImplementedError
    
    def setoutputsize(size, column=None):
        raise NotImplementedError
        
class SqliteServerStarter(threading.Thread):
    def __init__ (self):
        threading.Thread.__init__(self)
        self.finished = threading.Event()
 
    def stop (self):
        try:
            print "Shutting down SqliteServer"
            self.sqliteserver.shutdown()
        except: pass
        self.finished.set()
        self.join()

    def run (self):
        try:
            #Go into the main listener loop
            print "Starting SqliteServer on port 60004"
            self.sqliteserver = SqliteServer(None, 60004)
        except:
            print "SqliteServer already started on port 60004"

if LoadCustomSetting("SqliteServer")=="1":
    SqliteServerTask = SqliteServerStarter()
    SqliteServerTask.start()
# SqliteServer - End

# SYSINFO - Begin
def system_user():
    if sys.platform != 'darwin':
        import getpass
        return getpass.getuser()

def system_ram():
    #if sys.platform == 'win32':
    if os.name=='nt':
        #Windows Only
        import ctypes
        kernel32 = ctypes.windll.kernel32
        c_ulong = ctypes.c_ulong
        class MEMORYSTATUS(ctypes.Structure):
            _fields_ = [
                ('dwLength', c_ulong),
                ('dwMemoryLoad', c_ulong),
                ('dwTotalPhys', c_ulong),
                ('dwAvailPhys', c_ulong),
                ('dwTotalPageFile', c_ulong),
                ('dwAvailPageFile', c_ulong),
                ('dwTotalVirtual', c_ulong),
                ('dwAvailVirtual', c_ulong)
            ]
        memoryStatus = MEMORYSTATUS()
        memoryStatus.dwLength = ctypes.sizeof(MEMORYSTATUS)
        kernel32.GlobalMemoryStatus(ctypes.byref(memoryStatus))
        mem = memoryStatus.dwTotalPhys / (1024.0*1024.0)
        availRam = memoryStatus.dwAvailPhys / (1024.0*1024.0)
        if mem >= 1000:
            mem = mem/1000.0
            totalRam = str(mem) + ' GB'
        else:
            totalRam = str(mem) + ' MB'
        return totalRam+" Total - "+str(availRam)+" Avail"

def get_registry_value(key, subkey, value):
    #Windows Only
    import _winreg
    key = getattr(_winreg, key)
    handle = _winreg.OpenKey(key, subkey)
    (value, type) = _winreg.QueryValueEx(handle, value)
    return value

def system_cpu():
    #if sys.platform == 'win32':
    import os
    if os.name=='nt':
        #Windows Only
        import ctypes
        try:
            cputype = get_registry_value("HKEY_LOCAL_MACHINE","HARDWARE\\DESCRIPTION\\System\\CentralProcessor\\0","ProcessorNameString")
        except:
            import wmi, pythoncom
            pythoncom.CoInitialize()
            c = wmi.WMI()
            for i in c.Win32_Processor ():
                cputype = i.Name
            pythoncom.CoUninitialize()
        if cputype == 'AMD Athlon(tm)':
            c = wmi.WMI()
            for i in c.Win32_Processor ():
                cpuspeed = i.MaxClockSpeed
            cputype = 'AMD Athlon(tm) %.2f Ghz' % (cpuspeed / 1000.0)
        elif cputype == 'AMD Athlon(tm) Processor':
            import wmi
            c = wmi.WMI()
            for i in c.Win32_Processor ():
                cpuspeed = i.MaxClockSpeed
            cputype = 'AMD Athlon(tm) %s' % cpuspeed
        else:
            pass
        return cputype
    
def system_os():
    #if sys.platform == 'win32':
    import os
    if os.name=='nt':
        #Windows Only
        os=""
        try: os = get_registry_value("HKEY_LOCAL_MACHINE","SOFTWARE\\Microsoft\\WINDOWS NT\\CurrentVersion","ProductName")
        except: pass
        sp=""
        try: sp = get_registry_value("HKEY_LOCAL_MACHINE","SOFTWARE\\Microsoft\\WINDOWS NT\\CurrentVersion","CSDVersion")
        except: pass
        build=""
        try: build = get_registry_value("HKEY_LOCAL_MACHINE","SOFTWARE\\Microsoft\\WINDOWS NT\\CurrentVersion","CurrentBuildNumber")
        except: pass
        return "%s %s (build %s)" % (os, sp, build)
    
#if sys.platform == 'win32':
if os.name=='nt':
    sysinfofile = open('sysinfo.txt','w')
    now = datetime.datetime.utcnow()
    ts=now.strftime("%Y-%m-%d %H:%M:%S")
    sysinfofile.write(ts+'\n')
    #sysinfofile.write("System User: "+system_user()+'\n')
    sysinfofile.write("CPU: "+system_cpu()+'\n')
    sysinfofile.write("RAM: "+system_ram()+'\n')
    sysinfofile.write("OS version: "+str(system_os())+'\n')
    sysinfofile.close()
# SYSINFO - End

# IDLE TIME on Windows - Begin
#if sys.platform == 'win32':
if os.name=='nt':
    from ctypes import Structure, windll, c_uint, sizeof, byref

    class LASTINPUTINFO(Structure):
        _fields_ = [
            ('cbSize', c_uint),
            ('dwTime', c_uint),
        ]

    def get_idle_duration():
        lastInputInfo = LASTINPUTINFO()
        lastInputInfo.cbSize = sizeof(lastInputInfo)
        windll.user32.GetLastInputInfo(byref(lastInputInfo))
        millis = windll.kernel32.GetTickCount() - lastInputInfo.dwTime
        return millis / 1000.0
# IDLE TIME on Windows - End

def GetRtfValue(testo):
    if testo!="":
        testo=testo.replace(u"\\"  ,"")
        testo=testo.replace(u"\r\n","\\par ")
        testo=testo.replace(u"\r"  ,"\\par ")
        testo=testo.replace(u"\n"  ,"\\par ")
        testo=testo.replace(u"\t"  ,"\\par ")
        testo=testo.replace(u"`"   ,"\\'92")
        testo=testo.replace(u"\'"  ,"\\'92")
        testo=testo.replace(u'"'   ,"\\'94")
        testo=testo.replace(u"à"   ,"\\'e0")
        testo=testo.replace(u"è"   ,"\\'e8")
        testo=testo.replace(u"é"   ,"\\'e9")
        testo=testo.replace(u"ì"   ,"\\'ec")
        testo=testo.replace(u"ò"   ,"\\'f2")
        testo=testo.replace(u"ù"   ,"\\'f9")
    return testo

class MultiReplace:
    def __init__(self, repl_dict):
        #import string,re
        # "compile" replacement dictionary

        # assume char to char mapping
        charmap = map(chr, range(256))
        for k, v in repl_dict.items():
            if len(k) != 1 or len(v) != 1:
                self.charmap = None
                break
            charmap[ord(k)] = v
        else:
            self.charmap = string.join(charmap, "")
            return

        # string to string mapping; use a regular expression
        keys = repl_dict.keys()
        keys.sort() # lexical order
        keys.reverse() # use longest match first
        pattern = string.join(map(re.escape, keys), "|")
        self.pattern = re.compile(pattern)
        self.dict = repl_dict

    def replace(self, str):
        # apply replacement dictionary to string
        if self.charmap:
            return string.translate(str, self.charmap)
        def repl(match, get=self.dict.get):
            item = match.group(0)
            return get(item, item)
        return self.pattern.sub(repl, str)

def HB_Decrypt(trans,key,words,encoding):
    alldone=False
    while not alldone:
        try:
            jobs = [(word, job_server.submit(HB_SyncroDecrypt,(key,word,encoding), (), ("pyRijndael","base64"))) for word in words]
            for word, job in jobs:
                res=job()
                if res!="!ERROR!":
                    trans[word]=res
                else:
                    res=HB_SyncroDecrypt(key,word,encoding)
                    trans[word]=res
            alldone=True
        except: 
            print "HB_Decrypt failed. Retrying."
            time.sleep(1)
        
    return

def HB_SyncroDecrypt(key,word,encoding):
    if word=="":
        return ""
    else:
        try:
            import sys    
            sys.path.insert(0, '.')
            import pyRijndael
            if encoding!="":
                return unicode(pyRijndael.DecryptData(key,base64.b64decode(word)),encoding)
            else:
                return pyRijndael.DecryptData(key,base64.b64decode(word))
        except:
            return "!ERROR!"

def HB_DecryptOne(key,word,encoding):
    if word=="":
        return ""
    else:
        retry=True
        while retry:
            try:
                job=job_server.submit(HB_SyncroDecrypt,(key,word,encoding), (), ("pyRijndael","base64"))
                retry=False
            except: time.sleep(1)
        return job()

def HB_EncryptOne(key,word):
    if word=="":
        return ""
    else:
        retry=True
        while retry:
            try:
                job=job_server.submit(HB_SyncroEncrypt,(key,word), (), ("pyRijndael","base64"))
                retry=False
            except: time.sleep(1)
        return job()

def HB_SyncroEncrypt(key,word):
    if word=="":
        return ""
    else:
        import sys    
        sys.path.insert(0, '.')
        import pyRijndael
        return base64.b64encode(pyRijndael.EncryptData(key,word))

def HB_ZipString(content_in):
    job=job_server.submit(HB_SyncroZipString,(content_in,), (), ())
    return job()

def HB_SyncroZipString(content_in):
    from gzip import GzipFile
    from StringIO import StringIO
    sio = StringIO()
    gzf = GzipFile(fileobj=sio, mode='wb')
    gzf.write(content_in)
    gzf.close()
    return sio.getvalue()

def create_permutations(invar,minflags,maxflags,permutations):
    flag="$"
    if minflags<=maxflags:
        for i in range(len(invar)+1):
            cando=True
            if i>0: 
                if invar[i-1]==flag: cando=False
            if i<len(invar)-1: 
                if invar[i+1]==flag: cando=False
            if i<len(invar):
                if invar[i]==flag: cando=False
            if cando:
                newvar=invar[:i]+flag+invar[i:]
                if newvar not in permutations: 
                    permutations[newvar]=minflags
                    create_permutations(newvar,minflags+1,maxflags,permutations)

def HB_PrintCRF(text,contents,REPORTS_PATH,filename,strip="",templatetype="rtf",DIRTY_TAGS=False,DEBUG_MODE=False):
    import re
    import string
    cleankeys=[]
    couples={}
    ontology_id=-1
    if text.find("$#=")>=0:
        if text.find("$#=0$")>=0: ontology_id=0
        elif text.find("$#=1$")>=0: ontology_id=1
        elif text.find("$#=2$")>=0: ontology_id=2
        elif text.find("$#=3$")>=0: ontology_id=3
        elif text.find("$#=4$")>=0: ontology_id=4
        elif text.find("$#=5$")>=0: ontology_id=5
        elif text.find("$#=6$")>=0: ontology_id=6
        elif text.find("$#=7$")>=0: ontology_id=7
        elif text.find("$#=8$")>=0: ontology_id=8
        elif text.find("$#=9$")>=0: ontology_id=9
        couples["$#="+str(ontology_id)+"$"]=""

    for keycont in contents:
        dcurValue=contents[keycont]
        if ontology_id!=-1:
            if "|" in dcurValue:
                try: dcurValue=dcurValue.split("|")[ontology_id]
                except: pass
        curName=keycont
        if keycont.find("#")>=0: 
            curName = keycont[:keycont.find("#")] + "__" + keycont[keycont.find("#")+1:]
        notstrippedcurName=curName
        if strip!="": 
            #curName=curName.replace(strip,"")
            a=len(strip)
            if curName[-a:]==strip: curName=curName[:-a]
        if curName!="":
            findString="$"+curName+"$"
            notstrippedfindString="$"+notstrippedcurName+"$"
            cleankeys.append(curName)
            if curName!=notstrippedcurName:
                cleankeys.append(notstrippedcurName)
            rtfValue=""
            if templatetype=="abw":
                rtfValue=dcurValue
            else:   #rtf
                rtfValue=GetRtfValue(dcurValue)
            if rtfValue!="": 
                couples[findString]=rtfValue
                couples[notstrippedfindString]=rtfValue
            # 05/09/2013 - da controllare se ha senso la prossima riga
            if findString not in couples: couples[findString]=""
            if DEBUG_MODE: print curName+" ---> "+couples[findString]
    r = MultiReplace(couples)
    if templatetype=="abw":
        text=r.replace(unicode(text,"utf-8"))
        #text=unicode(text,"utf-8")
    else:   #rtf
        text=r.replace(text)
        #pass
    if DIRTY_TAGS:
        #dirty tags - begin
        for key in cleankeys:
            maxdirty=2
            permutations={}
            create_permutations(key,1,maxdirty,permutations)
            newper=[]
            for nflags in range(1,maxdirty+1):
                for elm in permutations:
                    if permutations[elm]==nflags:
                        newper.append(elm)
            for elm in newper:
                dirtykey=string.replace(re.escape(elm),"\\$","[^$]+?")
                tofind="\$"+dirtykey+"\$"
                #print dirtykey
                p=re.compile(tofind)
                for findString in p.findall(text):
                    if len(findString)>(2*len(key)) and couples.has_key("$"+key+"$"):
                        text=string.replace(text,findString,couples["$"+key+"$"])
        #dirty tags - end
    if not DEBUG_MODE:
        #p=re.compile('\$.*?\$')
        p=re.compile('\$[^$]+?\$')
        for findString in p.findall(text):
            #if len(findString)<75:
            #    import random
            #    random.seed()
            #    word="dglmnprtz"[int(random.random()*9)]+"ino"
            text=string.replace(text,findString,"")
    return text,templatetype,filename

def HB_CB_PrintCRF(text_templatetype_filename):
    text=text_templatetype_filename[0]
    templatetype=text_templatetype_filename[1]
    filename=text_templatetype_filename[2]
    
    #now = datetime.datetime.utcnow()
    #ts=now.strftime("%Y%m%d%H%M%S")

    #reportname="report_"+ts+"."+templatetype
    #print "writing report: "+reportname
    
    outfile = open(filename, 'w')
    if templatetype=="abw":
        outfile.write(text.encode('utf-8', 'replace'))
    else:   #rtf
        outfile.write(text.encode('latin-1', 'replace'))
    outfile.close()

    if not assoc_open(filename):
        wx.MessageBox("File '"+filename+"' created.", "Report", wx.ICON_INFORMATION | wx.OK, None)

def AsyncKillAll(hwnd,pp_fnames):
    import sys,os
    pid=os.getpid()
    import time
    time.sleep(10)
    import os
    if os.name=='nt':
        sys.path.insert(0, 'HeavyBase_library1.zip')
        import wmi
        c = wmi.WMI ()
        for process in c.Win32_Process(name="pythonw.exe"):
            if process.ProcessId!=pid:
                if os.path.abspath(".") in process.ExecutablePath:
                    try: process.Terminate()
                    except: pass
        for process in c.Win32_Process(name="python.exe"):
            if process.ProcessId!=pid:
                if os.path.abspath(".") in process.ExecutablePath:
                    try: process.Terminate()
                    except: pass
        if hwnd!=0:
            import win32gui
            try: win32gui.DestroyWindow(hwnd)
            except: pass
    else:
        import subprocess,signal
        process_name="python"
        ps     = subprocess.Popen("ps -eaf | grep "+process_name, shell=True, stdout=subprocess.PIPE)
        processes = ps.stdout.read()
        ps.stdout.close()
        ps.wait()
        print processes
        for process in processes.split("\n"):
            if (os.path.abspath(".") in process) or ("HeavyBase" in process) or ("ppserver" in process) or ("ntlmaps" in process):
                arr=process.strip().split(" ")
                first=True
                for elm in arr:
                    if not first:
                        if elm!="":
                            if int(elm)!=pid:
                                command=""
                                if sys.platform != 'darwin':
                                    command="pwdx "+elm
                                else:
                                    #ps = subprocess.Popen("lsof -a -p "+elm+" -d cwd -Fn | cut -c2- | grep -v "+elm, shell=True, stdout=subprocess.PIPE)
                                    command="lsof -a -p "+elm+" -d cwd -Fn | cut -c2- | grep -v "+elm
                                    #command="lsof -a -p "+elm+" -d cwd -Fn"
                                print command
                                ps = subprocess.Popen(command, shell=True, stdout=subprocess.PIPE)
                                pdir = ps.stdout.read()
                                ps.stdout.close()
                                ps.wait()
                                print pdir
                                if os.path.abspath(".") in pdir:
                                    print "killing "+elm
                                    try: os.kill(int(elm), signal.SIGTERM)
                                    except:
                                        try: os.kill(int(elm), signal.SIGKILL)
                                        except:
                                            try: os.kill(int(elm), signal.SIGABORT)
                                            except: pass
                            break
                    first=False
    ## pp begin
    for fname in pp_fnames:
        if "/" in fname:
            fname_nopath=fname.split("/")[1]
        else:
            fname_nopath=fname
        try:
            os.unlink(fname_nopath)
            os.unlink(fname_nopath+"c")
        except: pass
    ## pp end
    #self.Close()
    log(PROGNAME+" killed")
    print "System clean."
    #autorestart - begin
    try: 
        os.unlink("autorestart.cmd")
        import subprocess
        if os.name=='nt':
            newheavybase = subprocess.Popen(["pythonw",PROGNAME+".py"])
        else:
            newheavybase = subprocess.Popen(["python",PROGNAME+".py"])
    except: pass
    #autorestart - end
        
def ifelse(cond, true_part, false_part):
    if cond:
        return true_part()
    else:
        return false_part()

import threading
class ThreadLooper (threading.Thread):
 
    def __init__ (self, sleep_interval, function, args=[], kwargs={}):
 
        threading.Thread.__init__(self)
        self.sleep_interval = sleep_interval
        self.function = function
        self.args = args
        self.kwargs = kwargs
        self.finished = threading.Event()
 
    def stop (self):
 
        self.finished.set()
        self.join()
 
    def run (self):
        while not self.finished.isSet():
            self.finished.wait(self.sleep_interval)
            self.function(*self.args, **self.kwargs)

class PreDecrypt(threading.Thread):
 
    def __init__ (self, parentFrame, childFrame, digestkey, id_keyheader, mainfldlist):
        threading.Thread.__init__(self)
        self.parentFrame=parentFrame
        self.childFrame=childFrame
        self.digestkey=digestkey
        self.id_keyheader=id_keyheader
        self.mainfldlist=mainfldlist
 
    def run (self):
        DoPreDecrypt(self.parentFrame,self.childFrame,self.digestkey,self.id_keyheader,self.mainfldlist)
        #print "PreDecrypt Ended"

def DoPreDecrypt(parentFrame,childFrame,digestkey,id_keyheader,mainfldlist):
    parentFrame.preDecrypt_waitstate=3
    
    parentFrame.syncroTaskPause=True
    import time
    while parentFrame.syncroTaskRunning:
        time.sleep(1)

    import time
    con2 = sqlite.connect(":memory:")
    cur2 = con2.cursor()
    cur2.execute('PRAGMA temp_store =  MEMORY')
    cur2.execute('PRAGMA journal_mode = OFF')

    con1 = sqlite.connect(DATABASE, isolation_level=None, timeout=60.0)
    cur1 = con1.cursor()
    cur1.execute('PRAGMA temp_store =  MEMORY')
    cur1.execute('PRAGMA journal_mode = OFF')

    cur2.execute("CREATE TABLE contents_dictionary (data BLOB)")
    cur1.execute("SELECT data FROM contents_dictionary WHERE id_header="+str(id_keyheader))
    for row in cur1:
        cData=row[0]
        cur2.execute("INSERT INTO contents_dictionary (data) VALUES ('"+cData+"')")        

    words=[]
    for key in parentFrame.trans.keys():
        words.append(key)
        
    cur2.execute("SELECT data FROM contents_dictionary")
    for row in cur2:
        cData=row[0]
        if not cData in words:
            words.append(cData)
            if len(words)>=500:
                HB_Decrypt(parentFrame.trans,digestkey,words,"latin-1")
                words=[]
    if len(words)>0:
        HB_Decrypt(parentFrame.trans,digestkey,words,"latin-1")        
        words=[]
                
    parentFrame.preDecrypt_waitstate=2

    cur2.execute("DROP TABLE contents_dictionary")
    cur2.execute("CREATE TABLE contents_dictionary (data BLOB)")
    cur1.execute("SELECT data FROM contents_dictionary WHERE id_header in ("+mainfldlist+")")
    for row in cur1:
        cData=row[0]
        cur2.execute("INSERT INTO contents_dictionary (data) VALUES ('"+cData+"')")        

    cur2.execute("SELECT data FROM contents_dictionary")
    for row in cur2:
        cData=row[0]
        if not parentFrame.trans.has_key(cData):
            words.append(cData)
            if len(words)>=500:
                HB_Decrypt(parentFrame.trans,digestkey,words,"latin-1")        
                words=[]
    if len(words)>0:
        HB_Decrypt(parentFrame.trans,digestkey,words,"latin-1")        
        words=[]

    parentFrame.preDecrypt_waitstate=1
    
    
    fileexists=False
    try:
        hfile=open("quickreport.py","r")
        hfile.close()
        fileexists=True
    except: pass
    if fileexists:
        ShowQuickReport(DATABASE, con1, cur1, con2, cur2, digestkey, parentFrame, childFrame)
    cur2.close()
    con2.close()
    parentFrame.preDecrypt_waitstate=0
    parentFrame.syncroTaskPause=False


class QuickReporter(threading.Thread):
 
    def __init__ (self, parentFrame, childFrame, digestkey):
        threading.Thread.__init__(self)
        self.parentFrame=parentFrame
        self.childFrame=childFrame
        self.digestkey=digestkey
 
    def run (self):
        fileexists=False
        try:
            hfile=open("quickreport.py","r")
            hfile.close()
            fileexists=True
        except: pass
        if fileexists:
            con2 = sqlite.connect(":memory:")
            cur2 = con2.cursor()

            con1 = sqlite.connect(DATABASE, isolation_level=None, timeout=60.0)
            cur1 = con1.cursor()
            #cur1.execute('PRAGMA count_changes = 0')
            cur1.execute('PRAGMA temp_store =  MEMORY')
            cur1.execute('PRAGMA journal_mode = MEMORY')
            ShowQuickReport(DATABASE, con1, cur1, con2, cur2, self.digestkey, self.parentFrame, self.childFrame)
            cur1.close()
            con1.close()
            cur2.close()
            con2.close()
        
def ShowQuickReport(DATABASE, con1, cur1, con2, cur2, digestkey, parentFrame, childFrame):
    if parentFrame.quickreportRunning: return
    parentFrame.quickreportRunning=True
    if DEBUG_MODE:
        childFrame.quickreport_output=quickreport.Run(DATABASE, digestkey, parentFrame.trans, childFrame.id_group)
    else:
        try: childFrame.quickreport_output=quickreport.Run(DATABASE, digestkey, parentFrame.trans, childFrame.id_group)
        except: pass
    parentFrame.quickreportRunning=False

def SyncroP2P(parentFrame,childFrame):
    try: 
        #f=open("restore.cmd","r")
        #f.close()
        os.unlink("restore.cmd")
        print "restoring..."
        parentFrame.Show()
    except: pass
        
    if parentFrame.syncroTaskStop: 
        return
    if parentFrame.syncroTaskWaitSave: 
        parentFrame.p2p_message="synchro waiting for saving record"
        return
    if parentFrame.syncroTaskRunning or parentFrame.syncroTaskPause or parentFrame.syncroTaskUpdating: 
        #parentFrame.p2p_message="synchro waiting for prev. synchro tasks"
        #parentFrame.p2p_message="synchro paused"
        return
    if parentFrame.p2p_state=="" and len(parentFrame.p2p_defUpdate)>0: 
        parentFrame.p2p_message="synchro waiting for prev. updates"
        return
    
    parentFrame.syncroTaskRunning=True
    
    try: a=parentFrame.timer
    except:
        parentFrame.syncroTaskStop=True
        print "SyncroP2P Stopped"
        sys.exit()

    p2p={}
    p2p["state"]                = parentFrame.p2p_state
    p2p["defUpdate"]            = parentFrame.p2p_defUpdate
    p2p["id_instance"]          = parentFrame.id_instance
    p2p["message"]              = parentFrame.p2p_message
    p2p["md5Rows"]              = parentFrame.p2p_md5Rows
    p2p["lstRows"]              = parentFrame.p2p_lstRows
    p2p["deltaRows"]            = parentFrame.p2p_deltaRows
    p2p["deltaContents_index"]  = parentFrame.p2p_deltaContents_index
    p2p["deltaHeaders"]         = parentFrame.p2p_deltaHeaders
    p2p["deltaDict"]            = parentFrame.p2p_deltaDict
    p2p["partner"]              = parentFrame.p2p_partner
    p2p["partner_state"]        = parentFrame.p2p_partner_state
    p2p["partner_md5Rows"]      = parentFrame.p2p_partner_md5Rows
    p2p["partner_lstRows"]      = parentFrame.p2p_partner_lstRows
    p2p["partner_deltaRows"]    = parentFrame.p2p_partner_deltaRows
    p2p["partner_deltaContents_index"] = parentFrame.p2p_partner_deltaContents_index
    p2p["partner_deltaHeaders"] = parentFrame.p2p_partner_deltaHeaders
    p2p["partner_deltaDict"]    = parentFrame.p2p_partner_deltaDict
    p2p["chkRowsId"]            = parentFrame.p2p_chkRowsId
    p2p["keysDict"]             = parentFrame.p2p_keysDict
    p2p["arrSettings"]          = parentFrame.arrSettings
    p2p["cursyncrotable_id"]    = parentFrame.cursyncrotable_id
    p2p["syncrotables"]         = parentFrame.syncrotables  # input only
    p2p["syncrotableskeys"]     = parentFrame.syncrotableskeys  # input only
    p2p["syncrotablesflds"]     = parentFrame.syncrotablesflds  # input only
    p2p["syncrotablesupd"]      = parentFrame.syncrotablesupd
    p2p["chkRowsId_tables"]     = parentFrame.chkRowsId_tables
    p2p["lstRows_tables"]       = parentFrame.lstRows_tables
    p2p["md5Rows_tables"]       = parentFrame.md5Rows_tables
    p2p["lastConnectionUrl"]    = parentFrame.p2p_lastConnectionUrl
    p2p["syncroTaskStop"]       = parentFrame.syncroTaskStop
    p2p["max_idHeadDict"]       = parentFrame.max_idHeadDict
    p2p["liveupdate_ts"]        = parentFrame.p2p_liveupdate_ts
        
    ##DoSyncroP2P(p2p)
    #retry=True
    #while retry==True and parentFrame.syncroTaskStop==False:
        #try:
            #job=job_server_local.submit(DoSyncroP2P,(p2p,DATABASE), (GetSystemProxy,CustomRouter,ProxyTransport,GetSqlValue,HTTPconn,IEconn,TryConnect,getFileListMD5,getFileList,walker,ifelse), ("xmlrpclib","os","singletonmixin"))
            #p2p=job()
            #retry=False
        #except: 
            #time.sleep(1)
            #print "SyncroP2P closed unexpectedly"
    now = datetime.datetime.utcnow()
    ts=now.strftime("%d/%m/%Y, %H:%M:%S")
    anyerror=False    
    #try:
    job=job_server_local.submit(DoSyncroP2P,(p2p,DATABASE,DEBUG_MODE), (GetSystemProxy,CustomRouter,ProxyTransport,GetSqlValue,HTTPconn,IEconn,TryConnect,getFileListMD5,getFileList,walker,ifelse,log), ("xmlrpclib","os","singletonmixin"))
    p2p=job()
    #except: 
        #print ts+" - SyncroP2P closed unexpectedly"
        #parentFrame.p2p_message="Out of sync. Please restart the application"
        #anyerror=True

    #anyerror=False    
    if not anyerror:
        try:
            parentFrame.p2p_state                = p2p["state"]
        except:
            anyerror=True
            
    if not anyerror:
        parentFrame.p2p_defUpdate            = p2p["defUpdate"]
        parentFrame.id_instance              = p2p["id_instance"]
        parentFrame.p2p_message              = p2p["message"]
        parentFrame.p2p_md5Rows              = p2p["md5Rows"]
        parentFrame.p2p_lstRows              = p2p["lstRows"]
        parentFrame.p2p_deltaRows            = p2p["deltaRows"]
        parentFrame.p2p_deltaContents_index  = p2p["deltaContents_index"]
        parentFrame.p2p_deltaHeaders         = p2p["deltaHeaders"]
        parentFrame.p2p_deltaDict            = p2p["deltaDict"]
        parentFrame.p2p_partner              = p2p["partner"]
        parentFrame.p2p_partner_state        = p2p["partner_state"]
        parentFrame.p2p_partner_md5Rows      = p2p["partner_md5Rows"]
        parentFrame.p2p_partner_lstRows      = p2p["partner_lstRows"]
        parentFrame.p2p_partner_deltaRows    = p2p["partner_deltaRows"]
        parentFrame.p2p_partner_deltaContents_index = p2p["partner_deltaContents_index"]
        parentFrame.p2p_partner_deltaHeaders = p2p["partner_deltaHeaders"]
        parentFrame.p2p_partner_deltaDict    = p2p["partner_deltaDict"]
        parentFrame.p2p_chkRowsId            = p2p["chkRowsId"]
        parentFrame.p2p_keysDict             = p2p["keysDict"]
        parentFrame.arrSettings              = p2p["arrSettings"]
        parentFrame.cursyncrotable_id        = p2p["cursyncrotable_id"]
        parentFrame.syncrotablesupd          = p2p["syncrotablesupd"]

        parentFrame.chkRowsId_tables         = p2p["chkRowsId_tables"]
        parentFrame.lstRows_tables           = p2p["lstRows_tables"]
        parentFrame.md5Rows_tables           = p2p["md5Rows_tables"]
        parentFrame.p2p_lastConnectionUrl    = p2p["lastConnectionUrl"]
        parentFrame.syncroTaskStop           = p2p["syncroTaskStop"]
        parentFrame.max_idHeadDict           = p2p["max_idHeadDict"]
        parentFrame.p2p_liveupdate_ts        = p2p["liveupdate_ts"]
    try:
        p2pkeys=p2p.keys()
        for key in p2pkeys:
            del p2p[key]
        del p2p
    except: pass
    
    parentFrame.syncroTaskRunning=False

def DoSyncroP2P(p2p,DATABASE,DEBUG_MODE):
    DEBUG_SYNCHRO_SIMPLE=False
    DEBUG_SYNCHRO_BROADCAST=True
    DEBUG_SYNCHRO_UPDHP=False
    
    import socket
    socket.setdefaulttimeout(10)

    p2p["syncroTaskStop"]=False
    DATABASE_PATH=DATABASE[:DATABASE.rfind(os.path.sep)+1]
    import datetime,time,base64,zlib,hashlib
    import sqlite3 as sqlite
    try:
        con = sqlite.connect(DATABASE, isolation_level=None, timeout=60.0)
    except:
        p2p["message"]="synchro standby"
        return p2p
    cur = con.cursor()
    #cur.execute('PRAGMA count_changes = 0')
    cur.execute('PRAGMA temp_store =  MEMORY')
    cur.execute('PRAGMA journal_mode = OFF')
    cur.execute('PRAGMA query_only = 1')
    
    now = datetime.datetime.utcnow()
    ts=now.strftime("%d/%m/%Y, %H:%M:%S")
    #New reset after idle
    if p2p["state"]!="":
        cur.execute("SELECT setting_value FROM settings WHERE setting_key='p2p_syncro'")
        row=cur.fetchone()
        if row!=None:
            recorded_datetime=row[0].split(";")[1]
            p2p_lastupdate=datetime.datetime.strptime(recorded_datetime,"%d/%m/%Y, %H:%M:%S")
            timetowait=180-(now-p2p_lastupdate).seconds
            if timetowait<0: 
                for table in p2p["syncrotablesupd"]:
                    p2p["syncrotablesupd"][table]=True
                p2p["state"]=""
                p2p["message"]="synchro reset"
                p2p["defUpdate"].append("DELETE FROM settings WHERE setting_key='p2p_syncro'")
                p2p["defUpdate"].append("INSERT INTO settings (setting_key,setting_value) VALUES ('p2p_syncro','"+str(p2p["id_instance"])+";"+ts+"')")
                return p2p
        
    #if DEBUG_MODE: log(p2p["state"]+"\t"+datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S"))
    
    if p2p["state"] in ("","A"):
        lock_ok=False
        cur.execute("SELECT setting_value FROM settings WHERE setting_key='p2p_syncro'")
        row=cur.fetchone()
        timetowait=0
        if row==None:
            lock_ok=True
            #cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('p2p_syncro','"+p2p["id_instance"]+";"+ts+"')")
            p2p["defUpdate"].append("INSERT INTO settings (setting_key,setting_value) VALUES ('p2p_syncro','"+str(p2p["id_instance"])+";"+ts+"')")
        else:
            log_ts=False
            if row[0]=="":
                lock_ok=True
                log_ts=True
            else:
                recorded_id_instance=row[0].split(";")[0]
                recorded_datetime=row[0].split(";")[1]
                p2p_lastupdate=datetime.datetime.strptime(recorded_datetime,"%d/%m/%Y, %H:%M:%S")
                if recorded_id_instance==p2p["id_instance"]:
                    timetowait=90-(now-p2p_lastupdate).seconds
                    lock_ok=True
                    if timetowait<0: 
                        log_ts=True
                else:
                    timetowait=180-(now-p2p_lastupdate).seconds
                    #if (now-p2p_lastupdate).seconds>300: 
                    if timetowait<0: 
                        lock_ok=True
                        log_ts=True
            if lock_ok and log_ts:
                p2p["message"]="synchro resync"
                p2p["defUpdate"].append("UPDATE settings SET setting_value='"+str(p2p["id_instance"])+";"+ts+"' WHERE setting_key='p2p_syncro'")
                for table in p2p["syncrotablesupd"]:
                    p2p["syncrotablesupd"][table]=True
                p2p["state"]=""

        if not lock_ok:
            #p2p["message"]="synchro waiting for lock: "+`timetowait`+"s"
            p2p["message"]="synchro standby"
            #p2p_syncroTaskRunning=False
            return p2p
    
    if p2p["state"]=="":
        #reset - begin
        p2p["md5Rows"]=""
        p2p["lstRows"]=""
        #p2p["chkRowsId_tables"]={}
        #p2p["lstRows_tables"]={}
        p2p["deltaRows"]=""
        p2p["deltaContents_index"]=""
        p2p["deltaHeaders"]=""
        p2p["deltaDict"]=""
        
        p2p["partner"]=""
        p2p["partner_state"]=""
        p2p["partner_md5Rows"]=""      
        p2p["partner_lstRows"]=""      
        p2p["partner_deltaRows"]=""
        p2p["partner_deltaContents_index"]=""
        p2p["partner_deltaHeaders"]=""
        p2p["partner_deltaDict"]=""
        #reset - end
        
        #history days
        history_days=""
        cur.execute("select setting_value from settings where setting_key='history_days'")
        row=cur.fetchone()
        if row!=None: 
            history_days=str(row[0])
            intdays=int(history_days)
            if intdays==0:
                history_days=""
            else:
                now = datetime.datetime.utcnow()
                olddate=now-datetime.timedelta(days=intdays)
                strolddate=olddate.strftime("%Y-%m-%d")
        
        #pre-syncro - begin
        if p2p["syncrotablesupd"]["rows"]:
            p2p["syncrotablesupd"]["rows"]=False
            strqy="SELECT * FROM rows_checksum"
            if history_days!="": strqy+=" WHERE status=0 OR date_time>='"+strolddate+"'"
            cur.execute(strqy)
            cols = [i[0] for i in cur.description]; hc={}; i=0
            for col in cols: hc[col]=i; i=i+1
            p2p["chkRowsId"]={}
            arrLstRows=[]
            firstRow=True
            for row in cur:
                if os.path.isfile("quit_synchro.cmd"): 
                    os.unlink("quit_synchro.cmd")
                    p2p["syncrotablesupd"]["rows"]=True
                    return p2p
                if not firstRow: arrLstRows.append("\n")
                firstRow=False
                firstCol=True
                for col in cols:
                    if not firstCol: arrLstRows.append("\t")
                    firstCol=False
                    rawfield=row[hc[col]]
                    if row[hc[col]]!=None:
                        field=str(row[hc[col]])
                    else:
                        field="#NULL#"
                    arrLstRows.append(field)
                # parentFrame.p2p_chkRows[str(row[0])+"-"+str(row[1])+"-"+str(row[2])+"-"+str(row[8])+"-"+str(row[9])+"-"+str(row[10])+"-"+str(row[11])]=True
                p2p["chkRowsId"][str(row[0])+"-"+str(row[1])+"-"+str(row[2])]=str(row[8])+","+str(row[9])+","+str(row[10])+","+str(row[4])
            p2p["lstRows"] = ''.join(arrLstRows)
            p2p["lstRows_tables"]["rows"]=p2p["lstRows"]
            p2p["chkRowsId_tables"]["rows"]=p2p["chkRowsId"]
            p2p["md5Rows_tables"]["rows"] = hashlib.md5(p2p["lstRows_tables"]["rows"]).hexdigest()
        #p2p["md5Rows"] = hashlib.md5(p2p["lstRows_tables"]["rows"]).hexdigest()
        # caricamento lista file attached
        if p2p["syncrotablesupd"]["attach"]:
            p2p["syncrotablesupd"]["attach"]=False
            arrLstRows, p2p["md5Rows_tables"]["attach"] = getFileListMD5(DATABASE_PATH+"attach",DATABASE_PATH)
            if DEBUG_MODE: log(arrLstRows)
            p2p["lstRows_tables"]["attach"]=''.join(arrLstRows)
            p2p["chkRowsId_tables"]["attach"]={}
            for filename in p2p["lstRows_tables"]["attach"].split("\n"):
                p2p["chkRowsId_tables"]["attach"][filename]=True
        # caricamento di tutte le altre tabelle da sincronizzare
        chkRowsId=lstRows={}
        for table in p2p["syncrotables"]:
            if table!="rows" and table!="attach":
                if p2p["syncrotablesupd"][table]:
                    p2p["syncrotablesupd"][table]=False
                    #print "loading "+table
                    cur.execute("SELECT "+p2p["syncrotablesflds"][table]+" FROM "+table+" ORDER BY "+p2p["syncrotablesflds"][table])
                    cols = [i[0] for i in cur.description]; hc={}; i=0
                    for col in cols: hc[col]=i; i=i+1
                    chkRowsId={}
                    arrLstRows=[]
                    firstRow=True
                    for row in cur:
                        strRow=[]
                        if not firstRow: arrLstRows.append("\n")
                        firstRow=False
                        firstCol=True
                        for col in cols:
                            if not firstCol: 
                                strRow.append("\t")
                                arrLstRows.append("\t")
                            firstCol=False
                            rawfield=row[hc[col]]
                            if row[hc[col]]!=None:
                                try:
                                    field=str(row[hc[col]])
                                except:
                                    #print row[hc[col]]
                                    field=row[hc[col]].encode('ascii', 'replace')
                                    #print field
                                #field=str(row[hc[col]])
                                #field=field.replace("\n"," ").replace("\t"," ").strip()
                            else:
                                field="#NULL#"
                            strRow.append(field)
                            field=base64.b64encode(field)
                            arrLstRows.append(field)
                        chkkey=""
                        for keypos in p2p["syncrotableskeys"][table].split(","):
                            if chkkey!="": chkkey=chkkey+"-"
                            chkkey=chkkey+str(row[int(keypos)])
                        chkRowsId[chkkey]=''.join(strRow)
                    lstRows = ''.join(arrLstRows)
                    p2p["chkRowsId_tables"][table]=chkRowsId
                    p2p["lstRows_tables"][table]=lstRows
                    p2p["md5Rows_tables"][table] = hashlib.md5(p2p["lstRows_tables"][table]).hexdigest()
                #p2p["md5Rows"] += "-" + hashlib.md5(p2p["lstRows_tables"][table]).hexdigest()
        #pre-syncro - end
        
        #Timestamp update for big recordsets - Begin
        now = datetime.datetime.utcnow()
        ts=now.strftime("%d/%m/%Y, %H:%M:%S")
        cur.execute("SELECT setting_value FROM settings WHERE setting_key='p2p_syncro'")
        row=cur.fetchone()
        if row!=None:
            recorded_id_instance=row[0].split(";")[0]
            recorded_datetime=row[0].split(";")[1]
            p2p_lastupdate=datetime.datetime.strptime(recorded_datetime,"%d/%m/%Y, %H:%M:%S")
            if recorded_id_instance==p2p["id_instance"]:
                timetowait=60-(now-p2p_lastupdate).seconds
                if timetowait<0: 
                    p2p["defUpdate"].append("UPDATE settings SET setting_value='"+str(p2p["id_instance"])+";"+ts+"' WHERE setting_key='p2p_syncro'")
        #Timestamp update for big recordsets - End

        p2p["md5Rows"]=""
        for table in p2p["syncrotables"]:
            if table=="rows":
                p2p["md5Rows"] = p2p["md5Rows_tables"]["rows"]
            else:
                p2p["md5Rows"] += "-" + p2p["md5Rows_tables"][table]

        # 22/07/2013
        p2p["state"]="A"

    #network settings - begin
    serverurl = p2p["arrSettings"]['heavybase_server']
    projectname = p2p["arrSettings"]['project_name']
    currelease = p2p["arrSettings"]['heavybase_release']
    #print serverurl
    username="heavybase"
    password="heavybase"
    
    #p2p["message"]="connecting to "+serverurl        
    p2p["message"]="connecting..."
    connectionMode=p2p["arrSettings"]['Network_Connection_Mode']
    arrSettings=p2p["arrSettings"]
    curConnMode, p2p_server, res = TryConnect(connectionMode,arrSettings,serverurl,projectname,username,password)
    #network settings - end
    if curConnMode!="": 
        #p2p["lastConnectionUrl"]=serverurl
        if p2p["state"] in ("","A"): 
            #liveupdate - begin
            liveupdate=False
            now = datetime.datetime.utcnow()
            ts=now.strftime("%d/%m/%Y, %H:%M:%S")
            if p2p["liveupdate_ts"]=="": liveupdate=True
            else:
                recorded_datetime=p2p["liveupdate_ts"]
                p2p_lastliveupdate=datetime.datetime.strptime(recorded_datetime,"%d/%m/%Y, %H:%M:%S")
                timetowait=120-(now-p2p_lastliveupdate).seconds
                if timetowait<0: liveupdate=True
            if liveupdate:
                p2p["liveupdate_ts"]=ts
                if DEBUG_MODE: log("liveupdate")
                # check version
                updateserverok=True
                newrelease=res
                res_a=res.split(".")
                if len(res_a)==3:
                    try:
                        res_n=int(res_a[0])*10000+int(res_a[1])*100+int(res_a[2])
                        cur_a=currelease.split(".")
                        cur_n=int(cur_a[0])*10000+int(cur_a[1])*100+int(cur_a[2])
                    except:
                        p2p["message"]="invalid server response - update failed"
                        updateserverok=False
                    if updateserverok:
                        if cur_n>=res_n:
                            p2p["message"]="system update verified."
                        else:
                            p2p["message"]="current release is "+currelease+", new release is "+res+". downloading updates."
                            oldrelease=currelease
                            currelease=res
                            # download updates
                            if DEBUG_MODE:
                                print "checkupdates with project_name = "+projectname+"--"+oldrelease
                                log("checkupdates with project_name = "+projectname+"--"+oldrelease)
                            res = p2p_server.checkUpdates(projectname+"--"+oldrelease,username,password)
                            print "To be downloaded: "+res
                            download_ok=False
                            for filename in res.split("|"):
                                encoded = p2p_server.downloadFile(projectname,username,password,filename)
                                if encoded=="":
                                    p2p["message"]=filename+" not found."
                                else:
                                    data=""
                                    download_ok=False
                                    tryno=0
                                    while tryno<50:
                                        try: 
                                            data = zlib.decompress(base64.b64decode(encoded))
                                            print filename+" downloaded."
                                            download_ok=True
                                            break
                                        except:
                                            print filename+" wrong. file size: "+`len(encoded)`
                                            try: encoded = p2p_server.downloadFile(projectname,username,password,filename)
                                            except: pass
                                            tryno=tryno+1

                                    if download_ok:
                                        if filename.lower().find(".rtf")>0 or filename.lower().find(".abw")>0:
                                            filename=".."+os.path.sep+"reports"+os.path.sep+filename
                                        if filename.lower().find(".xls")>0 or filename.lower().find(".doc")>0 or filename.lower().find(".pdf")>0:
                                            filename=".."+os.path.sep+"doc"+os.path.sep+filename
                                        if filename.lower().find(".sh")>0 or filename.lower().find(".bat")>0 or filename.lower().find(".command")>0:
                                            filename=".."+os.path.sep+filename
                                        outfile = open(filename, 'wb')
                                        outfile.write(data)
                                        outfile.close()
                                        p2p["message"]=filename+" updated."
                                        print filename+" written."
                                        # #zip files handling
                                        # import zipfile
                                        # zfilename=filename
                                        # if zipfile.is_zipfile(zfilename):
                                        #     zfile = zipfile.ZipFile( zfilename, "r" )
                                        #     for info in zfile.infolist():
                                        #         fname = info.filename
                                        #         data = zfile.read(fname)
                                        #         fout = open(fname, "w")
                                        #         fout.write(data)
                                        #         fout.close()
                                        #     os.unlink(zfilename)
                                    else:
                                        print filename+" download canceled."
                                        break
                            #p2p["defUpdate"].append("UPDATE settings SET setting_value='"+currelease+"' WHERE setting_key='heavybase_release'")
                            if download_ok:
                                p2p["arrSettings"]['heavybase_release']=newrelease
                                restart=False
                                try: 
                                    os.unlink("restart.cmd")
                                    restart=True
                                except: pass
                                if restart:
                                    p2p["message"]="system updated. please restart the application"
                                    print "SyncroP2P Stopped"
                                    p2p["syncroTaskStop"]=True
                                    #sys.exit()
                                    return p2p
                                else: 
                                    p2p["message"]="system updated."
                            else:
                                print "system update canceled"
                                p2p["message"]="system update canceled"
                                #p2p["state"]=""
                else:
                    p2p["message"]="invalid server response - update failed"
                    updateserverok=False
                    #p2p["state"]=""
                #liveupdate - end
                if updateserverok: p2p["state"]="A"
            else:
                p2p["state"]="A"
    else: 
        return p2p

    ##chiusura connessione a hub principale e riapertura verso round robin service
    #if curConnMode!="use MSIE": p2p_server=None
    #connectionMode=p2p["arrSettings"]['Network_Connection_Mode']
    #"use xmlrpc"
    #arrSettings=p2p["arrSettings"]
    #curConnMode, p2p_server, res = TryConnect(connectionMode,arrSettings,serverurl,projectname,username,password)
    #if curConnMode=="": return p2p
    #p2p["lastConnectionUrl"]=serverurl
    
    #syncro - begin
    #print p2p["id_instance"]+"-curstate="+p2p["state"]
    encoded=""
    if p2p["state"]=="A":
        p2p["partner"]=""
        encoded = base64.b64encode(zlib.compress(p2p["md5Rows"]))
    elif p2p["state"]=="B":
        encoded = base64.b64encode(zlib.compress(p2p["lstRows"]))
    elif p2p["state"]=="B1":
        encoded = "NOP"
    elif p2p["state"]=="C":
        encoded_a = base64.b64encode(zlib.compress(p2p["deltaRows"]))
        encoded_b = base64.b64encode(zlib.compress(p2p["deltaContents_index"]))
        encoded_c = base64.b64encode(zlib.compress(p2p["deltaHeaders"]))
        encoded = encoded_a + "," + encoded_b + "," + encoded_c
    elif p2p["state"]=="C1":
        encoded = "NOP"
    elif p2p["state"]=="D":
        encoded = base64.b64encode(zlib.compress(p2p["deltaDict"]))
    elif p2p["state"]=="D1":
        encoded = "NOP"
    elif p2p["state"]=="E":
        encoded = "NOP"

    if p2p["state"] in ("A","B","B1","C","C1","D","D1","E"):
        res=",,RES"

        # effettiva release in esecuzione - inizio
        dbrelease="0.0.0"
        cur.execute("select setting_value from settings where setting_key='heavybase_release'")
        row=cur.fetchone()
        if row!=None: dbrelease=row[0]
        # effettiva release in esecuzione - fine
        try:
            p2p["message"]="p2p waiting for connection"
            res = p2p_server.p2p(projectname+"--"+dbrelease, username, password, p2p["id_instance"], p2p["state"], p2p["partner"], encoded)
            p2p["message"]="p2p connected"
        except:
            p2p["message"]="p2p not connected. retrying"
            try:
                res = p2p_server.p2p(projectname+"--"+dbrelease, username, password, p2p["id_instance"], p2p["state"], p2p["partner"], encoded)
                p2p["message"]="p2p connected"
            except:
                p2p["message"]="p2p not connected."

        #Format: PartnerID, PartnerState, PartnerData (only PartnerData encrypted and base64 encoded; can be multiple)
        if len(res.split(","))<3:
            p2p["message"]="Error: "+res
        elif res.split(",")[2]!="NOP" and res.split(",")[2]!="RES":
            p2p["partner"]=res.split(",")[0]
            p2p["partner_state"]=res.split(",")[1]
            #if p2p["partner_state"]=="A" and p2p["state"]=="A":
            if p2p["state"]=="A":
                #qui assumo che i due md5 siano diversi, lo garantisce il server
                err=False
                try: data=zlib.decompress(base64.b64decode(res.split(",")[2]))
                except: err=True
                if not err:
                    arr_partner_md5Rows=data.split("-")
                    arr_md5Rows=p2p["md5Rows"].split("-")
                    # ricerca di quale tabella allineare
                    p2p["cursyncrotable_id"]=0  # per default allineo sempre ROWS
                    for i in range(len(data.split("-"))):
                        #if DEBUG_MODE:
                        #    print arr_md5Rows[i]+" - " +arr_partner_md5Rows[i]
                        if arr_md5Rows[i]!=arr_partner_md5Rows[i]:
                            p2p["cursyncrotable_id"] = i
                            syncrotablename=p2p["syncrotables"][p2p["cursyncrotable_id"]]
                            if DEBUG_MODE:
                                print "Synchronizing "+syncrotablename
                            p2p["lstRows"]         = p2p["lstRows_tables"][syncrotablename]
                            p2p["chkRowsId"]       = p2p["chkRowsId_tables"][syncrotablename]
                            break
                    p2p["partner_md5Rows"]=data
                    p2p["state"]="B"
                else: res=",,RES"
            #elif p2p["partner_state"] in ("B","B1") and p2p["state"] in ("B","B1","C"):
            elif p2p["state"] in ("B","B1"):
                err=False
                try: data=zlib.decompress(base64.b64decode(res.split(",")[2]))
                except: err=True
                if not err:
                    p2p["partner_lstRows"]=data
                    syncrotablename=p2p["syncrotables"][p2p["cursyncrotable_id"]]
                    #if p2p["cursyncrotable_id"]==0:
                    if syncrotablename=="rows":
                        #ROWS,HEADERS,CONTENTS syncronization
                        #differential Rows
                        partner_chkRows={}
                        partner_chkRowsId={}
                        rowsToBeSent={}
                        if len(p2p["partner_lstRows"])>0:
                            for prow in p2p["partner_lstRows"].split("\n"):
                                pcol=prow.split("\t")
                                partner_chkRows[prow]=True
                                partner_chkRowsId[pcol[0]+"-"+pcol[1]+"-"+pcol[2]]=pcol[8]+","+pcol[9]+","+pcol[10]+","+pcol[4]
                        dRows=[]
                        if len(p2p["lstRows"])>0:
                            dRowsLimit=100
                            dRowsCount=0
                            firstRow=True
                            for prow in p2p["lstRows"].split("\n"):
                                if dRowsCount>=dRowsLimit: break
                                pcol=prow.split("\t")
                                if not partner_chkRows.has_key(prow):
                                    addrow=False
                                    if not partner_chkRowsId.has_key(pcol[0]+"-"+pcol[1]+"-"+pcol[2]): 
                                        addrow=True
                                    else:
                                        if partner_chkRowsId[pcol[0]+"-"+pcol[1]+"-"+pcol[2]]!=pcol[8]+","+pcol[9]+","+pcol[10]+","+pcol[4]:
                                            date_time=partner_chkRowsId[pcol[0]+"-"+pcol[1]+"-"+pcol[2]].split(",")[3]
                                            my_date_time=pcol[4]
                                            status=partner_chkRowsId[pcol[0]+"-"+pcol[1]+"-"+pcol[2]].split(",")[0]
                                            my_status=pcol[0]
                                            if date_time=="": date_time="''"
                                            if my_date_time=="": my_date_time="''"
                                            if date_time.find("#NULL#")<0:
                                                if date_time.find(":")<0 or date_time.find("-")<0 or date_time.find(" ")<0: date_time="''"
                                                if date_time[0]!="'": date_time="'"+date_time
                                                if date_time[-1]!="'": date_time=date_time+"'"
                                            if my_date_time.find("#NULL#")<0:
                                                if my_date_time.find(":")<0 or my_date_time.find("-")<0 or my_date_time.find(" ")<0: my_date_time="''"
                                                if my_date_time[0]!="'": my_date_time="'"+my_date_time
                                                if my_date_time[-1]!="'": my_date_time=my_date_time+"'"
                                            if date_time < my_date_time: addrow = True
                                            elif date_time == my_date_time:
                                                if int(status) == 0:
                                                    if int(my_status) != 0: addrow = True
                                                    elif int(p2p["partner"]) < int(p2p["id_instance"]): addrow = True
                                                else:
                                                    if int(my_status) != 0:
                                                        if int(p2p["partner"]) < int(p2p["id_instance"]): addrow = True
                                            #addrow=True
                                    if addrow:
                                        key=pcol[0]+"-"+pcol[1]+"-"+pcol[2]
                                        if key not in rowsToBeSent:
                                            rowsToBeSent[key]=True
                                            dRowsCount=dRowsCount+1
                                            if not firstRow: dRows.append("\n")
                                            firstRow=False
                                            dRows.append(prow)
                        p2p["deltaRows"]=''.join(dRows)
                        #if DEBUG_MODE:
                        #    print "sending "+str(dRowsCount)+" records"
                        #    print "deltaRows = "+p2p["deltaRows"]
                        #differential Contents_index & list of id_dictionary
                        #partner_chkDict={}
                        p2p["keysDict"]={}
                        lstIdHeaders=[]
                        dContents_index=[]
                        firstRow=True
                        for keyrow in rowsToBeSent:
                            # print "keyrow = "+keyrow
                            arr=keyrow.split("-")
                            cur.execute("SELECT id_row,id_user,id_instance,id_header,id_dictionary,id_cycle FROM contents_index WHERE id_row="+arr[0]+" AND id_user="+arr[1]+" AND id_instance="+arr[2])
                            cols = [i[0] for i in cur.description]; hc={}; i=0
                            for col in cols: hc[col]=i; i=i+1
                            for row in cur:
                                if not firstRow: 
                                    dContents_index.append("\n")
                                    lstIdHeaders.append(",")
                                firstRow=False
                                firstCol=True
                                for col in cols:
                                    if not firstCol: dContents_index.append("\t")
                                    firstCol=False
                                    dContents_index.append(str(row[hc[col]]))
                                p2p["keysDict"][str(row[hc["id_header"]])+"-"+str(row[hc["id_dictionary"]])]=True
                                lstIdHeaders.append(str(row[hc["id_header"]]))
                        p2p["deltaContents_index"]=''.join(dContents_index)
                        # print "deltaContents_index = "+p2p["deltaContents_index"]
                        strIdHeaders=''.join(lstIdHeaders)
                        #differential Headers
                        dHeaders=[]
                        cur.execute("SELECT id_header,child_of,description,label,pos,hap,stype,id_section,deftype,cyclic,doubleinput,typedesc,subtypedesc,textlength,subtextlength,textalign,freetext,defaultvalue,validation,onchange,date_time,status FROM headers WHERE id_header in ("+strIdHeaders+")")
                        cols = [i[0] for i in cur.description]; hc={}; i=0
                        for col in cols: hc[col]=i; i=i+1
                        firstRow=True
                        for row in cur:
                            if not firstRow: 
                                dHeaders.append("\n")
                            firstRow=False
                            firstCol=True
                            for col in cols:
                                if not firstCol: dHeaders.append("\t")
                                firstCol=False
                                text=""
                                if col not in ("defaultvalue","validation","onchange"):
                                    if row[hc[col]]!=None:
                                        text=str(row[hc[col]])
                                    else:
                                        text="#NULL#"
                                else:
                                    if row[hc[col]]!=None:
                                        if isinstance(text, (int, long, float, complex)): 
                                            text=str(row[hc[col]])
                                        else:
                                            text=row[hc[col]]
                                    else:
                                        text="#NULL#"
                                    text=base64.b64encode(text)
                                dHeaders.append(text)
                        p2p["deltaHeaders"]=''.join(dHeaders)
                        #differential Dictionary
                        dDict=[]
                        cur.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary WHERE id_header in ("+strIdHeaders+")")
                        cols = [i[0] for i in cur.description]; hc={}; i=0
                        for col in cols: hc[col]=i; i=i+1
                        firstRow=True
                        for row in cur:
                            #if p2p["keysDict"].has_key(row["id_header"]+"-"+row["id_dictionary"]) and not partner_chkDict.has_key(row["id_header"]+"-"+row["id_dictionary"]):
                            if p2p["keysDict"].has_key(str(row[hc["id_header"]])+"-"+str(row[hc["id_dictionary"]])):
                                if not firstRow: 
                                    dDict.append("\n")
                                firstRow=False
                                firstCol=True
                                for col in cols:
                                    if not firstCol: dDict.append("\t")
                                    firstCol=False
                                    dDict.append(str(row[hc[col]]))
                        p2p["deltaDict"]=''.join(dDict)
                    elif syncrotablename=="attach":
                        #differential Files
                        #print "sue"
                        #print p2p["partner_lstRows"]
                        partner_chk={}
                        filesToBeSent={}
                        if len(p2p["partner_lstRows"])>0:
                            for pfile in p2p["partner_lstRows"].split("\n"):
                                partner_chk[pfile]=True                        
                        dFileNames=[]
                        dFiles=[]
                        if len(p2p["lstRows"])>0:
                            dFilesLimit=5
                            dFilesCount=0
                            firstFile=True
                            #print "mie"
                            #print p2p["lstRows"]
                            for pfile in p2p["lstRows"].split("\n"):
                                if pfile!="":
                                    if dFilesCount>=dFilesLimit: break
                                    if not partner_chk.has_key(pfile):
                                        #if not firstFile: dFiles.append("\n")
                                        #firstFile=False
                                        #lettura file - inizio
                                        #print "to be sent: "+pfile
                                        f_in = open(pfile, 'rb')
                                        content_in = f_in.read()
                                        f_in.close()
                                        #dFiles.append([base64.b64encode(pfile),"\t",base64.b64encode(content_in)])
                                        dFileNames.append(base64.b64encode(pfile))
                                        dFiles.append(base64.b64encode(content_in))
                                        #lettura file - fine
                                        filesToBeSent[pfile]=True
                                        dFilesCount=dFilesCount+1
                        p2p["deltaRows"]='\n'.join(dFileNames)
                        p2p["deltaContents_index"]=''
                        p2p["deltaHeaders"]=''
                        p2p["deltaDict"]='\n'.join(dFiles)
                    else:
                        #Other tables syncronization
                        #differential Rows
                        #print "p2p[\"cursyncrotable_id\"]="+`p2p["cursyncrotable_id"]`
                        syncrotablename=p2p["syncrotables"][p2p["cursyncrotable_id"]]
                        partner_chkRows={}
                        partner_chkRowsId={}
                        rowsToBeSent={}
                        if len(p2p["partner_lstRows"])>0:
                            for prow in p2p["partner_lstRows"].split("\n"):
                                pcol=prow.split("\t")
                                partner_chkRows[prow]=True
                                chkkey=""
                                for keypos in p2p["syncrotableskeys"][syncrotablename].split(","):
                                    if chkkey!="": chkkey=chkkey+"-"
                                    chkkey=chkkey+pcol[int(keypos)]
                                partner_chkRowsId[chkkey]=prow
                        dRows=[]
                        if len(p2p["lstRows"])>0:
                            dRowsLimit=100
                            dRowsCount=0
                            firstRow=True
                            for prow in p2p["lstRows"].split("\n"):
                                if dRowsCount>=dRowsLimit: break
                                pcol=prow.split("\t")
                                if not partner_chkRows.has_key(prow):
                                    chkkey=""
                                    for keypos in p2p["syncrotableskeys"][syncrotablename].split(","):
                                        if chkkey!="": chkkey=chkkey+"-"
                                        chkkey=chkkey+pcol[int(keypos)]
                                    addrow=False
                                    if not partner_chkRowsId.has_key(chkkey): 
                                        addrow=True
                                    else:
                                        if partner_chkRowsId[chkkey]!=prow:
                                            addrow=True
                                    if addrow:
                                        rowsToBeSent[chkkey]=True
                                        dRowsCount=dRowsCount+1
                                    
                                        if not firstRow: dRows.append("\n")
                                        firstRow=False
                                        dRows.append(prow)
                        p2p["deltaRows"]=''.join(dRows)
                        #print p2p["deltaRows"]
                        p2p["deltaContents_index"]=''
                        p2p["deltaHeaders"]=''
                        p2p["deltaDict"]=''
                    p2p["state"]="C"
                else: res=",,RES"
            #elif p2p["partner_state"] in ("C","C1") and p2p["state"] in ("C","C1","D"):
            elif p2p["state"] in ("C","C1"):
                if len(res.split(","))!=5:
                    print "SyncroP2P Error: state="+p2p["state"]+" - partner_state="+p2p["partner_state"]
                    print "Received: \n"+res
                    #if len(res.split(","))>=3: print zlib.decompress(base64.b64decode(res.split(",")[2]))
                    #if len(res.split(","))>=4: print zlib.decompress(base64.b64decode(res.split(",")[3]))
                    #if len(res.split(","))>=5: print zlib.decompress(base64.b64decode(res.split(",")[4]))
                        
                err=False
                try:
                    if len(res.split(","))>=3: 
                        p2p["partner_deltaRows"]=zlib.decompress(base64.b64decode(res.split(",")[2]))
                        #if DEBUG_MODE: print "\nReceived partner_deltaRows: \n"+p2p["partner_deltaRows"]
                    if len(res.split(","))>=4: 
                        p2p["partner_deltaContents_index"]=zlib.decompress(base64.b64decode(res.split(",")[3]))
                        #if DEBUG_MODE: print "\nReceived partner_deltaContents_index: \n"+p2p["partner_deltaContents_index"]
                    if len(res.split(","))>=5: 
                        p2p["partner_deltaHeaders"]=zlib.decompress(base64.b64decode(res.split(",")[4]))
                        #if DEBUG_MODE: print "\nReceived partner_deltaHeaders: \n"+p2p["partner_deltaHeaders"]
                except: err=True
                if not err:
                    p2p["state"]="D"
                else: res=",,RES"
            #elif p2p["partner_state"] in ("D","D1") and p2p["state"] in ("D","D1"):
            elif p2p["state"] in ("D","D1"):
                err=False
                try: data=zlib.decompress(base64.b64decode(res.split(",")[2]))
                except: 
                    err=True
                    if DEBUG_MODE:
                        print "Error Receiving data from state D."
                if not err:
                    #if DEBUG_MODE:
                    #    print "Received:\n"+data
                    p2p["partner_deltaDict"]=data
                    p2p["defUpdate"].append("BEGIN TRANSACTION")
                    syncrotablename=p2p["syncrotables"][p2p["cursyncrotable_id"]]
                    if syncrotablename=="rows":
                        #ROWS,HEADERS,CONTENTS syncronization
                        #registrazione modifiche (ordine: headers,dictionary,contents_index,rows)
                        #headers
                        if len(p2p["deltaHeaders"])>0:
                            arr_partner_deltaHeaders=p2p["deltaHeaders"].split("\n")
                            for lstrow in arr_partner_deltaHeaders:
                                lst=lstrow.split("\t")
                                cur.execute("SELECT id_header,child_of,description FROM headers WHERE id_header="+lst[0]+" AND child_of="+lst[1]+" AND description='"+lst[2]+"'")
                                row=cur.fetchone()
                                if row==None:
                                    idx=0
                                    normlst=[]
                                    for elm in lst:
                                        if idx<=16: 
                                            normlst.append(ifelse(elm!="#NULL#", lambda:"'"+elm+"'", lambda:"null"))
                                        else:
                                            tmp=base64.b64decode(elm)
                                            normlst.append(ifelse(tmp!="#NULL#", lambda:"'"+GetSqlValue(tmp)+"'", lambda:"null"))
                                        idx=idx+1
                                    strlst=""
                                    for elm in normlst:
                                        if strlst!="": strlst=strlst+","
                                        strlst=strlst+elm
                                    #id_header,child_of,description,label,pos,hap,stype,id_section,deftype,cyclic,doubleinput,typedesc,subtypedesc,textlength,subtextlength,textalign,freetext,defaultvalue,validation,onchange=[lst[0],lst[1],lst[2],lst[3],lst[4],lst[5],lst[6],lst[7],lst[8],lst[9],lst[10],lst[11],lst[12],lst[13],lst[14],lst[15],lst[16],GetSqlValue(base64.b64decode(lst[17])),GetSqlValue(base64.b64decode(lst[18])),GetSqlValue(base64.b64decode(lst[19]))]
                                    #p2p["defUpdate"].append("INSERT INTO headers (id_header,child_of,description,label,pos,hap,stype,id_section,deftype,cyclic,doubleinput,typedesc,subtypedesc,textlength,subtextlength,textalign,freetext,defaultvalue,validation,onchange) VALUES ("+id_header+","+child_of+",'"+description+"','"+label+"',"+pos+",'"+hap+"','"+stype+"',"+id_section+",'"+deftype+"','"+cyclic+"','"+doubleinput+"','"+typedesc+"','"+subtypedesc+"',"+textlength+","+subtextlength+",'"+textalign+"','"+freetext+"','"+defaultvalue+"','"+validation+"','"+onchange+"')")
                                    p2p["defUpdate"].append("INSERT INTO headers (id_header,child_of,description,label,pos,hap,stype,id_section,deftype,cyclic,doubleinput,typedesc,subtypedesc,textlength,subtextlength,textalign,freetext,defaultvalue,validation,onchange,date_time,status) VALUES ("+strlst+")")
                        #dict
                        hashDict={}
                        partnerDict={}
                        if len(p2p["partner_deltaDict"])>0:
                            #max_idHeadDict={}
                            #cur.execute("SELECT id_header,max(id_dictionary) AS maxIdDict FROM contents_dictionary GROUP BY id_header")
                            #for row in cur:
                                #max_idHeadDict[`row[0]`]=row[1]
                            # Correttore di sicurezza per il caso multiutente sullo stesso DB - Inizio
                            cur.execute("SELECT id_header,max(id_dictionary) AS maxIdDict FROM contents_dictionary GROUP BY id_header")
                            for row in cur:
                                if p2p["max_idHeadDict"].has_key(str(row[0])):
                                    if p2p["max_idHeadDict"][str(row[0])]<row[1]:
                                        p2p["max_idHeadDict"][str(row[0])]=row[1]
                                else:
                                    p2p["max_idHeadDict"][str(row[0])]=row[1]
                            # Correttore di sicurezza per il caso multiutente sullo stesso DB - Fine
                            max_idHeadDict=p2p["max_idHeadDict"]
                            
                            cur.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary")
                            for row in cur:
                                hashDict[str(row[0])+"-"+row[2]]=row[1]
                            
                            arr_partner_deltaDict=p2p["partner_deltaDict"].split("\n")
                            queue=[]
                            for lstrow in arr_partner_deltaDict:
                                lst=lstrow.split("\t")
                                if len(lst)>2:
                                    if not hashDict.has_key(lst[0]+"-"+lst[2]):
                                        if max_idHeadDict.has_key(lst[0]):
                                            max_idHeadDict[lst[0]]=max_idHeadDict[lst[0]]+1
                                        else:
                                            max_idHeadDict[lst[0]]=1
                                        #p2p["defUpdate"].append("INSERT INTO contents_dictionary (id_header,id_dictionary,data) VALUES ("+lst[0]+","+`max_idHeadDict[lst[0]]`+",'"+lst[2]+"')")
                                        if len(queue)>=500:
                                            p2p["defUpdate"].append(''.join(queue))
                                            queue=[]
                                        if len(queue)==0: queue.append("INSERT INTO contents_dictionary (id_header,id_dictionary,data) SELECT "+lst[0]+","+str(max_idHeadDict[lst[0]])+",'"+lst[2]+"'")
                                        else: queue.append(" UNION SELECT "+lst[0]+","+str(max_idHeadDict[lst[0]])+",'"+lst[2]+"'")
                                        hashDict[lst[0]+"-"+lst[2]]=max_idHeadDict[lst[0]]
                                    partnerDict[lst[0]+"-"+lst[1]]=lst[2]
                            p2p["max_idHeadDict"]=max_idHeadDict
                            if len(queue)>0: p2p["defUpdate"].append(''.join(queue))
                        #contents_index
                        if len(p2p["partner_deltaContents_index"])>0:
                            arr_partner_deltaContents_index=p2p["partner_deltaContents_index"].split("\n")
                            queue=[]
                            for lstrow in arr_partner_deltaContents_index:
                                lst=lstrow.split("\t")
                                if partnerDict.has_key(lst[3]+"-"+lst[4]):
                                    if hashDict.has_key(lst[3]+"-"+partnerDict[lst[3]+"-"+lst[4]]):
                                        id_dictionary=hashDict[lst[3]+"-"+partnerDict[lst[3]+"-"+lst[4]]]
                                        #cur.execute("INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary) VALUES ("+lst[0]+","+lst[1]+","+lst[2]+","+`id_dictionary`+")")
                                        cur.execute("SELECT * FROM contents_index WHERE id_row="+lst[0]+" AND id_user="+lst[1]+" AND id_instance="+lst[2]+" AND id_header="+lst[3]+" AND id_cycle="+lst[5])
                                        row=cur.fetchone()
                                        if row==None:
                                            #p2p["defUpdate"].append("INSERT INTO contents_index (id_row,id_user,id_instance,id_header,id_dictionary,id_cycle) VALUES ("+lst[0]+","+lst[1]+","+lst[2]+","+lst[3]+","+`id_dictionary`+","+lst[5]+")")
                                            if len(queue)>=500:
                                                p2p["defUpdate"].append(''.join(queue))
                                                queue=[]
                                            if len(queue)==0: queue.append("INSERT INTO contents_index (id_row,id_user,id_instance,id_header,id_dictionary,id_cycle) SELECT "+lst[0]+","+lst[1]+","+lst[2]+","+lst[3]+","+str(id_dictionary)+","+lst[5])
                                            else: queue.append(" UNION SELECT "+lst[0]+","+lst[1]+","+lst[2]+","+lst[3]+","+str(id_dictionary)+","+lst[5])
                            if len(queue)>0: p2p["defUpdate"].append(''.join(queue))
                        #rows
                        if len(p2p["partner_deltaRows"])>0:
                            arr_partner_deltaRows=p2p["partner_deltaRows"].split("\n")
                            for lstrow in arr_partner_deltaRows:
                                lst=lstrow.split("\t")
                                #print "processing: "+str(lst)
                                idx=0
                                normlst=[]
                                for elm in lst:
                                    if idx<(len(lst)-1):
                                        normlst.append(ifelse(elm!="#NULL#", lambda:"'"+elm+"'", lambda:"null"))
                                    idx=idx+1
                                strlst=""
                                for elm in normlst:
                                    if strlst!="": strlst=strlst+","
                                    strlst=strlst+elm
                                anyerror=False
                                id_row,id_user,id_instance,id_header,date_time,rap,hpath,id_locking,status,status_user,status_instance,rec=[lst[0],lst[1],lst[2],lst[3],lst[4],lst[5],lst[6],lst[7],lst[8],lst[9],lst[10],lst[11]]
                                try:
                                    id_row=str(int(ifelse(id_row!="#NULL#", lambda:"0"+id_row, lambda:"0")))
                                    id_user=str(int(ifelse(id_user!="#NULL#", lambda:"0"+id_user, lambda:"0")))
                                    id_instance=str(int(ifelse(id_instance!="#NULL#", lambda:"0"+id_instance, lambda:"0")))
                                    id_header=str(int(ifelse(id_header!="#NULL#", lambda:"0"+id_header, lambda:"0")))
                                    id_locking=str(int(ifelse(id_locking!="#NULL#", lambda:"0"+id_locking, lambda:"0")))
                                    status=str(int(ifelse(status!="#NULL#", lambda:status, lambda:"0")))
                                    status_user=str(int(ifelse(status_user!="#NULL#", lambda:status_user, lambda:"0")))
                                    status_instance=str(int(ifelse(status_instance!="#NULL#", lambda:status_instance, lambda:"0")))
                                except: anyerror=True
                                #if DEBUG_MODE:
                                #    if anyerror: 
                                #        print "Error!"
                                #        print id_row,id_user,id_instance,id_header,date_time,rap,hpath,id_locking,status,status_user,status_instance,rec
                                #        print strlst
                                if not anyerror:
                                    #if DEBUG_MODE:
                                    #    print id_row,id_user,id_instance,id_header,date_time,rap,hpath,id_locking,status,status_user,status_instance,rec
                                    #id_row,id_user,id_instance=[lst[0],lst[1],lst[2]]
                                    if not p2p["chkRowsId"].has_key(id_row+"-"+id_user+"-"+id_instance):
                                        #p2p["defUpdate"].append("INSERT INTO rows (id_row,id_user,id_instance,id_header,date_time,rap,hpath,id_locking,status,status_user,status_instance) VALUES ("+id_row+","+id_user+","+id_instance+","+id_header+",'"+date_time+"','"+rap+"','"+hpath+"','"+id_locking+"',"+status+","+status_user+","+status_instance+")")
                                        sql="INSERT INTO rows (id_row,id_user,id_instance,id_header,date_time,rap,hpath,id_locking,status,status_user,status_instance) VALUES ("+strlst+")"
                                        #if DEBUG_MODE: print sql
                                        p2p["defUpdate"].append(sql)
                                    else:
                                        #arrVal=parentFrame.p2p_chkRows[id_row+"-"+id_user+"-"+id_instance].split(",")
                                        arrVal=p2p["chkRowsId"][id_row+"-"+id_user+"-"+id_instance].split(",")
                                        my_status=arrVal[0]
                                        my_status_user=arrVal[1]
                                        my_status_instance=arrVal[2]
                                        my_date_time=arrVal[3]
                                        doupdate=False
                                        if (status!=my_status or status_user!=my_status_user or status_instance!=my_status_instance or date_time!=my_date_time):
                                            if date_time=="": date_time="''"
                                            if my_date_time=="": my_date_time="''"
                                            if date_time!="#NULL#":
                                                if date_time.find(":")<0 or date_time.find("-")<0 or date_time.find(" ")<0: date_time="''"
                                                if date_time[0]!="'": date_time="'"+date_time
                                                if date_time[-1]!="'": date_time=date_time+"'"
                                            orig_my_date_time=my_date_time
                                            if my_date_time!="#NULL#":
                                                if my_date_time.find(":")<0 or my_date_time.find("-")<0 or my_date_time.find(" ")<0: my_date_time="''"
                                                if my_date_time[0]!="'": my_date_time="'"+my_date_time
                                                if my_date_time[-1]!="'": my_date_time=my_date_time+"'"
                                            #if DEBUG_MODE: print "confronto tra ["+date_time+"] e ["+my_date_time+"], status="+str(status)+", mystatus="+str(my_status)+", instance="+p2p["partner"]+", my_instance="+p2p["id_instance"]
                                            if date_time > my_date_time: doupdate = True
                                            elif date_time == my_date_time:
                                                if int(my_status) == 0:
                                                    if int(status) != 0: doupdate = True
                                                    elif int(p2p["partner"]) > int(p2p["id_instance"]): doupdate = True
                                                else:
                                                    #if int(status) == 0: doupdate = True
                                                    #elif int(p2p["partner"]) > int(p2p["id_instance"]): doupdate = True
                                                    if int(status) != 0:
                                                        if int(p2p["partner"]) > int(p2p["id_instance"]): doupdate = True
                                            if doupdate:
                                                if my_status=="": my_status="''"
                                                if my_status_user=="": my_status_user="''"
                                                if my_status_instance=="": my_status_instance="''"
                                                qydate_time=ifelse(date_time!="#NULL#", lambda:date_time, lambda:"null")
                                                qymy_date_time=ifelse(my_date_time!="#NULL#", lambda:"= '"+orig_my_date_time+"'", lambda:"is null")
                                                sql="UPDATE rows SET status="+status+", status_user="+status_user+", status_instance="+status_instance+", date_time="+qydate_time+" WHERE id_row="+id_row+" AND id_user="+id_user+" AND id_instance="+id_instance+" AND status="+my_status+" AND status_user="+my_status_user+" AND status_instance="+my_status_instance+" AND date_time "+qymy_date_time
                                                #sql="UPDATE rows SET status="+status+", status_user="+status_user+", status_instance="+status_instance+", date_time=date_time||' -> '||"+qydate_time+" WHERE id_row="+id_row+" AND id_user="+id_user+" AND id_instance="+id_instance+" AND status="+my_status+" AND status_user="+my_status_user+" AND status_instance="+my_status_instance+" AND date_time "+qymy_date_time
                                                #if DEBUG_MODE: print sql
                                                p2p["defUpdate"].append(sql)
                    elif syncrotablename=="attach":
                        #print p2p["partner_deltaRows"]
                        if len(p2p["partner_deltaRows"])>0:
                            arr_partner_deltaRows=p2p["partner_deltaRows"].split("\n")
                            arr_partner_deltaDict=p2p["partner_deltaDict"].split("\n")
                            i=0
                            for elm in arr_partner_deltaRows:
                                filename=DATABASE_PATH+base64.b64decode(elm)
                                # separator: always unix like
                                filecontents=base64.b64decode(arr_partner_deltaDict[i])
                                attach_dir=filename[:filename.rfind("/")]
                                try: os.makedirs(attach_dir)
                                except: pass
                                fout = open(filename,"wb")
                                fout.write(filecontents)
                                fout.close()
                                i=i+1
                    else:
                        #Other tables syncronization
                        if len(p2p["partner_deltaRows"])>0:
                            syncrotablename=p2p["syncrotables"][p2p["cursyncrotable_id"]]
                            #print syncrotablename
                            #print p2p["partner_deltaRows"]
                            arr_partner_deltaRows=p2p["partner_deltaRows"].split("\n")
                            for lstrow in arr_partner_deltaRows:
                                #print str(lstrow)
                                good=False
                                if lstrow.find("\t")>0: good=True
                                prelst=lstrow.split("\t")
                                lst=[]
                                for elm in prelst: lst.append(base64.b64decode(elm))
                                if good and len(lst)>2:
                                    date_time = lst[-2]
                                    status    = lst[-1]
                                    chkkey=""
                                    for keypos in p2p["syncrotableskeys"][syncrotablename].split(","):
                                        if chkkey!="": chkkey=chkkey+"-"
                                        chkkey=chkkey+lst[int(keypos)]
                                    if not p2p["chkRowsId"].has_key(chkkey):
                                        normlst=[]
                                        for elm in lst:
                                            normlst.append(ifelse(elm!="#NULL#", lambda:"'"+GetSqlValue(elm)+"'", lambda:"null"))
                                        strlst=""
                                        for elm in normlst:
                                            if strlst!="": strlst=strlst+","
                                            strlst=strlst+elm
                                        p2p["defUpdate"].append("INSERT INTO "+syncrotablename+" ("+p2p["syncrotablesflds"][syncrotablename]+") VALUES ("+strlst+")")
                                    else:
                                        arrVal=p2p["chkRowsId"][chkkey].split("\t")
                                        my_date_time = arrVal[-2]
                                        my_status    = arrVal[-1]
                                        #print "my_date_time="+my_date_time+" --- date_time="+date_time
                                        doupdate=False
                                        if (status!=my_status or date_time!=my_date_time):
                                            if date_time=="": date_time="''"
                                            if my_date_time=="": my_date_time="''"
                                            if date_time!="#NULL#":
                                                if date_time.find(":")<0 or date_time.find("-")<0 or date_time.find(" ")<0: date_time="''"
                                                if date_time[0]!="'": date_time="'"+date_time
                                                if date_time[-1]!="'": date_time=date_time+"'"
                                            orig_my_date_time=my_date_time
                                            if my_date_time!="#NULL#":
                                                if my_date_time.find(":")<0 or my_date_time.find("-")<0 or my_date_time.find(" ")<0: my_date_time="''"
                                                if my_date_time[0]!="'": my_date_time="'"+my_date_time
                                                if my_date_time[-1]!="'": my_date_time=my_date_time+"'"
                                            if date_time > my_date_time: doupdate = True
                                            elif date_time == my_date_time:
                                                if my_status=="#NULL#": my_status="0"
                                                if status=="#NULL#": status="0"
                                                if int(my_status) == 0:
                                                    if int(status) != 0: doupdate = True
                                                    elif int(p2p["partner"]) > int(p2p["id_instance"]): doupdate = True     # caso impossibile, tenuto x simmetria con rows
                                                else:
                                                    #if int(status) != 0 and int(p2p["partner"]) > int(p2p["id_instance"]): doupdate = True
                                                    if int(status) != 0:
                                                        if int(p2p["partner"]) > int(p2p["id_instance"]): doupdate = True
                                            if doupdate:
                                                if my_status=="": my_status="''"
                                                sql="UPDATE "+syncrotablename+" SET "
                                                strlst=""
                                                i=0
                                                for fld in p2p["syncrotablesflds"][syncrotablename].split(","):
                                                    iskey=False
                                                    for keypos in p2p["syncrotableskeys"][syncrotablename].split(","):
                                                        if int(keypos)==i: iskey=True
                                                    if not iskey:
                                                        if strlst!="": strlst=strlst+", "
                                                        #print str(i)+" --- "+str(lst)
                                                        tmp=""
                                                        try: tmp=str(lst[i])
                                                        except: pass
                                                        val=ifelse(tmp!="#NULL#", lambda:"'"+GetSqlValue(tmp)+"'", lambda:"null")
                                                        #val="'"+tmp+"'"
                                                        strlst=strlst+fld+"="+val
                                                    i=i+1
                                                sql=sql+strlst+" WHERE "
                                                arrFld=p2p["syncrotablesflds"][syncrotablename].split(",")
                                                strlst=""
                                                for keypos in p2p["syncrotableskeys"][syncrotablename].split(","):
                                                    if strlst!="": strlst=strlst+" AND "
                                                    my_val=arrVal[int(keypos)]
                                                    qymy_val=ifelse(my_val!="#NULL#", lambda:"= '"+my_val+"'", lambda:"is null")
                                                    strlst=strlst+arrFld[int(keypos)]+" "+qymy_val
                                                #qymy_date_time=ifelse(my_date_time!="#NULL#", lambda:"= '"+my_date_time+"'", lambda:"is null")
                                                qymy_date_time=ifelse(my_date_time!="#NULL#", lambda:"= '"+orig_my_date_time+"'", lambda:"is null")
                                                qymy_status=ifelse(my_status!="#NULL#", lambda:"= '"+my_status+"'", lambda:"is null")     # caso si spera impossibile, tenuto x simmetria
                                                sql=sql+strlst+" AND status "+qymy_status+" AND date_time "+qymy_date_time
                                                #print sql
                                                p2p["defUpdate"].append(sql)
                    p2p["defUpdate"].append("COMMIT TRANSACTION")
                    p2p["syncrotablesupd"][syncrotablename]=True
                    p2p["state"]="E"
                else: res=",,RES"
        elif res.split(",")[2]=="NOP":
            if p2p["state"]=="A":
                #p2p["state"]=""
                pass
            if p2p["state"]=="B":
                p2p["state"]="B1"
            if p2p["state"]=="C":
                p2p["state"]="C1"
            if p2p["state"]=="D":
                p2p["state"]="D1"
            if p2p["state"]=="E":
                if p2p["partner_state"]=="E":
                    p2p["state"]=""
                    p2p["message"]="Syncronization completed"
                else:
                    p2p["message"]="waiting the other end for completing the syncronization"
        
        if len(res.split(","))<3: p2p["state"]=""
        elif res.split(",")[2] == "RES": p2p["state"]=""

        #if p2p["state"]=="":
        #    #cur.execute("UPDATE settings SET setting_value='' WHERE setting_key='p2p_syncro'")
        #    p2p["defUpdate"].append("UPDATE settings SET setting_value='' WHERE setting_key='p2p_syncro'")

    #if connectionMode!="use MSIE" or os.name!='nt':
    if curConnMode!="use MSIE": p2p_server=None

#    print p2p["id_instance"]+"-newstate="+p2p["state"]

    #cur.close()
    #con.close()
    return p2p
    #syncro - end


def md5_for_file(filename):
    block_size=2**20
    f=file(filename,'rb')
    md5 = hashlib.md5()
    while True:
        data = f.read(block_size)
        if not data:
            break
        md5.update(data)
    return md5.hexdigest()

def getFileListMD5(curpath, DATABASE_PATH=""):
    block_size=2**20
    lst=getFileList(curpath)
    import hashlib
    md5 = hashlib.md5()
    for elm in lst:
        f=file(elm,'rb')
        while True:
            data = f.read(block_size)
            if not data:
                break
            md5.update(data)
    strlst='\n'.join(lst)
    if DATABASE_PATH!="": strlst=strlst.replace(DATABASE_PATH,"")
    strlst=strlst.replace(os.path.sep,"/")
    return strlst, md5.hexdigest()
# SYNCRO FS - End

# CRC Components - Begin
class CrcAlgorithm:
    def __init__(self, width, polynomial, name=None, seed=0,
                 lsbFirst=False, lsbFirstData=None, xorMask=0):
        if width > 0:
            try:
                polyMask = long(polynomial)
            except TypeError:
                polynomial = list(polynomial)
                polynomial.sort()
                polynomial.reverse()
                polynomial = tuple(polynomial)
            else:
                if lsbFirst:
                    polyMask = crc_reflect(polyMask, width)
                polynomial = (width,)
                for i in range(width-1,-1,-1):
                    if (polyMask >> i) & 1:
                        polynomial += (i,)
            if polynomial[:1] != (width,):
                ValueError("mismatch between width and polynomial degree")
        self.width = width
        self.polynomial = polynomial
        self.name = name
        self.seed = seed
        self.lsbFirst = lsbFirst
        self.lsbFirstData = lsbFirstData
        self.xorMask = xorMask
        if not hasattr(width, "__rlshift__"):
            raise ValueError
    def __repr__(self):
        info = ""
        if self.name is not None:
            info = ' "%s"' % str(self.name)
        result = "<%s.%s%s @ %#x>" % (
            self.__class__.__module__,
            self.__class__.__name__,
            info, id(self))
        return result
    def calcString(self, s, value=None):
        r = CrcRegister(self, value)
        r.takeString(s)
        return r.getFinalValue()
    def calcWord(self, word, width, value=None):
        r = CrcRegister(self, value)
        r.takeWord(word, width)
        return r.getFinalValue()
    def crc_reflect(self):
        ca = CrcAlgorithm(0, 0)
        ca._initFromOther(self)
        ca.lsbFirst = not self.lsbFirst
        if self.lsbFirstData is not None:
            ca.lsbFirstData = not self.lsbFirstData
        if ca.name:
            ca.name += " reflected"
        return ca
    def reverse(self):
        ca = CrcAlgorithm(0, 0)
        ca._initFromOther(self)
        ca.polynomial = [ (self.width - e) for e in self.polynomial ]
        ca.polynomial.sort()
        ca.polynomial.reverse()
        ca.polynomial = tuple(ca.polynomial)
        if ca.name:
            ca.name += " reversed"
        return ca
    def _initFromOther(self, other):
        self.width = other.width
        self.polynomial = other.polynomial
        self.name = other.name
        self.seed = other.seed
        self.lsbFirst = other.lsbFirst
        self.lsbFirstData = other.lsbFirstData
        self.xorMask = other.xorMask
class CrcRegister:
    def __init__(self, crcAlgorithm, value=None):
        self.crcAlgorithm = crcAlgorithm
        p = crcAlgorithm
        self.bitMask = (1 << p.width) - 1
        word = 0
        for n in p.polynomial:
            word |= 1 << n
        self.polyMask = word & self.bitMask
        if p.lsbFirst:
            self.polyMask = crc_reflect(self.polyMask, p.width)
        if p.lsbFirst:
            self.inBitMask = 1 << (p.width - 1)
            self.outBitMask = 1
        else:
            self.inBitMask = 1
            self.outBitMask = 1 << (p.width - 1)
        if p.lsbFirstData is not None:
            self.lsbFirstData = p.lsbFirstData
        else:
            self.lsbFirstData = p.lsbFirst
        self.reset()
        if value is not None:
            self.value = value ^ p.xorMask
    def __str__(self):
        return crc_formatBinaryString(self.value, self.crcAlgorithm.width)
    def reset(self):
        self.value = long(self.crcAlgorithm.seed)
    def takeBit(self, bit):
        outBit = ((self.value & self.outBitMask) != 0)
        if self.crcAlgorithm.lsbFirst:
            self.value >>= 1
        else:
            self.value <<= 1
        self.value &= self.bitMask
        if outBit ^ bool(bit):
            self.value ^= self.polyMask
    def takeWord(self, word, width=8):
        if self.lsbFirstData:
            bitList = range(0,width)
        else:
            bitList = range(width-1,-1,-1)
        for n in bitList:
            self.takeBit((word >> n) & 1)
    def takeString(self, s):
        for c in s:
            self.takeWord(ord(c))
    def getValue(self):
        return self.value
    def getFinalValue(self):
        p = self.crcAlgorithm
        return self.value ^ p.xorMask
def crc_reflect(value, width):
    return sum(
        ((value >> x) & 1) << (width - 1 - x)
        for x in range(width))
def crc_formatBinaryString(value, width):
    return "".join(
        "01"[(value >> i) & 1] for i in range(width-1,-1,-1))
#: Used in RFC-2440 and MIL STD 188-184.
CRC24 = CrcAlgorithm(
    name         = "CRC-24",
    width        = 24,
    polynomial   = (24, 23, 18, 17, 14, 11, 10, 7, 6, 5, 4, 3, 1, 0),
    seed         = 0xB704CE,
    lsbFirst     = False,
    xorMask      = 0)
# CRC Components - End

def StringCRC(txt):
    import zlib,string
    return string.rjust(str(zlib.crc32(txt.strip().lower()) & 0xffffffff),10,"0")

def LongCRC(txt):
    import zlib
    return zlib.crc32(txt.strip().lower()) & 0xffffffff

def xor_crypt_string(data, key):
    from itertools import izip, cycle
    return ''.join(chr(ord(x) ^ ord(y)) for (x,y) in izip(data, cycle(key)))

def UniqueLongCRC(txt,cur):
    #longCRC=LongCRC(txt)
    longCRC=CRC24.calcString(txt)
    crcfound=True
    while crcfound:
        crcfound=False
        cur.execute("SELECT id_header FROM headers WHERE id_header="+str(longCRC))
        if cur.fetchone()!=None:
            longCRC=longCRC+1
            crcfound=True
    return longCRC

class GeneralTranscodeDialog(wx.Dialog):
    def __init__(self, parent, title, idCrf, lstfields, lstsizes, recordset):
        super(GeneralTranscodeDialog, self).__init__(parent=parent, 
            title=title, size=(900, 400))

        self.parent=parent
        self.recordset=recordset
        self.lstfields=lstfields
        
        panel = wx.Panel(self)

        vbox = wx.BoxSizer(wx.VERTICAL)

        sb = wx.StaticBox(panel)
        sbs = wx.StaticBoxSizer(sb, orient=wx.VERTICAL)

        hbox1 = wx.BoxSizer(wx.HORIZONTAL)
        self.fields={}
        for i in range(len(lstfields)):
            vbox1 = wx.BoxSizer(wx.VERTICAL)
            lbl = wx.StaticText(panel, label=lstfields[i])
            self.fields[lstfields[i]] = wx.TextCtrl(panel, size=wx.Size(lstsizes[i], 25))
            self.fields[lstfields[i]].Bind(wx.EVT_KEY_UP, self.OnFieldEnterPressed)
            vbox1.Add(lbl, border=5)
            vbox1.Add(self.fields[lstfields[i]], border=5)
            hbox1.Add(vbox1, border=5)
        sbs.Add(hbox1, border=5)

        self.btnFind = wx.Button(panel, label=TT("Search"))
        self.btnFind.Bind(wx.EVT_BUTTON, self.OnFind)
        sbs.Add(self.btnFind, border=5)

        self.lstFound = wx.ListBox(panel, style=wx.HSCROLL)
        self.lstFound.SetFont(wx.Font( 10, wx.MODERN, wx.NORMAL, wx.NORMAL))
        self.lstFound.Bind(wx.EVT_LISTBOX_DCLICK, self.OnConfirm)
        sbs.Add(self.lstFound, proportion=1, flag=wx.ALL|wx.EXPAND, border=5)

        panel.SetSizer(sbs)
       
        hbox4 = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(self, label='Ok')
        closeButton = wx.Button(self, label=TT("Cancel"))
        hbox4.Add(okButton)
        hbox4.Add(closeButton, flag=wx.LEFT, border=5)

        vbox.Add(panel, proportion=1, flag=wx.ALL|wx.EXPAND, border=5)
        vbox.Add(hbox4, flag= wx.ALIGN_CENTER|wx.TOP|wx.BOTTOM, border=10)

        self.SetSizer(vbox)
        
        okButton.Bind(wx.EVT_BUTTON, self.OnConfirm)
        closeButton.Bind(wx.EVT_BUTTON, self.OnClose)

        self.arrfound=[]
        
    def OnFieldEnterPressed(self, e):
        if e.GetKeyCode() in (wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER):
            self.OnFind(e)
        
    def OnFind(self, e):
        wx.BeginBusyCursor()
        import string
        self.lstFound.Clear()
        
        self.arrfound=[]
        for row in self.recordset:
            found=False
            atleast1=False
            for i in range(len(self.fields)):
                pattern=self.fields[self.lstfields[i]].GetValue().lower().strip().encode('ascii', 'replace')
                if pattern!="":
                    atleast1=True
                    if pattern in row[i].lower().strip().encode('ascii', 'replace'): 
                        found=True
                        break
            if found or (atleast1==False):
                self.arrfound.append(row)
                
        orderBy=2
        self.arrfound.sort(key=lambda x:(x[orderBy:]))
        maxsize={}
        for elm in self.arrfound:
            for col in range(0,len(elm)):
                if not maxsize.has_key(col): maxsize[col]=0
                if maxsize[col]<len(elm[col]):
                    maxsize[col]=len(elm[col])
        for elm in self.arrfound:
            outrow=[]
            for col in range(0,len(elm)):
                outrow.append(elm[col].ljust(maxsize[col]))
            self.lstFound.Append(" | ".join(outrow))
            
        wx.EndBusyCursor()
            
    def OnClose(self, e):
        self.Destroy()

    def OnConfirm(self, e):
        self.parent.res=self.arrfound[self.lstFound.GetSelections()[0]]
        self.Destroy()

    
class ExcelTranscodeDialog(wx.Dialog):
    def __init__(self, parent, title, filename, idxpage, idxcode, idxdesc):
        super(ExcelTranscodeDialog, self).__init__(parent=parent, 
            title=title, size=(600, 400))

        self.parent=parent
        
        self.filename=filename
        self.idxpage=idxpage
        self.idxcode=idxcode
        self.idxdesc=idxdesc
        
        panel = wx.Panel(self)

        vbox = wx.BoxSizer(wx.VERTICAL)

        #sb = wx.StaticBox(panel, label=TT("Search"))
        sb = wx.StaticBox(panel)
        sbs = wx.StaticBoxSizer(sb, orient=wx.VERTICAL)

        hbox1 = wx.BoxSizer(wx.HORIZONTAL)        
        self.lblCode = wx.StaticText(panel, label=TT("Code"))
        self.txtCode = wx.TextCtrl(panel, size=wx.Size(100, 25))
        self.txtCode.Bind(wx.EVT_KEY_UP, self.OnFieldEnterPressed)
        self.lblDesc = wx.StaticText(panel, label=TT("Description"))
        self.txtDesc = wx.TextCtrl(panel, size=wx.Size(360, 25))
        self.txtDesc.Bind(wx.EVT_KEY_UP, self.OnFieldEnterPressed)
        hbox1.Add(self.lblCode)
        hbox1.Add(self.txtCode)
        hbox1.Add(self.lblDesc)
        hbox1.Add(self.txtDesc)
        sbs.Add(hbox1, border=5)

        #hbox2 = wx.BoxSizer(wx.HORIZONTAL)
        self.btnFind = wx.Button(panel, label=TT("Search"))
        self.btnFind.Bind(wx.EVT_BUTTON, self.OnFind)
        #hbox2.Add(self.btnFind)
        sbs.Add(self.btnFind, border=5)

        self.lstFound = wx.ListBox(panel, style=wx.HSCROLL)
        self.lstFound.SetFont(wx.Font( 10, wx.MODERN, wx.NORMAL, wx.NORMAL))
        self.lstFound.Bind(wx.EVT_LISTBOX_DCLICK, self.OnConfirm)
        sbs.Add(self.lstFound, proportion=1, flag=wx.ALL|wx.EXPAND, border=5)

        panel.SetSizer(sbs)
       
        hbox4 = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(self, label='Ok')
        closeButton = wx.Button(self, label=TT("Cancel"))
        hbox4.Add(okButton)
        hbox4.Add(closeButton, flag=wx.LEFT, border=5)

        vbox.Add(panel, proportion=1, flag=wx.ALL|wx.EXPAND, border=5)
        vbox.Add(hbox4, flag= wx.ALIGN_CENTER|wx.TOP|wx.BOTTOM, border=10)

        self.SetSizer(vbox)
        
        okButton.Bind(wx.EVT_BUTTON, self.OnConfirm)
        closeButton.Bind(wx.EVT_BUTTON, self.OnClose)

        self.arrfound=[]
        self.book = xlrd.open_workbook(self.filename)
        self.sh=self.book.sheet_by_index(self.idxpage-1)
        
    def OnFieldEnterPressed(self, e):
        if e.GetKeyCode() in (wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER):
            self.OnFind(e)
        
    def OnFind(self, e):
        wx.BeginBusyCursor()
        import string
        self.lstFound.Clear()
        code=self.txtCode.GetValue().lower().strip()
        desc=self.txtDesc.GetValue().lower().strip()
        self.arrfound=[]
        for y in range(1,self.sh.nrows):
            shcode=str(self.sh.cell_value(y,self.idxcode-1)).strip()
            shdesc=self.sh.cell_value(y,self.idxdesc-1).encode('ascii', 'replace').strip()
            if code in shcode.lower() and desc in shdesc.lower():
                self.arrfound.append((shcode,shdesc))
        maxcodesize=0
        for elm in self.arrfound:
            if maxcodesize<len(elm[0]): maxcodesize=len(elm[0])
        for elm in self.arrfound:
            self.lstFound.Append(string.ljust(elm[0],maxcodesize)+"  "+elm[1])
        wx.EndBusyCursor()
            
    def OnClose(self, e):
        self.Destroy()

    def OnConfirm(self, e):
        self.parent.res=self.arrfound[self.lstFound.GetSelections()[0]]
        self.Destroy()

def create(parent):
    #03/09/2008
    #return HeavyBaseFrame(parent)
    return ParentFrame(parent)

class ParentFrame(wx.aui.AuiMDIParentFrame):
    def __init__(self, parent):
        wx.aui.AuiMDIParentFrame.__init__(self, parent, -1,
                                          title=PROGNAME,
                                          size=(wx.Size(SCREEN_X, SCREEN_Y)),
                                          style=wx.DEFAULT_FRAME_STYLE)

        print "Loading "+PROGNAME+" "+RELEASE+" user interface."
        self.SetBackgroundColour(wx.Colour(240,240,240))
        import os
        if os.name=='nt' and (IsRunningFromRemovable()==False):
            self.Bind(wx.EVT_CLOSE, self.OnHide)
        else:
            self.Bind(wx.EVT_CLOSE, self.OnClose)

        import base64
        icon_b64 = ""
        if PROGNAME.lower()!="p2pdb":
            if LoadCustomSetting("demo_mode")!="1":
                icon_b64 += "AAABAAEAICAEAAAAAADoAgAAFgAAACgAAAAgAAAAQAAAAAEABAAAAAAAAAIAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACAAACAAAAAgIAAgAAAAIAAgACAgAAAwMDAAICAgAAAAP8AAP8AAAD//wD/AAAA/wD/AP//AAD///8AzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzMzO7MzM7szMzMzMzO7MzO7MzMzM7szM7szMzMzMzMzMzMzMzMzM7MzMzMzMzO7szMzMzMzMzszMzszM7OzMzMzO7szMzuzMzuzMzMzMzO7O4RER7uERERHMzOzO7AARERERERERERERHuzu7MABERERERERERERERzAAAAAERERERERERERERERAAAAARERERAREREREREREQAAABEREQEREAEAAAEREREQAAAREBERd393d3d3ERAQEAABEREXd3/3d3d//3ERERAAARF3d3//93d///93d3AAAB"
                icon_b64 += "AHd////3d////3d3d3cAEAd3////93////d3////AQAHf/H//3f////3f///8BAAARER//93////93////AAAAB3dxERf/////9////wAAAHd3d//xEREP//ER//8AAAB3d///AHd3EREXcRERAAAAd////wB3d3f/93d38AAAAA////AHd////3f///AAAAABER//B3////9///8AAAAAAAAREQd////wERH/8AAAAAAAEAAA////8AABERAAAAAAABmQABERD/AAAQAAAAAAAAAAAAAAAREQAAEAAAAAAAAAAAAAAAEAAAABmQAAAAAAAAAAAAABmQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA+AAAAfACAAHwEN8A4IAACuAAAADgAAAD7AAAANgAAAC4AAABeAAAAfwAAAH4ABAB+AYAAfgGAAP8DAAD/AQAB/+EAgP/vgPD/44T3///w9///9/H///H/w=="
            else:
                icon_b64 += "AAABAAEAICD/AAAAAACoCAAAFgAAACgAAAAgAAAAQAAAAAEACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD///8AgICAAAAAAAAAAIAAAICAAMDAwAAAAP8AAP//AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                icon_b64 += "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                icon_b64 += "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAQEBAQEBAgICAgICAgICAgMCAgAAAAAAAAAAAAECAgICAgICAgMDAwMDAwMCAgICAgICAgMCAAAA"
                icon_b64 += "AAABAgICAgICAwMDAwMDAgICBAMDAgICAgICAgIBAAAAAQICAgICAwMCBAQEBAQEBAQEBAQCAwMDAgICAgIAAAABAgICAgIDAwQEBAQEBAQEBAQEBAQEAwMDAgICAgAAAAAAAQIDAwQEBAQEBAQEBAQEBAQEBAQEBAMDAgICAAAAAAAAAwIEBAQEBAQEBAQEBAQEBAQEBAQEAgMCAgIAAAAAAAADBAQEBAQEBAQEBAQEBAQCAwQEBAQEAgMCAAAAAAAAAwMEBAQEBAQEBAQEBAICAwMDBAQEBAQCAwMAAAAAAAADAgQEBAQEBAQEAgMEBAQBAQEDAwQEBAQDAwAAAAAAAAMDBAQEBAQEBAQDBAEBAQEBAQQDBAQEBAMDAAAAAAAAAwIEBAQEBAQEBAQBBAQEBAQEBAEDAgQEBAIDAAAAAAADAgQEBAQEBAQEAQEEBAQEBAQEBAQCBAQEBAMAAAAAAAIEBAQEBAQEBAQEBAQEBAQEBAQEAQQEBAQEAwAA"
                icon_b64 += "AAAAAgQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQDAAAAAAICAgICAgIEBAQEBAQEBAQEBAQEBAQEBAQEBAMAAAAAAgIBBQUBAQICBAQEBAQEBAQEBAQEBAQEBAQEAwAAAAACAgEBAQUFBQUCAgMCBAQEBAQBAQEEBAQEBAQDAAYGBgYCAgICAgIBAQEBAgQCAgICAwIEAQEBBAQEBAMGBwUFBwYGBAIEBAICAQIDAwMAAAAAAwMDBAQEAgQEAwYGBwcHBgYGBAQEAgICAgMAAAAAAAAAAAMDAwQEBAQAAAAGBgcHBgYGAgQCAgMDAAAAAAAAAAAAAAAEBAQEAwAAAAAGBgYGBgYCBAICAwMAAAAAAAAAAAAAAAACBAIDAAAAAAYGBgYEAgQEBAIDAwAAAAAAAAAAAAAAAAMEAwAAAAAAAwYGBAICBAQEAgMDAAAAAAAAAAAAAAAAAwIDAAAAAAMEAQIAAAAEBAQEAwMAAAAAAAAAAAAAAAAAAwAAAAAA"
                icon_b64 += "AwAFAQMAAAIEBAIDAwAAAAAAAAAAAAAAAAACAAAAAAAAAQUDAQUDBAQCAgMDAAAAAAAAAAAAAAAAAAAAAAAAAAADAgQCAgQEBAICAwMAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAgIEAgICAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMDAgICAwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
        else:
            if LoadCustomSetting("demo_mode")!="1":
                icon_b64 += "AAABAAEAMDD/AAAAAACoDgAAFgAAACgAAAAwAAAAYAAAAAEACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAmZmZALu7uwDMzMwA3d3dACIiIgBmZmYA7u7uAP///wCqqqoAREREAMz//wCZ/8wAZsyZAGbMzACZzMwAERERAMz/zABm/8wAAMyZAAD/zACZzJkAAMzMADP/zAAAzGYAAJlmADPMmQAzzMwAAP+ZADMzMwCZ//8AiIiIAAD//wAAmZkAZv//ADP//wAA/2YAAGYzAAAARAAAMzMAAGb/AAAzmQAzM2YAZpmZAABmZgAAmcwAM2b/AGZmzAAzmf8AZpnMAABmzABmmf8AADP/AAAAMwAzZpkAAAAiAABmmQAz/5kAADNmAGb/mQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                icon_b64 += "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                icon_b64 += "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQID"
                icon_b64 += "AwQEBAQDAwMDAgIBBQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABgIDBAcICAgIBwMDAwMDAwMDAwkKAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADAwcICAgIBwsMDQ4ODg0NDwIDAwQECRAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMECAgIBxERCxITFBQUFBMTExMNFQQHBAMFAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQICAcRDBELDBYUFBcXFBQTExMYGRoEBwcDCgAAAAAAAAAAAAAAAAAAAAAAAAADBAcICwwRERELGxYUFBQUFBQUHBMTGRkNAwcEAx0AAAAAAAAAAAAAAAAAAAAAAAMEBwgLEQsLCwseFBQUFBQUFBQUFBwTExMaDgMEBAIAAAAAAAAAAAAAAAAAAAAAHwMEBwsDAxELERENExMWFCAUFBQUFxwTExMaDg4EBAQfAAAAAAAAAAAAAAAAAAAAAwQEBxECAwMRCwsh"
                icon_b64 += "ISEhGRMgIBQUFxQTExMaDg4PBAQCAAAAAAAAAAAAAAAAAAADAwQECwsEBAQLCx4gICAgFhkZEyAgFBQWExwcEg4OBAQEHwAAAAAAAAAAAAAAAAADBAQHCwsLAgQLCyIgIxQhExQUJBgWIBQUExQUEhISDwQEAgAAAAAAAAAAAAAAAAADBAQLCwsLEQQLCyMgIBQgFCETIBMZGRYUFBYbEhISDwQEAgAAAAAAAAAAAAAAAAADBAQLCwsLCwQLCyMgFCEUICAUIRMUEyUZExQTFxISEgMDAwAAAAAAAAAAAAAAAAAEBAQLCwsLCwsECyMgFhMZEyAgIBMZExYZJRkUFxISEgMDAwAAAAAAAAAAAAAAAAAEAwQLCwsLCwsRCyMTExQUExkTICAUExkcEyYnGhISEg8DAwAAAAAAAAAAAAAAAAAEAwQLCwsLCwsLDxYTFBQUFBQTGRMUFBQcGSgpKisSEgMEAwAAAAAAAAAAAAAAAAAEAwQLCwsLCwsLDywW"
                icon_b64 += "IBQUIBQUFhkZExQULSgoLi8OEgQEAwAAAAAAAAAAAAAAAAADAwMLCwsLCwsLCyElEyAUFxQgFBQTJRkhKCgoMDESEhEEAwAAAAAAAAAAAAAAAAADBAMLCwsLCwsLCxcXGSwWIBQWICAgFBkyKCgoMzESEhEEAwAAAAAAAAAAAAAAAAADBAMRCwsLCwsLCxcXFxoZGSAgExYgFDIoKCgoMw4SEgQEAwAAAAAAAAAAAAAAAAADAwMDCwsLCwsLCxcTHBwUGCUZExMZKTQ0KCguMRISDAQEAwAAAAAAAAAAAAAAAAAAAwQDCwsLERERCxITExMcFBQZJSU1Mi0hISE2DRISEQQDAwAAAAAAAAAAAAAAAAAAAwMDAwsMERERERETExwcHBQUHBMnNzgWFiwNORoMBAQDAAAAAAAAAAAAAAAAAAAAAAMDAxELEQwMDAsSExMcHBwcHBwcGSc6LBo7DQ8EBAMDAAAAAAAAAAAAAAAAAAAAAAMDAwMHCAsLCwsL"
                icon_b64 += "ExMcHBwcExMTHBMZDQwEBAQEBAMAAAAAAAAAAAAAAAAAAAAAAAADAwMDBwgICAgIBxUPDw8PDw8PAgIDBwcEBAQEAwAAAAAAAAAAAAAAAAAAAAAAAAAAAwMDAwQICAgICAQCAgICAgICAgIEBAQEBAMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMEBwgICAgEAgICAgICAgQHBAQEAwMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADAwMDAwQHCAgIBwICAgIDBAcHBAQDAwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMDBwgICAMCAgIEBwQDAwMAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQEBwgICAICAgIDBwQDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQEBwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQE"
                icon_b64 += "BwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIECAcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBwgICAICAgIECAcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQH"
                icon_b64 += "BwgICAICAgIECAcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwQHBAQEAwkBAQkDBwgDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAwMDAwQEAwQDAwMDAgIJAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABwQIBwQDAgICAwQHCAcCAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAQEBAMCAgICAwQEBwcDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMDAwMDAwMDAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA/+AAAAH/AAD/wAAAAD8AAP+AAAAADwAA/wAAAAAHAAD+AAAAAAMAAPwAAAAAAQAA/gAAAAABAAD8AAAAAAAAAPwA"
                icon_b64 += "AAAAAAAA/AAAAAAAAAB4AAAAAAAAADgAAAAAAAAAOAAAAAABAAD4AAAAAAEAAPAAAAAAAwAA8AAAAAAHAAD4AAAAAA8AAPgAAAAAPwAA+AAAAAD/AAD4AAAAAP8AAPgAAAAA/wAA+AAAAAD/AAD8AAAAAP8AAPwAAAAB/wAA/gAAAAH/AAD+AAAAA/8AAP8AAAAH/wAA/4AAAA//AAD/wAAAH/8AAP/gAAA//wAA//gAAP//AAD//gAD//8AAP/+AAP//wAA//4AA///AAD//gAD//8AAP/+AAP//wAA//4AA///AAD//gAD//8AAP/+AAP//wAA//4AA///AAD//gAD//8AAP/+AAP//wAA//4AA///AAD//gAD//8AAP/+AAP//wAA//4AA///AAD//wAD//8AAP///////wAA"
            else:
                icon_b64 += "AAABAAEAICD/AAAAAACoCAAAFgAAACgAAAAgAAAAQAAAAAEACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAD///8AgICAAAAAAAAAAIAAAICAAMDAwAAAAP8AAP//AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                icon_b64 += "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
                icon_b64 += "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABAQEBAQEBAgICAgICAgICAgMCAgAAAAAAAAAAAAECAgICAgICAgMDAwMDAwMCAgICAgICAgMCAAAA"
                icon_b64 += "AAABAgICAgICAwMDAwMDAgICBAMDAgICAgICAgIBAAAAAQICAgICAwMCBAQEBAQEBAQEBAQCAwMDAgICAgIAAAABAgICAgIDAwQEBAQEBAQEBAQEBAQEAwMDAgICAgAAAAAAAQIDAwQEBAQEBAQEBAQEBAQEBAQEBAMDAgICAAAAAAAAAwIEBAQEBAQEBAQEBAQEBAQEBAQEAgMCAgIAAAAAAAADBAQEBAQEBAQEBAQEBAQCAwQEBAQEAgMCAAAAAAAAAwMEBAQEBAQEBAQEBAICAwMDBAQEBAQCAwMAAAAAAAADAgQEBAQEBAQEAgMEBAQBAQEDAwQEBAQDAwAAAAAAAAMDBAQEBAQEBAQDBAEBAQEBAQQDBAQEBAMDAAAAAAAAAwIEBAQEBAQEBAQBBAQEBAQEBAEDAgQEBAIDAAAAAAADAgQEBAQEBAQEAQEEBAQEBAQEBAQCBAQEBAMAAAAAAAIEBAQEBAQEBAQEBAQEBAQEBAQEAQQEBAQEAwAA"
                icon_b64 += "AAAAAgQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQDAAAAAAICAgICAgIEBAQEBAQEBAQEBAQEBAQEBAQEBAMAAAAAAgIBBQUBAQICBAQEBAQEBAQEBAQEBAQEBAQEAwAAAAACAgEBAQUFBQUCAgMCBAQEBAQBAQEEBAQEBAQDAAYGBgYCAgICAgIBAQEBAgQCAgICAwIEAQEBBAQEBAMGBwUFBwYGBAIEBAICAQIDAwMAAAAAAwMDBAQEAgQEAwYGBwcHBgYGBAQEAgICAgMAAAAAAAAAAAMDAwQEBAQAAAAGBgcHBgYGAgQCAgMDAAAAAAAAAAAAAAAEBAQEAwAAAAAGBgYGBgYCBAICAwMAAAAAAAAAAAAAAAACBAIDAAAAAAYGBgYEAgQEBAIDAwAAAAAAAAAAAAAAAAMEAwAAAAAAAwYGBAICBAQEAgMDAAAAAAAAAAAAAAAAAwIDAAAAAAMEAQIAAAAEBAQEAwMAAAAAAAAAAAAAAAAAAwAAAAAA"
                icon_b64 += "AwAFAQMAAAIEBAIDAwAAAAAAAAAAAAAAAAACAAAAAAAAAQUDAQUDBAQCAgMDAAAAAAAAAAAAAAAAAAAAAAAAAAADAgQCAgQEBAICAwMAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAgIEAgICAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMDAgICAwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAMDAwMDAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
            
        icon = base64.b64decode(icon_b64)
        fout = open("ship.ico","wb")
        fout.write(icon)
        fout.close()
        ICON = wx.Icon('ship.ico', wx.BITMAP_TYPE_ICO)
        self.SetIcon(ICON)
        #os.unlink("ship.ico")

        mb = self.MakeMenuBar()
        self.mb=mb
        self.SetMenuBar(self.mb)
        #self.CreateStatusBar()
        self.statusbar = wx.StatusBar(self, -1)
        self.statusbar.SetFieldsCount(3)
        self.statusbar.SetStatusWidths([100,400,-1])
        self.SetStatusBar(self.statusbar)
        self.progbar = ProgressStatusBar(self, self.statusbar, 2, 0, 8)

        #self.Fit()
        self.SetMinSize(self.GetSize())
        if sys.platform[:3] == 'win':
            self.Maximize(True)

        self.arrSettings={}

        # Database initialization - Begin
        cur.execute('BEGIN TRANSACTION')
        dbrelease="0.0.0"
        try:
            cur.execute("SELECT setting_value FROM settings WHERE setting_key='heavybase_release'")
            row = cur.fetchone()
            if row!=None: dbrelease=row[0]
            if len(dbrelease.split("."))!=3: dbrelease="0.0.0"
        except: pass
        arr_dbrelease=[int(dbrelease.split(".")[0]),int(dbrelease.split(".")[1]),int(dbrelease.split(".")[2])]
        if arr_dbrelease<[2,4,0]:
            if arr_dbrelease<[2,2,0]:
                try:
                    cur.execute('CREATE TABLE groups (id_group INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, group_level INTEGER, member_of INTEGER, shortcode TEXT, description TEXT, acl TEXT)')
                    cur.execute('CREATE TABLE users (id_user INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_group INTEGER NOT NULL DEFAULT 0, uap TEXT, stype TEXT, rights INTEGER, username TEXT NOT NULL, password TEXT NOT NULL, fullname TEXT, notes TEXT, home TEXT, lastaccess TEXT, firstaccess TEXT)')
                    cur.execute('CREATE TABLE eforms (id_eform INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_header INTEGER, pos INTEGER, description TEXT, href TEXT)')
                    cur.execute('CREATE TABLE headers (id_header INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, child_of INTEGER NOT NULL, description TEXT NOT NULL, label TEXT, pos INTEGER NOT NULL DEFAULT 0, hap TEXT, stype TEXT, id_section INTEGER, deftype INTEGER, cyclic INTEGER, doubleinput INTEGER, typedesc TEXT, subtypedesc TEXT, textlength TEXT, subtextlength TEXT, textalign TEXT, freetext INTEGER, defaultvalue TEXT, validation TEXT, onchange TEXT)')
                    #cur.execute('CREATE TABLE rows (id_row INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER, rap TEXT, hpath TEXT)')
                    cur.execute('CREATE TABLE rows (id_row INTEGER NOT NULL, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER NOT NULL, rap TEXT, hpath TEXT)')
                    #cur.execute('CREATE TABLE contents (id_content INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_row INTEGER NOT NULL, id_header INTEGER NOT NULL, data BLOB, id_user INTEGER NOT NULL)')
                    cur.execute('CREATE TABLE contents (id_row INTEGER NOT NULL, id_header INTEGER NOT NULL, data BLOB, id_user INTEGER NOT NULL)')
                    cur.execute('CREATE TABLE transcode (id INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, child_of INTEGER, code TEXT, description)')
                    cur.execute('CREATE TABLE attachments (id_attachment INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_row INTEGER NOT NULL, id_user INTEGER NOT NULL, filename TEXT, contenttype TEXT, size TEXT, timestamp TEXT, checksum TEXT, stream BLOB)')
                    cur.execute('CREATE TABLE logbook (id_log INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_session TEXT, id_user INTEGER NOT NULL, logon TEXT, logoff TEXT, host TEXT)')
                    cur.execute('CREATE TABLE logstat (id_logstat INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_log INTEGER, timestamp TEXT, statement BLOB)')
                    cur.execute('CREATE TABLE settings (id_setting INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, setting_key TEXT, setting_value TEXT)')
                    cur.execute('CREATE TABLE reports_profiles (id_profile INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, description TEXT, ylabel TEXT, xlabel TEXT, outcome TEXT, report_type TEXT, id_header INTEGER)')
                    cur.execute('CREATE TABLE reports_details (id_detail INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_profile INTEGER, inout TEXT, variable_name TEXT, variable_value TEXT, variable_label TEXT, id_header INTEGER)')
                except: pass
                try:
                    cur.execute('ALTER TABLE headers ADD onchange TEXT')
                    cur.execute('ALTER TABLE reports_profiles ADD report_type TEXT')
                except: pass
                try:
                    cur.execute('CREATE TABLE exports(id_export INTEGER NOT NULL PRIMARY KEY AUTOINCREMENT, id_user INTEGER NOT NULL, timestamp TEXT, filename TEXT, synced TEXT)')
                except: pass
                #Distributed version
                try:
                    cur.execute("SELECT max(id_content) FROM contents")
                    # se arrivo qui' il formato delle tabelle rows e contents e' quello vecchio
                    try: 
                        cur.execute('DROP TABLE rows_new')
                    except: pass
                    try: 
                        cur.execute('DROP TABLE contents_new')
                    except: pass
                    cur.execute('CREATE TABLE rows_new (id_row INTEGER NOT NULL, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER NOT NULL, rap TEXT, hpath TEXT)')
                    cur.execute('CREATE TABLE contents_new (id_row INTEGER NOT NULL, id_header INTEGER NOT NULL, data BLOB, id_user INTEGER NOT NULL)')
                    cur.execute('INSERT INTO rows_new (id_row,id_header,status,date_time,id_user,rap,hpath) SELECT id_row,id_header,status,date_time,id_user,rap,hpath FROM rows')
                    cur.execute('INSERT INTO contents_new (id_row,id_header,data,id_user) SELECT id_row,id_header,data,id_user FROM contents')
                    cur.execute('ALTER TABLE rows RENAME TO rows_old')
                    cur.execute('ALTER TABLE contents RENAME TO contents_old')
                    cur.execute('ALTER TABLE rows_new RENAME TO rows')
                    cur.execute('ALTER TABLE contents_new RENAME TO contents')
                    cur.execute('DROP TABLE rows_old')
                    cur.execute('DROP TABLE contents_old')

                    #cur.execute('CREATE INDEX idx_headers ON headers (id_header,pos,description)')
                    #cur.execute('CREATE INDEX idx_rows ON rows (id_row, id_user)')
                    #cur.execute('CREATE INDEX idx_contents ON contents (id_row, id_user)')
                except:
                    try:
                        cur.execute('DROP TABLE rows_old')
                        cur.execute('DROP TABLE contents_old')
                    except: pass
                try:
                    cur.execute('ALTER TABLE reports_profiles ADD id_header INTEGER')
                    cur.execute('ALTER TABLE reports_details ADD id_header INTEGER')
                except: pass

                #V.2.1.0
                try:
                    cur.execute('ALTER TABLE rows ADD status_user INTEGER')
                    cur.execute('ALTER TABLE rows ADD id_locking INTEGER')
                    cur.execute('CREATE TABLE locking (id_locking INTEGER NOT NULL, id_user INTEGER NOT NULL, locking_date_time TEXT, locking_description TEXT, unlocking_date_time TEXT, unlocking_description TEXT, status INTEGER)')
                    cur.execute('CREATE TABLE queries (id_query INTEGER NOT NULL, id_row INTEGER NOT NULL, id_user INTEGER NOT NULL, problem_datetime TEXT, problem_description TEXT, solution_datetime TEXT, solution_description TEXT, query_no TEXT, outcome_cod TEXT, status INTEGER)')
                except: pass
                try:
                    #Dictionary - Begin
                    cur.execute('CREATE TABLE contents_index (id_row INTEGER NOT NULL, id_user INTEGER NOT NULL, id_header INTEGER NOT NULL, id_dictionary INTEGER NOT NULL)')
                    cur.execute('CREATE TABLE contents_dictionary (id_header INTEGER NOT NULL, id_dictionary INTEGER NOT NULL, data BLOB)')

                    wx.MessageBox("Il sistema sta per procedere alla conversione del database ad un nuovo formato.\nL'operazione puo' richiedere da alcuni secondi ad alcuni minuti.\nSi prega di non interrompere il processo e attendere la pagina di Login.", "Update", wx.ICON_INFORMATION | wx.OK, None)
                    #chiave di crittografia standard x tutti gli studi vecchi
                    cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('do_encrypt','heavybase')")
                except: pass

                cur.execute("SELECT id_row,id_user,id_header,data FROM contents")
                for row in cur:
                    contents_id_row=row[0]
                    contents_id_user=row[1]
                    contents_id_header=row[2]
                    contents_data=row[3]
                    #data dictionary - Begin
                    id_dictionary=0
                    cur2 = con.cursor()
                    cur2.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary WHERE id_header="+str(contents_id_header)+" AND data='"+GetSqlValue(contents_data)+"'")
                    row2 = cur2.fetchone()
                    if row2!=None:
                        id_dictionary=row2[1]
                    if id_dictionary==0:
                        cur2.execute("SELECT MAX(id_dictionary) FROM contents_dictionary WHERE id_header="+str(contents_id_header))
                        row2 = cur2.fetchone()
                        if row2!=None:
                            id_dictionary=row2[0]
                        if id_dictionary==None: id_dictionary=0
                        id_dictionary=id_dictionary+1
                        cur2.execute("INSERT INTO contents_dictionary (id_header,id_dictionary,data) VALUES ("+str(contents_id_header)+","+str(id_dictionary)+",'"+GetSqlValue(contents_data)+"')")
                    #data dictionary - End
                    cur2.execute("INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary) VALUES ("+str(contents_id_row)+","+str(contents_id_user)+","+str(contents_id_header)+","+str(id_dictionary)+")")
                #Dictionary - End
                cur.execute('ALTER TABLE contents RENAME TO contents_old')
                cur.execute('CREATE VIEW contents AS SELECT contents_index.id_row AS id_row, contents_index.id_header AS id_header, contents_dictionary.data AS data, contents_index.id_user AS id_user FROM contents_index INNER JOIN contents_dictionary ON contents_index.id_dictionary=contents_dictionary.id_dictionary AND  contents_index.id_header=contents_dictionary.id_header')
                cur.execute('DROP TABLE contents_old')
                #except: pass

                #V.2.1.3
                try:
                    cur.execute('CREATE UNIQUE INDEX contents_index_i0 on contents_index (id_row,id_user,id_header)')
                except: pass
            #end test pre 2.2.0

            #V.2.3.5
            if LoadCustomSetting("Network_Connection_Mode")=="": SaveCustomSetting("Network_Connection_Mode","automatic")

            #V.2.4.0
            try:
                cur.execute('ALTER TABLE rows ADD id_instance INTEGER')
                cur.execute('ALTER TABLE rows ADD status_instance INTEGER')
                cur.execute('ALTER TABLE contents_index ADD id_instance INTEGER')
                cur.execute("UPDATE rows SET id_instance=0, status_instance=0")
                cur.execute("UPDATE contents_index SET id_instance=0")
                cur.execute("DROP VIEW contents")
                cur.execute('CREATE VIEW contents AS SELECT contents_index.id_row AS id_row, contents_index.id_header AS id_header, contents_dictionary.data AS data, contents_index.id_user AS id_user, contents_index.id_instance as id_instance FROM contents_index INNER JOIN contents_dictionary ON contents_index.id_dictionary=contents_dictionary.id_dictionary AND  contents_index.id_header=contents_dictionary.id_header')            

                cur.execute('CREATE INDEX rows_index_i0 on rows (id_row,id_user,id_header,id_instance)')
                cur.execute('DROP INDEX contents_index_i0')
                cur.execute('CREATE INDEX contents_index_i0 on contents_index (id_row,id_user,id_header,id_instance)')
            except: pass
        import datetime
        now = datetime.datetime.utcnow()
        ts=now.strftime("%d/%m/%Y, %H:%M:%S")
        ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
        if arr_dbrelease<[5,6,0]:
            if DEBUG_MODE: print "checking database for releases pre 5.6.0" #sarebbe 5.6.6
            #V2.6.2
            try: 
                cur.execute('ALTER TABLE groups ADD date_time TEXT')
                cur.execute('ALTER TABLE groups ADD status INTEGER')
                cur.execute('UPDATE groups SET date_time=\''+ts_db+'\', status=0')
                cur.execute('ALTER TABLE users ADD date_time TEXT')
                cur.execute('ALTER TABLE users ADD status INTEGER')
                cur.execute('UPDATE users SET date_time=\''+ts_db+'\', status=0')
            except: pass
            #V2.9.0
            try: 
                cur.execute('CREATE INDEX contents_dictionary_i1 on contents_dictionary (id_header,id_dictionary)')
                cur.execute('CREATE INDEX contents_index_i1 on contents_index (id_header,id_dictionary)')
            except: pass
            #V2.9.1
            try: 
                cur.execute('ALTER TABLE eforms ADD cyclic INTEGER')
                cur.execute('ALTER TABLE contents_index ADD id_cycle INTEGER')
                #se arrivo qui' e' perche' i cicli non c'erano ancora
                cur.execute('UPDATE contents_index SET id_cycle=0')
                cur.execute('DROP VIEW contents')
                cur.execute('CREATE VIEW contents AS SELECT contents_index.id_row AS id_row, contents_index.id_header AS id_header, contents_dictionary.data AS data, contents_index.id_user AS id_user, contents_index.id_instance AS id_instance, contents_index.id_cycle AS id_cycle FROM contents_index INNER JOIN contents_dictionary ON contents_index.id_header=contents_dictionary.id_header AND contents_index.id_dictionary=contents_dictionary.id_dictionary')
            except: pass
            #V2.9.9
            try: 
                cur.execute('ALTER TABLE eforms ADD onactivate TEXT')
            except: pass
            #V3.0.0
            cur.execute('UPDATE users SET firstaccess=substr(firstaccess,7,4)||"-"||substr(firstaccess,4,2)||"-"||substr(firstaccess,1,2)||substr(firstaccess,12,length(firstaccess)-12+1) WHERE firstaccess like "__/__/____%"')
            cur.execute('UPDATE users SET lastaccess=substr(lastaccess,7,4)||"-"||substr(lastaccess,4,2)||"-"||substr(lastaccess,1,2)||substr(lastaccess,12,length(lastaccess)-12+1) WHERE lastaccess like "__/__/____%"')
            cur.execute('UPDATE users SET date_time=substr(date_time,7,4)||"-"||substr(date_time,4,2)||"-"||substr(date_time,1,2)||substr(date_time,12,length(date_time)-12+1) WHERE date_time like "__/__/____%"')
            cur.execute('UPDATE groups SET date_time=substr(date_time,7,4)||"-"||substr(date_time,4,2)||"-"||substr(date_time,1,2)||substr(date_time,12,length(date_time)-12+1) WHERE date_time like "__/__/____%"')
            cur.execute('UPDATE rows SET date_time=substr(date_time,7,4)||"-"||substr(date_time,4,2)||"-"||substr(date_time,1,2)||substr(date_time,12,length(date_time)-12+1) WHERE date_time like "__/__/____%"')

            #V3.1.4
            try: cur.execute('CREATE INDEX rows_index_i1 on rows (status,status_user,status_instance,id_row,id_user,id_header,id_instance,date_time)')
            except: pass
            #V3.1.5 - V4.1.4 (aggiunta DROP)
            try: cur.execute("DROP VIEW rows_checksum")
            except: pass
            try: cur.execute("CREATE VIEW rows_checksum AS SELECT rows.id_row AS id_row, rows.id_user AS id_user, rows.id_instance AS id_instance, rows.id_header AS id_header, rows.date_time AS date_time, rows.rap AS rap, rows.hpath AS hpath, rows.id_locking AS id_locking, rows.status AS status, rows.status_user AS status_user, rows.status_instance AS status_instance, count(*) AS rec FROM rows LEFT JOIN contents_index ON rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance GROUP BY rows.status, rows.status_user, rows.status_instance, rows.id_row, rows.id_user, rows.id_instance, rows.date_time ORDER BY rows.status, rows.status_user, rows.status_instance, rows.id_row, rows.id_user, rows.id_instance, rows.date_time")
            except: pass
            
            #V3.2.0
            try: 
                cur.execute('ALTER TABLE eforms ADD label TEXT')
                cur.execute('UPDATE eforms SET label=description')
                cur.execute('UPDATE eforms SET description="panel" || pos')
            except: pass
            try: 
                cur.execute('ALTER TABLE eforms ADD id_group INTEGER')
                cur.execute('UPDATE eforms SET id_group=0')
            except: pass
            try: 
                cur.execute('ALTER TABLE eforms ADD date_time TEXT')
                cur.execute('ALTER TABLE eforms ADD status INTEGER')
                cur.execute('UPDATE eforms SET status=0')
            except: pass
            try:
                cur.execute('CREATE TABLE random_treatments (id_treatment INTEGER, id_group INTEGER, stratum TEXT, treatment TEXT, trialcode TEXT, id_row INTEGER, id_user INTEGER, id_instance INTEGER)')
                cur.execute('CREATE TABLE random_packages (id_package INTEGER, id_group INTEGER, treatment TEXT, package TEXT, trialcode TEXT, id_row INTEGER, id_user INTEGER, id_instance INTEGER)')
            except: pass
            try: 
                cur.execute('ALTER TABLE random_treatments ADD date_time TEXT')
                cur.execute('ALTER TABLE random_treatments ADD status INTEGER')
                cur.execute('UPDATE random_treatments SET status=0')
                cur.execute('ALTER TABLE random_packages ADD date_time TEXT')
                cur.execute('ALTER TABLE random_packages ADD status INTEGER')
                cur.execute('UPDATE random_packages SET status=0')
            except: pass
            try:
                cur.execute('CREATE INDEX headers_index_i1 ON headers (child_of,description)')
            except: pass
            #V3.8.4
            try: 
                cur.execute('ALTER TABLE queries ADD id_instance INTEGER')
            except: pass
            try:
                cur.execute('ALTER TABLE logbook ADD date_time TEXT')
                cur.execute('ALTER TABLE logbook ADD status INTEGER')
                cur.execute('ALTER TABLE logstat ADD date_time TEXT')
                cur.execute('ALTER TABLE logstat ADD status INTEGER')
            except: pass
            #V4.0.5
            try:
                cur.execute('select id_eform from eforms')
                row = cur.fetchone()
                if row!=None: 
                    curw = con.cursor()
                    cur.execute('ALTER TABLE eforms ADD height INTEGER')
                    cur.execute('delete from settings where setting_key="virtualpagesize"')
                    id_eform=[]
                    cur.execute('select id_eform from eforms order by pos')
                    for row in cur:
                        id_eform.append(int(row[0]))
                    cur.execute('select setting_key,setting_value from settings where setting_key like "panel%_height"')
                    for row in cur:
                        idx=int(row[0][5:row[0].find("_height")])
                        if idx<=len(id_eform):
                            curw.execute('update eforms set height='+str(row[1])+' where id_eform='+str(id_eform[idx-1]))
                    cur.execute('delete from settings where setting_key like "panel%_height"')
                else:
                    cur.execute('ALTER TABLE eforms ADD height INTEGER')
            except: pass
            #V4.4.2
            try:
                cur.execute('CREATE TABLE efields (id_eform INTEGER, classname TEXT, style TEXT, label TEXT, id_header INTEGER, name TEXT, pos_x INTEGER, pos_y INTEGER, size_x INTEGER, size_y INTEGER)')
            except: pass
            try: 
                cur.execute('ALTER TABLE headers ADD date_time TEXT')
                cur.execute('ALTER TABLE headers ADD status INTEGER')
                cur.execute('UPDATE headers SET status=0')
            except: pass
            #4.8.4
            try:
                cur.execute('ALTER TABLE efields ADD maxlength INTEGER')
            except: pass
            #5.1.5
            try:
                cur.execute('ALTER TABLE users ADD otp TEXT')
                cur.execute('UPDATE users SET otp=\'\'')
            except: pass
        try:
            cur.execute('CREATE INDEX idx_rows_checksum ON rows (status, status_user, status_instance, id_row, id_user, id_instance, date_time)')
        except: pass
        SaveCustomSetting("heavybase_release",RELEASE)
        
        # old hystory
        history_days=LoadCustomSetting("history_days")
        if history_days!="":
            intdays=int(history_days)
            if intdays==0:
                history_days=""
            else:
                import datetime
                now = datetime.datetime.utcnow()
                olddate=now-datetime.timedelta(days=intdays)
                strolddate=olddate.strftime("%Y-%m-%d")
                # wrong timestamp
                cur.execute("update rows set date_time=substr(date_time,1,19)")
                cur.execute("delete from rows where status<>0 and date_time<'"+strolddate+"'")
                print "substituted records older than "+strolddate+" removed."
                
        curw = con.cursor()
        if LoadCustomSetting("dupes_checking")!="0":
            print "Checking database..."
            # Database corrections - Begin
            cur.execute('UPDATE rows SET status=0 WHERE status is null')
            cur.execute('UPDATE rows SET status=0 WHERE status=""')
            cur.execute('UPDATE rows SET status_user=0 WHERE status_user is null')
            cur.execute('UPDATE rows SET status_user=0 WHERE status_user=""')
            cur.execute('UPDATE rows SET status_instance=0 WHERE status_instance is null')
            cur.execute('UPDATE rows SET status_instance=0 WHERE status_instance=""')
            cur.execute('UPDATE contents_index SET id_cycle=0 WHERE id_cycle is null')

            cur.execute('UPDATE rows SET hpath=null WHERE hpath=""')
            cur.execute('UPDATE rows SET hpath=null WHERE hpath="NULL"')
            cur.execute('UPDATE rows SET id_locking=null WHERE id_locking=""')
            cur.execute('UPDATE rows SET id_locking=null WHERE id_locking="NULL"')

            cur.execute('UPDATE rows SET date_time="'+ts_db+'" WHERE date_time=""')
            
            #empty id_header
            cur.execute("SELECT description,child_of FROM headers WHERE id_header IS NULL")
            tmpcorr=[]
            for row in cur:
                tmpcorr.append(str(row[1])+","+row[0])
            for elm in tmpcorr:
                longCRC=UniqueLongCRC(elm,cur)
                cur.execute("UPDATE headers SET id_header="+str(longCRC)+" WHERE child_of="+elm.split(",")[0]+" AND description='"+elm.split(",")[1]+"'")

            # not correctly initialized tables
            cur.execute("UPDATE headers SET status=0 WHERE status is null")
            cur.execute("UPDATE headers SET date_time='0000-00-00 00:00:00' WHERE date_time is null")
            cur.execute("UPDATE eforms SET status=0 WHERE status is null")
            cur.execute("UPDATE eforms SET date_time='0000-00-00 00:00:00' WHERE date_time is null")
            
            #rap become empty
            cur.execute("SELECT count(*) as conta from rows where rap=''")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>0:
                    cur2 = con.cursor()
                    cur.execute("SELECT r1.id_row,r1.id_user,r1.id_instance,r2.rap FROM rows AS r1, rows AS r2 WHERE r1.id_row=r2.status AND r1.id_user=r2.status_user AND r1.id_instance=r2.status_instance AND r1.rap='' AND r2.rap<>''")
                    for row in cur:
                        cur2.execute("UPDATE rows SET rap='"+row[3]+"' WHERE id_row="+str(row[0])+" AND id_user="+str(row[1])+" AND id_instance="+str(row[2]))
                    
            # duplicate headers descriptions
            corrections=True
            while corrections:
                corrections=False
                cur.execute("select min(id_header) as minIdHeader, max(id_header) as maxIdHeader, headers.description from headers, (select description from (select description,count(*) as quanti from headers group by description) as s1 where quanti>1) as s2 where headers.description=s2.description group by headers.description")
                for row in cur:
                    corrections=True
                    #print "update contents_index set id_header="+str(row[0])+" where id_header="+str(row[1])
                    curw.execute("update contents_index set id_header="+str(row[0])+" where id_header="+str(row[1]))
                    curw.execute("delete from headers where id_header="+str(row[1]))
            #duplicate rows records
            cur.execute("select max(dup) as maxdup from (select id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance, count(*) as dup from rows group by id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance)")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>1:
                    try: cur.execute('DROP TABLE rows_nodup')
                    except: pass
                    cur.execute('CREATE TABLE rows_nodup (id_row INTEGER, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER, rap TEXT, hpath TEXT, status_user INTEGER, id_locking INTEGER, id_instance INTEGER, status_instance INTEGER)')
                    cur.execute('INSERT INTO rows_nodup (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT DISTINCT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows')
                    cur.execute('DELETE FROM rows')
                    cur.execute('INSERT INTO rows (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows_nodup')
                    cur.execute('DROP TABLE rows_nodup')
                    print "dupes from rows removed."
            #duplicate contents_index records
            cur.execute("select max(dup) as maxdup from (select id_row, id_header, id_user, id_dictionary, id_instance, id_cycle, count(*) as dup from contents_index group by id_row, id_header, id_user, id_dictionary, id_instance, id_cycle)")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>1:
                    try: cur.execute('DROP TABLE contents_index_nodup')
                    except: pass
                    cur.execute('CREATE TABLE contents_index_nodup (id_row INTEGER, id_user INTEGER, id_header INTEGER, id_dictionary INTEGER, id_instance INTEGER, id_cycle INTEGER)')
                    cur.execute('INSERT INTO contents_index_nodup (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT DISTINCT id_row, id_header, id_user, id_dictionary, id_instance, id_cycle FROM contents_index')
                    cur.execute('DELETE FROM contents_index')
                    cur.execute('INSERT INTO contents_index (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT id_row, id_header, id_user, id_dictionary, id_instance, id_cycle FROM contents_index_nodup')
                    cur.execute('DROP TABLE contents_index_nodup')
                    print "dupes from contents_index removed."
                    
            #perfect duplicate contents_dictionary records
            cur.execute("select max(dup) as maxdup from (select id_header, id_dictionary, data, count(*) as dup from contents_dictionary group by id_header, id_dictionary, data)")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>1:
                    try: cur.execute('DROP TABLE contents_dictionary_nodup')
                    except: pass
                    cur.execute('CREATE TABLE contents_dictionary_nodup (id_header INTEGER, id_dictionary INTEGER, data BLOB)')
                    cur.execute('INSERT INTO contents_dictionary_nodup (id_header, id_dictionary, data) SELECT DISTINCT id_header, id_dictionary, data FROM contents_dictionary')
                    cur.execute('DELETE FROM contents_dictionary')
                    cur.execute('INSERT INTO contents_dictionary (id_header, id_dictionary, data) SELECT id_header, id_dictionary, data FROM contents_dictionary_nodup')
                    cur.execute('DROP TABLE contents_dictionary_nodup')
                    print "dupes from contents_dictionary removed."
                    
            #not perfect duplicate contents_dictionary records (same indexes, different value, unknown problem)
            cur.execute("select max(dup) as maxdup from (select id_header, id_dictionary, count(*) as dup from contents_dictionary group by id_header, id_dictionary)")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>1:
                    try: cur.execute('DROP TABLE contents_dictionary_nodup')
                    except: pass
                    cur.execute('CREATE TABLE contents_dictionary_nodup (id_header INTEGER, id_dictionary INTEGER, data BLOB)')
                    chk={}
                    cur.execute('SELECT id_header,id_dictionary,data FROM contents_dictionary')
                    for row in cur:
                        if not chk.has_key(str(row[0])+"-"+str(row[1])):
                            curw.execute('INSERT INTO contents_dictionary_nodup (id_header, id_dictionary, data) VALUES ('+str(row[0])+','+str(row[1])+',\''+row[2]+'\')')
                            chk[str(row[0])+"-"+str(row[1])]=True
                    cur.execute('DELETE FROM contents_dictionary')
                    cur.execute('INSERT INTO contents_dictionary (id_header, id_dictionary, data) SELECT id_header, id_dictionary, data FROM contents_dictionary_nodup')
                    cur.execute('DROP TABLE contents_dictionary_nodup')
                    print "inconsistent dupes from contents_dictionary removed."
            #rows with status>=0 without contents_index (unknown problem, only lonely rows with status<0 are allowed)
            cur.execute("select count(*) from rows left join contents_index on rows.id_row=contents_index.id_row and rows.id_user=contents_index.id_user and rows.id_instance=contents_index.id_instance where rows.status>=0 and contents_index.id_row is null")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>0:
                    cur.execute('CREATE TABLE rows_notalone (id_row INTEGER, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER, rap TEXT, hpath TEXT, status_user INTEGER, id_locking INTEGER, id_instance INTEGER, status_instance INTEGER)')
                    cur.execute('INSERT INTO rows_notalone (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows WHERE status<0')
                    cur.execute('INSERT INTO rows_notalone (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT DISTINCT rows.id_row, rows.id_header, rows.status, rows.date_time, rows.id_user, rows.rap, rows.hpath, rows.status_user, rows.id_locking, rows.id_instance, rows.status_instance FROM rows inner join contents_index on rows.id_row=contents_index.id_row and rows.id_user=contents_index.id_user and rows.id_instance=contents_index.id_instance')
                    cur.execute('DELETE FROM rows')
                    cur.execute('INSERT INTO rows (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows_notalone')
                    cur.execute('DROP TABLE rows_notalone')
                    print "table rows fixed."
            #contents_index without rows (previous syncronizations interrupted)
            cur.execute("select count(*) from contents_index left join rows on contents_index.id_row=rows.id_row and contents_index.id_user=rows.id_user and contents_index.id_instance=rows.id_instance where rows.id_row is null")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>0:
                    cur.execute('CREATE TABLE contents_index_notalone (id_row INTEGER, id_user INTEGER, id_header INTEGER, id_dictionary INTEGER, id_instance INTEGER, id_cycle INTEGER)')
                    cur.execute('INSERT INTO contents_index_notalone (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT DISTINCT contents_index.id_row, contents_index.id_header, contents_index.id_user, contents_index.id_dictionary, contents_index.id_instance, contents_index.id_cycle FROM contents_index inner join rows on contents_index.id_row=rows.id_row and contents_index.id_user=rows.id_user and contents_index.id_instance=rows.id_instance')
                    cur.execute('DELETE FROM contents_index')
                    cur.execute('INSERT INTO contents_index (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT id_row, id_header, id_user, id_dictionary, id_instance, id_cycle FROM contents_index_notalone')
                    cur.execute('DROP TABLE contents_index_notalone')
                    print "table contents_index fixed."
            #contents_dictionary without contents_index (previous syncronizations interrupted)
            cur.execute("select count(*) from contents_dictionary left join contents_index on contents_dictionary.id_header=contents_index.id_header and contents_dictionary.id_dictionary=contents_index.id_dictionary where contents_index.id_header is null")
            row = cur.fetchone()
            if row!=None: 
                if row[0]>0:
                    cur.execute('CREATE TABLE contents_dictionary_notalone (id_header INTEGER, id_dictionary INTEGER, data BLOB)')
                    cur.execute('INSERT INTO contents_dictionary_notalone (id_header,id_dictionary,data) SELECT DISTINCT contents_dictionary.id_header,contents_dictionary.id_dictionary,contents_dictionary.data from contents_dictionary inner join contents_index on contents_dictionary.id_header=contents_index.id_header and contents_dictionary.id_dictionary=contents_index.id_dictionary')
                    cur.execute('DELETE FROM contents_dictionary')
                    cur.execute('INSERT INTO contents_dictionary (id_header,id_dictionary,data) SELECT id_header,id_dictionary,data FROM contents_dictionary_notalone');
                    cur.execute('DROP TABLE contents_dictionary_notalone')
                    print "table contents_dictionary fixed."

        cur.execute('COMMIT TRANSACTION')
        # Database corrections - End

        #con.commit()
        self.max_idHeadDict={}
        cur.execute("SELECT id_header,max(id_dictionary) AS maxIdDict FROM contents_dictionary GROUP BY id_header")
        for row in cur:
            self.max_idHeadDict[str(row[0])]=row[1]
        # Database initialization - End

        # Host ID
        import uuid
        cur_host=str(uuid.getnode())
        id_host=cur_host+":"+DATABASE_PATH
        cur.execute("SELECT setting_value FROM settings where setting_key='id_host'")
        row = cur.fetchone()
        if row!=None: 
            cur.execute("UPDATE settings SET setting_value='"+id_host+"' WHERE setting_key='id_host'")
        else:
            cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('id_host','"+id_host+"')")

        #p2p
        import random
        random.seed()
        self.id_instance=str(int(random.random()*10000000000))[0:10]
        print "id_instance="+self.id_instance

        self.p2p_state=""
        self.p2p_message=""
        self.p2p_server=None

        self.p2p_md5Rows=""
        self.p2p_lstRows=""
        self.p2p_deltaRows=""
        self.p2p_deltaContents_index=""
        self.p2p_deltaHeaders=""
        self.p2p_deltaDict=""
        self.p2p_partner=""
        self.p2p_partner_state=""
        self.p2p_partner_md5Rows=""
        self.p2p_partner_lstRows=""
        self.p2p_partner_deltaRows=""
        self.p2p_partner_deltaContents_index=""
        self.p2p_partner_deltaHeaders=""
        self.p2p_partner_deltaDict=""
        self.p2p_chkRowsId=""
        self.p2p_keysDict=""
        self.p2p_liveupdate_ts=""
        self.p2p_lastConnectionUrl=""
        # tabelle da tenere sincronizzate a parte rows,headers,contents_index,contents_dictionary
        # N.B. date_time e status, campi obbligatori per tutte le tabelle da sincronizzare, devono essere necessariamente il penultimo e ultimo campo elencato
        self.cursyncrotable_id = ""
        self.syncrotables = ["rows","random_treatments","random_packages","eforms","headers","groups","users","attach","logbook","logstat"]
        self.syncrotableskeys = {}
        self.syncrotablesflds = {}
        self.syncrotableskeys["random_treatments"] = "0"    #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["random_treatments"] = "id_treatment,id_group,stratum,treatment,trialcode,id_row,id_user,id_instance,date_time,status"
        self.syncrotableskeys["random_packages"] = "0"  #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["random_packages"] = "id_package,id_group,treatment,package,trialcode,id_row,id_user,id_instance,date_time,status"
        self.syncrotableskeys["eforms"] = "0"  #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["eforms"] = "id_eform,id_header,pos,description,href,cyclic,onactivate,label,id_group,height,date_time,status"
        self.syncrotableskeys["headers"] = "0"  #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["headers"] = "id_header,child_of,description,label,hap,stype,deftype,cyclic,doubleinput,typedesc,subtypedesc,textlength,subtextlength,textalign,freetext,defaultvalue,validation,onchange,date_time,status"
        self.syncrotableskeys["groups"] = "0"   #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["groups"] = "id_group,group_level,member_of,shortcode,description,acl,date_time,status"
        self.syncrotableskeys["users"]  = "0"   #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["users"]  = "id_user,id_group,uap,stype,rights,username,password,fullname,notes,home,date_time,status"
        self.syncrotableskeys["logbook"] = "0"  #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["logbook"] = "id_log,id_session,id_user,logon,logoff,host,date_time,status"
        self.syncrotableskeys["logstat"] = "0"  #comma separated positions of the keys in the fields list, starting by zero
        self.syncrotablesflds["logstat"] = "id_logstat,id_log,timestamp,statement,date_time,status"
        self.syncrotablesupd = {}
        self.syncrotablesupd["rows"]=self.syncrotablesupd["random_treatments"]=self.syncrotablesupd["random_packages"]=self.syncrotablesupd["eforms"]=self.syncrotablesupd["headers"]=self.syncrotablesupd["groups"]=self.syncrotablesupd["users"]=self.syncrotablesupd["attach"]=self.syncrotablesupd["logbook"]=self.syncrotablesupd["logstat"]=True
        self.chkRowsId_tables = {}
        self.lstRows_tables = {}
        self.md5Rows_tables = {}
        
        self.p2p_defUpdate=[]
        self.preDecrypt_waitstate=0
        self.LogoutCount=0
        
        #self.OnTimer(None)
        self.timer = wx.Timer(self, -1)
        self.timer.Start(1000,True) #One shot
        self.Bind(wx.EVT_TIMER, self.OnTimer)

        #Backward compatibility - begin
        self.reloadSettings()
        if GetSetting(self,"heavybase_server")=="xmlrpc.heavybase.org:8080": 
            SaveCustomSetting("heavybase_server","xmlrpc.heavybase.org:80")
        self.reloadSettings()
        #Backward compatibility - end

        #self.pp_server = PP_ServerService(self)
        #self.pp_server.start()

        self.SetStatusText("Start P2P")
        #self.syncroTask = ThreadLooper (3, SyncroP2P, (self,self))
        self.syncroTask = ThreadLooper (0.5, SyncroP2P, (self,self))
        self.syncroTask.start()
        self.syncroTaskStop = False
        self.syncroTaskPause = False
        self.syncroTaskUpdating = False
        self.syncroTaskRunning = False
        self.syncroTaskWaitSave = False
        self.quickreportRunning = False
        self.autoCloseRequest = False
        self.autoRestartRequest = False
        
        self.trans={}

        self.lastSearch=[]
        self.lastSearch_lstfields=[]

        self.child = HeavyBaseFrame(self, TT("Data management"))
        self.child.Show()
        
        title=GetSetting(self,"project_title")
        if title!="": title=title+" - "
        title=title+PROGNAME+" "+RELEASE
        self.SetTitle(title)
        
        self.minAutologout=GetSetting(self,"autologout_minutes")
        #self.minAutologout="1"
        if self.minAutologout!="":
            print "Activating Autologout"
            self.autologout_minutes=int(self.minAutologout)
            self.IdleBegin=0
            self.IdleTS=0
            self.IdleTriggered=False
            self.Bind(wx.EVT_IDLE, self.OnIdle)
        
        self.DEMO_MODE=False
        tmp=GetSetting(self,"demo_mode")
        if tmp=="1": self.DEMO_MODE=True
        if self.DEMO_MODE: print "Switching to DEMO MODE."
        
        print "Initialization completed."

    def OnIdle(self, event):
        autologout_triggered=False
        #if sys.platform != 'win32':
        if os.name!='nt':
            import time
            curTS=time.time()
            if self.IdleTS==0: self.IdleTS=curTS
            #if DEBUG_MODE: print str(curTS-self.IdleTS)
            if curTS-self.IdleTS<0.0005:
                self.IdleTriggered=False
                self.IdleBegin=curTS
                #if DEBUG_MODE: print "Idle resetted"
            #if DEBUG_MODE: print str(curTS-self.IdleBegin)
            secAutologout=self.autologout_minutes*60
            if ((curTS-self.IdleBegin)>secAutologout) and (self.IdleTriggered==False):
                self.IdleTriggered=True
                self.LogoutCount=self.LogoutCount+1
                autologout_triggered=True
            self.IdleTS=curTS
        else:
            #print str(get_idle_duration())+"\n"+str(self.autologout_minutes*60)
            idle_duration=get_idle_duration()
            if idle_duration<1:
                self.IdleTriggered=False
            if (idle_duration>float(self.autologout_minutes*60)) and (self.IdleTriggered==False):
                self.IdleTriggered=True
                self.LogoutCount=self.LogoutCount+1
                autologout_triggered=True
        if autologout_triggered:
            if self.syncroTaskStop:
                print "Autorestart."
                self.autoRestartRequest=True
                if self.DoCloseHeavyBase(): self.Destroy()
            else:
                print "Autologout."
        
    def MakeMenuBar(self):
        mb = wx.MenuBar()

        menuFile = wx.Menu()
        item = menuFile.Append(-1, TT("New Data management"))
        self.Bind(wx.EVT_MENU, self.OnDataEntry, item)
        item = menuFile.AppendSeparator()
        item = menuFile.Append(-1, TT("Network settings"))
        self.Bind(wx.EVT_MENU, self.OnSettings, item)
        item = menuFile.Append(-1, TT("Export DB"))
        self.Bind(wx.EVT_MENU, self.OnExportDb, item)
        item = menuFile.Append(-1, TT("Help Desk: Send Screenshot"))
        self.Bind(wx.EVT_MENU, self.OnSendScreenshot, item)
        item = menuFile.AppendSeparator()
#        if DEBUG_MODE:
        imp = wx.Menu()
        item = imp.Append(-1, TT('Fields and labels list'))
        self.Bind(wx.EVT_MENU, self.OnDatasetEfields, item)
        item = imp.Append(-1, TT('Data structure document'))
        self.Bind(wx.EVT_MENU, self.OnDatasetStructure, item)
        item = menuFile.AppendMenu(-1, TT('Database Administration'), imp)
        item = menuFile.AppendSeparator()
        
        item = menuFile.Append(-1, TT("Exit"))
        self.Bind(wx.EVT_MENU, self.OnClose, item)
        mb.Append(menuFile, TT("Activity"))

        return mb

    def OnDatasetEfields(self,evt):
        try: w=xlwt.Workbook()
        except: w=Workbook()
        ws=w.add_sheet('efields')
        qy="select * from efields"
        cur.execute(qy)
        cols = [i[0] for i in cur.description]; hc={}; i=0
        for col in cols: hc[col]=i; i=i+1
        
        for x in range(len(cols)):
            ws.write(0,x,cols[x])
        y=1
        for row in cur:
            for x in range(len(cols)):
                ws.write(y,x,row[x])
            y=y+1
        now = datetime.datetime.now()
        ts=now.strftime("%Y%m%d%H%M%S")
        filename=REPORTS_PATH+"efields_"+ts+".xls"
        w.save(filename)
        if not assoc_open(filename):
            wx.MessageBox("File\n'"+filename+"'\ncreated.", TT("Fields and labels list"), wx.ICON_INFORMATION | wx.OK, self)
        
    def OnDatasetStructure(self,evt):
        intestazione_ist="____________________"
        intestazione_ist_lunga="_______________________________________________"
        intestazione_ist_indir="______________________________"
        intestazione_lab="_______________________________________"
        intestazione_lsi="Life Science Informatics"
        intestazione_head_info="________________________________________________________________________"
        intestazione_head_stat="________________________________________________________________________"
        intestazione_head_lab ="________________________________________________________________________"
        if GetSetting(self,"project_name")[:6]=="irfmn_":
            intestazione_ist="Istituto Mario Negri"
            intestazione_ist_lunga="Istituto di Ricerche Farmacologiche Mario Negri"
            intestazione_ist_indir="via La Masa, 19 - 20156 Milano"
            intestazione_lab="Laboratorio di Ricerca Clinica - L.S.I."
            intestazione_lsi="Life Science Informatics"
            intestazione_head_info="Dr. Davide Poli (Capo unita' Informatica e Gestione degli studi clinici)"
            intestazione_head_stat="Dr.ssa Eliana Rulli (Capo Unita' Statistica per la Ricerca Clinica)"
            intestazione_head_lab="Dr.ssa Irene Floriani (Capo Laboratorio Ricerca Clinica)"
        nome_studio=GetSetting(self,"project_title")
        if nome_studio=="": nome_studio="___________"
        if GetSetting(self,"doc_ist")!="":       intestazione_ist       = GetSetting(self,"doc_ist")
        if GetSetting(self,"doc_ist_long")!="":  intestazione_ist_lunga = GetSetting(self,"doc_ist_long")
        if GetSetting(self,"doc_ist_addr")!="":  intestazione_ist_indir = GetSetting(self,"doc_ist_addr")
        if GetSetting(self,"doc_ist_lab")!="":   intestazione_lab       = GetSetting(self,"doc_ist_lab")
        if GetSetting(self,"doc_logo_sw")!="":   intestazione_lsi       = GetSetting(self,"doc_logo_sw")
        if GetSetting(self,"doc_head_info")!="": intestazione_head_info = GetSetting(self,"doc_head_info")
        if GetSetting(self,"doc_head_stat")!="": intestazione_head_stat = GetSetting(self,"doc_head_stat")
        if GetSetting(self,"doc_head_lab")!="":  intestazione_head_lab  = GetSetting(self,"doc_head_lab")

        now = datetime.datetime.now()
        ts=now.strftime("%d/%m/%Y")
        

        rb=RtfBuilder()
        buf=[]
        # COPERTINA
        buf.append(rb.IntestazioneRTF(intestazione_lab, "Struttura Dati", "Studio "+nome_studio, intestazione_ist, "Data Report: "+ts, intestazione_lsi, intestazione_lab, "Struttura dati", "Studio "+nome_studio, "\n"+intestazione_ist, "Data Report: "+ts))
        buf.append(rb.NuovoTesto("\nStudio\n\n", False, True, 14, "Arial", "Center"))
        buf.append(rb.NuovoTesto(nome_studio+"\n\n", True, False, 14, "Arial", "Center"))
        buf.append(rb.NuovoTesto(intestazione_ist_lunga+"\n\n", False, True, 12, "Arial", "Center"))
        buf.append(rb.NuovoTesto("\n\nSTRUTTURA DATI\n", True, False, 18, "Arial", "Center"))
        buf.append(rb.NuovoTesto("\n\nCodice Eudract: ________________________________"+"\n", False, True, 12, "Arial", "Center"))
        buf.append(rb.NuovoTesto("\n\n\n\n\n\nAggiornato al "+ts+"\n\n\n", False, False, 12, "Arial", "Center"))
        buf.append(rb.NuovaTbl_Inizio(2))
        buf.append(rb.NuovaTbl_Colonna(True, True, 10, "Promotore"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 10, "Autore"))
        buf.append(rb.NuovaTbl_Riga(2))
        buf.append(rb.NuovaTbl_Colonna(True, False, 10, intestazione_lab+"\n"+intestazione_ist_lunga+"\n"+intestazione_ist_indir))
        buf.append(rb.NuovaTbl_Colonna(False, False, 10, "_____________________________\n"+intestazione_lab+"\n"+intestazione_ist_lunga+"\nTel. __________ Fax __________"))
        buf.append(rb.NuovaTbl_Fine())
        buf.append(rb.NuovoBreak(0))
        # SECONDA PAGINA
        buf.append(rb.NuovoTesto("Obiettivo del documento\n", True, False, 12, "Arial", "Center"))
        buf.append(rb.NuovoTesto("Questo documento ha lo scopo di definire le proprieta' di tutte le variabili contenute nella Scheda Raccolta Dati (e-CRF) dello studio "+nome_studio+"\n\n", False, False, 12, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("Contenuto del documento\n", True, False, 12, "Arial", "Center"))
        buf.append(rb.NuovoTesto("Il documento contiene le seguenti  informazioni:\n - Nome delle variabili\n - Descrizione delle variabili\n - Formato delle variabili\n - Il tipo di variabile\n - Ricodifica delle variabili (nome, formato e codice)\n - Numero di pagina della e-CRF dove e' collocata la variabile\n\nN.B. Le righe a sfondo grigio individuano quelle variabili ausiliarie che verranno definite e codificate alla fine dello studio, prima della chiusura del database."+"\n\n", False, False, 12, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("Informazioni generali sulle variabili\n", True, False, 12, "Arial", "Center"))
        buf.append(rb.NuovoTesto("* trialcode (= Codice identificativo del paziente) e' la variabile chiave del dataset ed e' composta dall'unione di un codice a 3 lettere per la tipologia della patologia oggetto dello Studio, un numero progressivo dello Studio per questa patologia, uno spazio, la variabile group_shortcode (codice centro), uno spazio, un progressivo del paziente nello studio all'interno del centro, uno spazio, un carattere alfanumerico che costituisce un checksum del trial_code per controllarne la correttezza, un carattere alfanumerico random.\n", False, False, 12, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("* I nomi delle variabili devono essere al massimo composti da 32 caratteri, devono iniziare con una lettera dell'alfabeto(a-z) o con un underscore(_) e non possono contenere spazi, lettere accentate o altri caratteri.\n", False, True, 12, "Arial", "Center"))
        buf.append(rb.NuovoTesto("* Le variabili possono essere di vario tipo:\n - TB (text box) campo libero\n - DC (drop down menu closed) menu a tendina chiuso\n - DO (drop down menu open) menu a tendina aperto con possibilita' di digitare opzioni diverse da quelle previste. La lunghezza di queste variabili non e' stata definita: verra' considerata come lunghezza l'opzione con piu' caratteri.\n - DATE = data (DD/MM/YYYY)\n - CB (check box)"+"\n\n", False, False, 12, "Arial", "Justify"))
        buf.append(rb.NuovaTbl_Fine())
        buf.append(rb.NuovoBreak(0))


        # PAGINA CODIFICHE
        codbuf=[]
        codbuf.append(rb.NuovoTesto("Elenco drop-down menu closed (DC)", True, False, 10, "Arial", "Justify"))
        fld_options={}
        format_id={}
        format_count=0
        format_name={}
        maxlength={}
        cur.execute("select name,label from efields where classname='wxChoice' order by id_eform,pos_y,pos_x")
        for row in cur:
            fld_ref=""
            for key in fld_options:
                if fld_options[key]==row[1]:
                    fld_ref=key
                    break
            maxlength[row[0]]=0
            if fld_ref=="":
                codbuf.append(rb.NuovoTesto("\n", False, False, 10, "Arial", "Center"))
                codbuf.append(rb.NuovaTbl_Inizio(1))
                codbuf.append(rb.NuovaTbl_Colonna(True, True, 8, row[0]))
                opts=row[1].split(",")
                for elm in opts:
                    elm=elm.replace("\"","")
                    if elm!="":
                        codbuf.append(rb.NuovaTbl_Riga(1))
                        codbuf.append(rb.NuovaTbl_Colonna(True, False, 8, elm))
                        if len(elm)>maxlength[row[0]]: maxlength[row[0]]=len(elm)
                codbuf.append(rb.NuovaTbl_Fine())
                format_name[row[0]]=row[0].upper()
                format_count+=1
                format_id[row[0]]="F"+str(format_count)
            else:
                codbuf.append(rb.NuovoTesto("\n"+row[0]+" (see "+fld_ref+")", True, False, 8, "Arial", "Justify"))
                format_name[row[0]]=format_name[fld_ref]
                format_id[row[0]]=format_id[fld_ref]
                maxlength[row[0]]=maxlength[fld_ref]
            fld_options[row[0]]=row[1]
        codbuf.append(rb.NuovoBreak(0))
        codbuf.append(rb.NuovoTesto("Elenco drop-down menu opened (DO)", True, False, 10, "Arial", "Justify"))
        fld_options={}
        cur.execute("select name,label from efields where classname='wxComboBox' order by id_eform,pos_y,pos_x")
        for row in cur:
            codbuf.append(rb.NuovoTesto("\n", False, False, 10, "Arial", "Center"))
            codbuf.append(rb.NuovaTbl_Inizio(1))
            codbuf.append(rb.NuovaTbl_Colonna(True, True, 8, row[0]))
            opts=row[1].split(",")
            for elm in opts:
                elm=elm.replace("\"","")
                if elm!="":
                    codbuf.append(rb.NuovaTbl_Riga(1))
                    codbuf.append(rb.NuovaTbl_Colonna(True, False, 8, elm))
            codbuf.append(rb.NuovaTbl_Riga(1))
            codbuf.append(rb.NuovaTbl_Colonna(True, False, 8, "..."))
            codbuf.append(rb.NuovaTbl_Fine())
            fld_options[row[0]]=row[1]
            format_name[row[0]]=row[0].upper()
            format_count+=1
            format_id[row[0]]="F"+str(format_count)
        codbuf.append(rb.NuovoBreak(0))

        # PAGINA FORMATI
        codbuf.append(rb.NuovoTesto("Lista dei formati", True, False, 12, "Arial", "Justify"))
        keys=[]
        for key in fld_options:
            keys.append(key)
        keys.sort()
        for key in keys:
            opts=fld_options[key].split(",")
            if len(opts)>0 and fld_options[key]!="\"\"":
                codbuf.append(rb.NuovoTesto("\n\n"+key, False, True, 10, "Arial", "Justify"))
                codbuf.append(rb.NuovaTbl_Inizio(2))
                codbuf.append(rb.NuovaTbl_Colonna(True, True, 8, "Codice"))
                codbuf.append(rb.NuovaTbl_Colonna(False, True, 8, "Etichetta"))
                ct=0
                for elm in opts:
                    elm=elm.replace("\"","")
                    if elm!="":
                        codbuf.append(rb.NuovaTbl_Riga(2))
                        ct+=1
                        codbuf.append(rb.NuovaTbl_Colonna(True, False, 8, str(ct)))
                        codbuf.append(rb.NuovaTbl_Colonna(False, False, 8, elm))
                codbuf.append(rb.NuovaTbl_Fine())


        # TERZA PAGINA
        cyclic_page={}
        ncycles={}

        buf.append(rb.NuovoTesto("Copertina (Pagina Ricerca Pazienti)\n", False, True, 12, "Arial", "Center"))
        buf.append(rb.NuovaTbl_Inizio(8))
        buf.append(rb.NuovaTbl_Colonna(True, True, 8, "Variable"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Description"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Format"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Type"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Re-coded name"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Re-coded format"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Re-coded code"))
        buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Page"))
        fieldtype={}
        fieldtype["wxTextCtrl"]="TB"
        fieldtype["wxChoice"]="DC"
        fieldtype["wxComboBox"]="DO"
        fieldtype["wxCheckBox"]="CB"
        cur.execute("select efields.name, efields.classname, headers.label, headers.typedesc, headers.textlength from efields,headers where efields.id_header=headers.id_header and id_eform=0 order by pos_y,pos_x")
        id_eform=0
        cyclic_page[0]=0
        for row in cur:
            buf.append(rb.NuovaTbl_Riga(8))
            buf.append(rb.NuovaTbl_Colonna(True, False, 8, row[0]+("__(n)" if cyclic_page[id_eform] else "")))
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, (row[2] if row[2] is not None else "")))
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, ((row[3] if "=" not in row[3] else row[3].split("=")[1]) if row[3] is not None else "$"+(row[4] if row[4] is not None else ("255" if row[0] not in maxlength else str(maxlength[row[0]]))))))
            data_type=("DATE" if row[3]=="date" and row[1]=="wxTextCtrl" else fieldtype[row[1]])
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, data_type))
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, row[0]+("__(n)" if cyclic_page[id_eform] else "")+("_cod" if row[0] in format_name else "")))
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, ("Date" if data_type=="DATE" else (format_id[row[0]] if row[0] in format_id else ""))))
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, (format_name[row[0]] if row[0] in format_name else "")))
            buf.append(rb.NuovaTbl_Colonna(False, False, 8, "Tutte"))
        buf.append(rb.NuovaTbl_Fine())
        buf.append(rb.NuovoBreak(0))

        # QUARTA PAGINA
        eform_id=[]
        eform_title={}
        cur.execute("select id_eform,label,cyclic from eforms where pos>0 order by pos")
        for row in cur:
            eform_id.append(row[0])
            eform_title[row[0]]=row[1]
            if row[2]>1: 
                cyclic_page[row[0]]=True
                ncycles[row[0]]=row[2]
            else: cyclic_page[row[0]]=False
        ctpage=0
        for id_eform in eform_id:
            ctpage+=1
            buf.append(rb.NuovoTesto(eform_title[id_eform]+"\n", False, True, 12, "Arial", "Center"))
            buf.append(rb.NuovaTbl_Inizio(8))
            buf.append(rb.NuovaTbl_Colonna(True, True, 8, "Variable"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Description"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Format"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Type"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Re-coded name"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Re-coded format"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Re-coded code"))
            buf.append(rb.NuovaTbl_Colonna(False, True, 8, "Page"))
            cur.execute("select efields.name, efields.classname, headers.label, headers.typedesc, headers.textlength from efields,headers where efields.id_header=headers.id_header and id_eform="+str(id_eform)+" order by pos_y,pos_x")
            for row in cur:
                if row[1] not in ("wxButton"):
                    #print row[0]
                    buf.append(rb.NuovaTbl_Riga(8))
                    buf.append(rb.NuovaTbl_Colonna(True, False, 8, row[0]+("__(n)" if cyclic_page[id_eform] else "")))
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, (row[2] if row[2] is not None else "")))
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, ((row[3] if "=" not in row[3] else row[3].split("=")[1]) if row[3] is not None else "$"+(row[4] if row[4] is not None else ("255" if row[0] not in maxlength else str(maxlength[row[0]]))))))
                    data_type=("DATE" if row[3]=="date" and row[1]=="wxTextCtrl" else fieldtype[row[1]])
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, data_type))
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, row[0]+("__(n)" if cyclic_page[id_eform] else "")+("_cod" if row[0] in format_name else "")))
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, ("Date" if data_type=="DATE" else (format_id[row[0]] if row[0] in format_id else ""))))
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, (format_name[row[0]] if row[0] in format_name else "")))
                    buf.append(rb.NuovaTbl_Colonna(False, False, 8, str(ctpage)))
            buf.append(rb.NuovaTbl_Fine())
            
            if cyclic_page[id_eform]:
                buf.append(rb.NuovoTesto("\nTutte le variabili sono cicliche e finiscono per '_n' dove 'n' assume i valori da 1 a "+str(ncycles[id_eform])+" e denota la variabile associata all'n-esimo follow-up", False, False, 10, "Arial", "Justify"))
            
            buf.append(rb.NuovoBreak(0))
            
        for elm in codbuf:
            buf.append(elm)
            
        buf.append(rb.NuovoBreak(0))
            
        # PAGINA CHIUSURA
        buf.append(rb.NuovoTesto("Letto, firmato e sottoscritto\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("Data Manager (nome cognome) ____________________________________\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("(firma) _______________________________     Data _______________\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("Letto, firmato e revisionato\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto(intestazione_head_info+"\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("(firma) _______________________________     Data _______________\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto(intestazione_head_stat+"\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("(firma) _______________________________     Data _______________\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("Letto, firmato e approvato\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto(intestazione_head_lab+"\n\n", False, False, 10, "Arial", "Justify"))
        buf.append(rb.NuovoTesto("(firma) _______________________________     Data _______________\n\n", False, False, 10, "Arial", "Justify"))

        buf.append(rb.TerminaRtf())

        #CREAZIONE FILE E VISUALIZZAZIONE RISULTATO
        now = datetime.datetime.now()
        ts=now.strftime("%Y%m%d%H%M%S")
        filename=REPORTS_PATH+"data_structure_"+ts+".rtf"
        outfile = open(filename, 'w')
        outfile.write("\n".join(buf).encode('latin-1', 'replace'))
        outfile.close()
        if not assoc_open(filename):
            wx.MessageBox("File\n'"+filename+"'\ncreated.", TT("Data structure document"), wx.ICON_INFORMATION | wx.OK, self)


    def OnTimer(self,evt):
        if self.syncroTaskPause:
            self.SetStatusText("(paused)")
        else: 
            import time
            starttime=time.time()
            toolong=False
            while len(self.p2p_defUpdate)>0 and not toolong:
                #rimozione ridondanze e inutilita' - begin
                if len(self.p2p_defUpdate)>1:
                    if self.p2p_defUpdate[0]=="BEGIN TRANSACTION" and self.p2p_defUpdate[1]=="COMMIT TRANSACTION":
                        self.p2p_defUpdate.remove(self.p2p_defUpdate[0])
                        self.p2p_defUpdate.remove(self.p2p_defUpdate[0])
                    tsclean=True
                    while tsclean:
                        tsclean=False
                        if len(self.p2p_defUpdate)>1:
                            if "UPDATE settings SET setting_value" in self.p2p_defUpdate[0] and "WHERE setting_key='p2p_syncro'" in self.p2p_defUpdate[0] and "UPDATE settings SET setting_value" in self.p2p_defUpdate[1] and "WHERE setting_key='p2p_syncro'" in self.p2p_defUpdate[1]:
                                id_instance1=self.p2p_defUpdate[0].split("'")[1].split(";")[0]
                                id_instance2=self.p2p_defUpdate[1].split("'")[1].split(";")[0]
                                ts1=self.p2p_defUpdate[0].split("'")[1].split(";")[1]
                                ts2=self.p2p_defUpdate[1].split("'")[1].split(";")[1]
                                if id_instance1==id_instance2:
                                    if abs((datetime.datetime.strptime(ts2,"%d/%m/%Y, %H:%M:%S")-datetime.datetime.strptime(ts1,"%d/%m/%Y, %H:%M:%S")).seconds)<10:
                                        self.p2p_defUpdate.remove(self.p2p_defUpdate[0])
                                        tsclean=True
                #rimozione ridondanze e inutilita' - end
                if len(self.p2p_defUpdate)>0:
                    self.syncroTaskUpdating=True
                    try:
                        cur.execute(self.p2p_defUpdate[0])
                        #if DEBUG_MODE: print self.p2p_defUpdate[0]
                        self.p2p_defUpdate.remove(self.p2p_defUpdate[0])
                    except:
                        if "COMMIT TRANSACTION" not in self.p2p_defUpdate[0]:
                            print "Deferred query error: "+self.p2p_defUpdate[0]+" - requeueing..."
                            toolong=True
                        else:
                            self.p2p_defUpdate.remove(self.p2p_defUpdate[0])
                    time.sleep(0.01)
                    if (time.time()-starttime) > 1.0: toolong=True

            #statusbar - inizio
            description=""
            #if self.p2p_state!="": description=description+" --- P2P status: "+self.p2p_state
            #if self.p2p_message!="": description=description+" ("+self.p2p_message+")"
            if self.p2p_message!="": description=self.p2p_message
            if len(self.p2p_defUpdate)>0: 
                description=description+" (updating)"
                #if DEBUG_MODE: print str(len(self.p2p_defUpdate))+" sql queries pending, first one: "+self.p2p_defUpdate[0]
            else:
                self.syncroTaskUpdating=False
            if self.preDecrypt_waitstate>0:
                description=description+" "+str("*") * self.preDecrypt_waitstate
            #self.SetStatusText(time.asctime() + description,1)
            self.SetStatusText(description,1)
            graphstate={'':0,'A':1,'B':2,'B1':3,'C':4,'C1':5,'D':6,'D1':7,'E':8}
            #if "connected" in description or DEBUG_MODE:
            self.progbar.SetValue(graphstate[self.p2p_state])
            #statusbar - fine
                
        #if self.syncroTaskStop and (not self.autoCloseRequest) and (self.frameLogin.GetPosition().x==0 or self.frameSearch.GetPosition().x==0 or self.frameDataExtraction.GetPosition().x==0):
        if self.syncroTaskStop and (not self.autoCloseRequest):
            if wx.MessageBox(TT("Heavybase has been updated and needs to be restarted. Can it be closed now?"), TT("System updated"), wx.ICON_INFORMATION | wx.YES_NO)==wx.YES:
                if self.DoCloseHeavyBase(): self.Destroy()
            self.autoCloseRequest = True
            
        #if sys.platform == 'win32':
        if os.name=='nt':
            if self.minAutologout!="": self.OnIdle(None)
        
        self.timer.Start(500,True)

    def OnDataEntry(self, evt):
        wx.BeginBusyCursor()
        self.child = HeavyBaseFrame(self, TT("Data management"))
        self.child.Show()
        #self.child.Update()
        wx.EndBusyCursor()

    def OnSettings(self, evt):
        test=None
        try:
            test=self.settings.settingsPanel
        except:
            pass
        if test==None:
            self.syncroTaskPause=True
            self.settings = SettingsFrame(self, TT("Network settings"))
            self.settings.Show()
    
    def OnExportDb(self, evt):
        wildcard = "HB files (*.hb)|*.hb"
        dlg = wx.FileDialog(
            self, message="Save file as ...", defaultFile=PROGNAME.lower()+".hb", wildcard=wildcard, style=wx.SAVE | wx.OVERWRITE_PROMPT
            )
        dlg.SetFilterIndex(0)
        export_ok=False
        if dlg.ShowModal() == wx.ID_OK:
            path = dlg.GetPath()
            
            f_in = open(PROGNAME.lower()+".db", 'rb')
            content_in = f_in.read()
            f_in.close()
            zipstring=HB_ZipString(content_in)
            f_out = open(path, 'wb')
            f_out.write(zipstring)
            f_out.close()
            export_ok=True
        dlg.Destroy()
        if export_ok: wx.MessageBox("File '"+path+"' created.", TT("Export DB"), wx.ICON_INFORMATION | wx.OK, self)
        
    def OnSendScreenshot(self,evt):
        serverurl = GetSetting(self,'heavybase_server')
        projectname = GetSetting(self,'project_name')
        currelease = GetSetting(self,'heavybase_release')
        username='heavybase'
        password='heavybase'
        connectionMode=GetSetting(self,'Network_Connection_Mode')
        arrSettings=self.arrSettings
        curConnMode, p2p_server, res = TryConnect(connectionMode,arrSettings,serverurl,projectname,username,password)
        if curConnMode!='':
            wx.BeginBusyCursor()
            print curConnMode
            #online
            import uuid,zlib
            cur_host=str(uuid.getnode())
            #now = datetime.datetime.utcnow()
            #ts=now.strftime("%Y%m%d%H%M%S")
            #filename="shot_"+cur_host+"_"+ts+".png"
            filename="shot_"+cur_host+".png"
            filetype=wx.BITMAP_TYPE_PNG
            SaveScreenshot(filename,filetype)
            # Begin Please wait
            frame = wx.MiniFrame(None, -1, TT('Help Desk'), size=wx.Size(240,80))
            panel = wx.Panel(frame)
            lbl=wx.StaticText(label=TT('Wait'), parent=panel, style=wx.ALIGN_CENTRE_HORIZONTAL|wx.EXPAND)
            lbl.SetFont(wx.Font( 24, wx.SWISS, wx.NORMAL, wx.BOLD))
            sizer = wx.BoxSizer(wx.VERTICAL)
            sizer.Add(lbl, 0, wx.ALL, 10)
            panel.SetSizer(sizer)
            panel.Layout()
            frame.Show(1)
            frame.CenterOnParent()
            if os.name=='nt':
                import win32gui
                win32gui.PumpWaitingMessages()
            else:
                import time
                time.sleep(0.5)
            frame.Update()
            # End Please wait
            infile = open(filename,"rb")
            contents=infile.read()
            infile.close()
            os.unlink(filename)
            encoded = base64.b64encode(zlib.compress(contents))
            print "file name: "+filename+" - size: "+str(len(encoded))
            try:
                # screenshot
                res = p2p_server.uploadFile(projectname,username,password,filename,encoded)
                # logfile
                #if sys.platform == 'win32' and ("pythonw" in sys.executable) and not DEBUG_MODE:
                if os.name=='nt' and ("pythonw" in sys.executable) and not DEBUG_MODE:
                    encoded=""
                    try:
                        LOGFILE.close()
                        sys.stdout=sys.__stdout__
                        sys.stderr=sys.__stderr__
                        filename="lastlog.txt"
                        infile = open(filename,"r")
                        contents=infile.read()
                        infile.close()
                        encoded = base64.b64encode(zlib.compress(contents))
                        LOGFILE = open('lastlog.txt','w')
                        sys.stdout = LOGFILE
                        sys.stderr = LOGFILE
                    except: pass
                    if len(encoded)>0:
                        res = p2p_server.uploadFile(projectname,username,password,"lastlog_"+cur_host+".txt",encoded)
                    encoded=""
                    try:
                        filename="sysinfo.txt"
                        infile = open(filename,"r")
                        contents=infile.read()
                        infile.close()
                        encoded = base64.b64encode(zlib.compress(contents))
                    except: pass
                    if len(encoded)>0:
                        res = p2p_server.uploadFile(projectname,username,password,"sysinfo_"+cur_host+".txt",encoded)
                encoded=""
                #ok
                frame.Destroy()
                wx.EndBusyCursor()
                wx.MessageBox(TT("Screenshot sent successfully"), TT("Help Desk"), wx.ICON_INFORMATION | wx.OK, self)
            except:
                #timeout
                frame.Destroy()
                wx.EndBusyCursor()
                wx.MessageBox(PROGNAME+" "+TT("Timeout. Please retry later"), TT("Help Desk"), wx.ICON_EXCLAMATION | wx.OK, self)
            if curConnMode!='use MSIE': p2p_server=None
        else:
            #offline
            frame.Destroy()
            wx.EndBusyCursor()
            wx.MessageBox(PROGNAME+" "+TT("Offline. Please retry later"), TT("Help Desk"), wx.ICON_EXCLAMATION | wx.OK, self)

    def DoCloseHeavyBase(self):
        if self.autoRestartRequest:
            outfile = open("autorestart.cmd", 'w')
            outfile.write("\n")
            outfile.close()
            close_ok=True
        else:
            if self.syncroTaskStop and (not self.autoCloseRequest):
                close_ok=True
            else:
                close_ok=False
                if self.syncroTaskUpdating and len(self.p2p_defUpdate)>0:
                    dlg = wx.MessageDialog(self, 
                        TT("Update in progress, this application should not be closed now. Do you wish to close it anyway?"),
                        TT("Confirm Exit"), wx.OK|wx.CANCEL|wx.ICON_ERROR)
                    result = dlg.ShowModal()
                    dlg.Destroy()
                    if result == wx.ID_OK: close_ok=True
                else:
                    dlg = wx.MessageDialog(self, 
                        TT("Do you really want to close this application?"),
                        TT("Confirm Exit"), wx.OK|wx.CANCEL|wx.ICON_QUESTION)
                    result = dlg.ShowModal()
                    dlg.Destroy()
                    if result == wx.ID_OK: close_ok=True
        if close_ok:
            print "Shutdown in progress..."
#            if os.name=='nt':
#                sys.stdout=sys.__stdout__
#                sys.stderr=sys.__stderr__
            hwnd=0
            if os.name=='nt':
                if IsRunningFromRemovable()==False:
                    hwnd=self.taskbarIconHwnd
            job=job_server_local.submit(AsyncKillAll,(hwnd,pp_fnames), (log,), ())

            self.syncroTask.stop()
            print "Syncro Task closed"
            try:
                if os.name=='nt':
                    #p2p_server = IEconn.getInstance(GetSetting(self,"heavybase_server"))
                    #p2p_server.close()
                    singletonmixin.forgetAllSingletons()
                    print "MSIE closed."
            except:
                print "ERROR closing MSIE."
                
            try:
                job_server_local.destroy()
                job_server.destroy()
                print "Multiprocessing closed."
            except:
                print "ERROR closing Multiprocessing instances."
            try:
                pp_server.terminate()
                #pp_server.wait()
                print "PP_SERVER closed."
            except:
                print "ERROR closing PP_SERVER."

            try:
                ntlmaps_server.terminate()
                ###ntlmaps_server.kill()
                #ntlmaps_server.wait()
                print "NTLMAPS_SERVER closed."
            except:
                print "ERROR closing NTLMAPS_SERVER."

            if LoadCustomSetting("HeavyBaseService")!="0":
                HeavyBaseServiceTask.stop()
                print PROGNAME+"Service closed."
    #            try:
    #                heavybase_service.terminate()
    #                #heavybase_service.wait()
    #                print "HeavyBaseService closed."
    #            except:
    #                print "ERROR closing HeavyBaseService."

            if LoadCustomSetting("HeavyBaseServer")=="1":
                HeavyBaseServerTask.stop()
                print PROGNAME+"Server closed."
                
            if LoadCustomSetting("SqliteServer")=="1":
                SqliteServerTask.stop()
                print "SqliteServer closed."
                
            ## pp begin
            for fname in pp_fnames:
                if "/" in fname:
                    fname_nopath=fname.split("/")[1]
                else:
                    fname_nopath=fname
                try:
                    os.unlink(fname_nopath)
                    os.unlink(fname_nopath+"c")
                except: pass
            ## pp end
            #self.Close()
            
            # Cancellazione file inutile di AbiWord
            filename=REPORTS_PATH+"readme.abw"
            if os.path.exists(filename): os.unlink(filename)

            try: os.unlink("ship.ico")
            except: pass
            
            self.destroyTaskBar()
            return True
        else:
            return False

    def destroyTaskBar(self):
        if os.name=='nt':
            if IsRunningFromRemovable()==False:
                import win32gui
                try: win32gui.DestroyWindow(self.taskbarIconHwnd)
                except: pass
            if  ("pythonw" in sys.executable) and not DEBUG_MODE:
                sys.stdout=sys.__stdout__
                sys.stderr=sys.__stderr__
                try: LOGFILE.close()
                except: pass
        
    def reloadSettings(self):
        cur.execute("SELECT setting_key,setting_value FROM settings")
        for row in cur:
            self.arrSettings[row[0]]=row[1]

    def OnHide(self, event):
        if self.syncroTaskStop:
            self.Onclose(event)
        else:
            dlg = wx.MessageDialog(self, 
                        TT("Do you want to close the application window?"),
                        TT("Confirm Exit"), wx.OK|wx.CANCEL|wx.ICON_QUESTION)
            result = dlg.ShowModal()
            dlg.Destroy()
            if result == wx.ID_OK:
                self.LogoutCount=self.LogoutCount+1
                self.Hide()
            
    def OnClose(self, event):
        self.syncroTaskPause=True
        if self.DoCloseHeavyBase():
            self.Destroy()
        else:
            self.syncroTaskPause=False
            try: event.Veto()
            except: pass

[wx.ID_FRAMEBACK] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMELOGIN] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMESEARCH] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMERECORDSET] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMECRF] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONLOGIN] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONLOGOUT] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONSEARCH] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONADDNEW] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONPRINT] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONEXPORT] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONCLONE] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONSAVE] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONCOMMIT] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONCANCEL] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMEBUTTONDELETE] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_FRAMELSTFOUND] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_REPORTCHOICE] = map(lambda _init_ctrls: wx.NewId(), range(1))
[wx.ID_REPORTCHOICEBUTTON] = map(lambda _init_ctrls: wx.NewId(), range(1))

import urllib
from urllib import unquote, splittype, splithost
import xmlrpclib
class ProxyTransport(xmlrpclib.Transport):
    def __init__(self,proxy): 
        self.proxyurl = proxy 
        self._use_datetime = 0
                
    def request(self, host, handler, request_body, verbose=0):
        type, r_type = splittype(self.proxyurl)
        phost, XXX = splithost(r_type)
        puser_pass = None
        if '@' in phost:
            user_pass, phost = phost.split('@', 1)
            if ':' in user_pass:
                user, password = user_pass.split(':', 1)
                puser_pass = base64.encodestring('%s:%s' % (unquote(user),
                                                unquote(password))).strip()
        
        urlopener = urllib.FancyURLopener({'http':'http://%s'%phost})
        if not puser_pass:
            urlopener.addheaders = [('Content-Type', 'text/xml'),
                                    ('User-agent', self.user_agent)]
        else:
            urlopener.addheaders = [('Content-Type', 'text/xml'),
                                    ('User-agent', self.user_agent),
                                    ('Proxy-authorization', 'Basic ' + puser_pass) ]

        host = unquote(host)
        f = urlopener.open("http://%s%s"%(host,handler), request_body)

        self.verbose = verbose 
        return self.parse_response(f)

import singletonmixin
class IEconn(singletonmixin.Singleton):
    def __init__(self,serverurl):
        serverurl=serverurl.replace("xmlrpc","httprpc")     #backward compatibility
        self.serverurl="http://"+serverurl.split(":")[0]+"/scooter"

#        import win32com
#        import win32api
#        import win32con
#        import exceptions
#        import win32com.client
#        import win32gui
#        from wx.lib.activexwrapper import MakeActiveXClass
#        IEmodule=win32com.client.gencache.EnsureModule('{EAB22AC0-30C1-11CF-A7EB-0000C05BAE0B}', 0,1,1)
#        InternetExplorerActiveXClass = MakeActiveXClass(IEmodule.WebBrowser,eventObj = panel)
#        self.WebBrowser = InternetExplorerActiveXClass(panel, -1)
#        self.WebBrowser.SetClientSize(wx.Size(0,0))
#        self.WebBrowser.Navigate2(self.serverurl)
#        while self.WebBrowser.ReadyState!=4:
#            win32gui.PumpWaitingMessages()
#            time.sleep(0.5)
        import time
        import win32gui
        import win32com.client
        ie_mod=win32com.client.gencache.EnsureModule('{EAB22AC0-30C1-11CF-A7EB-0000C05BAE0B}' ,0, 1, 1)
        class IE_Events(ie_mod.DWebBrowserEvents2):
             def OnNavigateComplete2(self, pDisp, URL):
                 #print 'OnNavigateComplete2:', URL
                 pass

        self.WebBrowser=win32com.client.DispatchWithEvents('InternetExplorer.Application',IE_Events)
        self.WebBrowser.Visible=False
        self.WebBrowser.Navigate(self.serverurl)
        while self.WebBrowser.ReadyState!=4:
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)

    def close(self):
        self.WebBrowser.Quit()

    def __del__(self):
        self.WebBrowser.Quit()

    def getRelease(self, project_name, username, password):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="getRelease"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_filename.value=""
        doc.frm.heavybase_encoded.value=""
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        return doc.all.results.innerHTML
    def checkUpdates(self, project_name, username, password):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="checkUpdates"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_filename.value=""
        doc.frm.heavybase_encoded.value=""
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        #print doc.all.results.innerHTML
        return doc.all.results.innerHTML
    def downloadFile(self, project_name, username, password, filename):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="downloadFile"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_filename.value=filename
        doc.frm.heavybase_encoded.value=""
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        return doc.all.results.innerHTML
    def uploadFile(self, project_name, username, password, filename, encoded):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="uploadFile"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_filename.value=filename
        doc.frm.heavybase_encoded.value=encoded
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        return doc.all.results.innerHTML
    def p2p(self, project_name, username, password, id_instance, state, partner, encoded):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="p2p"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_id_instance.value=id_instance
        doc.frm.heavybase_state.value=state
        doc.frm.heavybase_partner.value=partner
        doc.frm.heavybase_encoded.value=encoded
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        #print doc.all.results.innerHTML
        return doc.all.results.innerHTML
    def sendMail(self, project_name, username, password, sender, comma_sep_receivers, subject, body):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="sendMail"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_sender.value=sender
        doc.frm.heavybase_comma_sep_receivers.value=comma_sep_receivers
        doc.frm.heavybase_subject.value=subject
        doc.frm.heavybase_body.value=body
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        return doc.all.results.innerHTML
    def findPeers(self, project_name, username, password, id_instance, internal_ip, use_proxy, client_ip):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="findPeers"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.frm.heavybase_id_instance.value=id_instance
        doc.frm.heavybase_heavybase_sender.value=sender
        doc.frm.heavybase_internal_ip.value=internal_ip
        doc.frm.heavybase_use_proxy.value=use_proxy
        doc.frm.heavybase_client_ip.value=client_ip
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        return doc.all.results.innerHTML
    def getServiceInstance(self):
        import time,win32gui
        doc=self.WebBrowser.Document
        doc.frm.heavybase_command.value="getServiceInstance"
        doc.frm.heavybase_project_name.value=project_name
        doc.frm.heavybase_username.value=username
        doc.frm.heavybase_password.value=password
        doc.all.results.innerHTML="wait"
        doc.frm.submit()
        wait=True
        while wait==True:
            try:
                if doc.all.results.innerHTML!="wait": wait=False
            except: pass
            win32gui.PumpWaitingMessages()
            time.sleep(0.5)
        return doc.all.results.innerHTML

class HTTPconn():
    def __init__(self, serverurl, useproxy, proxy_host, proxy_port, proxy_username, proxy_password):
        serverurl=serverurl.replace("xmlrpc","httprpc")     #backward compatibility
        self.host=serverurl.split(":")[0]
        self.port=80
        self.serverurl="http://"+self.host+":"+`self.port`+"/scooter"
        self.useproxy=useproxy
        self.proxy_host=proxy_host
        self.proxy_port=proxy_port
        self.proxy_username=proxy_username
        self.proxy_password=proxy_password

    def post_multipart(self, fields, files):
        """
        Post fields and files to an http host as multipart/form-data.
        fields is a sequence of (name, value) elements for regular form fields.
        files is a sequence of (name, filename, value) elements for data to be uploaded as files
        Return the server's response page.
        """
        import urllib2
        import urllib
        import time
        import httplib 
        content_type, body = self.encode_multipart_formdata(fields, files)
        if self.useproxy:
            h = httplib.HTTP(self.proxy_host, self.proxy_port)
            h.putrequest('POST', self.serverurl)
            h.putheader('Host', self.host)
            h.putheader('content-type', content_type)
            h.putheader('content-length', str(len(body)))
            if self.proxy_username!="":
                import base64
                auth = base64.encodestring(self.proxy_username + ":" + self.proxy_password)
                h.putheader('Proxy-Authorization', '''Basic %s''' % auth)
        else:
            h = httplib.HTTP(self.host, self.port)
            h.putrequest('POST', '/scooter')
            h.putheader('Host', self.host)
            h.putheader('content-type', content_type)
            h.putheader('content-length', str(len(body)))
        h.endheaders()
        h.send(body)
        h.getreply()
        return h.file.read()

    def encode_multipart_formdata(self, fields, files):
        """
        fields is a sequence of (name, value) elements for regular form fields.
        files is a sequence of (name, filename, value) elements for data to be uploaded as files
        Return (content_type, body) ready for httplib.HTTP instance
        """
        BOUNDARY = '----------ThIs_Is_tHe_bouNdaRY_$'
        CRLF = '\r\n'
        L = []
        if fields:
            for (key, value) in fields:
                L.append('--' + BOUNDARY)
                L.append('Content-Disposition: form-data; name="%s"' % key)
                L.append('')
                L.append(value)
        if files:
            for (key, filename, value) in files:
                L.append('--' + BOUNDARY)
                L.append('Content-Disposition: form-data; name="%s"; filename="%s"' % (key, filename))
                L.append('Content-Type: %s' % get_content_type(self,filename))
                L.append('')
                L.append(value)
        L.append('--' + BOUNDARY + '--')
        L.append('')
        body = CRLF.join(L)
        content_type = 'multipart/form-data; boundary=%s' % BOUNDARY
        return content_type, body

    def get_content_type(self, filename):
        import mimetypes
        return mimetypes.guess_type(filename)[0] or 'application/octet-stream'

    def connection(self,q):
        #print  post_multipart(HOST, PORT, q, (('query','query','Query'), ), )
        res=""
        try:
            #print q
            res=self.post_multipart(q, None)
            #print "res="+res
            if res!="":
                tag_begin="<div name='results' id='results' style='visibility:hidden'>"
                tag_end="</div>"
                res=res[res.find(tag_begin)+len(tag_begin):]
                res=res[:res.find(tag_end)]
        except: 
            pass
        return res        
    
    def getRelease(self, project_name, username, password):
        q=[]
        q.append(('heavybase_command', 'getRelease'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        return self.connection(q)    

    def checkUpdates(self, project_name, username, password):
        q=[]
        q.append(('heavybase_command', 'checkUpdates'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        return self.connection(q)    

    def downloadFile(self, project_name, username, password, filename):
        q=[]
        q.append(('heavybase_command', 'downloadFile'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        q.append(('heavybase_filename', filename))
        return self.connection(q)    

    def uploadFile(self, project_name, username, password, filename, encoded):
        q=[]
        q.append(('heavybase_command', 'downloadFile'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        q.append(('heavybase_filename', filename))
        q.append(('heavybase_encoded', encoded))
        return self.connection(q)    

    def p2p(self, project_name, username, password, id_instance, state, partner, encoded):
        q=[]
        q.append(('heavybase_command', 'p2p'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        q.append(('heavybase_id_instance', id_instance))
        q.append(('heavybase_state', state))
        q.append(('heavybase_partner', partner))
        q.append(('heavybase_encoded', encoded))
        return self.connection(q)    

    def sendMail(self, project_name, username, password, sender, comma_sep_receivers, subject, body):
        q=[]
        q.append(('heavybase_command', 'sendMail'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        q.append(('heavybase_sender', sender))
        q.append(('heavybase_comma_sep_receivers', comma_sep_receivers))
        q.append(('heavybase_subject', subject))
        q.append(('heavybase_body', body))
        return self.connection(q)    

    def findPeers(self, project_name, username, password, id_instance, internal_ip, use_proxy, client_ip):
        q=[]
        q.append(('heavybase_command', 'findPeers'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        q.append(('heavybase_id_instance', id_instance))
        q.append(('heavybase_internal_ip', internal_ip))
        q.append(('heavybase_use_proxy', use_proxy))
        q.append(('heavybase_client_ip', client_ip))
        return self.connection(q)    
    
    def getServiceInstance(self, project_name, username, password):
        q=[]
        q.append(('heavybase_command', 'getServiceInstance'))
        q.append(('heavybase_project_name', project_name))
        q.append(('heavybase_username', username))
        q.append(('heavybase_password', password))
        return self.connection(q)    

class CustomRouter():
    def __init__(self,arrSettings,serverurl): 
        self.serverurl = serverurl
        self.customTransport=None
        self.Network_Proxy_Setting   = arrSettings.get("Network_Proxy_Setting")
        self.Network_Proxy_Host      = arrSettings.get("Network_Proxy_Host")
        self.Network_Proxy_Port      = arrSettings.get("Network_Proxy_Port")
        self.Network_Proxy_Username  = arrSettings.get("Network_Proxy_Username")
        self.Network_Proxy_Password  = arrSettings.get("Network_Proxy_Password")
        self.Network_Connection_Mode = arrSettings.get("Network_Connection_Mode")
        self.useproxy=False

    def GetSettings(self):    
        self.useproxy=False
        if self.Network_Proxy_Setting=="" or self.Network_Proxy_Setting=="automatic":
            autoProxySetting=GetSystemProxy()
            if autoProxySetting[0]==1:
                #proxy=autoProxySetting[1][7:]
                #self.Network_Proxy_Host=proxy.split(":")[0]
                #self.Network_Proxy_Port=proxy.split(":")[1]
                self.Network_Proxy_Host="localhost"
                self.Network_Proxy_Port="60001"
                #self.Network_Proxy_Username=""
                #self.Network_Proxy_Password=""
                self.useproxy=True
        elif self.Network_Proxy_Setting=="manual proxy":
            if self.Network_Proxy_Host!="" and self.Network_Proxy_Port!="":
                self.useproxy=True
        elif self.Network_Proxy_Setting=="NTLM proxy":
            self.Network_Proxy_Host="localhost"
            self.Network_Proxy_Port="60001"
            self.useproxy=True

    def DoOpen(self):
        self.GetSettings()
        self.customTransport=None
        if self.useproxy==True:
            ProxyURL=self.Network_Proxy_Host+":"+str(self.Network_Proxy_Port)
            if self.Network_Proxy_Username!="" or self.Network_Proxy_Password!="":
                ProxyURL=self.Network_Proxy_Username+":"+self.Network_Proxy_Password+"@"+ProxyURL
            self.customTransport = ProxyTransport("http://"+ProxyURL)
        else:
            pass

    def DoClose(self):
        pass

def TryConnect(connectionMode,arrSettings,serverurl,projectname,username,password,id_instance=""):
    if connectionMode=="" or connectionMode=="automatic":
        if os.name=='nt': 
            tryConnModes=["use xmlrpc","use http","use MSIE"]
        else:
            tryConnModes=["use xmlrpc","use http"]
    else:
        tryConnModes=[connectionMode]

    customRouting=CustomRouter(arrSettings,serverurl)
    customRouting.GetSettings()

    p2p_server=None
    curConnMode=""
    res=""
    for connectionMode in tryConnModes:
        if connectionMode=="use xmlrpc":
            import xmlrpclib, zlib
            #customRouting=CustomRouter(arrSettings,serverurl)
            customRouting.DoOpen()
            p2p_server = xmlrpclib.ServerProxy("http://"+customRouting.serverurl, transport = customRouting.customTransport)
        elif connectionMode=="use http":
            #customRouting=CustomRouter(arrSettings,serverurl)
            #customRouting.GetSettings()
            p2p_server = HTTPconn(serverurl, customRouting.useproxy, customRouting.Network_Proxy_Host, customRouting.Network_Proxy_Port, 
                                              customRouting.Network_Proxy_Username, customRouting.Network_Proxy_Password)
        elif connectionMode=="use MSIE":
            try:
                p2p_server = IEconn.getInstance(serverurl)
                win32gui.PumpWaitingMessages()
                time.sleep(0.1)
            except:
                try:
                    import win32gui
                    singletonmixin.forgetAllSingletons()
                    win32gui.PumpWaitingMessages()
                    time.sleep(0.1)
                    p2p_server = IEconn.getInstance(serverurl)
                    win32gui.PumpWaitingMessages()
                    time.sleep(0.1)
                except: pass
        try: res = p2p_server.getRelease(projectname,username,password)
        except: 
            if connectionMode=="use MSIE":
                try:
                    p2p_server.WebBrowser.Quit()        #Tentativo x nascondere IE a Windows Vista
                    singletonmixin.forgetAllSingletons()
                    win32gui.PumpWaitingMessages()
                    time.sleep(0.1)
                    p2p_server = IEconn.getInstance(serverurl)
                    win32gui.PumpWaitingMessages()
                    time.sleep(0.1)
                    res = p2p_server.getRelease(projectname,username,password)
                except: pass
            else: pass

        if len(res.split("."))==3:
            curConnMode=connectionMode

            if id_instance!="" and False:
                #Internal IP discovery - begin
                test_ip_host=serverurl.split(":")[0]
                test_ip_port=serverurl.split(":")[1]
                if customRouting.useproxy: 
                    test_ip_host=customRouting.Network_Proxy_Host
                    test_ip_port=customRouting.Network_Proxy_Port
                import socket
                s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
                s.connect((test_ip_host,int(test_ip_port)))
                internal_ip=s.getsockname()
                client_ip=""
                use_proxy=""
                if customRouting.useproxy: use_proxy="1"
                peers=p2p_server.findPeers(projectname,username,password,id_instance, internal_ip, use_proxy, client_ip)
                print str(peers)
                #Internal IP discovery - End
            break
    return curConnMode,p2p_server,res

# ANALYSER - BEGIN
def custom_listdir(path):
    dirs = sorted([d for d in os.listdir(path) if os.path.isdir(path + os.path.sep + d)])
    dirs.extend(sorted([f for f in os.listdir(path) if os.path.isfile(path + os.path.sep + f)]))
    return dirs

def initMergeDb(cur):
    cur.execute('PRAGMA auto_vacuum = 1')
    try:
        cur.execute('CREATE TABLE headers (id_header INTEGER, child_of INTEGER, description TEXT, label TEXT, pos INTEGER, hap TEXT, stype TEXT, id_section INTEGER, deftype INTEGER, cyclic INTEGER, doubleinput INTEGER, typedesc TEXT, subtypedesc TEXT, textlength TEXT, subtextlength TEXT, textalign TEXT, freetext INTEGER, defaultvalue TEXT, validation TEXT, onchange TEXT)')
        cur.execute('CREATE INDEX idx_headers ON headers (id_header,pos,description)')
        cur.execute('CREATE TABLE rows (id_row INTEGER, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER, rap TEXT, hpath TEXT, id_instance INTEGER)')
        cur.execute('CREATE TABLE contents (id_row INTEGER, id_header INTEGER, data BLOB, id_user INTEGER, id_instance INTEGER)')
    except:
        pass

def square(con_r,con_w,name,fields,id_group,idApp=0,keyfields="",adminmode=False,reqstatus=0):
    #if keyfields!="": keyfields=keyfields.replace("-","___")
    cur_r = con_r.cursor()
    cur_r.execute('PRAGMA temp_store =  MEMORY')
    cur_r.execute('PRAGMA journal_mode = OFF')
    cur_w = con_w.cursor()
    cur_w.execute('PRAGMA temp_store =  MEMORY')
    cur_w.execute('PRAGMA journal_mode = OFF')
    #init
    if idApp==0:
        cur_r.execute("SELECT id_header FROM headers WHERE child_of=0")
        row = cur_r.fetchone()
        if row!=None: 
            idApp=row[0]
            
    #keyname="key"
    #cur_r.execute("SELECT setting_value FROM settings WHERE setting_key='keyname'")
    #row = cur.fetchone()
    #if row!=None: keyname=row[0]
            
    #contents_dictionary
    validheaders={}
    dictionary={}
    cur_r.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary")
    for row in cur_r:
        dictionary[str(row[0])+"-"+str(row[1])]=row[2]
        validheaders[row[0]]=True
    #hashes
    hashHead={}
    descHead={}
    #cur_r.execute("SELECT id_header,description FROM headers WHERE child_of="+str(idApp)+" ORDER BY pos,description,id_header")
    cur_r.execute("SELECT id_header,description FROM headers WHERE child_of="+str(idApp))
    for row in cur_r:
        hashHead[row[1]]=row[0]
        descHead[row[0]]=row[1]
    idHead=[]
    if keyfields!="":
        # Test - Begin
        # #Valid fields - Begin
        # tmpFields=[]
        # for field in fields:
        #     if validheaders.has_key(hashHead[field.split("#")[0]]): tmpFields.append(field)
        # fields=tmpFields
        # Test - End
        
        arrkeyfields=keyfields.split(",")
        tmpFields=[]
        for field in arrkeyfields:
            if validheaders.has_key(hashHead[field.split("#")[0]]): tmpFields.append(field)
        arrkeyfields=tmpFields
        keyfields=",".join(arrkeyfields)
        
        for keyfield in arrkeyfields:
            if keyfield not in fields: fields.append(keyfield)
        #print keyfields
    #Valid fields - End
    for i in range(0,len(fields)):
        fldPartName=fields[i]
        if fldPartName.find("#")>=0: fldPartName = fldPartName[:fldPartName.find("#")]
        #idHead.append(hashHead[fields[i]])
        if hashHead.has_key(fldPartName):
            idHead.append(hashHead[fldPartName])
    #create table
    fieldsList=""
    fieldsListId=""
    try: cur_w.execute("DROP TABLE "+name+"_full")
    except: pass
    try: cur_w.execute("DROP TABLE "+name)
    except: pass
    qy=" (id_row INTEGER, id_user INTEGER, id_instance INTEGER, date_time TEXT"
    for i in range(0,len(fields)):
        fldName=fields[i]
        if fldName.find("#")>=0: fldName = fldName[:fldName.find("#")] + "__" + fldName[fldName.find("#")+1:]
        qy=qy+", "
        #qy=qy+fields[i]+" BLOB"
        qy=qy+fldName.replace("-","___")+" BLOB"
        #if fieldsList!="": 
            #fieldsList=fieldsList+","
            #fieldsListId=fieldsListId+","
        #fieldsList=fieldsList+fldName
        fldPartName=fields[i]
        if fldPartName.find("#")>=0: fldPartName = fldPartName[:fldPartName.find("#")]
        #fieldsListId=fieldsListId+`hashHead[fields[i]]`
        if hashHead.has_key(fldPartName):
            if fieldsList!="": 
                fieldsList=fieldsList+","
                fieldsListId=fieldsListId+","
            fieldsList=fieldsList+fldName
            fieldsListId=fieldsListId+str(hashHead[fldPartName])
    qy=qy+")"
    qynow="CREATE TABLE "+name+"_full"+qy
    qylater="CREATE TABLE "+name+qy
    #if DEBUG_MODE: print qy
    cur_w.execute(qynow)
    #populate table
    qy=""
    hashValues={}
    curKey=""
    newKey=""
    #sql="SELECT contents_index.id_row, contents_index.id_header, contents_index.id_dictionary, contents_index.id_user, rows.rap, contents_index.id_instance, contents_index.id_cycle, rows.date_time FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header in ("+fieldsListId+") AND rows.id_header="+`idApp`+" AND contents_index.id_cycle=0 ORDER BY contents_index.id_row, contents_index.id_user, contents_index.id_instance, contents_index.id_header"
    if reqstatus==0:
        sql="SELECT contents_index.id_row, contents_index.id_header, contents_index.id_dictionary, contents_index.id_user, rows.rap, contents_index.id_instance, contents_index.id_cycle, rows.date_time FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header in ("+fieldsListId+") AND rows.id_header="+str(idApp)+" ORDER BY contents_index.id_row, contents_index.id_user, contents_index.id_instance, contents_index.id_header, contents_index.id_cycle"
    elif reqstatus==-1:
        sql="SELECT contents_index.id_row, contents_index.id_header, contents_index.id_dictionary, contents_index.id_user, rows1.rap as rap, contents_index.id_instance, contents_index.id_cycle, rows1.date_time as date_time FROM rows as rows1, rows as rows2, contents_index WHERE rows1.id_row=contents_index.id_row AND rows1.id_user=contents_index.id_user AND rows1.id_instance=contents_index.id_instance AND contents_index.id_header in ("+fieldsListId+") AND rows1.id_header="+str(idApp)+" AND rows1.status=rows2.id_row AND rows1.status_user=rows2.id_user AND rows1.status_instance=rows2.id_instance AND rows2.id_header="+str(idApp)+" AND rows2.status=-1 ORDER BY contents_index.id_row, contents_index.id_user, contents_index.id_instance, contents_index.id_header, contents_index.id_cycle"
    elif reqstatus==1:
        sql="SELECT contents_index.id_row, contents_index.id_header, contents_index.id_dictionary, contents_index.id_user, rows.rap, contents_index.id_instance, contents_index.id_cycle, rows.date_time FROM rows,contents_index WHERE rows.status>0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header in ("+fieldsListId+") AND rows.id_header="+str(idApp)+" ORDER BY contents_index.id_row, contents_index.id_user, contents_index.id_instance, contents_index.id_header, contents_index.id_cycle"
    cur_r.execute(sql);

    for row in cur_r:
        currap=row[4]
        if currap==None: currap=""
        if currap=="" or str(id_group) in currap.split(";")[0].split(","):
            newKey=str(row[0])+"-"+str(row[3])+"-"+str(row[5])
            if newKey!=curKey:
                curKey=newKey
                if len(hashValues)>0:
                    for i in range(0,len(fields)):
                        fldPartName=fields[i]
                        fldPostName="0"
                        if fldPartName.find("#")>=0: 
                            fldPostName = fldPartName[fldPartName.find("#")+1:]
                            fldPartName = fldPartName[:fldPartName.find("#")]
                        if hashHead.has_key(fldPartName):
                            if hashValues.has_key(str(hashHead[fldPartName])+"-"+fldPostName):
                                qy=qy+", '"+hashValues[str(hashHead[fldPartName])+"-"+fldPostName]+"'"
                            else:
                                qy=qy+", ''"
                    qy=qy+")"
                    #valid_record=False
                    #if adminmode:
                        #valid_record=True
                    #else:
                        #if (keyname in hashValues) and ("group_shortcode" in hashValues):
                            #if hashValues[keyname]!="" and hashValues["group_shortcode"]!="":
                                #valid_record=True
                    #if valid_record:
                    try: cur_w.execute(qy)
                    except: print "ERROR in query:\n"+qy
                hashValues={}
                qy=""
                qy=qy+"INSERT INTO "+name+"_full"+" (id_row,id_user,id_instance,date_time,"+fieldsList.replace("-","___")+") VALUES ("
                qy=qy+str(row[0])+", "+str(row[3])+", '"+str(row[5])+"', '"+str(row[7])+"'"
            keyval=str(row[1])+"-"+str(row[6])
            if not hashValues.has_key(keyval):
                if dictionary.has_key(str(row[1])+"-"+str(row[2])):
                    hashValues[keyval]=GetSqlValue(dictionary[str(row[1])+"-"+str(row[2])])
                else:
                    hashValues[keyval]=""

    if len(hashValues)>0:
        for i in range(0,len(fields)):

            fldPartName=fields[i]
            fldPostName="0"
            if fldPartName.find("#")>=0: 
                fldPostName = fldPartName[fldPartName.find("#")+1:]
                fldPartName = fldPartName[:fldPartName.find("#")]

            if hashHead.has_key(fldPartName):
                if hashValues.has_key(str(hashHead[fldPartName])+"-"+fldPostName):
                    qy=qy+", '"+hashValues[str(hashHead[fldPartName])+"-"+fldPostName]+"'"
                else:
                    qy=qy+", ''"
        qy=qy+")"
        try:
            cur_w.execute(qy)
        except:
            print "ERROR in query:\n"+qy
    #nodups - begin
    cur_w.execute(qylater)
    if not adminmode:
        newfieldsList=""
        for elm in fieldsList.split(","):
            if newfieldsList!="": newfieldsList+=","
            newfieldsList += name+"_full."+elm
        
        if keyfields=="":
            qylater="INSERT INTO "+name+" (id_row,id_user,id_instance,date_time,"+fieldsList.replace("-","___")+") SELECT id_row,id_user,id_instance,"+name+"_full.date_time,"+newfieldsList.replace("-","___")+" FROM "+name+"_full, (SELECT max(date_time) as max_date_time,"+fieldsList.replace("-","___")+" FROM "+name+"_full"+" GROUP BY "+fieldsList.replace("-","___")+") AS midtable"
        else:
            tmpGroupBy=""
            if keyfields!="": tmpGroupBy=" GROUP BY "+keyfields
            qylater="INSERT INTO "+name+" (id_row,id_user,id_instance,date_time,"+fieldsList.replace("-","___")+") SELECT id_row,id_user,id_instance,"+name+"_full.date_time,"+newfieldsList.replace("-","___")+" FROM "+name+"_full, (SELECT max(date_time) as max_date_time,"+keyfields.replace("-","___")+" FROM "+name+"_full"+tmpGroupBy+") AS midtable"
            
        qylater+=" WHERE "+name+"_full.date_time=midtable.max_date_time"
            
        qylatersuffix=""
        flds=[]
        if keyfields=="":
            flds=fieldsList.split(",")
        else:
            flds=keyfields.split(",")
        for elm in flds:
            qylatersuffix += " AND "+name+"_full."+elm.replace("-","___")+"=midtable."+elm.replace("-","___")
        
        qylater+=qylatersuffix
    else:
        qylater="INSERT INTO "+name+" SELECT * FROM "+name+"_full"
    #print qylater
    cur_w.execute(qylater)
        
    cur_w.execute("DROP TABLE "+name+"_full")
    #nodups - end
    cur_w.close()
    cur_r.close()
    
def log(string,mode="a"):
    import datetime
    now = datetime.datetime.utcnow()
    ts=now.strftime("%Y%m%d%H%M%S")
    try:
        logfile = open("log.txt", mode)
        try:
            logfile.write(ts+"\t"+string+"\n")
        finally:
            logfile.close()
    except IOError:
        pass
    
# Copia Dati - Inizio
class DataCopier():
    def __init__(self, parent, DATABASE, digestkey, trans, id_group, id_crf):

        self.parent=parent
        self.DATABASE=DATABASE
        self.digestkey=digestkey
        self.trans=trans
        self.id_group=id_group
        self.id_crf=id_crf
        
    def col(self,row,hc,colname):
        cData=""
        try:cData=row[hc[colname]]
        except: pass
        if type(cData).__name__=="int": 
            dData=cData
        else:
            if cData!="":
                if not self.trans.has_key(cData):
                    try: dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(cData)),"latin-1")
                    except: dData=cData
                    self.trans[cData]=dData
                else:
                    dData=self.trans[cData]
                #dData=dData.encode('ascii', 'replace')
            else:
                dData=""
        return dData
        
    def datavalida(self,data):
        ret=False
        try:
            import time
            a=time.strptime(data, "%d/%m/%Y")
            ret=True
        except: pass
        return ret

    def DoCopy(self, group_shortcode,cod_paz,data_ins, testfld,curcycle,maxcycle,keyfields,copyfldList):
        wx.BeginBusyCursor()
        con = sqlite.connect(self.DATABASE, isolation_level=None, timeout=60.0)
        cur = con.cursor()
        con_m = sqlite.connect(":memory:")
        cur_m = con_m.cursor()
        
        res={}
            
        availfields=[]
        cur = con.cursor()
        cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND pos is not null GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
        for row in cur:
            availfields.append(row[0])
        if len(availfields)==0:
            wx.MessageBox("No forms fullfilled.", "Data Copier", wx.ICON_ERROR | wx.OK, self)
        else:
            lstfields=[]
            for elm in keyfields:
                lstfields.append(elm)
            for i in range(1,maxcycle+1):
                lstfields.append(testfld+"#"+str(i))
                for elm in copyfldList:
                    fld=elm+"#"+str(i)
                    if fld in availfields:
                        lstfields.append(fld)
                        #print fld
                        
            square(con,con_m,"search",lstfields,self.id_group,self.id_crf,','.join(keyfields))
            hc={}; i=4
            for col in lstfields: hc[col]=i; i=i+1
            
            try:
                found=False
                cur_m.execute("SELECT data_ins FROM search WHERE group_shortcode='"+group_shortcode+"' AND cod_paz='"+cod_paz+"'")
                cols = [i[0] for i in cur_m.description]; hc={}; i=0
                for col in cols: hc[col]=i; i=i+1
                for row in cur_m:
                    cdata_ins=row[hc["data_ins"]]
                    ddata_ins=""
                    try: ddata_ins=self.col(row,hc,"data_ins")
                    except: pass
                    if len(ddata_ins)==10:
                        ddata_ins=ddata_ins[6:10]+"-"+ddata_ins[3:5]+"-"+ddata_ins[0:2]
                    cur_m.execute("UPDATE search SET date_time='"+ddata_ins+"' WHERE group_shortcode='"+group_shortcode+"' AND cod_paz='"+cod_paz+"' AND data_ins='"+cdata_ins+"'")
            except: pass
            
            cur_m.execute("SELECT * FROM search WHERE group_shortcode='"+group_shortcode+"' AND cod_paz='"+cod_paz+"' ORDER BY date_time DESC")
            cols = [i[0] for i in cur_m.description]; hc={}; i=0
            for col in cols: hc[col]=i; i=i+1
            cycle=curcycle-1
            found=False
            for row in cur_m:
                while cycle>=1:
                    testval=""
                    try: testval=self.col(row,hc,testfld+"__"+str(cycle))
                    except: pass
                    print "retrieving cycle "+str(cycle)+" - testing field "+testfld+"__"+str(cycle)+" = "+row[hc[testfld+"__"+str(cycle)]]+" = "+testval
                    if self.datavalida(testval):
                        #print "found."
                        found=True
                        for elm in copyfldList:
                            res[elm]=self.col(row,hc,elm+"__"+str(cycle))
                            #print elm
                        break
                    cycle=cycle-1
                if found: break
                cycle=maxcycle
        
        wx.EndBusyCursor()
        return res
# Copia Dati - Fine

# Funzioni generali - Inizio
def Dummy():
    return True

def GetSqlValue(value):
    sqlValue=""
    for i in range(len(value)):
        if value[i]=="'":
            sqlValue+="''"
        else:
            sqlValue+=value[i]
    return sqlValue
    
def date1_le_date2(data1,data2):
    #Verifica se data1 e' precedente o uguale a data2
    d1=data1.ljust(10)
    d2=data2.ljust(10)
    d1=d1[6:10]+d1[3:5]+d1[0:2]
    d2=d2[6:10]+d2[3:5]+d2[0:2]
    if d1<=d2:
        return True
    else:
        return False

def valid_date_or_null(data):
    if data=="": return True
    else: return valid_date(data)
    
def valid_date(data):
    ret=False
    data=data.strip()
    if len(data)==10:
        try:
            if int("0"+data[0:2])>=1 and int("0"+data[0:2])<=31 and int("0"+data[3:5])>=1 and int("0"+data[3:5])<=12 and int("0"+data[6:10])>=1800 and int("0"+data[6:10])<=2100:
                import time
                a=time.strptime(data, "%d/%m/%Y")
                ret=True
        except: pass
    return ret
    
def date_diff(date1,date2):
    import datetime
    ret=None
    try:
        dd, mm, yyyy = (int(x) for x in date1.split('/'))
        date1b = datetime.date(yyyy, mm, dd)
        dd, mm, yyyy = (int(x) for x in date2.split('/'))
        date2b = datetime.date(yyyy, mm, dd)
        diff=date2b-date1b
        ret=diff.days
    except: pass
    return ret
    
def date_calc(date1,ndays):
    import datetime    
    ret=None
    try:
        dd, mm, yyyy = (int(x) for x in date1.split('/'));
        dateb = datetime.date(yyyy, mm, dd);
        datec=dateb+datetime.timedelta(days=ndays)
        ret=datec.strftime("%d/%m/%Y")
    except: pass
    return ret

def levenshtein_distance(first, second):
    if len(first) > len(second):
        first, second = second, first
    if len(second)==0:
        return len(first)
    #first=first.strip().lower()
    #second=second.strip().lower()
    first_length = len(first) + 1
    second_length = len(second) + 1
    distance_matrix = [range(second_length) for x in range(first_length)]
    for i in range(1, first_length):
        for j in range(1, second_length):
            deletion = distance_matrix[i-1][j] + 1
            insertion = distance_matrix[i][j-1] + 1
            substitution = distance_matrix[i-1][j-1]
            if first[i-1] != second[j-1]:
                substitution += 1
            distance_matrix[i][j] = min(insertion, deletion, substitution)
    return distance_matrix[first_length-1][second_length-1]
# Funzioni generali - Fine

#Warnings - Begin
from wx.lib.mixins.listctrl import CheckListCtrlMixin

class WarningsCheckListCtrl(wx.ListCtrl, CheckListCtrlMixin):
    def __init__(self, parent):
        wx.ListCtrl.__init__(self, parent, -1, style=wx.LC_REPORT)
        CheckListCtrlMixin.__init__(self)
        self.parent=parent
        self.Bind(wx.EVT_LIST_ITEM_ACTIVATED, self.OnItemActivated)
    def OnItemActivated(self, evt):
        self.ToggleItem(evt.m_itemIndex)
    # this is called by the base class when an item is checked/unchecked
    def OnCheckItem(self, index, flag):
        data = self.GetItemData(index)
        #title = self.parent.warndata[data][1]
        title=TT('Warnings')
        what=""
        if flag:
            if index in self.parent.blocking:
                self.ToggleItem(index)
                what = "blocked"
                wx.MessageBox(self.parent.blocking[index], TT('Warnings'), wx.ICON_ERROR | wx.OK, self)
            else:
                what = "checked"
                self.parent.warndatachecked[data]=True
        else:
            what = "unchecked"
            self.parent.warndatachecked[data]=False
        if DEBUG_MODE: print 'item "%s", at index %d was %s\n' % (title, index, what)

class WarningsFrame(wx.Dialog):
    def __init__(self, parent, title, warndata):
        wx.Dialog.__init__(self, None, title=title, size=(900,500))

        self.parent=parent
        self.title=title
        self.warndata=warndata
        self.warndatachecked={}
        self.blocking={}
        
        self.list = WarningsCheckListCtrl(self)
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(self.list, 1, wx.EXPAND)
        
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(self, label=TT('Ok'))
        closeButton = wx.Button(self, label=TT('Cancel'))
        exportButton = wx.Button(self, label=TT('Export'))
        hbox.Add(okButton)
        hbox.Add(closeButton, flag=wx.LEFT, border=5)
        hbox.Add(exportButton, flag=wx.RIGHT, border=5)

        sizer.Add(hbox, flag= wx.ALIGN_CENTER|wx.TOP|wx.BOTTOM, border=10)
        self.SetSizer(sizer)

        self.list.InsertColumn(0, TT('Type'))
        self.list.InsertColumn(1, TT('Page'))
        self.list.InsertColumn(2, TT('Field'))
        self.list.InsertColumn(3, TT('Description'))

        itemIndex=0
        for key, data in self.warndata.iteritems():
            index = self.list.InsertStringItem(sys.maxint, data[0])
            if "_warnings_" in self.parent.contents:
                if str(key) in self.parent.contents["_warnings_"].split(","):
                    self.warndatachecked[key]=True
                    self.list.CheckItem(index)
            self.list.SetStringItem(index, 1, data[1])
            self.list.SetStringItem(index, 2, data[2])
            self.list.SetStringItem(index, 3, data[3])
            if len(data)>4:
                if data[4]!="0":
                    self.blocking[itemIndex]=data[4]
            self.list.SetItemData(index, key)
            itemIndex=itemIndex+1
      
        self.list.SetColumnWidth(0, wx.LIST_AUTOSIZE)
        self.list.SetColumnWidth(1, wx.LIST_AUTOSIZE)
        self.list.SetColumnWidth(2, wx.LIST_AUTOSIZE)
        self.list.SetColumnWidth(3, wx.LIST_AUTOSIZE)
        #self.list.SetColumnWidth(2, 100)

#        for elm in self.parent.contents["_warnings_"].split(",")
#            key=int(elm)
        #self.list.CheckItem(4)
        #self.list.CheckItem(7)

        self.Bind(wx.EVT_LIST_ITEM_SELECTED, self.OnItemSelected, self.list)
        self.Bind(wx.EVT_LIST_ITEM_DESELECTED, self.OnItemDeselected, self.list)
        
        self.Bind(wx.EVT_BUTTON, self.OnConfirm, okButton)
        self.Bind(wx.EVT_BUTTON, self.OnClose, closeButton)
        self.Bind(wx.EVT_BUTTON, self.OnExport, exportButton)
        
        
    def OnItemSelected(self, evt):
        #print 'item selected: %s\n' % evt.m_itemIndex
        pass
        
    def OnItemDeselected(self, evt):
        #print 'item deselected: %s\n' % evt.m_itemIndex
        pass
        
    def OnExport(self, event):
        event.Skip()
        wx.BeginBusyCursor()
        try: w=xlwt.Workbook()
        except: w=Workbook()
        ws=w.add_sheet(self.title)
        ws.write(0, 0, TT('Key'))
        ws.write(0, 1, TT('Type'))
        ws.write(0, 2, TT('Page'))
        ws.write(0, 3, TT('Field'))
        ws.write(0, 4, TT('Description'))
        ws.write(0, 5, TT('Note'))
        y=1
        for key, data in self.warndata.iteritems():
            ws.write(y,0,key)
            for x in range(4):
                ws.write(y,x+1,data[x])
            if len(data)>4:
                if data[4]!="0":
                    ws.write(y,5,data[4])
            y=y+1
        reportname="lastWarnings.xls"
        w.save(REPORTS_PATH + reportname)
        wx.EndBusyCursor()
        if not assoc_open(REPORTS_PATH+reportname):
            wx.MessageBox("File '"+reportname+"' created.", self.title, wx.ICON_INFORMATION | wx.OK, self)
        
    def OnClose(self, e):
        self.Destroy()

    def OnConfirm(self, e):
        #self.parent.res=self.arrfound[self.lstFound.GetSelections()[0]]
        checkeditems=[]
        if "_warnings_" in self.parent.contents:
            checkeditems=self.parent.contents["_warnings_"].split(",")
        allchecked=True
        for key in self.warndata:
            if self.warndatachecked.has_key(key):
                if self.warndatachecked[key]==False: 
                    allchecked=False
                    if str(key) in checkeditems: checkeditems.delete(str(key))
                    #break
                else:
                    if key not in checkeditems: checkeditems.append(str(key))
            else:
                allchecked=False
                #break
        self.parent.contents["_warnings_"]=",".join(checkeditems)
        if not allchecked:
            wx.MessageBox(TT('All the items must be checked'), self.title, wx.ICON_ERROR | wx.OK, self)
        else:
            self.parent.outcome=True
            self.Destroy()
#Example:
# warnings = {
# 1 : ["blocco", "anagrafica", "data di nascita", "il campo non e' compilato"],
# 2 : ["avviso", "anagrafica", "data di nascita", "il paziente non e' maggiorenne"],
# }
# top = WarningsFrame("Warnings",warnings)
# top.Show()
#
#error
#warn
#memo
#
#errore
#avviso
#memo
#Warnings - End

#Metadata - Begin
class VarMetadataTabPanel(wx.Panel):
    def __init__(self, parent, hbclass, dialog, varname):
        wx.Panel.__init__(self, parent=parent, id=wx.ID_ANY)
        self.hbclass=hbclass
        self.dialog=dialog
        self.parent=parent
        self.varname=varname
        sizer = wx.BoxSizer(wx.VERTICAL)
        self.lbl1 = wx.StaticText( self, wx.ID_ANY, TT("Confidence"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.lbl1.Wrap( -1 )
        sizer.Add( self.lbl1, 0, wx.ALL, 2 )
        confidenceChoices = [ TT("Valid data"), TT("To be confirmed"), TT("Not available, type 1"), TT("Not available, type 2") ]
        self.confidence = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, confidenceChoices, 0 )
        if self.varname+"__metadata_confidence" in self.hbclass.contents:
            self.confidence.SetSelection(int(self.hbclass.contents[self.varname+"__metadata_confidence"]))
        else:
            self.confidence.SetSelection( 0 )
        sizer.Add(self.confidence, 0, wx.ALL, 2 )
        self.lbl2 = wx.StaticText( self, wx.ID_ANY, TT("Free text memo / Reason for change"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.lbl2.Wrap( -1 )
        sizer.Add( self.lbl2, 0, wx.ALL, 2 )
        self.txtFreeText = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), wx.HSCROLL | wx.VSCROLL | wx.TE_MULTILINE )
        if self.varname+"__metadata_freetext" in self.hbclass.contents:
            self.txtFreeText.SetValue(self.hbclass.contents[self.varname+"__metadata_freetext"])
        else:
            self.txtFreeText.SetValue("")
        sizer.Add(self.txtFreeText, 1, wx.EXPAND)
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(self, label=TT('Ok'))
        closeButton = wx.Button(self, label=TT('Cancel'))
        hbox.Add(okButton)
        hbox.Add(closeButton, flag=wx.LEFT, border=5)
        sizer.Add(hbox, flag= wx.ALIGN_CENTER|wx.TOP|wx.BOTTOM, border=30)
        self.SetSizer(sizer)
        
        self.Bind(wx.EVT_BUTTON, self.OnConfirm, okButton)
        self.Bind(wx.EVT_BUTTON, self.OnClose, closeButton)
        
    def OnClose(self, e):
        self.dialog.Destroy()

    def OnConfirm(self, e):
        if self.confidence.GetSelection()<1:
            if self.varname+"__metadata_confidence" in self.hbclass.contents:
                del self.hbclass.contents[self.varname+"__metadata_confidence"]
        else:
            self.hbclass.contents[self.varname+"__metadata_confidence"]=str(self.confidence.GetSelection())
        if not self.txtFreeText.GetValue():
            if self.varname+"__metadata_freetext" in self.hbclass.contents:
                del self.hbclass.contents[self.varname+"__metadata_freetext"]
        else:
            self.hbclass.contents[self.varname+"__metadata_freetext"]=self.txtFreeText.GetValue()
        self.dialog.Destroy()
            
class VarAnalysisTabPanel(wx.Panel):
    def __init__(self, parent, hbclass, dialog, varname):
        wx.Panel.__init__(self, parent=parent, id=wx.ID_ANY)
        self.hbclass=hbclass
        self.dialog=dialog
        self.parent=parent
        self.varname=varname
        self.parent.analysisPage=self
       
        sizer = wx.BoxSizer(wx.VERTICAL)
        self.lbl3 = wx.StaticText( self, wx.ID_ANY, TT("Type of selection"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.lbl3.Wrap( -1 )
        sizer.Add( self.lbl3, 0, wx.ALL, 2 )
        analysisChoices = [ TT("Current Variable ("+self.varname+") for all subjects"), TT("Selected variables for all subjects"), TT("Selected variables as a sequence for current subject") ]
        self.analysis = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, analysisChoices, 0 )
        self.analysis.SetSelection( 0 )
        sizer.Add(self.analysis, 0, wx.ALL, 2 )
        self.lbl4 = wx.StaticText( self, wx.ID_ANY, TT("Variables"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.lbl4.Wrap( -1 )
        sizer.Add( self.lbl4, 0, wx.ALL, 2 )
        if sys.platform == 'darwin':
            self.select_fields = wx.ListBox( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( -1,-1 ), [], wx.LB_EXTENDED )
        else:
            self.select_fields = wx.CheckListBox( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( -1,-1 ), [])
        sizer.Add(self.select_fields, 1, wx.EXPAND)
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        okButton = wx.Button(self, label=TT('Ok'))
        closeButton = wx.Button(self, label=TT('Cancel'))
        hbox.Add(okButton)
        hbox.Add(closeButton, flag=wx.LEFT, border=5)
        sizer.Add(hbox, flag= wx.ALIGN_CENTER|wx.TOP|wx.BOTTOM, border=30)
        self.SetSizer(sizer)
        
        self.Bind(wx.EVT_BUTTON, self.OnConfirm, okButton)
        self.Bind(wx.EVT_BUTTON, self.OnClose, closeButton)
        self.lbl4.Enable(False)
        self.select_fields.Enable(False)
        self.Bind(wx.EVT_CHOICE, self.selectReportType, self.analysis)
        
    def selectReportType(self, e):
        if self.analysis.GetSelection()==0:
            self.lbl4.Enable(False)
            self.select_fields.Enable(False)
        else:
            self.lbl4.Enable(True)
            self.select_fields.Enable(True)

    def FullFillList(self):
        if self.select_fields.GetCount()==0:
            # fullfill - begin
            varlist=[]
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.hbclass.id_crf)+" AND pos is not null GROUP BY description, id_cycle")
            for row in cur:
                varlist.append([levenshtein_distance(row[0],self.varname),row[0]])
            varlist.sort(key=lambda x:(x[0:]))
            self.select_fields.Clear()
            for elm in varlist: self.select_fields.Append(elm[1])
            # fullfill - end
        
    def OnClose(self, e):
        self.dialog.Destroy()

    def OnConfirm(self, e):
        from os.path import expanduser
        home = expanduser("~")
        now = datetime.datetime.now()
        ts=now.strftime("%Y%m%d%H%M%S")
        reportname="query_"+ts
        wildcard = "SPSS/PSPP data+syntax files (*.csv + *.sps)|*.sps"
        dlg = wx.FileDialog(self, message=TT("Choose a file name"), defaultDir=home, defaultFile=reportname, wildcard=wildcard, style=wx.SAVE | wx.FD_OVERWRITE_PROMPT)
        if dlg.ShowModal() == wx.ID_OK:
            filename=dlg.GetPath()
            filterindex=dlg.GetFilterIndex()
        dlg.Destroy()        
        if filename=="": return
        if filename[-4:]!=".sps": filename=filename+".sps"
        reportname=filename
        self.hbclass.PreSaveRecord()
        valueslist=[]
        if self.analysis.GetSelection() in (0,1):
            #single variable for all patients
            con_m = sqlite.connect(":memory:")
            cur_m = con_m.cursor()
            if self.analysis.GetSelection()==0:
                lstfields=["group_shortcode",self.hbclass.keyname,self.varname]
            else:
                lstfields=["group_shortcode",self.hbclass.keyname]
                if sys.platform == 'darwin':
                    for i in range(len(self.select_fields.GetSelections())):
                        lstfields.append(self.select_fields.GetSelections()[i])
                else:
                    for elm in self.select_fields.GetCheckedStrings():
                        lstfields.append(elm) 
            strlstfields=",".join(lstfields)
            square(con,con_m,"search",lstfields,self.hbclass.id_group,self.hbclass.id_crf)
            sql="SELECT "+strlstfields+" FROM search"
            cur_m.execute(sql)
            for row in cur_m:
                #valueslist.append(row[0])
                valueslist.append(row)
        else:
            #variables sequence for this patients
            if sys.platform == 'darwin':
                for i in range(len(self.select_fields.GetSelections())):
                    valueslist.append(self.select_fields.GetSelections()[i])
            else:
                for elm in self.select_fields.GetCheckedStrings():
                    valueslist.append(elm) 
        #spss exporter - begin
        arr1row=[]
        arrrows=[]
        #arrcols=[]
        #arrcols.append(self.varname)
        arr1row.append('GET DATA')
        arr1row.append('  /TYPE=TXT')
        arr1row.append('  /FILE="'+reportname.replace(".sps",".csv")+'"')
        arr1row.append('  /IMPORTCASES=ALL')
        arr1row.append('  /ARRANGEMENT=DELIMITED')
        arr1row.append('  /DELCASE=LINE')
        arr1row.append('  /FIRSTCASE=1')
        arr1row.append('  /DELIMITERS=","')
        arr1row.append('  /QUALIFIER=\'"\'')
        arr1row.append('  /ESCAPE')
        arr1row.append('  /VARIABLES=')
        varlist=[]
        if self.analysis.GetSelection() in (0,1): varlist=lstfields
        else: varlist=["variable_name","variable_value"]
        for varname in varlist:
            typedef=None
            if self.analysis.GetSelection() in (0,1):
                try: typedef=self.hbclass.arrHeadersTypedesc[varname]
                except: pass
            else:
                if varname==self.varname:
                    try: typedef=self.hbclass.arrHeadersTypedesc[self.varname]
                    except: pass
            spsstypedef="A255"
            if typedef==None:
                spsstypedef="A255"
            elif typedef.split("=")[0]=="mask":
                mask=typedef.split("=")[1]
                isnumber=True
                notnumbers=["/",":","-","N","A","a","C","X","&","*"]
                for elm in notnumbers:
                    if elm in mask:
                        isnumber=False
                        break
                if isnumber:
                    if "." in mask: spsstypedef="F10.3"
                    else: spsstypedef="F10.0"
                else:
                    spsstypedef="A255"
            elif typedef=="integer":
                spsstypedef="F10.0"
            elif typedef=="float":
                spsstypedef="F10.3"
            elif typedef=="number":
                mask=typedef[typedef.find("=")+1:]
                allowNegative=True
                intpart=0
                tmp=mask[:mask.find(".")]
                if len(tmp)>0:
                    if tmp[0]=="+": 
                        allowNegative=False
                        intpart=int(tmp[1:])
                    elif tmp[0]=="-": 
                        intpart=int(tmp[1:])
                    else:
                        intpart=int(tmp)
                decpart=int(mask[mask.find(".")+1:])    
                spsstypedef="F"+str(intpart)+"."+str(decpart)
            elif typedef=="date":
                spsstypedef="EDATE10"
            arr1row.append("    "+varname+" "+spsstypedef)
        arr1row.append('.')
            
        for row in valueslist:
            if self.analysis.GetSelection() in (0,1):
                currow=[]
                for col in row:
                    dData=""
                    tmp=col
                    if tmp==None: tmp=""
                    if tmp!="":
                        if not self.hbclass.parent.trans.has_key(tmp):
                            dData=HB_DecryptOne(self.hbclass.digestkey,tmp,"latin-1")
                            self.hbclass.parent.trans[tmp]=dData
                        else:
                            dData=self.hbclass.parent.trans[tmp]
                    data=dData.replace('"','""')
                    data=data.replace('\r\n','<CR>')
                    data=data.replace('\n','<CR>')
                    data=data.replace('\t','<TAB>')
                    data='"'+data+'"'
                    currow.append(data)
            else:
                varname=row
                currow=[]
                currow.append(varname)
                if varname.find("#")>=0: varname = varname[:varname.find("#")] + "__" + varname[varname.find("#")+1:]
                dData=""
                try: dData=self.hbclass.contents[varname]
                except: pass
                data=dData.replace('"','""')
                data=data.replace('\r\n','<CR>')
                data=data.replace('\n','<CR>')
                data=data.replace('\t','<TAB>')
                data='"'+data+'"'
                currow.append(data)
            arrrows.append(",".join(currow))
        
        filename1=reportname.replace(".sps",".csv")
        theFile = open(filename1, 'w')
        contents='\n'.join(arrrows)+'\n'
        theFile.write(contents.encode('ascii', 'replace'))
        theFile.close
        theFile = open(reportname, 'w')
        contents='\n'.join(arr1row)+'\n'
        theFile.write(contents.encode('ascii', 'replace'))
        theFile.close
        if not assoc_open(reportname):
            wx.MessageBox("File\n'"+reportname+"'\ncreated.", TT("Exporter"), wx.ICON_INFORMATION | wx.OK, self)
        #spss exporter - end
        
        self.dialog.Destroy()
        
class VarHistoryTabPanel(wx.Panel):
    def __init__(self, parent, hbclass, dialog, varname):
        wx.Panel.__init__(self, parent=parent, id=wx.ID_ANY)
        self.hbclass=hbclass
        self.dialog=dialog
        self.parent=parent
        self.varname=varname
        self.parent.historyPage=self
        
        sizer = wx.BoxSizer(wx.VERTICAL)
        self.txtHistory = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), wx.HSCROLL | wx.VSCROLL | wx.TE_MULTILINE | wx.TE_READONLY )
        self.txtHistory.SetBackgroundColour(wx.Colour(240,240,240))
        sizer.Add(self.txtHistory, 1, wx.EXPAND)
        self.SetSizer(sizer)
        
    def FullFillHistory(self):
        if self.txtHistory.GetValue()=="":
            id_header=self.hbclass.hashHeaders[self.varname]
            id_cycle=0
            if self.hbclass.ActivateCyclicity:
                if self.hbclass.Cyclicity[self.hbclass.ifrm+1]>0: 
                    id_cycle=wx.xrc.XRCCTRL(self.hbclass.frameRecordset,'cycles').GetSelection()+1
            hist=[]
            ref_row=self.hbclass.id_row
            ref_user=self.hbclass.id_user
            ref_instance=self.hbclass.id_rowinstance
            while(True):
                found=False
                old_val=""
                old_username=""
                old_fullname=""
                old_datetime=""
                cur.execute("SELECT data FROM headers,contents WHERE contents.id_row="+str(ref_row)+" AND contents.id_user="+str(ref_user)+" AND contents.id_instance="+str(ref_instance)+" AND headers.id_header="+str(id_header)+" AND headers.id_header=contents.id_header")
                row = cur.fetchone()
                if row!=None:
                    old_val=row[0]
                    cur.execute("SELECT users.username,users.fullname,rows.date_time,rows.status,rows.status_user,rows.status_instance FROM users,rows WHERE users.id_user=rows.id_user AND rows.id_row="+str(ref_row)+" AND rows.id_user="+str(ref_user)+" AND rows.id_instance="+str(ref_instance))
                    row = cur.fetchone()
                    if row!=None:
                        old_username=row[0]
                        old_fullname=row[1]
                        old_datetime=row[2]
                        old_row=row[3]
                        old_user=row[4]
                        old_instance=row[5]
                        found=True
                if not found:
                    break
                hist.append([old_val,old_username,old_fullname,old_datetime])
                ref_row=old_row
                ref_user=old_user
                ref_instance=old_instance
            arr=[]
            for elm in hist:
                txt=""
                cData=elm[0]
                dData=""
                if not self.hbclass.parent.trans.has_key(cData):
                    dData=HB_DecryptOne(self.hbclass.digestkey,cData,"latin-1")
                    self.hbclass.parent.trans[cData]=dData
                else:
                    dData=self.hbclass.parent.trans[cData]
                txt=dData
                cData=elm[1]
                dData=""
                if not self.hbclass.parent.trans.has_key(cData):
                    dData=HB_DecryptOne(self.hbclass.digestkey,cData,"latin-1")
                    self.hbclass.parent.trans[cData]=dData
                else:
                    dData=self.hbclass.parent.trans[cData]
                txt=txt+" -"+dData+"-"
                if elm[2]!="": txt=txt+" ("+elm[2]+")"
                txt=txt+" ["+elm[3]+"]"
                arr.append(txt)
                self.txtHistory.SetValue('\n'.join(arr))
      
class MetadataNotebook(wx.Notebook):
    def __init__(self, parent, hbclass, dialog, varname):
        wx.Notebook.__init__(self, parent, id=wx.ID_ANY, style=
                             wx.BK_DEFAULT|wx.EXPAND
                             #wx.BK_TOP 
                             #wx.BK_BOTTOM
                             #wx.BK_LEFT
                             #wx.BK_RIGHT
                             )
        self.hbclass=hbclass
        self.varname=varname
        self.parent=parent
        self.AddPage(VarMetadataTabPanel(self, hbclass, dialog, self.varname), TT("Metadata"))
        self.AddPage(VarHistoryTabPanel(self, hbclass, dialog, self.varname), TT("History"))
        self.AddPage(VarAnalysisTabPanel(self, hbclass, dialog, self.varname), TT("Analysis"))
        
        self.Bind(wx.EVT_NOTEBOOK_PAGE_CHANGED, self.OnPageChanged)
        
    def OnPageChanged(self, event):
        event.Skip()
        new = event.GetSelection()
        if new==2:
            wx.BeginBusyCursor()
            self.analysisPage.FullFillList()
            wx.EndBusyCursor()
        elif new==1:
            wx.BeginBusyCursor()
            self.historyPage.FullFillHistory()
            wx.EndBusyCursor()
            
class MetadataFrame(wx.Dialog):
    def __init__(self, parent, varname):
        wx.Dialog.__init__(self, None, title=varname, size=(500,500))
        self.hbclass=parent
        self.varname=varname

        panel = wx.Panel(self, size=(500,500))

        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(MetadataNotebook(panel, self.hbclass, self, self.varname), 1, wx.ALL|wx.EXPAND, 5)
        panel.SetSizer(sizer)
        panel.Layout()
        self.Layout()
        #self.Show()
#Metadata - End

#Popup menu - Begin
class MyPopupMenu(wx.Menu):
    def __init__(self, title, parent, obj, parent_event):
        wx.Menu.__init__(self, title=title)

        self.obj = obj
        self.title = title
        self.parent = parent
        self.parent_event = parent_event

        item = wx.MenuItem(self, wx.NewId(), TT("Metadata"))
        self.AppendItem(item)
        self.Bind(wx.EVT_MENU, self.OnItemMetadata, item)

        if obj.GetClassName() in ("wxComboBox") and GetSetting(self.parent.parent,"combo_learn")=="0":
            item = wx.MenuItem(self, wx.NewId(),TT("Load options"))
            self.AppendItem(item)
            self.Bind(wx.EVT_MENU, self.OnItemLoadCombo, item)

    def OnItemMetadata(self, event):
        self.parent.OnModifyMetadata(None)

    def OnItemLoadCombo(self, event):
        wx.BeginBusyCursor()
        self.parent.DoFulfillComboList(self.obj, self.title)
        wx.EndBusyCursor()
        self.parent.FulfillCombo(self.parent_event)
#Popup menu - End

#03/09/2008
#class HeavyBaseFrame(wx.Panel):
class HeavyBaseFrame(wx.aui.AuiMDIChildFrame):
    def _init_ctrls(self, prnt,childTitle):
        #03/09/2008
        #wx.Frame.__init__(self, id=wx.ID_FRAMEBACK, name='', parent=prnt, pos=wx.Point(0, 0), size=wx.Size(SCREEN_X, SCREEN_Y), style=wx.DEFAULT_FRAME_STYLE, title='HEAVyBASE')
        wx.aui.AuiMDIChildFrame.__init__(self, prnt, -1, title=childTitle)
        self.SetBackgroundColour(wx.Colour(240,240,240))

        self.parent=prnt

        mb = self.parent.MakeMenuBar()
        menu = wx.Menu()
        item = menu.Append(101, TT("Data entry mode"))
        self.menuDataEntryModeItem=item
        self.Bind(wx.EVT_MENU, self.OnDataEntryMode, item)
        item = menu.Append(102, TT("Data extraction mode"))
        self.menuDataExtractionModeItem=item
        self.Bind(wx.EVT_MENU, self.OnDataExtractionMode, item)
        item = menu.Append(103, TT("Data Analysis"))
        self.Bind(wx.EVT_MENU, self.OnStatistics, item)
        item = menu.Append(104, TT("Import external dataset"))
        self.Bind(wx.EVT_MENU, self.OnImportData, item)
        item = menu.AppendSeparator()
        item = menu.Append(105, TT("Current Record")+": "+TT("Audit Trail"))
        self.Bind(wx.EVT_MENU, self.OnAuditTrail, item)
        item = menu.Append(106, TT("Current Record")+": "+TT("Modify Access Rights"))
        self.Bind(wx.EVT_MENU, self.OnModifyAccessRights, item)
        #item = menu.AppendSeparator()
        #item = menu.Append(107, TT("Current Variable")+": "+TT("Metadata"))
        #self.Bind(wx.EVT_MENU, self.OnModifyMetadata, item)
        item = menu.AppendSeparator()
        item = menu.Append(108, TT("Change user password"))
        self.Bind(wx.EVT_MENU, self.OnChangeUserPassword, item)
#        item = menu.Append(-1, "Change cypher key")
#        self.Bind(wx.EVT_MENU, self.OnChangeCypherKey, item)
        item = menu.Append(109, TT("OTP Credentials Keyring"))
        self.Bind(wx.EVT_MENU, self.OnIdentities, item)
        mb.Append(menu, TT("Data management"))
        
        #menu = wx.Menu()
        #item = menu.Append(-1, "Audit &Trail")
        #self.menuAuditTrail=item
        #self.Bind(wx.EVT_MENU, self.OnDataEntryMode, item)
        #mb.Append(menu, "CRF")
        #self.menuAuditTrail.enable(False)
        
        self.mb=mb
        self.SetMenuBar(self.mb)
        
        #03/09/2008
        #self.SetClientSize(wx.Size(SCREEN_X, SCREEN_Y))
        #self.SetMinSize(wx.Size(SCREEN_X, SCREEN_Y))
        #self.Bind(wx.EVT_SIZE,self.OnWindowResize,id=wx.ID_FRAMEBACK)

        self.Bind(wx.EVT_SIZE, self.OnWindowResize)

        clientSize=self.GetClientSize()
        #03/09/2008
        #self.SetIcon(wx.Icon('ship.ico', wx.BITMAP_TYPE_ICO))
        self.frameLogin = wx.ScrolledWindow(id=wx.ID_FRAMELOGIN, name='frameLogin', parent=self, pos=wx.Point(0, 0), size=wx.Size(SCREEN_X, SCREEN_Y), style=wx.HSCROLL | wx.VSCROLL)
        self.lblUsername = wx.StaticText(label=TT('Username'), name='lblUsername', parent=self.frameLogin, pos=wx.Point(10, 20), size=wx.Size(130, 25), style=0)
        self.txtUsername = wx.TextCtrl(name='txtUsername', parent=self.frameLogin, pos=wx.Point(140, 20), size=wx.Size(200, 25), style=0, value='')
        self.txtUsername.Bind(wx.EVT_KEY_UP, self.OnLoginEnterPressed)
        self.lblPassword = wx.StaticText(label=TT('Password'), name='lblPassword', parent=self.frameLogin, pos=wx.Point(10, 50), size=wx.Size(130, 25), style=0)
        self.txtPassword = wx.TextCtrl(name='txtPassword', parent=self.frameLogin, pos=wx.Point(140, 50), size=wx.Size(200, 25), style=wx.TE_PASSWORD, value='')
        self.txtPassword.Bind(wx.EVT_KEY_UP, self.OnLoginEnterPressed)
        self.lblCryptokey = wx.StaticText(label=TT('Cypher key'), name='lblCryptokey', parent=self.frameLogin, pos=wx.Point(10, 80), size=wx.Size(130, 25), style=0)
        self.txtCryptokey = wx.TextCtrl(name='txtCryptokey', parent=self.frameLogin, pos=wx.Point(140, 80), size=wx.Size(200, 25), style=wx.TE_PASSWORD, value='')
        self.txtCryptokey.Bind(wx.EVT_KEY_UP, self.OnLoginEnterPressed)
        self.lblDatabase = wx.StaticText(label='Database', name='lblDatabase', parent=self.frameLogin, pos=wx.Point(410, 80), size=wx.Size(110, 25), style=0)
        self.choiceDatabase = wx.Choice(choices=[], name='choiceDatabase', parent=self.frameLogin, pos=wx.Point(520, 80), size=wx.Size(200, 25), style=0)
        #self.txtDatabase = wx.TextCtrl(name='txtDatabase', parent=self.frameLogin, pos=wx.Point(520, 80), size=wx.Size(200, 25), style=0, value='')
        self.choiceDatabase.Bind(wx.EVT_KEY_UP, self.OnLoginEnterPressed)
        self.buttonLogin = wx.Button(id=wx.ID_FRAMEBUTTONLOGIN, label='login', name='login', parent=self.frameLogin, pos=wx.Point(10,110), size=wx.Size(100, 30), style=0)
        self.buttonLogin.Bind(wx.EVT_BUTTON, self.OnButtonLoginButton, id=wx.ID_FRAMEBUTTONLOGIN)
        self.lblDisclaimer = wx.StaticText(label=DISCLAIMER, name='lblDisclaimer', parent=self.frameLogin, pos=wx.Point(5, SCREEN_Y-135), size=wx.Size(790, 35), style=0)

        self.frameSearch = wx.ScrolledWindow(id=wx.ID_FRAMESEARCH, name='frameSearch', parent=self, pos=wx.Point(5000, 0), size=wx.Size(SCREEN_X, SCREEN_Y), style=wx.HSCROLL | wx.VSCROLL)
        #self.lstFound = wx.ListBox(choices=[], id=wx.ID_FRAMELSTFOUND, name='lstFound', parent=self.frameSearch, pos=wx.Point(0, 170), size=wx.Size(clientSize.GetWidth(), clientSize.GetHeight()-350), style=wx.LB_MULTIPLE)
        self.lstFound = wx.ListBox(choices=[], id=wx.ID_FRAMELSTFOUND, name='lstFound', parent=self.frameSearch, pos=wx.Point(0, 170), size=wx.Size(clientSize.GetWidth(), clientSize.GetHeight()-350), style=wx.LB_EXTENDED)
        self.lstFound.SetFont(wx.Font( 10, wx.MODERN, wx.NORMAL, wx.NORMAL))
        self.lstFound.Bind(wx.EVT_LISTBOX_DCLICK, self.OnLstFoundDoubleClicked, id=wx.ID_FRAMELSTFOUND)
        self.lstFound.Bind(wx.EVT_LISTBOX, self.OnLstFoundClicked)
        self.buttonSearch = wx.Button(id=wx.ID_FRAMEBUTTONSEARCH, label='search', name='search', parent=self.frameSearch, pos=wx.Point(10,130), size=wx.Size(85, 30), style=0)
        self.buttonSearch.Bind(wx.EVT_BUTTON, self.OnButtonSearchButton, id=wx.ID_FRAMEBUTTONSEARCH)
        self.chkTolerance = wx.CheckBox (parent=self.frameSearch, pos=wx.Point(100,130), size=wx.Size(150, 25),label='Tolerance' )
        self.buttonExportSearch = wx.Button(label='export search results', name='exportSearch', parent=self.frameSearch, pos=wx.Point(self.lstFound.GetSize().GetWidth()/2-80,self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()), size=wx.Size(180, 25), style=0)
        self.buttonExportSearch.Bind(wx.EVT_BUTTON, self.OnButtonExportSearchButton)
        self.buttonAddnew = wx.Button(id=wx.ID_FRAMEBUTTONADDNEW, label='add new', name='addnew', parent=self.frameSearch, pos=wx.Point(10,self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()+10), size=wx.Size(120, 30), style=0)
        self.buttonAddnew.Bind(wx.EVT_BUTTON, self.OnButtonAddnewButton, id=wx.ID_FRAMEBUTTONADDNEW)
        self.buttonAutoMergeDupes = wx.Button(label='auto-merge dupes', name='autoMergeDupes', parent=self.frameSearch, pos=wx.Point(self.lstFound.GetSize().GetWidth()-390,self.lstFound.GetPosition().y-25), size=wx.Size(150, 25), style=0)
        self.buttonAutoMergeDupes.Bind(wx.EVT_BUTTON, self.OnButtonAutoMergeDupes)
        self.crfTypeChoice = wx.Choice(choices=['valid records','all records','deleted records','substituted records'], name='crfTypeChoice', parent=self.frameSearch, pos=wx.Point(self.lstFound.GetSize().GetWidth()-220,self.lstFound.GetPosition().y-25), size=wx.Size(150, 25), style=0)
        self.buttonShowDiff = wx.Button(label='diff', name='showDiff', parent=self.frameSearch, pos=wx.Point(self.lstFound.GetSize().GetWidth()-50,self.lstFound.GetPosition().y-25), size=wx.Size(50, 25), style=0)
        self.buttonShowDiff.Bind(wx.EVT_BUTTON, self.OnButtonShowDiff)
        self.buttonExport = wx.Button(id=wx.ID_FRAMEBUTTONEXPORT, label='export', name='export', parent=self.frameSearch, pos=wx.Point(550,self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()+10), size=wx.Size(200, 30), style=0)
        self.buttonExport.Bind(wx.EVT_BUTTON, self.OnButtonExportButton, id=wx.ID_FRAMEBUTTONEXPORT)
        self.txtQuickReport = wx.TextCtrl(name='txtQuickReport', parent=self.frameSearch, pos=wx.Point(0, self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()+50), size=wx.Size(clientSize.GetWidth(), 80), value='', style=wx.HSCROLL | wx.VSCROLL | wx.TE_MULTILINE | wx.TE_READONLY)
        self.txtQuickReport.SetFont(wx.Font( 10, wx.MODERN, wx.NORMAL, wx.NORMAL))
        self.txtQuickReport.SetBackgroundColour( wx.SystemSettings.GetColour( wx.SYS_COLOUR_BTNFACE ) )        
        self.lblReport = wx.StaticText(label='Report', name='lblReport', parent=self.frameSearch, pos=wx.Point(0, 480), size=wx.Size(60, 30), style=0)
        self.reportChoice = wx.Choice(choices=[], id=wx.ID_REPORTCHOICE, name='reportChoice', parent=self.frameSearch, pos=wx.Point(80, 480), size=wx.Size(500, 30), style=0)
        self.reportChoice.Bind(wx.EVT_CHOICE, self.OnReportChoice, id=wx.ID_REPORTCHOICE)
        self.reportChoice.Bind(wx.EVT_MOUSEWHEEL, self.OnComboMouseWheel)


        #self.buttonReportChoice = wx.Button(id=wx.ID_REPORTCHOICEBUTTON, label='build', name='reportButton', parent=self.frameSearch, pos=wx.Point(580,500), size=wx.Size(50, 30), style=0)
        #self.buttonReportChoice.Bind(wx.EVT_BUTTON, self.OnReportChoice, id=wx.ID_REPORTCHOICEBUTTON)
        self.buttonLogout = wx.Button(id=wx.ID_FRAMEBUTTONLOGOUT, label='logout', name='logout', parent=self.frameSearch, pos=wx.Point(SCREEN_X-100,500), size=wx.Size(85, 30), style=0)
        self.buttonLogout.Bind(wx.EVT_BUTTON, self.OnButtonLogoutButton, id=wx.ID_FRAMEBUTTONLOGOUT)

        self.frameRecordset = wx.ScrolledWindow(id=wx.ID_FRAMERECORDSET, name='frameRecordset', parent=self, pos=wx.Point(5000, 0), size=wx.Size(SCREEN_X, SCREEN_Y), style=wx.HSCROLL | wx.VSCROLL)
        #self.frameCrf = wx.ScrolledWindow(id=wx.ID_FRAMECRF, name='frameCrf', parent=self.frameRecordset, pos=wx.Point(0, 80), size=wx.Size(SCREEN_X, SCREEN_Y-80), style=wx.HSCROLL | wx.VSCROLL)
        self.frameCrf = wx.ScrolledWindow(id=wx.ID_FRAMECRF, name='frameCrf', parent=self.frameRecordset, pos=wx.Point(0, 100), size=wx.Size(SCREEN_X, SCREEN_Y-100), style=wx.HSCROLL | wx.VSCROLL)
        self.buttonClone = wx.Button(id=wx.ID_FRAMEBUTTONCLONE, label='clone/new', name='clone', parent=self.frameRecordset, pos=wx.Point(350,10), size=wx.Size(140, 30), style=0)
        self.buttonClone.Bind(wx.EVT_BUTTON, self.OnButtonCloneButton, id=wx.ID_FRAMEBUTTONCLONE)
        self.buttonSave = wx.Button(id=wx.ID_FRAMEBUTTONSAVE, label='save', name='save', parent=self.frameRecordset, pos=wx.Point(500,10), size=wx.Size(140, 30), style=0)
        self.buttonSave.Bind(wx.EVT_BUTTON, self.OnButtonSaveButton, id=wx.ID_FRAMEBUTTONSAVE)
        self.buttonCommit = wx.Button(id=wx.ID_FRAMEBUTTONCOMMIT, label='save and exit', name='commit', parent=self.frameRecordset, pos=wx.Point(650,10), size=wx.Size(140, 30), style=0)
        self.buttonCommit.Bind(wx.EVT_BUTTON, self.OnButtonCommitButton, id=wx.ID_FRAMEBUTTONCOMMIT)
        self.buttonCancel = wx.Button(id=wx.ID_FRAMEBUTTONCANCEL, label='quit, don\'t save', name='cancel', parent=self.frameRecordset, pos=wx.Point(350,50), size=wx.Size(140, 30), style=0)
        self.buttonCancel.Bind(wx.EVT_BUTTON, self.OnButtonCancelButton, id=wx.ID_FRAMEBUTTONCANCEL)
        self.buttonDelete = wx.Button(id=wx.ID_FRAMEBUTTONDELETE, label='delete record', name='delete', parent=self.frameRecordset, pos=wx.Point(500,50), size=wx.Size(140, 30), style=0)
        self.buttonDelete.Bind(wx.EVT_BUTTON, self.OnButtonDeleteButton, id=wx.ID_FRAMEBUTTONDELETE)
        self.buttonPrint = wx.Button(id=wx.ID_FRAMEBUTTONPRINT, label='print page', name='print', parent=self.frameRecordset, pos=wx.Point(650,50), size=wx.Size(140, 30), style=0)
        self.buttonPrint.Bind(wx.EVT_BUTTON, self.OnButtonPrintButton, id=wx.ID_FRAMEBUTTONPRINT)

        self.frameDataExtraction = wx.ScrolledWindow(name='frameDataExtraction', parent=self, pos=wx.Point(5000, 0), size=wx.Size(SCREEN_X, SCREEN_Y), style=wx.HSCROLL | wx.VSCROLL)
        self.dataExtractionPanel = DataExtractionPanel(self.frameDataExtraction)
        self.dataExtractionPanel.cmdMarkAll.Bind(wx.EVT_BUTTON, self.OnDataExtractionMarkAll)
        self.dataExtractionPanel.cmdUnmarkAll.Bind(wx.EVT_BUTTON, self.OnDataExtractionUnmarkAll)
        if sys.platform != 'darwin':
            self.dataExtractionPanel.filter_field1.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField1)
            self.dataExtractionPanel.filter_field2.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField2)
            self.dataExtractionPanel.filter_field3.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField3)
            self.dataExtractionPanel.filter_field4.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField4)
            self.dataExtractionPanel.filter_field5.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField5)
            self.dataExtractionPanel.filter_field6.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField6)
            self.dataExtractionPanel.filter_field7.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField7)
            self.dataExtractionPanel.filter_field8.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterField8)
            self.dataExtractionPanel.filter_value1.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue1)
            self.dataExtractionPanel.filter_value2.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue2)
            self.dataExtractionPanel.filter_value3.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue3)
            self.dataExtractionPanel.filter_value4.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue4)
            self.dataExtractionPanel.filter_value5.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue5)
            self.dataExtractionPanel.filter_value6.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue6)
            self.dataExtractionPanel.filter_value7.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue7)
            self.dataExtractionPanel.filter_value8.Bind(wx.EVT_SET_FOCUS, self.OnDataExtractionFilterValue8)
        else:
            self.dataExtractionPanel.filter_field1.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField1)
            self.dataExtractionPanel.filter_field2.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField2)
            self.dataExtractionPanel.filter_field3.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField3)
            self.dataExtractionPanel.filter_field4.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField4)
            self.dataExtractionPanel.filter_field5.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField5)
            self.dataExtractionPanel.filter_field6.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField6)
            self.dataExtractionPanel.filter_field7.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField7)
            self.dataExtractionPanel.filter_field8.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterField8)
            self.dataExtractionPanel.filter_value1.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue1)
            self.dataExtractionPanel.filter_value2.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue2)
            self.dataExtractionPanel.filter_value3.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue3)
            self.dataExtractionPanel.filter_value4.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue4)
            self.dataExtractionPanel.filter_value5.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue5)
            self.dataExtractionPanel.filter_value6.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue6)
            self.dataExtractionPanel.filter_value7.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue7)
            self.dataExtractionPanel.filter_value8.Bind(wx.EVT_ENTER_WINDOW, self.OnDataExtractionFilterValue8)

        self.dataExtractionPanel.cmdrun.Bind(wx.EVT_BUTTON, self.OnDataExtractionRunQuery)
        #self.dataExtractionPanel.optFieldOrder.Bind(wx.EVT_RADIOBOX, self.OnDataExtractionReorder)
        self.dataExtractionPanel.optFieldOrder.Bind(wx.EVT_CHOICE, self.OnDataExtractionReorder)


    #03/09/2008
    #def __init__(self, parent):
    def __init__(self, parent, childTitle):
        wx.BeginBusyCursor()

        self.projectname=""
        self.keyname=""
        self.keymode=""

        self.report_format=""

        self.id_crf=0
        self.id_row=0
        self.id_rowinstance=0
        self.id_user=0
        self.id_logged_user=0
        self.uap=""
        self.rap=""
        self.id_row=0
        self.id_group=0
        self.arrHeaders={}
        self.hashHeaders={}
        self.arrHeadersLabels={}
        self.arrHeadersEvents={}
        self.arrHeadersDefaults={}
        self.arrHeadersTypedesc={}
        self.arrHeadersSubtypedesc={}
        self.arrHeadersValidations={}
        self.arrHeadersComboPreload={}
        self.arrHeadersIdSection={}
        self.arrHeadersPos={}
        self.arrFound={}
        self.arrReports={}

        self.quickreport_output=""

        self.dont_trigger_field_change=False
        self.anyrecordchange=False
        
        self.lastfocusvariable=""
        self.lastfocusvariableobj=None
        
        self.waiting_predecrypt=False

        self._init_ctrls(parent, childTitle)
        #load profile

        self.mode_multi_database=False
        cur.execute("SELECT count(*) FROM headers WHERE child_of=0")
        row = cur.fetchone()
        if row!=None:
            if row[0]<2:
                self.lblDatabase.Hide()
                self.choiceDatabase.Hide()
            else:
                self.choiceDatabase.Clear()
                cur.execute("SELECT description FROM headers WHERE child_of=0 AND status=0 ORDER BY description")
                for row in cur:
                    self.choiceDatabase.Append(row[0])
                self.mode_multi_database=True
                if DEBUG_MODE: print "Activating Multi database mode"

        self.mode_shared_variables=False
        cur.execute("SELECT count(*) FROM headers WHERE deftype='S'")
        row = cur.fetchone()
        if row!=None:
            if row[0]>0:
                self.mode_shared_variables=True
                if DEBUG_MODE: print "Activating Shared variables mode"

        #customisations - begin
        #cur.execute("SELECT setting_key,setting_value FROM settings")
        #for row in cur:
        #    self.arrSettings[row[0]]=row[1]
        self.parent.reloadSettings()

        if (self.parent.arrSettings.has_key('lbl_login')):  self.buttonLogin.SetLabel(self.parent.arrSettings['lbl_login'])
        else: SaveCustomSetting('lbl_login',self.buttonLogin.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_search')): self.buttonSearch.SetLabel(self.parent.arrSettings['lbl_search'])
        else: SaveCustomSetting('lbl_search',self.buttonSearch.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_addnew')): self.buttonAddnew.SetLabel(self.parent.arrSettings['lbl_addnew'])
        else: SaveCustomSetting('lbl_addnew',self.buttonAddnew.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_export')): self.buttonExport.SetLabel(self.parent.arrSettings['lbl_export'])
        else: SaveCustomSetting('lbl_export',self.buttonExport.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_logout')): self.buttonLogout.SetLabel(self.parent.arrSettings['lbl_logout'])
        else: SaveCustomSetting('lbl_logout',self.buttonLogout.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_clone')):  self.buttonClone.SetLabel(self.parent.arrSettings['lbl_clone'])
        else: SaveCustomSetting('lbl_clone',self.buttonClone.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_save')):   self.buttonSave.SetLabel(self.parent.arrSettings['lbl_save'])
        else: SaveCustomSetting('lbl_save',self.buttonSave.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_commit')): self.buttonCommit.SetLabel(self.parent.arrSettings['lbl_commit'])
        else: SaveCustomSetting('lbl_commit',self.buttonCommit.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_cancel')): self.buttonCancel.SetLabel(self.parent.arrSettings['lbl_cancel'])
        else: SaveCustomSetting('lbl_cancel',self.buttonCancel.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_delete')): self.buttonDelete.SetLabel(self.parent.arrSettings['lbl_delete'])
        else: SaveCustomSetting('lbl_delete',self.buttonDelete.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_print')):  self.buttonPrint.SetLabel(self.parent.arrSettings['lbl_print'])
        else: SaveCustomSetting('lbl_print',self.buttonPrint.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_exportsearch')):  self.buttonExportSearch.SetLabel(self.parent.arrSettings['lbl_exportsearch'])
        else: SaveCustomSetting('lbl_exportsearch',self.buttonExportSearch.GetLabel())
        if (self.parent.arrSettings.has_key('lbl_tolerance')):  self.chkTolerance.SetLabel(self.parent.arrSettings['lbl_tolerance'])
        else: SaveCustomSetting('lbl_tolerance',self.chkTolerance.GetLabel())

        if GetSetting(self.parent,'show_buttonExport')=="0": self.buttonExport.Show(False)
        if GetSetting(self.parent,'show_buttonClone')=="0": self.buttonClone.Show(False)
        if GetSetting(self.parent,'show_buttonPrint')=="0": self.buttonPrint.Show(False)
        if GetSetting(self.parent,'show_buttonAddnew')=="0": self.buttonAddnew.Show(False)

        self.xmlResource1 = wx.xrc.EmptyXmlResource()
        self.xmlResource1.InsertHandler(MaskedCtrlXmlHandler())
        self.xmlResource1.InsertHandler(DateCtrlXmlHandler())
        self.xmlResource1.Load(PROFILE)

        self.xrcPanelSplash = self.xmlResource1.LoadPanel(name = 'panelSplash', parent = self.frameLogin)
        self.xrcPanelSplash.SetPosition(wx.Point(0,150))
        self.xrcPanelSplash.SetSize(wx.Size(1000,200))

#        # postponed
#        self.xrcPanel = self.xmlResource1.LoadPanel(name = 'panelHeader', parent = self.frameSearch)
#        wx.xrc.XRCCTRL(self.frameSearch,'panelHeader').SetPosition(wx.Point(0,0))

#        self.xrcPanel = self.xmlResource1.LoadPanel(name = 'panelSearch', parent = self.frameSearch)
#        self.PanelSearch.SetPosition(wx.Point(10,50))

        self.xrcPanelChoice = self.xmlResource1.LoadPanel(name = 'panelChoice', parent = self.frameRecordset)
        self.xrcPanelChoice.SetPosition(wx.Point(0,0))
        self.xrcPanelChoice.SetSize(wx.Size(500,100))
        wx.xrc.XRCCTRL(self.frameRecordset,'choice').Bind(wx.EVT_CHOICE, self.OnChoice, id=wx.xrc.XRCID('choice'))
        wx.xrc.XRCCTRL(self.frameRecordset,'choice').Bind(wx.EVT_MOUSEWHEEL, self.OnComboMouseWheel)

        #Database Init - Begin
        cur.execute("SELECT setting_value FROM settings WHERE setting_key='keyname'")
        row = cur.fetchone()
        if row==None:
            self.keyname="key"
            cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('keyname','"+self.keyname+"')")
        else:
            self.keyname=row[0]
        cur.execute("SELECT setting_value FROM settings WHERE setting_key='keymode'")
        row = cur.fetchone()
        if row==None:
            self.keymode="public"
            cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('keymode','"+self.keymode+"')")
        else:
            self.keymode=row[0]

        cur.execute("SELECT setting_value FROM settings WHERE setting_key='report_format'")
        row = cur.fetchone()
        if row==None:
            self.report_format="excel" #alternativa: "text"
            cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('report_format','"+self.report_format+"')")
        else:
            self.report_format=row[0]

        cur.execute("SELECT * FROM users")
        row = cur.fetchone()
        if row==None:
            cur.execute("INSERT INTO users (id_group,uap,username,password,fullname) VALUES (0,'','demo','demo','demo')")

        cur.execute("SELECT * FROM users")
        row = cur.fetchone()
        if row==None:
            cur.execute("INSERT INTO users (id_group,uap,username,password,fullname) VALUES (0,'','demo','demo','demo')")

        cur.execute("SELECT setting_value FROM settings WHERE setting_key='heavybase_server'")
        row = cur.fetchone()
        if row==None: cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('heavybase_server','xmlrpc.heavybase.org:80')")

        cur.execute("SELECT setting_value FROM settings WHERE setting_key='project_name'")
        row = cur.fetchone()
        if row==None: cur.execute("INSERT INTO settings (setting_key,setting_value) VALUES ('project_name','heavybase')")

        #Database Init - End

        self.LogoutCount=0
        self.timer = wx.Timer(self, -1)
        self.timer.Start(1000,True) #One shot
        self.Bind(wx.EVT_TIMER, self.OnChildTimer)

        self.EnableDisableButtons(1)
        wx.EndBusyCursor()

    def OnChildTimer(self,evt):
        self.txtQuickReport.SetValue(self.quickreport_output)
        if self.parent.LogoutCount>self.LogoutCount:
            self.lstFound.Clear()
            self.txtUsername.SetValue("")
            self.txtPassword.SetValue("")
            #self.txtCryptoKey.SetValue("")
            self.frameLogin.SetPosition(wx.Point(0,0))
            self.frameSearch.SetPosition(wx.Point(5000,0))
            self.frameRecordset.SetPosition(wx.Point(5000,0))
            self.EnableDisableButtons(1)
            self.LogoutCount=self.LogoutCount+1
            
        if GetSetting(self.parent,"synchro_decrypt")=="1":
            if self.frameLogin.GetPosition().x==0:
                if self.waiting_predecrypt and self.parent.preDecrypt_waitstate==0:
                    self.waitframe.SetPosition(wx.Point(5000,0))
                    self.waitframe.Destroy()
                    self.frameLogin.SetPosition(wx.Point(5000,0))
                    self.frameSearch.SetPosition(wx.Point(0,0))
                    self.frameRecordset.SetPosition(wx.Point(5000,0))
                    self.EnableDisableButtons(2)
                    self.waiting_predecrypt=False
                    try: wx.EndBusyCursor()
                    except: pass
        self.timer.Start(1000,True)

    def EnableDisableButtons(self,page):
        if page==1:
            # LOGIN
            self.buttonLogin.Enable(True)
            #
            self.buttonSearch.Enable(False)
            self.buttonExportSearch.Enable(False)
            self.buttonAddnew.Enable(False)
            self.buttonExport.Enable(False)
            self.reportChoice.Enable(False)
            #
            self.buttonClone.Enable(False)
            self.buttonSave.Enable(False)
            self.buttonCommit.Enable(False)
            self.buttonCancel.Enable(False)
            self.buttonDelete.Enable(False)
            self.buttonPrint.Enable(False)
        elif page==2:
            # SEARCH
            self.buttonLogin.Enable(False)
            #
            self.buttonSearch.Enable(True)
            self.buttonExportSearch.Enable(True)
            if self.uap!="": 
                self.buttonAddnew.Enable(True)
            else:
                self.buttonAddnew.Enable(False)
            self.buttonExport.Enable(True)
            self.reportChoice.Enable(True)
            #
            self.buttonClone.Enable(False)
            self.buttonSave.Enable(False)
            self.buttonCommit.Enable(False)
            self.buttonCancel.Enable(False)
            self.buttonDelete.Enable(False)
            self.buttonPrint.Enable(False)
        elif page==3:
            # CRF
            self.buttonLogin.Enable(False)
            #
            self.buttonSearch.Enable(False)
            self.buttonExportSearch.Enable(False)
            self.buttonAddnew.Enable(False)
            self.buttonExport.Enable(False)
            self.reportChoice.Enable(False)
            #
            self.buttonClone.Enable(True)
            self.buttonSave.Enable(True)
            self.buttonCommit.Enable(True)
            self.buttonCancel.Enable(True)
            self.buttonDelete.Enable(True)
            self.buttonPrint.Enable(True)
        elif page==4:
            # DATA EXTRACTION
            self.buttonLogin.Enable(False)
            #
            self.buttonSearch.Enable(False)
            self.buttonExportSearch.Enable(False)
            self.buttonAddnew.Enable(False)
            self.buttonExport.Enable(False)
            self.reportChoice.Enable(False)
            #
            self.buttonClone.Enable(False)
            self.buttonSave.Enable(False)
            self.buttonCommit.Enable(False)
            self.buttonCancel.Enable(False)
            self.buttonDelete.Enable(False)
            self.buttonPrint.Enable(False)
        self.parent.SetStatusText("",2)

    def OnDataEntryMode(self,event):
        if self.frameDataExtraction.GetPosition().x==0:
            self.frameSearch.SetPosition(wx.Point(0,0))
            self.frameDataExtraction.SetPosition(wx.Point(5000,0))
            self.EnableDisableButtons(2)
        else:
            if self.frameSearch.GetPosition().x==0:
                wx.MessageBox(TT("Already in Data entry mode"), TT("Data entry mode"), wx.ICON_ERROR | wx.OK, self)
            else:
                wx.MessageBox(TT("Data entry tool must be invoked from Search page"), TT("Data entry mode"), wx.ICON_ERROR | wx.OK, self)

    def OnDataExtractionMode(self,event):
        candoit=True
        permission=GetSetting(self.parent,"extraction_mode_allowed_groups")
        if permission!="":
            s1=","+permission+","
            s2=","+str(self.id_group)+","
            if s2 not in s1: candoit=False
        if not candoit:
            wx.MessageBox(TT("Module disabled"), TT("Data Analysis"), wx.ICON_ERROR | wx.OK, self)
            return
            
        if self.frameSearch.GetPosition().x==0:
            wx.BeginBusyCursor()
            self.frameDataExtraction.SetPosition(wx.Point(0,0))
            self.frameSearch.SetPosition(wx.Point(5000,0))
            self.EnableDisableButtons(4)

            if self.dataExtractionPanel.optFieldOrder.GetSelection()!=0: self.dataExtractionPanel.optFieldOrder.SetSelection( 0 )
            
            select_fields=self.dataExtractionPanel.select_fields
            if select_fields.GetCount()==0:
                self.lstDataExtFields=[]
                #cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
                cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end, id_cycle FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" and id_section is not null and pos is not null and id_cycle<99 GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
                for row in cur:
                    #if "#99" not in row[0]: #sporco sistematico
                    select_fields.Append(row[0])
                    self.lstDataExtFields.append(row[0])

                self.dataExtFilterValue={}

            wx.EndBusyCursor()
        else:
            if self.frameDataExtraction.GetPosition().x==0:
                wx.MessageBox(TT("Already in Data extraction mode"), TT("Data extraction mode"), wx.ICON_ERROR | wx.OK, self)
            else:
                wx.MessageBox(TT("Data extraction tool must be invoked from Search page"), TT("Data extraction mode"), wx.ICON_ERROR | wx.OK, self)

    def col(self,row,hc,colname):
        cData=""
        try:cData=row[hc[colname]]
        except: pass
        if type(cData).__name__=="int": 
            dData=cData
        else:
            if cData!="":
                if not self.parent.trans.has_key(cData):
                    try: dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(cData)),"latin-1")
                    except: dData=cData
                    self.parent.trans[cData]=dData
                else:
                    dData=self.parent.trans[cData]
                dData=dData.encode('ascii', 'replace')
            else:
                dData=""
        return dData
    
    def StandardDataAnalysis(self):
        # wx.MessageBox(TT("Module disabled"), TT("Data Analysis"), wx.ICON_ERROR | wx.OK, self)
        con = sqlite.connect(DATABASE, isolation_level=None, timeout=60.0)
        cur = con.cursor()
        con_m = sqlite.connect(":memory:")
        cur_m = con_m.cursor()
        
        cur = con.cursor()
        
        lstfields=[]
        #cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end,rows.id_row,rows.id_user,rows.id_instance FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND pos is not null GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
        cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end,rows.id_row,rows.id_user,rows.id_instance FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
        for row in cur:
            lstfields.append(row[0])
        if len(lstfields)==0:
            wx.MessageBox(TT("No Records found"), TT("Data Analysis"), wx.ICON_ERROR | wx.OK, self)
        else:
            filename=""
            from os.path import expanduser
            home = expanduser("~")
            filterindex=-1
            now = datetime.datetime.now()
            ts=now.strftime("%Y%m%d%H%M%S")
            reportname="validation_"+ts+".xls"
            wildcard = "Excel 2002 (*.xls)|*.xls"
            dlg = wx.FileDialog(self, message=TT("Choose a file name"), defaultDir=home, defaultFile=reportname, wildcard=wildcard, style=wx.SAVE | wx.FD_OVERWRITE_PROMPT)
            if dlg.ShowModal() == wx.ID_OK:
                filename=dlg.GetPath()
                filterindex=dlg.GetFilterIndex()
            dlg.Destroy()        
            if filename!="":
                wx.BeginBusyCursor()
                #inizializzazione foglio excel
                if filename[-4:].lower()!=".xls": filename+=".xls"
                try: w=xlwt.Workbook()
                except:
                    try:
                        w=Workbook()
                    except: 
                        wx.EndBusyCursor()
                        wx.MessageBox(TT("File format not available in this package"), TT("Data Analysis"), wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                ws=w.add_sheet('validation')
                y=0; x=0
                ws.write(0,x,"group_shortcode")
                x+=1
                ws.write(0,x,self.keyname)
                secondkeyname=GetSetting(self.parent,"secondkeyname")
                if secondkeyname!="": 
                    x+=1
                    ws.write(0,x,secondkeyname)
                x+=1
                ws.write(0,x,"#")
                x+=1
                ws.write(0,x,TT('OK'))
                x+=1
                ws.write(0,x,TT('Type'))
                x+=1
                ws.write(0,x,TT('Page'))
                x+=1
                ws.write(0,x,TT('Field'))
                x+=1
                ws.write(0,x,TT('Description'))
                y=0
                #apertura recordset
                keyfields="group_shortcode,"+self.keyname
                if secondkeyname!="": keyfields+=","+secondkeyname
                square(con,con_m,"search",lstfields,self.id_group,self.id_crf,keyfields=keyfields,adminmode=False)
                hc={}; i=4
                for col in lstfields: hc[col]=i; i=i+1
                #
                cur_m.execute("SELECT distinct * FROM search")
                for row in cur_m:
                    contents={}
                    for elm in lstfields:
                        contents[elm]=self.col(row,hc,elm)
                    warnings={}
                    try: 
                        import analysis
                        warnings=analysis.GetWarnings(contents)
                        if len(warnings)>0:
                            for key, data in warnings.iteritems():
                                y+=1; x=0
                                ws.write(y,x,self.col(row,hc,"group_shortcode"))
                                x+=1
                                ws.write(y,x,self.col(row,hc,self.keyname))
                                if secondkeyname!="": 
                                    x+=1
                                    ws.write(y,x,self.col(row,hc,secondkeyname))
                                x+=1
                                ws.write(y,x,key)
                                x+=1
                                if "_warnings_" in contents:
                                    if str(key) in contents["_warnings_"].split(","):
                                        ws.write(y,x,"X")
                                    else:
                                        ws.write(y,x,"O")
                                else:
                                    ws.write(y,x,"_")
                                x+=1
                                ws.write(y,x,data[0])
                                x+=1
                                ws.write(y,x,data[1])
                                x+=1
                                ws.write(y,x,data[2])
                                x+=1
                                ws.write(y,x,data[3])
                    except: pass
                        #wx.MessageBox(TT("No validation defined"), TT("Data Analysis"), wx.ICON_ERROR | wx.OK, self)
                        #return
                        
                ws=w.add_sheet('metadata')
                y=0; x=0
                ws.write(0,x,"group_shortcode")
                x+=1
                ws.write(0,x,self.keyname)
                secondkeyname=GetSetting(self.parent,"secondkeyname")
                if secondkeyname!="": 
                    x+=1
                    ws.write(0,x,secondkeyname)
                x+=1
                ws.write(0,x,"Page No.")
                x+=1
                ws.write(0,x,"Field")
                x+=1
                ws.write(0,x,"Value")
                x+=1
                ws.write(0,x,"Confidence")
                x+=1
                ws.write(0,x,"Free text")
                cur_m.execute("SELECT distinct * FROM search")
                for row in cur_m:
                    contents={}
                    for elm in lstfields:
                        contents[elm]=self.col(row,hc,elm)
                    for key in contents:
                        if "__metadata_" in key:
                            basename=key[:key.find("__metadata_")]
                            confidence=self.col(row,hc,basename+"__metadata_confidence")
                            freetext=self.col(row,hc,basename+"__metadata_freetext")
                            y+=1; x=0
                            ws.write(y,x,self.col(row,hc,"group_shortcode"))
                            x+=1
                            ws.write(y,x,self.col(row,hc,self.keyname))
                            if secondkeyname!="": 
                                x+=1
                                ws.write(y,x,self.col(row,hc,secondkeyname))
                            x+=1
                            ws.write(y,x,self.arrHeadersIdSection[basename])
                            x+=1
                            ws.write(y,x,basename)
                            x+=1
                            ws.write(y,x,self.col(row,hc,basename))
                            x+=1
                            ws.write(y,x,confidence)
                            x+=1
                            ws.write(y,x,freetext)
                w.save(filename)
                if not assoc_open(filename):
                    wx.MessageBox("File\n'"+reportname+"'\n"+TT("created"), TT("Data Analysis"), wx.ICON_INFORMATION | wx.OK, self)
                wx.EndBusyCursor()
    
    def OnStatistics(self, evt):
        if self.frameSearch.GetPosition().x==0:
            try:
                if DEBUG_MODE: print "Data analysis: Trying custom Report"
                import analysis
                report=analysis.ReportDialog(self, TT("Data Analysis"), DATABASE, self.digestkey, self.parent.trans, self.id_group, self.id_crf)
                report.Show()
            except: 
                if DEBUG_MODE: print "Data analysis: Standard Report"
                self.StandardDataAnalysis()
        else:
            wx.MessageBox(TT("Data analysis tool must be invoked from Search page"), TT("Data Analysis"), wx.ICON_ERROR | wx.OK, self)
            
    def AuditTrailRecursiveSearch(self,txtreport,newer,recfullname,recdatetime,oldfullname,olddatetime,ref_row,ref_user,ref_instance):
        recordfound=False
        cur.execute("SELECT id_row,id_user,id_instance,date_time FROM rows WHERE status="+str(ref_row)+" AND status_user="+str(ref_user)+" AND (status_instance="+str(ref_instance)+" OR status_instance=0)")
        for row in cur:
            txtreport.append("=================================================================")
            recordfound=True
            ref_row=row[0]
            ref_user=row[1]
            ref_instance=row[2]
            olddatetime=row[3]
            
            older={}
            cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(ref_row)+" AND contents.id_user="+str(ref_user)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(ref_instance))
            for row in cur:
                if not older.has_key(row[0]):
                    keycont=row[0]
                    if row[3]!=0: keycont=keycont+"#"+str(row[3])
                    older[keycont]=row[1]
            cur.execute("SELECT username,fullname FROM users WHERE id_user="+str(ref_user))
            row=cur.fetchone()
            if row!=None:
                cData=str(row[0])
                if not self.parent.trans.has_key(cData):
                    dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                    self.parent.trans[cData]=dData
                else:
                    dData=self.parent.trans[cData]
                oldfullname="'"+dData+"' ("+row[1]+")"
            
            tmp=[]               
            for newkey in newer.keys():
                if not older.has_key(newkey):
                    cData=newer[newkey]
                    if not self.parent.trans.has_key(cData):
                        #dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(cData)),"latin-1")
                        dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                        self.parent.trans[cData]=dData
                    else:
                        dData=self.parent.trans[cData]
                    tmp.append("added '"+newkey+"': '"+dData+"'")
                else:
                    if older[newkey]!=newer[newkey]:
                        cData=older[newkey]
                        if not self.parent.trans.has_key(cData):
                            #dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(cData)),"latin-1")
                            dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                            self.parent.trans[cData]=dData
                        else:
                            dData=self.parent.trans[cData]
                        tmpStr="changed '"+newkey+"': '"+dData+"'"
                        cData=newer[newkey]
                        if not self.parent.trans.has_key(cData):
                            #dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(cData)),"latin-1")
                            dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                            self.parent.trans[cData]=dData
                        else:
                            dData=self.parent.trans[cData]
                        tmpStr=tmpStr+" -> '"+dData+"'"
                        tmp.append(tmpStr)
            for oldkey in older.keys():
                if not newer.has_key(oldkey):
                    cData=older[oldkey]
                    if not self.parent.trans.has_key(cData):
                        #dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(cData)),"latin-1")
                        dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                        self.parent.trans[cData]=dData
                    else:
                        dData=self.parent.trans[cData]
                    tmp.append("removed '"+oldkey+"': '"+dData+"'")
            txtreport.append("\r\n* Update by user "+recfullname+" at "+recdatetime)
            #txtreport.append("\r\n* Update by user "+oldfullname+" at "+olddatetime)
            for elm in tmp:
                txtreport.append(elm)
            recdatetime=olddatetime
            recfullname=oldfullname
            newer=older
            self.AuditTrailRecursiveSearch(txtreport,newer,recfullname,recdatetime,oldfullname,olddatetime,ref_row,ref_user,ref_instance)
        if not recordfound:
            txtreport.append("\r\n* Record created by user "+oldfullname+" at "+olddatetime)
    
    def OnAuditTrail(self,evt):
        if self.frameRecordset.GetPosition().x==0:
            wx.BeginBusyCursor()
            txtreport=[]
            ref_row=self.id_row
            ref_user=self.id_user
            ref_instance=self.id_rowinstance
            
            laststatus=0
            recdatetime=""
            cur.execute("SELECT date_time,status FROM rows WHERE id_row="+str(ref_row)+" AND id_user="+str(ref_user)+" AND id_instance="+str(ref_instance))            
            row=cur.fetchone()
            if row!=None: 
                recdatetime=row[0]
                laststatus=row[1]
            olddatetime=recdatetime

            recfullname=""
            cur.execute("SELECT username,fullname FROM users WHERE id_user="+str(ref_user))
            row=cur.fetchone()
            if row!=None:
                cData=str(row[0])
                if not self.parent.trans.has_key(cData):
                    dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                    self.parent.trans[cData]=dData
                else:
                    dData=self.parent.trans[cData]
                recfullname="'"+dData+"' ("+row[1]+")"
            oldfullname=recfullname
            
            if laststatus==-1:
                txtreport.append("\r\n* Record deleted by user "+oldfullname+" at "+olddatetime)
                            
            newer={}
            cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(ref_row)+" AND contents.id_user="+str(ref_user)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(ref_instance))
            for row in cur:
                if not newer.has_key(row[0]):
                    keycont=row[0]
                    if row[3]!=0: keycont=keycont+"#"+str(row[3])
                    newer[keycont]=row[1]
            #while True:
            self.AuditTrailRecursiveSearch(txtreport,newer,recfullname,recdatetime,oldfullname,olddatetime,ref_row,ref_user,ref_instance)
                
            now = datetime.datetime.utcnow()
            ts=now.strftime("%Y%m%d%H%M%S")
            reportname="audit_trail_"+self.cur_group_shortcode+"_"+self.curkeyvalue+"_"+ts+".txt"
            theFile = open(REPORTS_PATH + reportname, 'w')
            currow=""
            for elm in txtreport:
                theFile.write(elm.encode('ascii', 'replace') + '\r\n')
            theFile.close
            wx.EndBusyCursor()
#            if sys.platform != 'darwin':
            if not assoc_open(REPORTS_PATH+reportname):
                wx.MessageBox("File '"+reportname+"' created.", "Audit Trail", wx.ICON_INFORMATION | wx.OK, self)
        else:
            wx.MessageBox("Audit Trail tool must be invoked from inside a CRF", "Audit Trail", wx.ICON_ERROR | wx.OK, self)

    def OnImportData(self,event):
        candoit=False
        permission=GetSetting(self.parent,"import_data_allowed_groups")
        if permission!="":
            s1=","+permission+","
            s2=","+str(self.id_group)+","
            if s2 in s1: candoit=True
        if not candoit:
            permission=GetSetting(self.parent,"sysadmin_groups")
            if permission!="":
                s1=","+permission+","
                s2=","+str(self.id_group)+","
                if s2 in s1: candoit=True
        if not candoit:
            id_group_admin=GetSetting(self.parent,"id_group_admin")
            if str(self.id_group)==str(id_group_admin):
                candoit=True
        if not candoit:
            wx.MessageBox(TT("Module disabled"), TT("Import Data"), wx.ICON_ERROR | wx.OK, self)
            return
            
        if self.frameSearch.GetPosition().x!=0:
            wx.MessageBox(TT("Import Data tool must be invoked from Search page"), TT("Import Data"), wx.ICON_ERROR | wx.OK, self)
            return
                
        filename=""
        wildcard = "Excel 2007 (*.xlsx)|*.xlsx;*.xlsm|Excel 2002 (*.xls)|*.xls"
        dlg = wx.FileDialog(self, message=TT("Choose a file"), defaultFile="", wildcard=wildcard, style=wx.OPEN)
        if dlg.ShowModal() == wx.ID_OK:
            filename=dlg.GetPath()
        dlg.Destroy()        
        if filename!="":
            filetype=filename[filename.find(".")+1:].lower()
            if DEBUG_MODE: print "file type: '"+filetype+"'"
            if filetype in ("xlsx", "xlsm", "xls"):
                if filetype in ("xlsx", "xlsm"):
                    try:
                        from openpyxl.reader.excel import load_workbook
                        from openpyxl.cell import get_column_letter
                    except:
                        wx.MessageBox(TT("File format not available in this package"), TT("Import Data"), wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                else:
                    pass
                if DEBUG_MODE: print "Start importer"
                wx.BeginBusyCursor()
                self.DoAutoMergeDupes()
                self.parent.syncroTaskPause=True
                con_m = sqlite.connect(":memory:")
                cur_m = con_m.cursor()
                now = datetime.datetime.utcnow()
                ts=now.strftime("%d/%m/%Y, %H:%M:%S")
                ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
                if filetype in ("xlsx", "xlsm"):
                    wb = load_workbook(filename = filename)
                    # for sheet_name in wb.get_sheet_names():
                    # ws = wb.get_sheet_by_name(name = sheet_name)
                    ws = wb.get_sheet_by_name(name = wb.get_sheet_names()[0])
                    if DEBUG_MODE: print "Excel 2007: Sheet 1 open"
                else:
                    wb = xlrd.open_workbook(filename)
                    ws = wb.sheet_by_index(0)
                    if DEBUG_MODE: print "Excel 2002: Sheet 1 open"
                #keyfields - begin
                hashrecords={}
                keypos={}
                lstfields=[]
                x=1
                while True:
                    fieldname=None
                    if filetype in ("xlsx", "xlsm"):
                        fieldname=ws.cell('%s%s'%(get_column_letter(x), 1)).value
                    else:
                        if x <= ws.ncols:
                            try:
                                if ws.cell_value(0,x-1)!=None:
                                    fieldname=str(ws.cell_value(0,x-1)).strip()
                                    if fieldname=="None": fieldname=None
                            except: pass
                    if fieldname!=None:
                        if DEBUG_MODE: print "Found fieldname="+str(fieldname)
                        if fieldname[0]=="(" and fieldname[-1]==")":
                            keyname=fieldname.replace("(","").replace(")","")
                            lstfields.append(keyname)
                            keypos[keyname]=x
                        x=x+1
                    else: break
                if len(lstfields)>0:
                    strlstfields=','.join(lstfields)
                    square(con,con_m,"search",lstfields,self.id_group,self.id_crf)
                    sql="SELECT id_row,id_user,id_instance,date_time,"+strlstfields+" FROM search"
                    cur_m.execute(sql)
                    cols = [i[0] for i in cur_m.description]; hc={}; i=0
                    for col in cols: hc[col]=i; i=i+1
                    for row in cur_m:
                        keyvalues=[]
                        for key in lstfields: 
                            cData=row[hc[key]]
                            dData=""
                            if not self.parent.trans.has_key(cData):
                                dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                                self.parent.trans[cData]=dData
                            else:
                                dData=self.parent.trans[cData]
                            keyvalues.append(dData)
                        strkeyvalues='|'.join(keyvalues)
                        hashrecords[strkeyvalues]=[row[0],row[1],row[2]]
                #keyfields - end
                y=2
                while True:
                    fieldvalue=None
                    if filetype in ("xlsx", "xlsm"):
                        fieldvalue=ws.cell('%s%s'%(get_column_letter(1), y)).value
                    else:
                        if y <= ws.nrows:
                            fieldvalue=ws.cell_value(y-1,0)
                    if fieldvalue!=None:
                        #print str(y)+" - "+ws.cell_value(y-1,6)
                        self.contents={}
                        self.rap=""
                        newrecord=True
                        if len(lstfields)>0:
                            keyvalues=[]
                            for fldName in lstfields:
                                elm=""
                                if filetype in ("xlsx", "xlsm"):
                                    # elm=str(ws.cell('%s%s'%(get_column_letter(keypos[fldName]), y)).value)
                                    elm=ws.cell('%s%s'%(get_column_letter(keypos[fldName]), y)).value
                                else:
                                    # elm=str(ws.cell_value(y-1,keypos[fldName]-1))
                                    elm=ws.cell_value(y-1,keypos[fldName]-1)
                                    #if fldName in ("(cod_sede)","(cod_paz)"): elm=str(int(elm))
                                try: elm=str(elm)
                                except: pass
                                if len(elm)>0:
                                    if elm[0]=="'": elm=elm[1:]
                                keyvalues.append(elm)
                            strkeyvalues='|'.join(keyvalues)
                            if hashrecords.has_key(strkeyvalues):
                                newrecord=False
                                #compact openrecord - begin
                                self.id_row=hashrecords[strkeyvalues][0]
                                self.id_user=hashrecords[strkeyvalues][1]
                                self.id_rowinstance=hashrecords[strkeyvalues][2]
                                self.rap=""
                                cur.execute("SELECT rap FROM rows WHERE id_row="+str(self.id_row)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance))
                                row = cur.fetchone()
                                if row!=None: self.rap=row[0]
                                self.cur_group_shortcode=""
                                contents={}
                                idKey=self.hashHeaders[self.keyname]
                                curKeyIdDic=0
                                cur.execute("SELECT id_dictionary FROM contents_index WHERE id_row="+str(self.id_row)+" AND id_user="+str(self.id_user)+" AND id_header="+str(idKey)+" AND id_instance="+str(self.id_rowinstance))
                                row = cur.fetchone()
                                if row!=None: curKeyIdDic=row[0]
                                if self.keymode=="private":
                                    query="SELECT id_row,id_user FROM rows WHERE id_row IN (SELECT MAX(rows.id_row) FROM rows,contents_index WHERE rows.id_header="+str(self.id_crf)+" AND rows.status=0 AND contents_index.id_header="+str(idKey)+" AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_dictionary="+str(curKeyIdDic)+" AND rows.id_user="+str(self.id_user)+")"
                                elif self.keymode=="group":
                                    query="SELECT id_row,id_user FROM rows WHERE id_row IN (SELECT MAX(rows.id_row) FROM users,rows,contents_index WHERE rows.id_header="+str(self.id_crf)+" AND rows.status=0 AND contents_index.id_header="+str(idKey)+" AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_dictionary="+str(curKeyIdDic)+" AND users.id_group="+str(self.id_group)+")"
                                else:
                                    query="SELECT id_row,id_user FROM rows WHERE id_row IN (SELECT MAX(rows.id_row) FROM rows,contents_index WHERE rows.id_header="+str(self.id_crf)+" AND rows.status=0 AND contents_index.id_header="+str(idKey)+" AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_dictionary="+str(curKeyIdDic)+")"
                                cur.execute(query)
                                row = cur.fetchone()
                                if row!=None:
                                    id_sharedrow=row[0]
                                    id_shareduser=row[1]
                                    cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(id_sharedrow)+" AND contents.id_user="+str(id_shareduser)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(self.id_rowinstance)+" AND deftype='S' ORDER BY contents.id_row, contents.id_user, contents.id_instance, contents.id_header, contents.id_cycle")
                                    for row in cur:
                                        keycont=row[0]
                                        if row[3]!=0: keycont=keycont+"#"+str(row[3])
                                        if not contents.has_key(keycont):
                                            if not self.parent.trans.has_key(row[1]):
                                                dData=HB_DecryptOne(self.digestkey,row[1],"latin-1")
                                                self.parent.trans[row[1]]=dData
                                            else:
                                                dData=self.parent.trans[row[1]]
                                            contents[keycont]=dData
                                self.curkeyvalue=""
                                cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(self.id_row)+" AND contents.id_user="+str(self.id_user)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(self.id_rowinstance)+" ORDER BY contents.id_row, contents.id_user, contents.id_instance, contents.id_header, contents.id_cycle")
                                for row in cur:
                                    keycont=row[0]
                                    if row[3]!=0: keycont=keycont+"#"+str(row[3])
                                    if not contents.has_key(keycont):
                                        if not self.parent.trans.has_key(row[1]):
                                            dData=HB_DecryptOne(self.digestkey,row[1],"latin-1")
                                            self.parent.trans[row[1]]=dData
                                        else:
                                            dData=self.parent.trans[row[1]]
                                        contents[keycont]=dData
                                        if row[2]==idKey: self.curkeyvalue=dData
                                self.contents={}
                                self.loadedcontents={}
                                for key in contents:
                                    self.contents[key]=contents[key]
                                    self.loadedcontents[key]=contents[key]
                                id_oldrow=self.id_row
                                #compact openrecord - end
                        x=1
                        while True:
                            fieldname=None
                            if filetype in ("xlsx", "xlsm"):
                                fieldname=ws.cell('%s%s'%(get_column_letter(x), 1)).value
                            else:
                                if x <= ws.ncols:
                                    try:
                                        if ws.cell_value(0,x-1)!=None:
                                            fieldname=str(ws.cell_value(0,x-1)).strip()
                                            if fieldname=="None": fieldname=None
                                    except: pass
                            if fieldname!=None:
                                elm=""
                                if filetype in ("xlsx", "xlsm"):
                                    elm=ws.cell('%s%s'%(get_column_letter(x), y)).value
                                else:
                                    elm=ws.cell_value(y-1,x-1)
                                try: elm=str(elm)
                                except: pass
                                if len(elm)>0:
                                    if elm[0]=="'": elm=elm[1:]
                                try: 
                                    fieldvalue=str(elm)
                                except:
                                    fieldvalue=elm.encode('latin-1', 'replace')
                                hasvalue=False
                                if fieldvalue!=None:
                                    if len(str(fieldvalue))>0: hasvalue=True
                                if hasvalue:
                                    self.contents[fieldname.replace("(","").replace(")","")]=str(fieldvalue)
                                else:
                                    if self.contents.has_key(fieldname.replace("(","").replace(")","")):
                                        del self.contents[fieldname.replace("(","").replace(")","")]
                                x=x+1
                            else: break
                            
                        #dummy save test - begin
                        dummysave=True
                        import unicodedata
                        if not newrecord:
                            checkedkeys={}
                            for keycont in self.contents:
                                if self.loadedcontents.has_key(keycont):
                                    #if self.contents[keycont]!=self.loadedcontents[keycont].encode('latin-1', 'replace'):
                                    newval=""
                                    try: newval=self.contents[keycont].decode('latin-1')
                                    except: newval=self.contents[keycont]
                                    if newval!=self.loadedcontents[keycont]:
                                        #print "updated "+keycont+": '"+self.loadedcontents[keycont]+"' -> '"+self.contents[keycont]+"'"
                                        #print self.loadedcontents[keycont]
                                        #print self.contents[keycont]
                                        dummysave=False
                                else:
                                    dummysave=False
                                    #print "added "+keycont+": '"+self.contents[keycont]+"'"
                                if not dummysave: break
                                checkedkeys[keycont]=True
                            if dummysave:
                                for keycont in self.loadedcontents:
                                    if not checkedkeys.has_key(keycont):
                                        dummysave=False
                                        #print "removed "+keycont+": '"+self.loadedcontents[keycont]+"'"
                        else:
                            dummysave=False
                        #dummy save test - end
                        if not dummysave:
                            #if DEBUG_MODE: print "Saving: "+str(contents)
                            #compact saverecord - begin
                            #tmprap=self.uap
                            tmprap=self.rap
                            if not (tmprap): tmprap=self.uap
                            if not (tmprap): tmprap=""
                            cur.execute("SELECT max(id_row) FROM rows")
                            row = cur.fetchone()
                            if row==None: self.id_row=1
                            else:
                                if row[0]==None: self.id_row=1 
                                else: self.id_row=row[0]+1
                                
                            import time
                            while self.parent.syncroTaskRunning:
                                time.sleep(1)
                                if DEBUG_MODE: print "waiting synchro standby..."
                            cur.execute("BEGIN TRANSACTION")
                            
                            cur.execute("INSERT INTO rows(id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(self.id_row)+","+str(self.id_logged_user)+","+str(self.id_crf)+",'"+tmprap+"','"+ts_db+"',0,0,0,"+str(self.parent.id_instance)+")")
                            self.SaveInDatabase()
                            #if len(lstfields)>0 and hashrecords.has_key(strkeyvalues):
                            if not newrecord:
                                cur.execute("UPDATE rows SET status="+str(self.id_row)+",status_user="+str(self.id_logged_user)+",status_instance="+str(self.parent.id_instance)+" WHERE id_row="+str(id_oldrow)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance)+" AND id_header="+str(self.id_crf))
                                
                            cur.execute("COMMIT TRANSACTION")
                            self.parent.syncrotablesupd["rows"]=True
                            #compact saverecord - end
                        y=y+1
                    else: break
                
                #reload headers - begin
                cur.execute("SELECT id_header,description,onchange,defaultvalue,validation,id_section,pos,typedesc,subtypedesc,label FROM headers WHERE child_of="+str(self.id_crf))
                for row in cur:
                    self.arrHeaders[row[0]]=row[1]
                    self.hashHeaders[row[1]]=row[0]
                #reload headers - end

                self.parent.syncroTaskPause=False
                wx.EndBusyCursor()
                wx.MessageBox(TT("All records imported successfully"), TT("Import Data"), wx.ICON_INFORMATION | wx.OK, self)
                
            elif filetype in ("txt","csv"):
                wx.MessageBox(TT("File format not available in this package"), TT("Import Data"), wx.ICON_EXCLAMATION | wx.OK, self)
                return
                
    
    def OnModifyAccessRights(self,event):
        if self.frameRecordset.GetPosition().x!=0:
            wx.MessageBox("Modify Access Rights tool must be invoked from inside a CRF", "Modify Access Rights", wx.ICON_ERROR | wx.OK, self)
            return
        if str(self.id_group)!=str(GetSetting(self.parent,"id_group_admin")):
            wx.MessageBox("Modify Access Rights tool can only be invoked by an administrator", "Modify Access Rights", wx.ICON_ERROR | wx.OK, self)
            return

        dlg = AccessRightsDlg(self)
        dlg.ShowModal()
        dlg.Destroy()  

    def OnModifyMetadata(self,event):
        if self.frameRecordset.GetPosition().x!=0:
            wx.MessageBox("Metadata tool must be invoked from inside a CRF", "Metadata", wx.ICON_ERROR | wx.OK, self)
            return
        if not self.IsInputFieldEditable(self.lastfocusvariableobj):
            wx.MessageBox("Not an user input field", "Metadata", wx.ICON_ERROR | wx.OK, self)
            return
        variablename=self.lastfocusvariable
        id_cycle=0
        if self.ActivateCyclicity:
            if self.Cyclicity[self.ifrm+1]>0: 
                id_cycle=wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1
        if variablename!="":
            if id_cycle!=0: variablename=variablename+"#"+str(id_cycle)
            top = MetadataFrame(self,variablename)
            top.ShowModal()
            try: top.Destroy()
            except: pass
            #colorazione campi - inizio
            confidence=variablename+"__metadata_confidence"
            if confidence in self.contents:
                if int(self.contents[confidence])==1: 
                    self.lastfocusvariableobj.SetBackgroundColour(wx.Colour(255, 0, 0))
                    try: self.lastfocusvariableobj.SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 0, 0), validBackgroundColour=wx.Colour(255, 0, 0), invalidBackgroundColour=wx.Colour(255, 0, 0))
                    except: pass
                if int(self.contents[confidence])==2: 
                    self.lastfocusvariableobj.SetBackgroundColour(wx.Colour(0, 255, 0))
                    try: self.lastfocusvariableobj.SetCtrlParameters(emptyBackgroundColour=wx.Colour(0, 255, 0), validBackgroundColour=wx.Colour(0, 255, 0), invalidBackgroundColour=wx.Colour(0, 255, 0))
                    except: pass
                if int(self.contents[confidence])==3: 
                    self.lastfocusvariableobj.SetBackgroundColour(wx.Colour(255, 255, 0))
                    try: self.lastfocusvariableobj.SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 0), validBackgroundColour=wx.Colour(255, 255, 0), invalidBackgroundColour=wx.Colour(255, 255, 0))
                    except: pass
            else:
                self.lastfocusvariableobj.SetBackgroundColour(wx.Colour(255, 255, 255))
                try: self.lastfocusvariableobj.SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 255), validBackgroundColour=wx.Colour(255, 255, 255), invalidBackgroundColour=wx.Colour(255, 255, 255))
                except: pass
            self.lastfocusvariableobj.Refresh()
            #colorazione campi - fine
        else:
            wx.MessageBox("No variable selected", "Metadata", wx.ICON_ERROR | wx.OK, self)
            return
        
    def ForceChangeUserPassword(self):
        self.userpasswordchanged=False
        self.dlg = wx.Dialog(None, -1, TT("PASSWORD EXPIRED")+" - "+TT("User")+": "+self.txtUsername.GetValue(), size=wx.Size(400,160))

        wx.StaticText(label=TT('New password'), parent=self.dlg, pos=wx.Point(10, 10), size=wx.Size(180, 25), style=0)
        self.txtNewPassword1 = wx.TextCtrl(parent=self.dlg, pos=wx.Point(180, 10), size=wx.Size(200, 25), style=wx.TE_PASSWORD, value='')
        wx.StaticText(label=TT('Repeat new password'), parent=self.dlg, pos=wx.Point(10, 50), size=wx.Size(180, 25), style=0)
        self.txtNewPassword2 = wx.TextCtrl(parent=self.dlg, pos=wx.Point(180, 50), size=wx.Size(200, 25), style=wx.TE_PASSWORD, value='')
        btnOk = wx.Button(label='OK', parent=self.dlg, pos=wx.Point(70,100), size=wx.Size(100,25))
        btnCancel = wx.Button(label=TT('Cancel'), parent=self.dlg, pos=wx.Point(230,100), size=wx.Size(100,25))
        btnOk.Bind(wx.EVT_BUTTON, self.OnConfirmUserPasswordChange)
        btnCancel.Bind(wx.EVT_BUTTON, self.OnCancelUserPasswordChange)
        self.dlg.ShowModal()
        self.dlg.Destroy()
        if not self.userpasswordchanged:
            self.frameLogin.SetPosition(wx.Point(0,0))
            self.frameSearch.SetPosition(wx.Point(5000,0))
            self.frameRecordset.SetPosition(wx.Point(5000,0))
            self.EnableDisableButtons(1)
            
    def OnChangeUserPassword(self,event):
        if self.frameSearch.GetPosition().x==0:
            self.userpasswordchanged=False
            # wx.MessageBox("Module disabled.", "Change user password", wx.ICON_ERROR | wx.OK, self)
            self.dlg = wx.Dialog(None, -1, TT("User")+": "+self.txtUsername.GetValue(), size=wx.Size(400,160))

            wx.StaticText(label=TT('New password'), parent=self.dlg, pos=wx.Point(10, 10), size=wx.Size(180, 25), style=0)
            self.txtNewPassword1 = wx.TextCtrl(parent=self.dlg, pos=wx.Point(180, 10), size=wx.Size(200, 25), style=wx.TE_PASSWORD, value='')
            wx.StaticText(label=TT('Repeat new password'), parent=self.dlg, pos=wx.Point(10, 50), size=wx.Size(180, 25), style=0)
            self.txtNewPassword2 = wx.TextCtrl(parent=self.dlg, pos=wx.Point(180, 50), size=wx.Size(200, 25), style=wx.TE_PASSWORD, value='')
            btnOk = wx.Button(label='OK', parent=self.dlg, pos=wx.Point(70,100), size=wx.Size(100,25))
            btnCancel = wx.Button(label=TT('Cancel'), parent=self.dlg, pos=wx.Point(230,100), size=wx.Size(100,25))
            btnOk.Bind(wx.EVT_BUTTON, self.OnConfirmUserPasswordChange)
            btnCancel.Bind(wx.EVT_BUTTON, self.OnCancelUserPasswordChange)
            self.dlg.ShowModal()
            self.dlg.Destroy()
        else:
            wx.MessageBox(TT("Change user password tool must be invoked from Search page"), TT("Change user password"), wx.ICON_ERROR | wx.OK, self)

    def OnConfirmUserPasswordChange(self,event):
        if len(self.txtNewPassword1.GetValue())==0:
            wx.MessageBox(TT("The password cannot be empty"), TT("Change user password"), wx.ICON_ERROR | wx.OK, self)
            return
        if self.txtNewPassword1.GetValue()!=self.txtNewPassword2.GetValue():
            wx.MessageBox(TT("The two passwords are different"), TT("Change user password"), wx.ICON_ERROR | wx.OK, self)
            return
        if len(self.txtNewPassword1.GetValue())<8:
            wx.MessageBox(TT("The password must be at least 8 characters long"), TT("Change user password"), wx.ICON_ERROR | wx.OK, self)
            return
        now = datetime.datetime.utcnow()
        ts=now.strftime("%d/%m/%Y, %H:%M:%S")
        ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
        cUsername=base64.b64encode(rijndael.EncryptData(self.digestkey,self.txtUsername.GetValue()))
        cPassword=base64.b64encode(rijndael.EncryptData(self.digestkey,self.txtPassword.GetValue()))
        cNewPassword=base64.b64encode(rijndael.EncryptData(self.digestkey,self.txtNewPassword1.GetValue()))
        cur.execute("UPDATE users SET password='"+cNewPassword+"', date_time='"+ts_db+"' WHERE username='"+cUsername+"' and password='"+cPassword+"' AND status=0")
        self.parent.syncrotablesupd["users"]=True
        self.userpasswordchanged=True
        self.dlg.Destroy()

    def OnCancelUserPasswordChange(self,event):
        self.dlg.Destroy()

    def OnChangeCypherKey(self,event):
        if self.frameSearch.GetPosition().x==0:
            wx.MessageBox(TT("Module disabled"), TT("Change cypher key"), wx.ICON_ERROR | wx.OK, self)
        else:
            wx.MessageBox(TT("Change cypher key tool must be invoked from Search page"), TT("Change cypher key"), wx.ICON_ERROR | wx.OK, self)

    def OnIdentities(self,event):
        import pickle
        if self.frameLogin.GetPosition().x==0:
            self.buttonLogin.Enable(False)
            #Pagina di Login - Scelta identita'
            from os.path import expanduser
            home = expanduser("~")
            wildcard = PROGNAME.lower()+" OTP key files (*.otp)|*.otp"
            dlg = wx.FileDialog(self, message=TT("OTP Credentials Keyring"), defaultDir=home, defaultFile=PROGNAME.lower()+".otp", wildcard=wildcard, style=wx.OPEN)
            if dlg.ShowModal() == wx.ID_OK:
                path = dlg.GetPath()
                dlg.Destroy()
                if not DEBUG_MODE:
                    if not IsRemovable(path):
                        wx.MessageBox(TT("The file must be on a removable device (eg. a USB pendrive)"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                        self.buttonLogin.Enable(True)
                        return
                    elif IsSameDevice(path,DATABASE_PATH):
                        wx.MessageBox(TT("The file and "+PROGNAME+" cannot be stored on the same device"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                        self.buttonLogin.Enable(True)
                        return
                pin=""
                pinok=False
                while not pinok:
                    dlg = wx.PasswordEntryDialog(self, TT('Enter your PIN (8 chars minimum)'),TT("Keyring PIN"))
                    dlg.SetValue("")
                    if dlg.ShowModal() == wx.ID_OK:
                        pin=dlg.GetValue()
                        dlg.Destroy()
                        if len(pin)>=8: pinok=True
                    else:
                        dlg.Destroy()
                        self.buttonLogin.Enable(True)
                        return
 
                otpkeys={}
                try: 
                    f_in = open(path, 'rb')
                    content_in = f_in.read()
                    f_in.close()
                    content_out=""
                    try:
                        content_out=HB_DecryptOne(sha256(pin).digest(),content_in,"")
                    except:
                        wx.MessageBox(TT("Invalid PIN"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                        self.buttonLogin.Enable(True)
                        return
                    #otpkeys=pickle.load(file(path, "r"))
                    try:
                        otpkeys=pickle.loads(content_out)
                    except:
                        wx.MessageBox(TT("Invalid PIN or Keyring"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                        self.buttonLogin.Enable(True)
                        return
                except: pass
                project_name=GetSetting(self.parent,"project_name")
                validkeys=[]
                for key in otpkeys:
                    if key.split(";")[0]==project_name:
                        cur.execute("select otp from users where id_user="+key.split(";")[1])
                        row = cur.fetchone()
                        if row!=None:
                            if row[0]==otpkeys[key][0]:
                                validkeys.append(otpkeys[key])
                found=False
                res=None
                if len(validkeys)==0:
                    wx.MessageBox(TT("No valid keys found"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                    self.buttonLogin.Enable(True)
                    return
                elif len(validkeys)==1:
                    res=validkeys[0]
                    found=True
                else:
                    choices=[]
                    for row in validkeys:
                        choices.append(row[2]+" - "+row[3]+" - "+row[5])
                    dlg = wx.SingleChoiceDialog(self, TT("OTP Credentials Keyring"), TT('Open file'), choices, wx.CHOICEDLG_STYLE)
                    if dlg.ShowModal() == wx.ID_OK:
                        res=validkeys[dlg.GetSelection()]
                        found=True
                    dlg.Destroy()
                if found:
                    dUsername=""
                    dPassword=""
                    cypherkey=""
                    idcrf=""
                    cur.execute("select username,password from users where id_user="+res[1]+" and status=0")
                    row = cur.fetchone()
                    if row!=None:
                        cypherkey=res[4]
                        self.digestkey=sha256(cypherkey).digest()
                        cUsername=row[0]
                        dUsername=HB_DecryptOne(self.digestkey,cUsername,"latin-1")
                        cPassword=row[1]
                        dPassword=HB_DecryptOne(self.digestkey,cPassword,"latin-1")
                        self.txtUsername.SetValue(dUsername)
                        self.txtPassword.SetValue(dPassword)
                        self.txtCryptokey.SetValue(cypherkey)
                        try: self.choiceDatabase.SetValue(res[5])
                        except: pass
                        
                        import random
                        random.seed()
                        otp=str(int(random.random()*10000000000))[0:10]
                        otpkeys[project_name+";"+res[1]]=[otp,res[1],res[2],res[3],res[4],res[5]]
                        content_in=pickle.dumps(otpkeys)
                        content_out=HB_EncryptOne(sha256(pin).digest(),content_in)
                        f_out = open(path, 'wb')
                        f_out.write(content_out)
                        f_out.close()
                        cur.execute("update users set otp='"+otp+"' where id_user="+res[1]+" and status=0")
                        
                        if DEBUG_MODE: print dUsername, dPassword, cypherkey, res[5], "new otp: "+otp
                        self.OnButtonLoginButton(event)
                        return
                    else:
                        self.buttonLogin.Enable(True)
                        return
                else:
                    self.buttonLogin.Enable(True)
                    return
            else:
                dlg.Destroy()
                self.buttonLogin.Enable(True)
                return
        elif self.frameSearch.GetPosition().x==0:
            #Pagina di Ricerca - Salvataggio identita'
            self.buttonSearch.Enable(False)
            from os.path import expanduser
            home = expanduser("~")
            wildcard = PROGNAME.lower()+" OTP key files (*.otp)|*.otp"
            dlg = wx.FileDialog(self, message="Save file as ...", defaultDir=home, defaultFile=PROGNAME.lower()+".otp", wildcard=wildcard, style=wx.SAVE)
            dlg.SetFilterIndex(0)
            if dlg.ShowModal() == wx.ID_OK:
                path = dlg.GetPath()
                dlg.Destroy()
                if not DEBUG_MODE:
                    if not IsRemovable(path):
                        wx.MessageBox(TT("The file must be on a removable device (eg. a USB pendrive)"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                        self.buttonSearch.Enable(True)
                        return
                    elif IsSameDevice(path,DATABASE_PATH):
                        wx.MessageBox(TT("The file and "+PROGNAME+" cannot be stored on the same device"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                        self.buttonSearch.Enable(True)
                        return
                pin=""
                pinok=False
                while not pinok:
                    dlg = wx.PasswordEntryDialog(self, TT('Enter your PIN (8 chars minimum)'),TT("Keyring PIN"))
                    dlg.SetValue("")
                    if dlg.ShowModal() == wx.ID_OK:
                        pin=dlg.GetValue()
                        dlg.Destroy()
                        if len(pin)>=8: pinok=True
                    else:
                        dlg.Destroy()
                        self.buttonSearch.Enable(True)
                        return
 
                otpkeys={}
                if(os.path.exists(path)):
                    try: 
                        f_in = open(path, 'rb')
                        content_in = f_in.read()
                        f_in.close()
                        content_out=""
                        try:
                            content_out=HB_DecryptOne(sha256(pin).digest(),content_in,"")
                        except:
                            wx.MessageBox(TT("Invalid PIN"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                            self.buttonSearch.Enable(True)
                            return
                        #otpkeys=pickle.load(file(path, "r"))
                        try:
                            otpkeys=pickle.loads(content_out)
                        except:
                            wx.MessageBox(TT("Invalid PIN or Keyring"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
                            self.buttonSearch.Enable(True)
                            return
                    except: pass
                project_name=GetSetting(self.parent,"project_name")
                id_user=self.id_logged_user
                id_group=0
                fullname=""
                cur.execute("select id_group,username,fullname from users where id_user="+str(self.id_logged_user)+" and status=0")
                row = cur.fetchone()
                if row!=None:
                    id_group=row[0]
                    fullname=row[2]
                    if fullname==None: fullname=""
                    if fullname=="":
                        cUsername=row[1]
                        dUsername=HB_DecryptOne(self.digestkey,cUsername,"latin-1")
                        fullname=dUsername
                shortcode=""
                groupname=""
                cur.execute("select shortcode, description from groups where id_group="+str(id_group)+" and status=0")
                row = cur.fetchone()
                if row!=None:
                    shortcode=row[0]
                    groupname=row[1]
                    if groupname==None: groupname=""
                    if groupname!="": groupname=" - ("+groupname+")"
                groupname=shortcode+groupname
                database=self.choiceDatabase.GetStringSelection()
                if database=="": database=CRF_DESC
                import random
                random.seed()
                otp=str(int(random.random()*10000000000))[0:10]
                otpkeys[project_name+";"+str(id_user)]=[otp,str(id_user),fullname,groupname,self.cryptokey,database]
                #pickle.dump(otpkeys,file(path, "w"))
                content_in=pickle.dumps(otpkeys)
                content_out=HB_EncryptOne(sha256(pin).digest(),content_in)
                f_out = open(path, 'wb')
                f_out.write(content_out)
                f_out.close()
                
                cur.execute("update users set otp='"+otp+"' where id_user="+str(self.id_logged_user)+" and status=0")
                wx.MessageBox(TT("File saved successfully"), TT("OTP Credentials Keyring"), wx.ICON_INFORMATION | wx.OK, self)
                self.buttonSearch.Enable(True)
                return
            else:
                dlg.Destroy()
                self.buttonSearch.Enable(True)
                return
        elif self.frameRecordset.GetPosition().x==0:
            #Pagina di CRF - Funzione disabilitata
            wx.MessageBox(TT("OTP Credentials Keyring")+" "+TT("must be invoked from Login or Search page"), TT("OTP Credentials Keyring"), wx.ICON_ERROR | wx.OK, self)
            return

    def OnDataExtractionMarkAll(self,event):
        lstfields=self.dataExtractionPanel.select_fields
        for i in range(lstfields.GetCount()):
            if sys.platform == 'darwin':
                lstfields.SetSelection(i, True)
            else:
                lstfields.Check(i, True)

    def OnDataExtractionUnmarkAll(self,event):
        lstfields=self.dataExtractionPanel.select_fields
        for i in range(lstfields.GetCount()):
            if sys.platform == 'darwin':
                lstfields.SetSelection(i, False)
            else:
                lstfields.Check(i, False)

    #def OnDataExtractionFilterFields(self,event):
    #    event.Skip()
    #    obj=event.GetEventObject()
    #    self.FulfillFilterFields(obj)

    def FulfillFilterFieldsNames(self,idx):
        obj=eval("self.dataExtractionPanel.filter_field"+`idx`)
        self.FulfillFilterFields(obj)

    def FulfillFilterFields(self,obj):
        wx.BeginBusyCursor()
        select_fields=self.dataExtractionPanel.select_fields
        obj.Clear()
        if sys.platform == 'darwin':
            for i in range(len(select_fields.GetSelections())):
                obj.Append(self.lstDataExtFields[select_fields.GetSelections()[i]]) 
        else:
            for elm in select_fields.GetCheckedStrings():
                obj.Append(elm) 
        wx.EndBusyCursor()
        
    def FulfillFilterFieldsValues(self,idx):
        if not self.dataExtFilterValue.has_key(idx): self.dataExtFilterValue[idx]=""
        obj=eval("self.dataExtractionPanel.filter_value"+`idx`)
        filter_field=eval("self.dataExtractionPanel.filter_field"+`idx`)
        if eval("self.dataExtractionPanel.filter_type"+`idx`+".GetStringSelection()")=="field":
            self.dataExtFilterValue[idx]=""
            self.FulfillFilterFields(obj)
        elif eval("self.dataExtractionPanel.filter_type"+`idx`+".GetStringSelection()")=="value":
            if self.dataExtFilterValue[idx]!=filter_field.GetValue():
                if not self.hashHeaders.has_key(filter_field.GetValue().split("#")[0]): 
                    obj.Clear()
                    return        
                self.dataExtFilterValue[idx]=filter_field.GetValue()
                self.FulfillFilterValues(obj,"filter_value"+`idx`,filter_field)

    def FulfillFilterValues(self,obj,objname,filter_field):
        wx.BeginBusyCursor()
        obj.Clear()
        lstfields=[]
        lstfields.append(filter_field.GetValue())
        strlstfields=filter_field.GetValue()
        con_m = sqlite.connect(":memory:")
        cur_m = con_m.cursor()
        square(con,con_m,"search",lstfields,self.id_group,self.id_crf)
        sql="SELECT DISTINCT "+(strlstfields.replace("-","___")).replace("#","__")+" FROM search"
        cur_m.execute(sql)
        cols = [i[0] for i in cur_m.description]; hc={}; i=0
        for col in cols: hc[(col.replace("___","-")).replace("__","#")]=i; i=i+1
        for row in cur_m:
            if not self.parent.trans.has_key(row[hc[strlstfields]]):
                #dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(row[hc[strlstfields]])),"latin-1")
                dData=HB_DecryptOne(self.digestkey,row[hc[strlstfields]],"latin-1")
                self.parent.trans[row[hc[strlstfields]]]=dData
            else:
                dData=self.parent.trans[row[hc[strlstfields]]]
            obj.Append(dData) 
        cur_m.close()
        con_m.close()
        wx.EndBusyCursor()

    def OnDataExtractionFilterField1(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(1)
    def OnDataExtractionFilterField2(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(2)
    def OnDataExtractionFilterField3(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(3)
    def OnDataExtractionFilterField4(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(4)
    def OnDataExtractionFilterField5(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(5)
    def OnDataExtractionFilterField6(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(6)
    def OnDataExtractionFilterField7(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(7)
    def OnDataExtractionFilterField8(self,event):
        event.Skip()
        self.FulfillFilterFieldsNames(8)

    def OnDataExtractionFilterValue1(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(1)
    def OnDataExtractionFilterValue2(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(2)
    def OnDataExtractionFilterValue3(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(3)
    def OnDataExtractionFilterValue4(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(4)
    def OnDataExtractionFilterValue5(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(5)
    def OnDataExtractionFilterValue6(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(6)
    def OnDataExtractionFilterValue7(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(7)
    def OnDataExtractionFilterValue8(self,event):
        event.Skip()
        self.FulfillFilterFieldsValues(8)

    def GetOrderedElm(self,headerName,value):
        elm=value
        if headerName==self.keyname or self.arrHeadersTypedesc[headerName.split("#")[0]]=="number":
            elm=elm.strip().replace(",",".")
            dec=elm.find(".")
            if dec>=0:
                elm=("0000000000"+elm[:dec])[-10:] + (elm[dec:]+"0000000000")[:11]
            else:
                elm=("0000000000"+elm)[-10:]
        if self.arrHeadersTypedesc[headerName.split("#")[0]]=="date":
            if len(elm)==10:
                elm=elm[6:10]+elm[3:5]+elm[0:2]
        return elm
    
    def OnDataExtractionReorder(self,event):
        wx.BeginBusyCursor()
        select_fields=self.dataExtractionPanel.select_fields
        if self.dataExtractionPanel.optFieldOrder.GetSelection()==0:
            select_fields.Clear()
            self.lstDataExtFields=[]
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND id_section is not null and pos is not null and id_cycle<99 GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
            for row in cur:
                #if "#99" not in row[0]: #sporco sistematico
                select_fields.Append(row[0])
                self.lstDataExtFields.append(row[0])
        elif self.dataExtractionPanel.optFieldOrder.GetSelection()==1:
            select_fields.Clear()
            self.lstDataExtFields=[]
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND id_section is not null and pos is not null and id_cycle<99 GROUP BY description, id_cycle ORDER BY id_section,id_cycle,pos")
            for row in cur:
                #if "#99" not in row[0]: #sporco sistematico
                select_fields.Append(row[0])
                self.lstDataExtFields.append(row[0])
        elif self.dataExtractionPanel.optFieldOrder.GetSelection()==2:
            select_fields.Clear()
            self.lstDataExtFields=[]
            maxCycleBySection={}
            cc=1
            cur.execute("SELECT id_eform,description,label,cyclic,onactivate,height,href FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
            for row in cur:
                if row[3]>1: maxCycleBySection[cc]=1
                cc+=1
            #cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end, id_section, id_cycle FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND id_section is not null and pos is not null GROUP BY description, id_cycle ORDER BY lower(description),id_cycle")
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end, id_section, id_cycle FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND id_section is not null and pos is not null and id_cycle<99 GROUP BY description")
            for row in cur:
                self.lstDataExtFields.append(row[0])
                if row[1] in maxCycleBySection:
                    if maxCycleBySection[row[1]]<row[2]: maxCycleBySection[row[1]]=row[2]
                else:
                    maxCycleBySection[row[1]]=row[2]
            for curName in self.hashHeaders:
                cyclicVar=False
                if self.ActivateCyclicity:
                    testName=curName
                    if "#" in curName: testName=curName[:curName.find("#")]
                    if testName in self.arrHeadersIdSection:
                        if self.arrHeadersIdSection[curName] in maxCycleBySection:
                            #cycles=self.Cyclicity[self.arrHeadersIdSection[curName]]
                            cycles=maxCycleBySection[self.arrHeadersIdSection[testName]]
                            if cycles>=1:
                                cyclicVar=True
                                for i in range(cycles):
                                    newName=curName+"#"+str(i+1)
                                    if newName not in self.lstDataExtFields:
                                        self.lstDataExtFields.append(newName)
                if not cyclicVar:
                    if curName in self.arrHeadersIdSection:
                        if curName not in self.lstDataExtFields:
                            self.lstDataExtFields.append(curName)
            #self.lstDataExtFields.sort(key=lambda a: a.lower() if "#" not in a else a[:a.find("#")].lower()+a[a.find("#")+1:].rjust(3,"0"))
            self.lstDataExtFields.sort(key=lambda a: a.lower() if "#" not in a else a.split("#")[0].lower()+a.split("#")[1].rjust(3,"0"))
            for elm in self.lstDataExtFields:
                select_fields.Append(elm)
        elif self.dataExtractionPanel.optFieldOrder.GetSelection()==3:
            select_fields.Clear()
            self.lstDataExtFields=[]
            maxCycleBySection={}
            cc=1
            cur.execute("SELECT id_eform,description,label,cyclic,onactivate,height,href FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
            for row in cur:
                if row[3]>1: maxCycleBySection[cc]=1
                cc+=1
            #cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end, id_section, id_cycle FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND id_section is not null and pos is not null GROUP BY description, id_cycle ORDER BY id_section,id_cycle,pos")
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end, id_section, id_cycle FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND id_section is not null and pos is not null and id_cycle<99 GROUP BY description")
            for row in cur:
                self.lstDataExtFields.append(row[0])
                if "#" in row[0]:
                    if row[1] in maxCycleBySection:
                        if maxCycleBySection[row[1]]<row[2]: maxCycleBySection[row[1]]=row[2]
                    else:
                        maxCycleBySection[row[1]]=row[2]
            for curName in self.hashHeaders:
                cyclicVar=False
                if self.ActivateCyclicity:
                    testName=curName
                    if "#" in curName: testName=curName[:curName.find("#")]
                    if testName in self.arrHeadersIdSection:
                        if self.arrHeadersIdSection[curName] in maxCycleBySection:
                            #cycles=self.Cyclicity[self.arrHeadersIdSection[curName]]
                            cycles=maxCycleBySection[self.arrHeadersIdSection[testName]]
                            if cycles>=1:
                                cyclicVar=True
                                for i in range(cycles):
                                    newName=curName+"#"+str(i+1)
                                    if newName not in self.lstDataExtFields:
                                        self.lstDataExtFields.append(newName)
                if not cyclicVar:
                    if curName in self.arrHeadersIdSection:
                        if curName not in self.lstDataExtFields:
                            self.lstDataExtFields.append(curName)
            #self.lstDataExtFields.sort(key=lambda a: a.lower() if "#" not in a else a[:a.find("#")].lower()+a[a.find("#")+1:].rjust(3,"0"))
            self.lstDataExtFields.sort(key=lambda a: str(self.arrHeadersIdSection[a]).rjust(3,"0")+str(self.arrHeadersPos[a]).rjust(4,"0")+a.lower() if "#" not in a else str(self.arrHeadersIdSection[a.split("#")[0]]).rjust(3,"0")+a.split("#")[1].rjust(3,"0")+str(self.arrHeadersPos[a.split("#")[0]]).rjust(4,"0")+a.split("#")[0].lower())
            for elm in self.lstDataExtFields:
                select_fields.Append(elm)
        elif self.dataExtractionPanel.optFieldOrder.GetSelection()==4:
            select_fields.Clear()
            self.lstDataExtFields=[]
            maxCycleBySection={}
            cc=1
            cur.execute("SELECT id_eform,description,label,cyclic,onactivate,height,href FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
            for row in cur:
                maxCycleBySection[cc]=row[3]
                cc+=1
            for curName in self.hashHeaders:
                cyclicVar=False
                if self.ActivateCyclicity:
                    testName=curName
                    if "#" in curName: testName=curName[:curName.find("#")]
                    if testName in self.arrHeadersIdSection:
                        if self.arrHeadersIdSection[curName] in maxCycleBySection:
                            #cycles=self.Cyclicity[self.arrHeadersIdSection[curName]]
                            cycles=maxCycleBySection[self.arrHeadersIdSection[testName]]
                            if cycles>=1:
                                cyclicVar=True
                                for i in range(cycles):
                                    newName=curName+"#"+str(i+1)
                                    if newName not in self.lstDataExtFields:
                                        self.lstDataExtFields.append(newName)
                if not cyclicVar:
                    if curName in self.arrHeadersIdSection:
                        if curName not in self.lstDataExtFields:
                            self.lstDataExtFields.append(curName)
            #self.lstDataExtFields.sort(key=lambda a: a.lower() if "#" not in a else a[:a.find("#")].lower()+a[a.find("#")+1:].rjust(3,"0"))
            self.lstDataExtFields.sort(key=lambda a: a.lower() if "#" not in a else a.split("#")[0].lower()+a.split("#")[1].rjust(3,"0"))
            for elm in self.lstDataExtFields:
                select_fields.Append(elm)
        elif self.dataExtractionPanel.optFieldOrder.GetSelection()==5:
            select_fields.Clear()
            self.lstDataExtFields=[]
            maxCycleBySection={}
            cc=1
            cur.execute("SELECT id_eform,description,label,cyclic,onactivate,height,href FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
            for row in cur:
                maxCycleBySection[cc]=row[3]
                cc+=1
            for curName in self.hashHeaders:
                cyclicVar=False
                if self.ActivateCyclicity:
                    testName=curName
                    if "#" in curName: testName=curName[:curName.find("#")]
                    if testName in self.arrHeadersIdSection:
                        if self.arrHeadersIdSection[curName] in maxCycleBySection:
                            #cycles=self.Cyclicity[self.arrHeadersIdSection[curName]]
                            cycles=maxCycleBySection[self.arrHeadersIdSection[testName]]
                            if cycles>=1:
                                cyclicVar=True
                                for i in range(cycles):
                                    newName=curName+"#"+str(i+1)
                                    if newName not in self.lstDataExtFields:
                                        self.lstDataExtFields.append(newName)
                if not cyclicVar:
                    if curName in self.arrHeadersIdSection:
                        if curName not in self.lstDataExtFields:
                            self.lstDataExtFields.append(curName)
            #self.lstDataExtFields.sort(key=lambda a: a.lower() if "#" not in a else a[:a.find("#")].lower()+a[a.find("#")+1:].rjust(3,"0"))
            self.lstDataExtFields.sort(key=lambda a: str(self.arrHeadersIdSection[a]).rjust(3,"0")+str(self.arrHeadersPos[a]).rjust(4,"0")+a.lower() if "#" not in a else str(self.arrHeadersIdSection[a.split("#")[0]]).rjust(3,"0")+a.split("#")[1].rjust(3,"0")+str(self.arrHeadersPos[a.split("#")[0]]).rjust(4,"0")+a.split("#")[0].lower())
            for elm in self.lstDataExtFields:
                select_fields.Append(elm)
        wx.EndBusyCursor()

    def StartExporter(self,sql, wherecond, lstfields,keyfields,adminmode,reqstatus):
        filename=""
        from os.path import expanduser
        home = expanduser("~")
        filterindex=-1
        now = datetime.datetime.now()
        ts=now.strftime("%Y%m%d%H%M%S")
        reportname="query_"+ts
        wildcard = "Excel 2002 (*.xls)|*.xls|Excel 2007 (*.xlsx)|*.xlsx|Tab Separated (*.txt)|*.txt|2 files Tab Separated (*.2.txt)|*.2.txt|SPSS/PSPP data+syntax files (*.csv + *.sps)|*.sps"
        dlg = wx.FileDialog(self, message=TT("Choose a file name"), defaultDir=home, defaultFile=reportname, wildcard=wildcard, style=wx.SAVE | wx.FD_OVERWRITE_PROMPT)
        if dlg.ShowModal() == wx.ID_OK:
            filename=dlg.GetPath()
            filterindex=dlg.GetFilterIndex()
        dlg.Destroy()        
        if filename!="":
            exts=["xls","xlsx","txt","2.txt","sps"]
            testfilename=filename.lower()
            if ".xls" not in testfilename and ".xlsx" not in testfilename and ".txt" not in testfilename and ".sps" not in testfilename:
                filename=filename+"."+exts[filterindex]
            #filetype=filename[filename.find(".")+1:].lower()
            filetype=exts[filterindex]
            if filetype not in ("xls","xlsx","txt","2.txt","sps"):
                wx.MessageBox(TT("file type invalid or not specified"), TT("Exporter"), wx.ICON_EXCLAMATION | wx.OK, self)
                self.parent.syncroTaskPause=False
                return
            else:
                if filetype=="xls":
                    if len(lstfields)>256:
                        self.parent.syncroTaskPause=False
                        wx.MessageBox(TT("Too many columns for a '.xls' file")+" ("+str(len(lstfields))+") Max=252", TT("Exporter"), wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                    try: w=xlwt.Workbook()
                    except:
                        try:
                            w=Workbook()
                        except: 
                            wx.EndBusyCursor()
                            self.parent.syncroTaskPause=False
                            wx.MessageBox(TT("File format not available in this package"), TT("Exporter"), wx.ICON_EXCLAMATION | wx.OK, self)
                            return
                
                if filetype=="xlsx":
                    if len(lstfields)>1024:
                        self.parent.syncroTaskPause=False
                        wx.MessageBox(TT("Too many columns for a '.xlsx' file")+" ("+str(len(lstfields))+") Max=1020", TT("Exporter"), wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                    format_available=True
                    try:
                        from openpyxl.workbook import Workbook
                        from openpyxl.writer.excel import ExcelWriter
                        from openpyxl.cell import get_column_letter
                        format_available=True
                    except:
                        wx.EndBusyCursor()
                        self.parent.syncroTaskPause=False
                        wx.MessageBox(TT("File format not available in this package"), TT("Export error"), wx.ICON_EXCLAMATION | wx.OK, self)
                        return


                self.Exporter(filename, filetype, sql, wherecond, lstfields,keyfields,adminmode,reqstatus)
        
    def Exporter(self, reportname, exporttype, sql, wherecond, lstfields,keyfields,adminmode,reqstatus):
        wx.BeginBusyCursor()
        con_m = sqlite.connect(':memory:')
        cur_m = con_m.cursor()
        #square(con,con_m,"search",lstfields,self.id_group,self.id_crf,keyfields=keyfields,adminmode=adminmode,reqstatus=reqstatus)
            
        if DEBUG_MODE:
            print "No. of variables exported: "+str(len(lstfields))
            #print str(lstfields)

        if exporttype=="xls":
            try: w=xlwt.Workbook()
            except: w=Workbook()
            ws=w.add_sheet('query')
            ctfields=[]
            x=0
            for lstfield in lstfields:
                ctfields.append(lstfield)
                ws.write(0,x,lstfield)
                x=x+1
            y=1
            square(con,con_m,"search",ctfields,self.id_group,self.id_crf,keyfields=keyfields,adminmode=adminmode,reqstatus=reqstatus)
            qy="select "
            if "SELECT DISTINCT" in sql: qy+="distinct "
            qy+="* from search"
            cur_m.execute(qy)
            cols = [i[0] for i in cur_m.description]; hc={}; i=0
            #print str(lstfields)
            for col in cols: hc[(col.replace("___","-")).replace("__","#")]=i; i=i+1
            
            
            for row in cur_m:
                dData={}
                #for x in range(len(cols)):
                for x in range(len(lstfields)):
                    colname=(cols[x+4].replace("___","-")).replace("__","#")
                    #print colname
                    tmp=row[x+4]
                    if tmp!=None:
                        if not self.parent.trans.has_key(tmp):
                            dData[colname]=HB_DecryptOne(self.digestkey,tmp,"latin-1")
                            self.parent.trans[tmp]=dData[colname]
                        else:
                            dData[colname]=self.parent.trans[tmp]
                    else:
                        dData[colname]=""
                show=True
                if wherecond!="":
                    show=False
                    try:
                        show=eval(wherecond)
                    except Exception,e:
                        wx.EndBusyCursor()
                        self.parent.syncroTaskPause=False
                        wx.MessageBox("where-condition error:\n"+wherecond, "Query error", wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                if show:
                    for x in range(len(lstfields)):
                    #for x in range(len(cols)):
                        colname=(cols[x+4].replace("___","-")).replace("__","#")
                        ws.write(y,x,dData[colname])
                    y=y+1
            w.save(reportname)

        elif exporttype=="xlsx":
            from openpyxl.workbook import Workbook
            from openpyxl.writer.excel import ExcelWriter
            from openpyxl.cell import get_column_letter
            wb = Workbook(optimized_write = True, encoding='utf-8')
            #ws = wb.worksheets[0]
            ws = wb.create_sheet()
            ws.title = "query"
            ws.append(['%s' % lstfield for lstfield in lstfields])
            ctfields=[]
            for lstfield in lstfields: ctfields.append(lstfield)
            y=2

            square(con,con_m,"search",ctfields,self.id_group,self.id_crf,keyfields=keyfields,adminmode=adminmode,reqstatus=reqstatus)
            qy="select "
            if "SELECT DISTINCT" in sql: qy+="distinct "
            qy+="* from search"
            cur_m.execute(qy)
            cols = [i[0] for i in cur_m.description]; hc={}; i=0
            for col in cols: hc[(col.replace("___","-")).replace("__","#")]=i; i=i+1
            for row in cur_m:
                dData={}
                #for x in range(len(cols)):
                for x in range(len(lstfields)):
                    colname=(cols[x+4].replace("___","-")).replace("__","#")
                    tmp=row[x+4]
                    if tmp!=None:
                        if not self.parent.trans.has_key(tmp):
                            dData[colname]=HB_DecryptOne(self.digestkey,tmp,"latin-1")
                            self.parent.trans[tmp]=dData[colname]
                        else:
                            dData[colname]=self.parent.trans[tmp]
                    else:
                        dData[colname]=""
                show=True
                if wherecond!="":
                    show=False
                    try:
                        show=eval(wherecond)
                    except Exception,e:
                        wx.EndBusyCursor()
                        self.parent.syncroTaskPause=False
                        wx.MessageBox("where-condition error:\n"+wherecond, "Query error", wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                if show:
                    #ws.append(['%s' % dData[cols[x]].encode('utf-8', 'replace') for x in range(len(cols))])
                    #ws.append(['%s' % dData[(cols[x].replace("___","-")).replace("__","#")] for x in range(len(cols))])
                    ws.append(['%s' % dData[(cols[x+4].replace("___","-")).replace("__","#")] for x in range(len(lstfields))])
            wb.save(filename = reportname)
                
        elif exporttype in ("txt","2.txt","sps"):
            arr1row=[]
            arrrows=[]
            arrcols=[]
            if exporttype in ("txt"):
                for lstfield in lstfields:
                    arrcols.append('"'+lstfield+'"')
                arrrows.append('\t'.join(arrcols))
            elif exporttype in ("2.txt"):
                for lstfield in lstfields:
                    arrcols.append(lstfield.replace("#","__"))
                arr1row.append('\t'.join(arrcols))
            elif exporttype in ("sps"):
                for lstfield in lstfields:
                    arrcols.append(lstfield.replace("#","__"))
                arr1row.append('GET DATA')
                arr1row.append('  /TYPE=TXT')
                arr1row.append('  /FILE="'+reportname.replace(".sps",".csv")+'"')
                arr1row.append('  /IMPORTCASES=ALL')
                arr1row.append('  /ARRANGEMENT=DELIMITED')
                arr1row.append('  /DELCASE=LINE')
                arr1row.append('  /FIRSTCASE=1')
                arr1row.append('  /DELIMITERS=","')
                arr1row.append('  /QUALIFIER=\'"\'')
                arr1row.append('  /ESCAPE')
                arr1row.append('  /VARIABLES=')
                for i in range(len(arrcols)):
                    col=arrcols[i]
                    base_col_name=lstfields[i]
                    if base_col_name.find("#")>=0: base_col_name=base_col_name[:base_col_name.find("#")]
                    typedef=self.arrHeadersTypedesc[base_col_name]
                    spsstypedef="A255"
                    if typedef==None:
                        spsstypedef="A255"
                    elif typedef.split("=")[0]=="mask":
                        mask=typedef.split("=")[1]
                        isnumber=True
                        notnumbers=["/",":","-","N","A","a","C","X","&","*"]
                        for elm in notnumbers:
                            if elm in mask:
                                isnumber=False
                                break
                        if isnumber:
                            if "." in mask: spsstypedef="F10.3"
                            else: spsstypedef="F10.0"
                        else:
                            spsstypedef="A255"
                    elif typedef=="integer":
                        spsstypedef="F10.0"
                    elif typedef=="float":
                        spsstypedef="F10.3"
                    elif typedef=="number":
                        mask=typedef[typedef.find("=")+1:]
                        allowNegative=True
                        intpart=0
                        tmp=mask[:mask.find(".")]
                        if len(tmp)>0:
                            if tmp[0]=="+": 
                                allowNegative=False
                                intpart=int(tmp[1:])
                            elif tmp[0]=="-": 
                                intpart=int(tmp[1:])
                            else:
                                intpart=int(tmp)
                        decpart=int(mask[mask.find(".")+1:])    
                        spsstypedef="F"+str(intpart)+"."+str(decpart)
                    elif typedef=="date":
                        spsstypedef="EDATE10"
                    arr1row.append("    "+col+" "+spsstypedef)
                arr1row.append('.')
                
            maxfields=1996
            #maxfields=3
            ctfields=lstfields
            dData_pool={}
            recordkey_pool=[]
            recordset_ok=True
            while len(ctfields)>0 and recordset_ok:
                chunkfields=ctfields[:min(len(ctfields),maxfields)]
                square(con,con_m,"search",chunkfields,self.id_group,self.id_crf,keyfields=keyfields,adminmode=adminmode,reqstatus=reqstatus)
                try:
                    qy="select "
                    if "SELECT DISTINCT" in sql: qy+="distinct "
                    qy+="* from search"
                    cur_m.execute(qy)
                    cols = [i[0] for i in cur_m.description]; hc={}; i=0
                    for col in cols: hc[(col.replace("___","-")).replace("__","#")]=i; i=i+1
                    ctfields=ctfields[min(len(ctfields),maxfields):]
                except Exception,e:
                    recordset_ok=False
                    wx.MessageBox(TT("field selection error")+":\n"+str(chunkfields), TT("Query error"), wx.ICON_EXCLAMATION | wx.OK, self)
                if recordset_ok:
                    ct=0
                    for row in cur_m:
                        dData={}
                        for x in range(len(chunkfields)):
                            colname=(cols[x+4].replace("___","-")).replace("__","#")
                            tmp=row[x+4]
                            if tmp!=None:
                                if not self.parent.trans.has_key(tmp):
                                    dData[colname]=HB_DecryptOne(self.digestkey,tmp,"latin-1")
                                    self.parent.trans[tmp]=dData[colname]
                                else:
                                    dData[colname]=self.parent.trans[tmp]
                            else:
                                dData[colname]=""
                        recordkey=str(row[0])+"-"+str(row[1])+"-"+str(row[2])+"-"+str(row[3])
                        recordkey_pool.append(recordkey)
                        if recordkey not in dData_pool: dData_pool[recordkey]={}
                        for key in dData:
                            dData_pool[recordkey][key]=dData[key]
                        ct=ct+1
            #for row in cur_pool[0]:
            for cur in range(ct):
                #dData={}
                dData={}
                for key in dData_pool[recordkey_pool[cur]]:
                    dData[key]=dData_pool[recordkey_pool[cur]][key]
                #for x in range(len(lstfields)):
                    #colname=(cols[x+4].replace("___","-")).replace("__","#")
                    #tmp=row[x+4]
                    #if tmp!=None:
                        #if not self.parent.trans.has_key(tmp):
                            #dData[colname]=HB_DecryptOne(self.digestkey,tmp,"latin-1")
                            #self.parent.trans[tmp]=dData[colname]
                        #else:
                            #dData[colname]=self.parent.trans[tmp]
                    #else:
                        #dData[colname]=""
                show=True
                if wherecond!="":
                    show=False
                    try:
                        show=eval(wherecond)
                    except Exception,e:
                        wx.EndBusyCursor()
                        self.parent.syncroTaskPause=False
                        wx.MessageBox("where-condition error:\n"+wherecond, "Query error", wx.ICON_EXCLAMATION | wx.OK, self)
                        return
                if show:
                    arrcols=[]
                    #for x in range(len(cols)):
                    for x in range(len(lstfields)):
                        #colname=(cols[x+4].replace("___","-")).replace("__","#")
                        colname=(lstfields[x].replace("___","-")).replace("__","#")
                        if exporttype in ("txt","sps"):
                            data=dData[colname].replace('"','""')
                            data=data.replace('\r\n','<CR>')
                            data=data.replace('\n','<CR>')
                            data=data.replace('\t','<TAB>')
                            arrcols.append('"'+data+'"')
                        else:
                            data=dData[colname]
                            data=data.replace('\r\n','<CR>')
                            data=data.replace('\n','<CR>')
                            data=data.replace('\t','<TAB>')
                            arrcols.append(data)
                    if exporttype in ("txt","2.txt"):
                        arrrows.append('\t'.join(arrcols))
                    elif exporttype in ("sps"):
                        arrrows.append(','.join(arrcols))

            filename1=""
            filename2=""
            if exporttype in ("txt"):
                theFile = open(reportname, 'w')
                contents='\n'.join(arrrows)
                #print contents
                theFile.write(contents.encode('ascii', 'replace'))
                theFile.close
            elif exporttype in ("2.txt"):
                filename1=reportname.replace(".2.txt",".headers.2.txt")
                theFile = open(filename1, 'w')
                contents='\n'.join(arr1row)
                theFile.write(contents.encode('ascii', 'replace'))
                theFile.close
                filename2=reportname.replace(".2.txt",".contents.2.txt")
                theFile = open(filename2, 'w')
                contents='\n'.join(arrrows)
                theFile.write(contents.encode('ascii', 'replace'))
                theFile.close
            elif exporttype in ("sps"):
                filename1=reportname.replace(".sps",".csv")
                theFile = open(filename1, 'w')
                contents='\n'.join(arrrows)
                theFile.write(contents.encode('ascii', 'replace'))
                theFile.close
                theFile = open(reportname, 'w')
                contents='\n'.join(arr1row)
                theFile.write(contents.encode('ascii', 'replace'))
                theFile.close
                    
        cur_m.close()
        con_m.close()
        wx.EndBusyCursor()
        self.parent.syncroTaskPause=False
        
        if exporttype!="":
            if exporttype!="2.txt":
                if not assoc_open(reportname):
                    wx.MessageBox("File\n'"+reportname+"'\ncreated.", TT("Exporter"), wx.ICON_INFORMATION | wx.OK, self)
            else:
                wx.MessageBox("File\n'"+filename1+"'\n'"+filename2+"'\ncreated.", TT("Exporter"), wx.ICON_INFORMATION | wx.OK, self)

    def OnDataExtractionRunQuery(self,event):
        wx.BeginBusyCursor()
        self.parent.syncroTaskPause=True
        import time
        while self.parent.syncroTaskUpdating:
            time.sleep(1)
            if DEBUG_MODE: print "waiting synchro standby..."
        lstfields=[]
        strlstfields=""
        select_fields=self.dataExtractionPanel.select_fields
        if sys.platform == 'darwin':
            for i in range(len(select_fields.GetSelections())):
                fldName = self.lstDataExtFields[select_fields.GetSelections()[i]]
                lstfields.append(fldName)
                if fldName.find("#")>=0: fldName = fldName[:fldName.find("#")] + "__" + fldName[fldName.find("#")+1:]
                if strlstfields!="": strlstfields=strlstfields+","
                strlstfields=strlstfields+fldName
        else:
            for i in range(len(select_fields.GetChecked())):
                fldName = self.lstDataExtFields[select_fields.GetChecked()[i]]
                lstfields.append(fldName)
                if fldName.find("#")>=0: fldName = fldName[:fldName.find("#")] + "__" + fldName[fldName.find("#")+1:]
                if strlstfields!="": strlstfields=strlstfields+","
                strlstfields=strlstfields+fldName
                
        if len(lstfields)==0:
            wx.MessageBox(TT("no fields selected."), TT("Exporter"), wx.ICON_EXCLAMATION | wx.OK, self)
            self.parent.syncroTaskPause=False
            wx.EndBusyCursor()
            return
        #if len(lstfields)>2000:
            #wx.MessageBox(TT("too many fields selected (max. 2000)."), TT("Exporter"), wx.ICON_EXCLAMATION | wx.OK, self)
            #self.parent.syncroTaskPause=False
            #wx.EndBusyCursor()
            #return
                
        Panel=self.dataExtractionPanel
        wherecond=""
        for i in range(1,8):
            #field1=eval("Panel.filter_field"+`i`+".GetValue()").replace("\\","\\\\").replace("\"","\\\"")
            field1=eval("Panel.filter_field"+`i`+".GetValue()").replace("\\","\\\\")
            andor_field=""
            if i>1: andor_field=eval("Panel.filter_andor"+`i-1`+".GetStringSelection()")
            if (i==1 and field1!="") or (i>1 and field1!="" and andor_field!=""):
                cond=eval("Panel.filter_cond"+`i`+".GetStringSelection()")
                if cond=="=": cond="=="
                if cond=="<>": cond="!="
                listmode=True
                if cond in ("==","!=","<",">","<=",">=","contains"): listmode=False
                if listmode:
                    elm1 = "dData[\"" + field1 + "\"]"
                else:
                    elm1 = "self.GetOrderedElm(\""+field1+"\",dData[\"" + field1 + "\"])" 
                #field2=eval("Panel.filter_value"+`i`+".GetValue()").replace("\\","\\\\").replace("\"","\\\"")
                field2=eval("Panel.filter_value"+`i`+".GetValue()").replace("\\","\\\\")
                elm2type=eval("Panel.filter_type"+`i`+".GetStringSelection()")
                elm2=""
                if elm2type=="field":
                    if listmode:
                        elm2 = "dData[\"" + field2 + "\"]"
                    else:
                        elm2 = "self.GetOrderedElm(\""+field2+"\",dData[\"" + field2 + "\"])"
                if elm2type=="value":
                    if listmode:
                        elm2 = field2
                        if elm2!=None:
                            if len(elm2)>1:
                                if elm2[0]!="(": elm2="("+elm2
                                if elm2[len(elm2)-1]!=")": elm2=elm2+")"
                                if elm2.find("\"")<0:
                                    elm2=elm2.replace("(","(\"")
                                    elm2=elm2.replace(")","\")")
                                    elm2=elm2.replace(",","\",\"")
                    else:
                        if cond=="contains":
                            elm2="\""+field2+"\""
                        else:
                            elm2 = "self.GetOrderedElm(\""+field1+"\",\""+field2+"\")"
                if i>1: wherecond = wherecond + " " + andor_field

                if cond!="contains":
                    wherecond = wherecond + " " + eval("Panel.filter_pre"+`i`+".GetStringSelection()") + elm1 + " " + cond + " " + elm2 + eval("Panel.filter_post"+`i`+".GetStringSelection()")
                else:
                    wherecond = wherecond + " " + eval("Panel.filter_pre"+`i`+".GetStringSelection()") + elm2 + " in " + elm1 + eval("Panel.filter_post"+`i`+".GetStringSelection()")

        wherecond = wherecond + " " + Panel.filter_freetext.GetValue()
        wherecond=wherecond.strip()
        
        #print wherecond
        #con_m = sqlite.connect(":memory:")
        #cur_m = con_m.cursor()
        
        #Campi chiave (quelli della pagina di ricerca)
        list=self.PanelSearch.GetChildren()
        keyfields=""
        for i in range(len(list)):
            if self.IsInputField(list[i]):
                if list[i].GetName()!="heavybase_memo1":
                    if keyfields!="": keyfields=keyfields+","
                    keyfields=keyfields+list[i].GetName()
        
        #Comandi amministratore
        candoit=False
        permission=GetSetting(self.parent,"sysadmin_groups")
        if permission!="":
            s1=","+permission+","
            s2=","+str(self.id_group)+","
            if s2 in s1: candoit=True
        else:
            id_group_admin=GetSetting(self.parent,"id_group_admin")
            if str(self.id_group)==str(id_group_admin):
                candoit=True
                        
        reqstatus=0
        adminmode=False
        if candoit: 
            crfTypeIdx=self.crfTypeChoice.GetSelection()
            if crfTypeIdx==1: adminmode=True
            elif crfTypeIdx==2: 
                reqstatus=-1
                adminmode=True
            elif crfTypeIdx==3: 
                reqstatus=1
                adminmode=True
        #print "adminmode="+str(adminmode)+"\treqstatus="+str(reqstatus)
        #square(con,con_m,"search",lstfields,self.id_group,self.id_crf,keyfields=keyfields,adminmode=adminmode,reqstatus=reqstatus)
        
        addclause=""
        if self.dataExtractionPanel.chkDistinct.IsChecked(): addclause="DISTINCT "
        sql="SELECT "+addclause+strlstfields.replace("-","___")+" FROM search"
        
        wx.EndBusyCursor()
        #self.StartExporter(sql, wherecond, lstfields, cur_m)
        self.StartExporter(sql, wherecond, lstfields,keyfields,adminmode,reqstatus)
    
    def PreSaveRecord(self):
        if self.ActivateCyclicity:
            if self.Cyclicity[self.ifrm+1]>0: self.SaveCycle(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1)

        #self.buttonPrint.Enable(False)
        for ifrm in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
            dosave=True
            if self.ActivateCyclicity:
                if self.Cyclicity[ifrm+1]>0: dosave=False
            if dosave: self.SaveFieldList(list,0)

    def CustomPrint(self,templatename,strip):
        self.PreSaveRecord()
        try: infile = open(REPORTS_PATH+templatename)
        except: infile = open(".."+os.path.sep+templatename)
        text=infile.read()
        infile.close()
        templatetype=templatename[-3:].lower()
        strDirtyTags=GetSetting(self.parent,"dirty_tags")
        DIRTY_TAGS=False
        if strDirtyTags=="1": DIRTY_TAGS=True
        binary_type_printout=GetSetting(self.parent,"binary_type_printout")
        if binary_type_printout=="": binary_type_printout="X"
        contents={}
        for key in self.contents:
            basevarname=key
            if "#" in basevarname: basevarname=basevarname[:basevarname.find("#")]
            if self.arrHeadersTypedesc[basevarname]!="binary":
                contents[key]=self.contents[key]
            else:
                contents[key]=binary_type_printout
        #scelta filename - inizio
        now = datetime.datetime.now()
        ts=now.strftime("%Y%m%d%H%M%S")
        reportname="report_"+ts+"."+templatetype
        wildchar=""
        if templatetype=="rtf":
            wildcard = "MS Word files (*.rtf)|*.rtf"
        elif templatetype=="abw":
            wildcard = "Abiword files (*.abw)|*.abw"
        dlg = wx.FileDialog(self, message=TT("Choose a file name"), defaultDir=REPORTS_PATH, defaultFile=reportname, wildcard=wildcard, style=wx.SAVE | wx.FD_OVERWRITE_PROMPT)
        filename=""
        if dlg.ShowModal() == wx.ID_OK:
            filename=dlg.GetPath()
            filterindex=dlg.GetFilterIndex()
        dlg.Destroy()        
        if filename!="":
            fs_error=True
            try:
                outfile = open(filename, 'w')
                outfile.write("test")
                outfile.close()
                os.unlink(filename)
                fs_error=False
            except:
                wx.MessageBox(TT("File not writable"), TT("Print eCRF"), wx.ICON_ERROR | wx.OK, self)
            if not fs_error:
                #scelta filename - fine
                job = job_server_local.submit(HB_PrintCRF,(text,contents,REPORTS_PATH,filename,strip,templatetype,DIRTY_TAGS,DEBUG_MODE), (GetRtfValue,MultiReplace,create_permutations), ("string","re"), callback=HB_CB_PrintCRF)
                #HB_PrintCRF(text,self.contents,REPORTS_PATH,strip,templatetype,DEBUG_MODE)
                #HB_CB_PrintCRF(templatetype)
                if GetSetting(self.parent,"big_print_templates")=="1":
                    wx.MessageBox(TT("Print preview preparation in progress"), TT("Print eCRF"), wx.ICON_INFORMATION | wx.OK, self)
        else:
            wx.MessageBox(TT("Print canceled"), TT("Print eCRF"), wx.ICON_EXCLAMATION | wx.OK, self)

    def CustomPrint2(self,fullpathtemplatename,contents,strip):
        infile = open(fullpathtemplatename)
        text=infile.read()
        infile.close()
        templatetype=fullpathtemplatename[-3:].lower()
        strDirtyTags=GetSetting(self.parent,"dirty_tags")
        DIRTY_TAGS=False
        if strDirtyTags=="1": DIRTY_TAGS=True
        #scelta filename - inizio
        now = datetime.datetime.now()
        ts=now.strftime("%Y%m%d%H%M%S")
        reportname="report_"+ts+"."+templatetype
        wildchar=""
        if templatetype=="rtf":
            wildcard = "MS Word files (*.rtf)|*.rtf"
        elif templatetype=="abw":
            wildcard = "Abiword files (*.abw)|*.abw"
        dlg = wx.FileDialog(self, message=TT("Choose a file name"), defaultDir=REPORTS_PATH, defaultFile=reportname, wildcard=wildcard, style=wx.SAVE | wx.FD_OVERWRITE_PROMPT)
        filename=""
        if dlg.ShowModal() == wx.ID_OK:
            filename=dlg.GetPath()
            filterindex=dlg.GetFilterIndex()
        dlg.Destroy()        
        if filename!="":
            #scelta filename - fine
            job = job_server_local.submit(HB_PrintCRF,(text,contents,REPORTS_PATH,filename,strip,templatetype,DIRTY_TAGS,DEBUG_MODE), (GetRtfValue,MultiReplace,create_permutations), ("string","re"), callback=HB_CB_PrintCRF)
        else:
            wx.MessageBox(TT("Print canceled"), TT("Print eCRF"), wx.ICON_INFORMATION | wx.OK, self)
        
    def PrepareCustomPrint(self,templatetype):
        templatename=""
        stripcycle=False
        if templatetype=="all":
            if os.path.isfile(REPORTS_PATH+"template.abw"): 
                templatename="template.abw"
            else: 
                templatename="template.rtf"
        else:
            if os.path.isfile(REPORTS_PATH+"template_"+self.FormName[self.ifrm]+".abw"):
                templatename="template_"+self.FormName[self.ifrm]+".abw"
            else:
                templatename="template_"+self.FormName[self.ifrm]+".rtf"
            stripcycle=True
        strip=""
        if stripcycle:
            objCycles=wx.xrc.XRCCTRL(self.frameRecordset,'cycles')
            idx = objCycles.GetSelection()+1
            strip="__"+str(idx)
        self.CustomPrint(templatename,strip)
        
    def StartPrintTemplateChooser(self):
        dia=wx.Dialog(None, -1, TT("Choose Print template"))
        hbox = wx.BoxSizer()
        sizer = wx.FlexGridSizer(1, 3, 10, 10)

        btn1 = wx.Button(dia, label=TT("All pages"), size=wx.Size(100,30))
        btn2 = wx.Button(dia, label=TT("Current page"), size=wx.Size(230,30))
        btn3 = wx.Button(dia, label=TT("Cancel"), size=wx.Size(80,30))

        sizer.AddMany([btn1, btn2, btn3])

        hbox.Add(sizer, 0, wx.ALL, 15)
        dia.SetSizer(hbox)
        dia.SetSize((460, 80))
        
        btn1.Bind(wx.EVT_BUTTON, lambda event: self.PrepareCustomPrint(event, "all"))
        btn2.Bind(wx.EVT_BUTTON, lambda event: self.PrepareCustomPrint(event, "page"))
        btn3.Bind(wx.EVT_BUTTON, lambda event: dia.Close(True))

        dia.ShowModal()
        dia.Destroy()
        
    def OnButtonPrintButton(self,event):
        self.buttonPrint.Enable(False)
        self.PreSaveRecord()
            
        bothavailable=False
        stripcycle=False
        templatename=""
        if os.path.isfile(REPORTS_PATH+"template.abw") or os.path.isfile(REPORTS_PATH+"template.rtf"):
            if os.path.isfile(REPORTS_PATH+"template_"+self.FormName[self.ifrm]+".abw") or os.path.isfile(REPORTS_PATH+"template_"+self.FormName[self.ifrm]+".rtf"):
                bothavailable=True
            else: 
                if os.path.isfile(REPORTS_PATH+"template.abw"): 
                    templatename="template.abw"
                else: 
                    templatename="template.rtf"
        else:
            if os.path.isfile(REPORTS_PATH+"template_"+self.FormName[self.ifrm]+".abw"):
                templatename="template_"+self.FormName[self.ifrm]+".abw"
            elif os.path.isfile(REPORTS_PATH+"template_"+self.FormName[self.ifrm]+".rtf"):
                templatename="template_"+self.FormName[self.ifrm]+".rtf"
            if templatename!="":
                if GetSetting(self.parent,'print_template_stripcycle')!="0":
                    stripcycle=True
        if bothavailable:
            self.StartPrintTemplateChooser()
        elif templatename!="":
            strip=""
            if stripcycle:
                objCycles=wx.xrc.XRCCTRL(self.frameRecordset,'cycles')
                idx = objCycles.GetSelection()+1
                strip="__"+str(idx)
            self.CustomPrint(templatename,strip)
        else:
            wx.MessageBox(TT("No Print template available"), TT("Print eCRF"), wx.ICON_ERROR | wx.OK, self)
        self.buttonPrint.Enable(True)
        
    def OnWindowResize(self, event):
        clientSize=self.GetClientSize()
        SCREEN_X=clientSize.GetWidth()
        SCREEN_Y=clientSize.GetHeight()

        self.frameLogin.SetSize(clientSize)
        self.lblDisclaimer.SetPosition(wx.Point(5,SCREEN_Y-40))

        self.frameSearch.SetSize(clientSize)
        self.lstFound.SetSize(wx.Size(clientSize.GetWidth(),clientSize.GetHeight()-350))
        self.buttonExportSearch.SetPosition(wx.Point(self.lstFound.GetSize().GetWidth()/2-70,self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()))
        self.buttonAddnew.SetPosition(wx.Point(10,self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()+10))
        self.buttonAutoMergeDupes.SetPosition(wx.Point(self.lstFound.GetSize().GetWidth()-390,self.lstFound.GetPosition().y-25))
        self.crfTypeChoice.SetPosition(wx.Point(self.lstFound.GetSize().GetWidth()-220,self.lstFound.GetPosition().y-25))
        self.buttonShowDiff.SetPosition(wx.Point(self.lstFound.GetSize().GetWidth()-50,self.lstFound.GetPosition().y-25))
        self.buttonExport.SetPosition(wx.Point(SCREEN_X-250,self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()+10))
        self.txtQuickReport.SetPosition(wx.Point(0, self.lstFound.GetPosition().y+self.lstFound.GetSize().GetHeight()+50))
        self.txtQuickReport.SetSize(wx.Size(clientSize.GetWidth(), 80))
        self.lblReport.SetPosition(wx.Point(10,SCREEN_Y-40))
        self.reportChoice.SetPosition(wx.Point(80,SCREEN_Y-40))
        #self.buttonReportChoice.SetPosition(wx.Point(580,SCREEN_Y-50))
        self.buttonLogout.SetPosition(wx.Point(SCREEN_X-85,SCREEN_Y-30))

        self.frameRecordset.SetSize(clientSize)
        #wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').SetSize(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSize())
        #wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').EnableScrolling(False,True)
        
        #self.frameCrf.SetSize(wx.Size(SCREEN_X, SCREEN_Y-80))
        #self.frameCrf.SetVirtualSize(wx.Size(SCREEN_X-80,self.frameCrf.GetVirtualSize().GetHeight()))
        self.frameCrf.SetSize(wx.Size(SCREEN_X, SCREEN_Y-100))
        self.frameCrf.SetVirtualSize(wx.Size(SCREEN_X-100,self.frameCrf.GetVirtualSize().GetHeight()))

        self.buttonClone.SetPosition(wx.Point(SCREEN_X-450,10))
        self.buttonSave.SetPosition(wx.Point(SCREEN_X-300,10))
        self.buttonCommit.SetPosition(wx.Point(SCREEN_X-150,10))
        self.buttonCancel.SetPosition(wx.Point(SCREEN_X-450,50))
        self.buttonDelete.SetPosition(wx.Point(SCREEN_X-300,50))
        self.buttonPrint.SetPosition(wx.Point(SCREEN_X-150,50))

        for i in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            panelName=""
            try: panelName=self.FormName[i]
            except: panelName='panel'+`i+1`
            panel=wx.xrc.XRCCTRL(self.frameCrf,panelName)
            if panel!=None:
                try: panel.SetSize(wx.Size(SCREEN_X,panel.GetSize().GetHeight()))
                except: pass

        self.frameDataExtraction.SetSize(clientSize)
        self.dataExtractionPanel.SetSize(clientSize)
        xspace=SCREEN_X-870
        yspace=SCREEN_Y-450
        ysize=self.dataExtractionPanel.select_fields.GetSize().GetHeight()-(self.dataExtractionPanel.filter_freetext.GetPosition().y-self.dataExtractionPanel.select_fields.GetPosition().y)
        xsize=self.dataExtractionPanel.filter_freetext.GetSize().GetWidth()
        self.dataExtractionPanel.filter_freetext.SetSize(wx.Size(xsize,ysize))

    def OnLoginEnterPressed(self, event):
        
        if event.GetKeyCode() in (wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER):
            self.OnButtonLoginButton(event)
        else:
            keycode = event.GetKeyCode()
            controlDown = event.CmdDown()
            altDown = event.AltDown()
            shiftDown = event.ShiftDown()
            if (keycode==wx.WXK_TAB) and (shiftDown and controlDown):
                shortcut_login=GetSetting(self.parent,'shortcut_login')
                if shortcut_login!="":
                    cur.execute("select username,password from users where id_user="+shortcut_login)
                    row = cur.fetchone()
                    if row!=None:
                        self.cryptokey=self.txtCryptokey.GetValue()
                        if self.cryptokey=="": return
                        self.digestkey=sha256(self.cryptokey).digest()
                        dUsername=HB_DecryptOne(self.digestkey,row[0],"latin-1")
                        dPassword=HB_DecryptOne(self.digestkey,row[1],"latin-1")
                        self.txtUsername.SetValue(dUsername)
                        self.txtPassword.SetValue(dPassword)
                        self.OnButtonLoginButton(event)
        
    def BuildFieldsList(self,id_eform,list,ct):
        fieldsList=[]
        for ielm in range(len(list)):
            if list[ielm].GetClassName() in ("wxPanel"):
                sublist=list[ielm].GetChildren()
                self.BuildFieldsList(id_eform,sublist,ct)
            else:
                if self.IsInputField(list[ielm]) or self.IsButton(list[ielm]):
                    ct=ct+1
                    label=""
                    if list[ielm].GetClassName() in ("wxTextCtrl"):
                        try: label=self.arrHeadersTypedesc[list[ielm].GetName()]
                        except: pass
                    else:
                        try: label=list[ielm].GetLabel()
                        except: pass
                    maxLength=0
                    try: maxLength=list[ielm].maxLength
                    except: pass
                    if list[ielm].GetClassName() in ("wxChoice", "wxComboBox"):
                        opts=[]
                        for elm in list[ielm].GetItems():
                            opts.append(elm)
                        label='"'+'","'.join(opts)+'"'
                    id_header=-1
                    try: id_header=self.hashHeaders[list[ielm].GetName()]
                    except: pass
                    font=list[ielm].GetFont()
                    strfont=str(font.GetPointSize())
                    weights={90: 'wxNORMAL', 91: 'wxLIGHT', 92: 'wxBOLD'}
                    strfont=strfont+","+weights[font.GetWeight()]
                    #cur.execute("INSERT INTO efields (id_efield, id_eform, class, style, label, id_header, name, pos_x, pos_y, size_x, size_y) VALUES ("+str(ct)+","+str(id_eform)+",'"+list[ielm].GetClassName()+"','"+self.GetSqlValue(str(list[ielm].GetFont()))+"','"+self.GetSqlValue(label)+"',"+str(id_header)+",'"+list[ielm].GetName()+"',"+str(list[ielm].GetPosition().x)+","+str(list[ielm].GetPosition().y)+","+str(list[ielm].GetSize().x)+","+str(list[ielm].GetSize().y)+")")
                    fieldsList.append("SELECT "+str(id_eform)+",'"+list[ielm].GetClassName()+"','"+self.GetSqlValue(strfont)+"','"+self.GetSqlValue(label)+"',"+str(id_header)+",'"+list[ielm].GetName()+"',"+str(list[ielm].GetPosition().x)+","+str(list[ielm].GetPosition().y)+","+str(list[ielm].GetSize().x)+","+str(list[ielm].GetSize().y)+","+str(maxLength))
                    if len(fieldsList)>=500:
                        qy="INSERT INTO efields (id_eform, classname, style, label, id_header, name, pos_x, pos_y, size_x, size_y, maxlength) "+' UNION '.join(fieldsList)
                        try: cur.execute(qy)
                        except: print "failed: "+qy
                        fieldsList=[]
                elif list[ielm].GetClassName() in ("wxStaticText"):
                    ct=ct+1
                    label=""
                    try: label=list[ielm].GetLabel()
                    except: pass
                    font=list[ielm].GetFont()
                    strfont=str(font.GetPointSize())
                    weights={90: 'wxNORMAL', 91: 'wxLIGHT', 92: 'wxBOLD'}
                    strfont=strfont+","+weights[font.GetWeight()]
                    fieldsList.append("SELECT "+str(id_eform)+",'"+list[ielm].GetClassName()+"','"+self.GetSqlValue(strfont)+"','"+self.GetSqlValue(label)+"',0,'',"+str(list[ielm].GetPosition().x)+","+str(list[ielm].GetPosition().y)+","+str(list[ielm].GetSize().x)+","+str(list[ielm].GetSize().y)+",-1")
                    if len(fieldsList)>=500:
                        qy="INSERT INTO efields (id_eform, classname, style, label, id_header, name, pos_x, pos_y, size_x, size_y, maxlength) "+' UNION '.join(fieldsList)
                        try: cur.execute(qy)
                        except: print "failed: "+qy
                        fieldsList=[]
        if len(fieldsList)>0:
            #cur.execute("INSERT INTO efields (id_efield, id_eform, classname, style, label, id_header, name, pos_x, pos_y, size_x, size_y) "+' UNION '.join(fieldsList))
            qy="INSERT INTO efields (id_eform, classname, style, label, id_header, name, pos_x, pos_y, size_x, size_y, maxlength) "+' UNION '.join(fieldsList)
            try:
                cur.execute(qy)
            except:
                print "failed: "+qy
        return ct

    def OnButtonLoginButton(self, event):
        #event.Skip()
        #db encryption - Begin
        cur.execute("SELECT setting_value FROM settings WHERE setting_key='do_encrypt'")
        row = cur.fetchone()
        if row!=None:
            wx.MessageBox("Il sistema sta per procedere alla crittografia del database.\nL'operazione puo' richiedere da alcuni secondi ad alcuni minuti.\nSi prega di non interrompere il processo e attendere il messaggio successivo.", "Update", wx.ICON_INFORMATION | wx.OK, None)
            wx.BeginBusyCursor()
            self.cryptokey=row[0]
            self.digestkey=sha256(self.cryptokey).digest()
            cur2 = con.cursor()
            cur.execute("SELECT id_user,username,password FROM users")
            for row in cur:
                cUsername=base64.b64encode(rijndael.EncryptData(self.digestkey,row[1]))
                cPassword=base64.b64encode(rijndael.EncryptData(self.digestkey,row[2]))
                cur2.execute("UPDATE users SET username='"+cUsername+"', password='"+cPassword+"' WHERE id_user="+str(row[0]))
            cur.execute('SELECT id_header,id_dictionary,data FROM contents_dictionary')
            for row in cur:
                cData=base64.b64encode(rijndael.EncryptData(self.digestkey,row[2]))
                cur2.execute("UPDATE contents_dictionary SET data='"+self.GetSqlValue(cData)+"' WHERE id_header="+str(row[0])+" AND id_dictionary="+str(row[1]))
            wx.EndBusyCursor()
            cur.execute("DELETE FROM settings WHERE setting_key='do_encrypt'")
        #db encryption - End

        self.cryptokey=self.txtCryptokey.GetValue()
        self.digestkey=sha256(self.cryptokey).digest()

        cUsername=base64.b64encode(rijndael.EncryptData(self.digestkey,self.txtUsername.GetValue()))
        cPassword=base64.b64encode(rijndael.EncryptData(self.digestkey,self.txtPassword.GetValue()))
        lastpasswordchange=""
        cur.execute("SELECT id_user,id_group,uap,date_time FROM users WHERE username='"+cUsername+"' AND password='"+cPassword+"' AND (status=0 OR status='')")
        row = cur.fetchone()
        if row==None:
            wx.MessageBox(TT("Bad Username, Password or Cypher key"), TT("Authentication failed"), wx.ICON_ERROR | wx.OK, self)
        else:
            self.id_logged_user=row[0]
            self.id_group=row[1]
            #print "id_group="+str(self.id_group)
            self.uap=row[2]
            lastpasswordchange=row[3]
            
            self.group_shortcode="-"
            cur.execute("SELECT shortcode FROM groups WHERE id_group="+str(self.id_group)+" AND status=0") 
            row = cur.fetchone()
            if row!=None:
                self.group_shortcode=row[0]

            #HEADERS INIT - BEGIN
            self.id_crf=0
            #database=self.txtDatabase.GetValue()
            database=self.choiceDatabase.GetStringSelection()
            if database=="": database=CRF_DESC
            if DEBUG_MODE: print "database: "+database
            cur.execute("SELECT id_header FROM headers WHERE description='"+database+"' AND child_of=0 and status=0 AND id_header IN (select child_of from headers where status=0)")
            row = cur.fetchone()
            if row==None:
                pass
                #if database==CRF_DESC:
                    #longCRC=UniqueLongCRC("0,"+database,cur)
                    #cur.execute("INSERT INTO headers (id_header,child_of,description,id_section) VALUES ("+str(longCRC)+",0,'"+database+"',0)")
                    #self.id_crf=longCRC
            else:
                self.id_crf=row[0]
            if self.id_crf==0:
                wx.MessageBox(TT("Database not found"), TT("Authentication failed"), wx.ICON_ERROR | wx.OK, self)
            else:
                wx.BeginBusyCursor()

                self.parent.syncroTaskPause=True
                import time
                while self.parent.syncroTaskUpdating:
                    time.sleep(1)
                    if DEBUG_MODE: print "waiting synchro standby..."
                cur.execute("BEGIN TRANSACTION")
                    
                self.waitframe = wx.MiniFrame(None, -1, TT("Initializing"), size=wx.Size(340,80))
                #frame = wx.MiniFrame(self.parent, -1, TT("Initializing"), size=wx.Size(340,80),style=wx.FRAME_FLOAT_ON_PARENT)
                panel = wx.Panel(self.waitframe)
                lbl=wx.StaticText(label=TT('Please wait'), parent=panel, style=wx.ALIGN_CENTRE_HORIZONTAL|wx.EXPAND)
                lbl.SetFont(wx.Font( 24, wx.SWISS, wx.NORMAL, wx.BOLD))
                sizer = wx.BoxSizer(wx.VERTICAL)
                sizer.Add(lbl, 0, wx.ALL, 10)
                panel.SetSizer(sizer)
                panel.Layout()
                self.waitframe.Show(1)
                #frame.Center(direction=wx.BOTH)
                self.waitframe.CenterOnParent()
                wx.Yield()
                if os.name=='nt':
                    import win32gui
                    win32gui.PumpWaitingMessages()
                else:
                    import time
                    time.sleep(0.5)
                self.waitframe.Update()
                

                cur.execute("SELECT id_header,description FROM headers WHERE child_of="+str(self.id_crf)+" AND description='"+self.keyname+"'")
                row = cur.fetchone()
                if row==None:
                    longCRC=UniqueLongCRC(str(self.id_crf)+","+self.keyname,cur)
                    cur.execute("INSERT INTO headers (id_header,child_of,description,date_time,status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'"+self.keyname+"','0000-00-00 00:00:00',0)")
                    
                cur.execute("SELECT id_header,description FROM headers WHERE child_of="+str(self.id_crf)+" AND description='group_shortcode'")
                row = cur.fetchone()
                if row==None:
                    longCRC=UniqueLongCRC(str(self.id_crf)+","+'group_shortcode',cur)
                    cur.execute("INSERT INTO headers (id_header,child_of,description,date_time,status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'group_shortcode','0000-00-00 00:00:00',0)")

                cur.execute("SELECT id_header,description FROM headers WHERE child_of="+str(self.id_crf)+" AND description='_warnings_'")
                row = cur.fetchone()
                if row==None:
                    longCRC=UniqueLongCRC(str(self.id_crf)+","+'_warnings_',cur)
                    cur.execute("INSERT INTO headers (id_header,child_of,description,date_time,status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'_warnings_','0000-00-00 00:00:00',0)")
                    
                #new reset - begin
                self.arrHeaders={}
                self.hashHeaders={}
                self.arrHeadersLabels={}
                self.arrHeadersEvents={}
                self.arrHeadersDefaults={}
                self.arrHeadersTypedesc={}
                self.arrHeadersSubtypedesc={}
                self.arrHeadersValidations={}
                self.arrHeadersComboPreload={}
                self.arrHeadersIdSection={}
                self.arrHeadersPos={}
                self.arrFound={}
                self.arrReports={}
                #new reset - end
                tmpHeadersIdSection={}
                tmpHeadersPos={}
                tmpHeadersAligned={}
                print "setting master header to "+str(self.id_crf)
                cur.execute("SELECT id_header,description,onchange,defaultvalue,validation,id_section,pos,typedesc,subtypedesc,label FROM headers WHERE child_of="+str(self.id_crf))
                for row in cur:
                    self.arrHeaders[row[0]]=row[1]
                    self.hashHeaders[row[1]]=row[0]
                    self.arrHeadersEvents[row[1]]=row[2]
                    if self.arrHeadersEvents[row[1]]=="": self.arrHeadersEvents[row[1]]=None
                    if self.arrHeadersEvents[row[1]]!=None:
                        ## Retrocompatibilita' - Inizio
                        if DEBUG_MODE: print "Initializing events handler for field: "+str(row[0])+"-"+row[1]
                        self.arrHeadersEvents[row[1]]=str(self.arrHeadersEvents[row[1]])
                        self.arrHeadersEvents[row[1]]=self.arrHeadersEvents[row[1]].replace("wx.xrc.XRCCTRL","XRCCTRL")
                        self.arrHeadersEvents[row[1]]=self.arrHeadersEvents[row[1]].replace("XRCCTRL","wx.xrc.XRCCTRL")
                        self.arrHeadersEvents[row[1]]=self.arrHeadersEvents[row[1]].replace("CustomRouter(parentFrame,serverurl)","CustomRouter(parentFrame.arrSettings,serverurl)")
                        ## Retrocompatibilita' - Fine
                    self.arrHeadersDefaults[row[1]]=row[3]
                    if self.arrHeadersDefaults[row[1]]=="": self.arrHeadersDefaults[row[1]]=None
                    self.arrHeadersTypedesc[row[1]]=row[7]
                    datectrls[row[1]]=str(row[7])
                    if self.arrHeadersTypedesc[row[1]]=="": self.arrHeadersTypedesc[row[1]]=None
                    self.arrHeadersSubtypedesc[row[1]]=row[8]
                    if self.arrHeadersSubtypedesc[row[1]]=="": self.arrHeadersSubtypedesc[row[1]]=None
                    self.arrHeadersValidations[row[1]]=row[4]
                    if self.arrHeadersValidations[row[1]]=="": self.arrHeadersValidations[row[1]]=None
                    tmpHeadersIdSection[row[1]]=row[5]
                    try: self.arrHeadersIdSection[row[1]]=int(row[5])
                    except: pass
                    try: self.arrHeadersPos[row[1]]=int(row[6])
                    except: pass
                    tmpHeadersPos[row[1]]=row[6]
                    #labels - begin
                    label=row[9]
                    if label==None: label=""
                    else: label=str(label)
                    self.arrHeadersLabels[row[1]]=label
                    #labels - end

                print "Headers initialized."
                #HEADERS INIT - END

                # Dupes checking - Begin
#                if GetSetting(self.parent,'dupes_checking')=="0":
                if int("0"+GetSetting(self.parent,'dupes_checking'))>=3:
                    print "Dupes checking started."
                    query=""
                    if self.keymode=="private":
                        query="SELECT users.id_user,id_dictionary,count(*) FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND users.id_user="+str(self.id_user)+" GROUP BY users.id_user,id_dictionary"
                    elif self.keymode=="group":
                        #query="SELECT id_dictionary,count(*) FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND users.id_group="+str(self.id_group)+" GROUP BY id_dictionary"
                        query="SELECT c1.id_dictionary,c2.id_dictionary,count(*) FROM users,rows,contents_index as c1, contents_index as c2 WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=c1.id_row AND rows.id_user=c1.id_user AND rows.id_instance=c1.id_instance AND rows.id_row=c2.id_row AND rows.id_user=c2.id_user AND rows.id_instance=c2.id_instance AND c1.id_header="+str(self.hashHeaders["group_shortcode"])+" AND c2.id_header="+str(self.hashHeaders[self.keyname])+" GROUP BY c1.id_dictionary,c2.id_dictionary"
                    else:
                        query="SELECT 0,id_dictionary,count(*) FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" GROUP BY id_dictionary"
                    repl=[]
                    cur.execute(query)
                    for row in cur:
                        if row[2]>1: repl.append(str(row[0])+"-"+str(row[1]))
                    if len(repl)==0:
                        print "Dupes checking: no errors."
                    else:
                        print "Dupes checking: fixing errors..."
                        if DEBUG_MODE: logfile = open('dupes_checking_log.txt','w')
                        
                        cur.execute("SELECT max(id_row) FROM ROWS")
                        row=cur.fetchone()
                        max_id_row=row[0]
                        #id_instance=9999999999
                        id_instance=self.parent.id_instance
                        for pat in repl:
                            lastdatetime=""
                            records=[]
                            query=""
                            if self.keymode=="private":
                                query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND users.id_user="+str(self.id_user)+" AND contents_index.id_dictionary="+pat.split("-")[1]
                            elif self.keymode=="group":
                                #query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND users.id_group="+str(self.id_group)+" AND contents_index.id_dictionary="+str(pat)
                                query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index as c1,contents_index as c2 WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=c1.id_row AND rows.id_user=c1.id_user AND rows.id_instance=c1.id_instance AND rows.id_row=c2.id_row AND rows.id_user=c2.id_user AND rows.id_instance=c2.id_instance AND c1.id_header="+str(self.hashHeaders["group_shortcode"])+" AND c2.id_header="+str(self.hashHeaders[self.keyname])+" AND c1.id_dictionary="+pat.split("-")[0]+" AND c2.id_dictionary="+pat.split("-")[1]
                            else:
                                query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND contents_index.id_dictionary="+pat.split("-")[1]
                            cur.execute(query)
                            for row in cur:
                                record={}
                                record["id_row"]=row[0]
                                record["id_user"]=row[1]
                                record["id_instance"]=row[2]
                                record["date_time"]=row[3]
                                if record["date_time"]>lastdatetime or lastdatetime=="": 
                                    lastdatetime=record["date_time"]
                                record["rap"]=row[4]
                                records.append(record)
                            id_user=0
                            rap=""
                            fields={}
                            fields_ts={}
                            lastdatetime=""
                            for item in records:
                                cur.execute("SELECT DISTINCT id_header,id_dictionary,id_cycle FROM contents_index WHERE id_row="+str(item["id_row"])+" AND id_user="+str(item["id_user"])+" AND id_instance="+str(item["id_instance"]))
                                for row in cur:
                                    curkey=str(row[0])+"-"+str(row[2])
                                    dopick=False
                                    if not fields.has_key(curkey):
                                        dopick=True
                                    elif item["date_time"]>fields_ts[curkey]:
                                        dopick=True
                                    if dopick:
                                        fields[curkey]=row[1]
                                        fields_ts[curkey]=item["date_time"]
                                        id_user=item["id_user"]
                                        rap=item["rap"]
                            import datetime
                            now = datetime.datetime.utcnow()
                            ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
                            if lastdatetime=="": lastdatetime=ts_db

                            max_id_row=max_id_row+1
                            #qy="INSERT INTO rows (id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(max_id_row)+","+str(id_user)+",1,'"+rap+"','"+ts_db+"',0,0,0,"+id_instance+")"
                            qy="INSERT INTO rows (id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(max_id_row)+","+str(id_user)+",1,'"+rap+"','"+lastdatetime+"',0,0,0,"+id_instance+")"
                            if DEBUG_MODE: logfile.write(qy+"\n")
                            #print qy
                            cur.execute(qy)
                            for elm in fields.keys():
                                id_header=elm.split("-")[0]
                                id_cycle=elm.split("-")[1]
                                #qy="INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) VALUES ("+str(max_id_row)+","+str(id_user)+","+str(id_header)+","+str(fields[id_header])+","+id_instance+",0)"
                                qy="INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) VALUES ("+str(max_id_row)+","+str(id_user)+","+id_header+","+str(fields[elm])+","+id_instance+","+id_cycle+")"
                                if DEBUG_MODE: logfile.write(qy+"\n")
                                #print qy
                                cur.execute(qy)
                            for item in records:
                                #qy="UPDATE rows SET status="+str(max_id_row)+",status_user="+str(id_user)+",status_instance="+str(id_instance)+" WHERE id_row="+str(item["id_row"])+" AND id_user="+str(item["id_user"])+" AND id_instance="+str(item["id_instance"])
                                qy="UPDATE rows SET status="+str(max_id_row)+",status_user="+str(id_user)+",status_instance="+id_instance+" WHERE id_row="+str(item["id_row"])+" AND id_user="+str(item["id_user"])+" AND id_instance="+str(item["id_instance"])
                                if DEBUG_MODE: logfile.write(qy+"\n")
                                #print qy
                                cur.execute(qy)

                        if DEBUG_MODE: logfile.close()
                        self.parent.syncrotablesupd["rows"]=True
                        print "Dupes checking: done fixing errors."
                else:
                    print "Dupes checking: skipping."
                # Dupes checking - End

                #direttive sulla ciclicita'
                self.IdForm=[]
                self.ActivateCyclicity=False
                self.Cyclicity={}   #conta da 1
                self.OnFormActivate={}
                self.FormName=[]
                self.FormTitle=[]
                self.FormHeight=[]
                self.FormHref=[]
                self.FormIdxDict=[]
                self.multidict=False
                cc=1
                #cur.execute("SELECT description,label,cyclic,onactivate,height FROM eforms WHERE id_header="+str(self.id_crf)+" AND (id_group=0 OR id_group="+str(self.id_group)+") AND status=0 AND pos>0 ORDER BY POS,id_eform")
                cur.execute("SELECT id_eform,description,label,cyclic,onactivate,height,href FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
                for row in cur:
                    #conta da uno. la chiave e' numerica.
                    if cc==1: 
                        try: wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(False)
                        except: pass
                        wx.xrc.XRCCTRL(self.frameRecordset,'cycles').Show(False)
                        if DEBUG_MODE: print "objCycles hidden"
                    self.IdForm.append(row[0])
                    self.FormName.append(row[1])
                    self.FormTitle.append(row[2])
                    self.Cyclicity[cc]=row[3]
                    if self.Cyclicity[cc]>0: self.ActivateCyclicity=True
                    self.OnFormActivate[cc]=row[4]
                    if self.OnFormActivate[cc]==None: self.OnFormActivate[cc]=""
                    self.OnFormActivate[cc]=str(self.OnFormActivate[cc])
                    self.OnFormActivate[cc]=self.OnFormActivate[cc].replace("wx.xrc.XRCCTRL","XRCCTRL")
                    self.OnFormActivate[cc]=self.OnFormActivate[cc].replace("XRCCTRL","wx.xrc.XRCCTRL")
                    formheight=1000
                    if row[5]!=None: formheight=row[5]
                    self.FormHeight.append(int("0"+str(formheight)))
                    href=""
                    if row[6]!=None: href=row[6]
                    self.FormHref.append(href)
                    for tag in href.split(","):
                        if tag.split("=")[0]=="#": 
                            self.multidict=True
                            self.FormIdxDict.append(int(tag.split("=")[1]))
                            break
                    cc=cc+1

                objChoice=wx.xrc.XRCCTRL(self.frameRecordset,'choice')
                if objChoice.GetCount()==0:
                    objChoice.Clear()
                    for i in range(len(self.FormTitle)):
                        objChoice.Append(self.FormTitle[i])
                else:
                    if len(self.FormName)==0:
                        for i in range(objChoice.GetCount()):
                            self.FormName.append('panel'+`i+1`)
                            self.FormHeight.append(int(GetSetting(self.parent,'panel'+`i+1`+'_height')))

                #Comandi amministratore
                candoit=False
                permission=GetSetting(self.parent,"sysadmin_groups")
                if permission!="":
                    s1=","+permission+","
                    s2=","+str(self.id_group)+","
                    if s2 in s1: candoit=True
                else:
                    id_group_admin=GetSetting(self.parent,"id_group_admin")
                    if str(self.id_group)==str(id_group_admin):
                        candoit=True
                if candoit:
                    if GetSetting(self.parent,'automergedupes')!="0":
                        self.buttonAutoMergeDupes.Show(True)
                    else: 
                        self.buttonAutoMergeDupes.Show(False)
                    self.crfTypeChoice.Show(True)
                    self.crfTypeChoice.SetStringSelection('valid records')
                    self.buttonShowDiff.Show(True)
                else:
                    self.buttonAutoMergeDupes.Show(False)
                    self.crfTypeChoice.Show(False)
                    self.buttonShowDiff.Show(False)
                #caricamento CRF
                crfloaded=False
                try:
                    tmp=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[0])
                    if tmp!=None: crfloaded=True
                except: pass
                if not crfloaded:
                    print "Loading CRF..."
                    
                    self.xrcPanel = self.xmlResource1.LoadPanel(name = 'panelHeader', parent = self.frameSearch)
                    wx.xrc.XRCCTRL(self.frameSearch,'panelHeader').SetPosition(wx.Point(0,0))

                    #check presenza search page customizzate x crf ed eventualmente x group
                    #cur.execute("SELECT description FROM eforms WHERE id_header="+str(self.id_crf)+" AND (id_group=0 OR id_group="+str(self.id_group)+") AND status=0 AND pos=0")
                    buildFieldsList=True
                    # cur.execute("SELECT * from efields")
                    # row = cur.fetchone()
                    # if row==None: buildFieldsList=True
                    ctFieldsList=0
                    idForm=0
                    cur.execute("SELECT id_eform,description FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos=0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+")))")
                    row = cur.fetchone()
                    if row!=None:
                        self.PanelSearch = self.xmlResource1.LoadPanel(name = row[1], parent = self.frameSearch)
                        idForm=row[0]
                    else:
                        self.PanelSearch = self.xmlResource1.LoadPanel(name = 'panelSearch', parent = self.frameSearch)

                    self.PanelSearch.SetPosition(wx.Point(10,40))
                    self.PanelSearch.SetSize(wx.Size(1000,90))

                    #PreDecrypt - data preparation
                    list=self.PanelSearch.GetChildren()
                    strlstfields=""
                    for i in range(len(list)):
                        if self.IsInputField(list[i]):
                            if list[i].GetName()!="heavybase_memo1":
                                curName=list[i].GetName()
                                if not self.hashHeaders.has_key(curName):
                                    longCRC=UniqueLongCRC(str(self.id_crf)+","+curName,cur)
                                    cur.execute("INSERT INTO headers (id_header,child_of,description,date_time,status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'"+curName+"','0000-00-00 00:00:00',0)")
                                    self.hashHeaders[curName]=longCRC
                                if strlstfields!="": strlstfields=strlstfields+","
                                strlstfields=strlstfields+str(self.hashHeaders[list[i].GetName()])
                    print "Starting predecrypt."
                    self.PreDecryptTask = PreDecrypt(self.parent,self,self.digestkey,self.hashHeaders[self.keyname],strlstfields)
                    self.PreDecryptTask.start()

                    list=self.PanelSearch.GetChildren()
                    for i in range(len(list)):
                        if self.IsInputField(list[i]):
                            if list[i].GetName()!="heavybase_memo1":
                                list[i].Bind(wx.EVT_KEY_UP, self.OnSearchEnterPressed)
                    if buildFieldsList: 
                        print "rebuilding fieldslist"
                        cur.execute("DELETE FROM efields WHERE id_eform="+str(idForm))
                        ctFieldsList=self.BuildFieldsList(idForm,list,ctFieldsList)
                        print "done."
                    self.CrfPageEnabled={}
                    if DEBUG_MODE: print "loading panels"
                    for i in range(objChoice.GetCount()):
                        self.CrfPageEnabled[i]=True
                        self.xrcPanel = self.xmlResource1.LoadPanel(name = self.FormName[i], parent = self.frameCrf)
                        if DEBUG_MODE: print ". "+self.FormName[i]
                        wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(False)
                        if i==0:
                            wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(10, 10))
                        else:
                            wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(5000, 0))
                        wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetSize(wx.Size(1000,self.FormHeight[i]))
                        if buildFieldsList:
                            idForm=i+1
                            try: idForm=self.IdForm[i]
                            except: pass
                            list=self.xrcPanel.GetChildren()
                            cur.execute("DELETE FROM efields WHERE id_eform="+str(idForm))
                            ctFieldsList=self.BuildFieldsList(idForm,list,ctFieldsList)
                        i=i+1
                    self.OnWindowResize(None)
                    if self.ActivateCyclicity:
                        wx.xrc.XRCCTRL(self.frameRecordset,'cycles').Bind(wx.EVT_RADIOBOX, self.OnChangeCycle, id=wx.xrc.XRCID('cycles'))
#                        wx.xrc.XRCCTRL(self.frameRecordset,'spinCycles').Bind(wx.EVT_SPIN_UP, self.OnScrollUpFrmCycles)
#                        wx.xrc.XRCCTRL(self.frameRecordset,'spinCycles').Bind(wx.EVT_SPIN_DOWN, self.OnScrollDnFrmCycles)

                        try:
                            frmCycles=wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles')
                            frmCycles.SetVirtualSize(wx.Size(440,120))
                            frmCycles.SetScrollbars(0,40,0,2)
                            frmCycles.Scroll(0,0)
                        except: pass

                    print "CRF Loaded."

                # Costruzione anticipata Headers - Inizio
                self.fieldNamesFound=[]
                ct=[0]
                for ifrm in range(objChoice.GetCount()):
                    list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
                    self.AlignHeaders(ifrm,list,ct,tmpHeadersIdSection,tmpHeadersPos,tmpHeadersAligned)
                
                for elm in self.hashHeaders:
                    if elm not in self.fieldNamesFound:
                        try: cur.execute("UPDATE headers SET id_section=null, pos=null WHERE description='"+elm+"'")
                        except: pass
                        if elm in self.arrHeadersIdSection: del self.arrHeadersIdSection[elm]
                        if elm in self.arrHeadersPos: del self.arrHeadersPos[elm]
                        
                self.HeadersUsrBySect={} #variabili degli utenti
                self.HeadersMonBySect={} #variabili dei monitor
                self.HeadersMonPattern={}
                self.HeadersAdmBySect={} #variabili degli amministratori
                cur.execute("SELECT description,id_section,deftype FROM headers WHERE child_of="+str(self.id_crf))
                for row in cur:
                    deftype=row[2]
                    if deftype==None: deftype=""
                    self.HeadersMonPattern[row[0]]=deftype
                    #if "M" in deftype:
                    if deftype!="" and deftype!="S" and deftype!="A":
                        if not self.HeadersMonBySect.has_key(row[1]): self.HeadersMonBySect[row[1]]=[]
                        self.HeadersMonBySect[row[1]].append(row[0])
                    else:
                        if not self.HeadersUsrBySect.has_key(row[1]): self.HeadersUsrBySect[row[1]]=[]
                        self.HeadersUsrBySect[row[1]].append(row[0])
                    if deftype=="A":
                        if not self.HeadersAdmBySect.has_key(row[1]): self.HeadersAdmBySect[row[1]]=[]
                        self.HeadersAdmBySect[row[1]].append(row[0])
                # Costruzione anticipata Headers - Fine

                print "Headers aligned."

                #Lista Reports disponibili - Inizio
                self.reportChoice.Append("")
                cur.execute("SELECT id_profile, description FROM reports_profiles ORDER BY id_profile")
                self.arrReports={}
                cc=0
                for row in cur:
                    if cc==0:
                        self.lblReport.Show(True)
                        self.reportChoice.Show(True)
                        #self.buttonReportChoice.Show(True)
                        self.reportChoice.Clear()
                        self.reportChoice.Append("")
                    cc=cc+1
                    self.reportChoice.Append(row[1])
                    self.arrReports[cc]=row[0]
                if cc==0:
                    self.lblReport.Show(False)
                    self.reportChoice.Show(False)
                    #self.buttonReportChoice.Show(False)
                #Lista Reports disponibili - Fine

                import datetime
                now = datetime.datetime.utcnow()
                ts=now.strftime("%d/%m/%Y, %H:%M:%S")
                ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
                cur.execute("SELECT firstaccess FROM users WHERE id_user="+str(self.id_logged_user))
                row = cur.fetchone()
                if row[0]==None:
                    cur.execute("UPDATE users SET firstaccess='"+ts_db+"' WHERE id_user="+str(self.id_logged_user))
                cur.execute("UPDATE users SET lastaccess='"+ts_db+"' WHERE id_user="+str(self.id_logged_user))

                #placeholder for old position of preDecrypt
                for srcField in self.PanelSearch.GetChildren():
                    if self.IsInputField(srcField):
                        if srcField.GetName()!="heavybase_memo1":
                            if srcField.GetClassName()=="wxTextCtrl":
                                srcField.SetEditable(True)
                                self.SetInputFieldValue(srcField,"")
                        else:
                            label_memo1=GetSetting(self.parent,"label_memo1")
                            if label_memo1!="": self.SetInputFieldValue(srcField,label_memo1)

                # test password expiration
                passexp=GetSetting(self.parent,'password_expiration')
                if passexp!="":
                    import time
                    a1=datetime.datetime(*time.strptime(lastpasswordchange, "%Y-%m-%d %H:%M:%S")[0:6])
                    b1=datetime.datetime(*time.strptime(ts_db, "%Y-%m-%d %H:%M:%S")[0:6])
                    if (b1-a1).days>int(passexp):
                        self.ForceChangeUserPassword()
                        
                if GetSetting(self.parent,"synchro_decrypt")!="1":
                    wx.EndBusyCursor()
                    self.frameLogin.SetPosition(wx.Point(5000,0))
                    self.frameSearch.SetPosition(wx.Point(0,0))
                    self.frameRecordset.SetPosition(wx.Point(5000,0))
                    self.EnableDisableButtons(2)
                    self.waitframe.SetPosition(wx.Point(5000,0))
                    self.waitframe.Destroy()
                else:
                    self.waiting_predecrypt=True
                    
                    
                retry_commit=True
                while retry_commit:
                    try:
                        if DEBUG_MODE: print "trying commit..."
                        cur.execute('COMMIT TRANSACTION')
                        retry_commit=False
                    except: 
                        outfile = open("quit_synchro.cmd", 'w')
                        outfile.write("\n")
                        outfile.close()
                        time.sleep(1)
                self.parent.syncroTaskPause=False
                
                print "Login completed."
                
                #Pool mode: 1 scheda x ogni utente
                if GetSetting(self.parent,'pool_mode')=="1": self.DoSearch()
                
    def AlignHeaders(self,ifrm,list,ct,tmpHeadersIdSection,tmpHeadersPos,tmpHeadersAligned):
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                self.AlignHeaders(ifrm,sublist,ct,tmpHeadersIdSection,tmpHeadersPos,tmpHeadersAligned)
            else:
                if self.IsInputField(list[i]) or list[i].GetClassName() in ("wxButton"):
                    curName=list[i].GetName()
                    if curName not in self.fieldNamesFound: self.fieldNamesFound.append(curName)
                    if not (self.hashHeaders.has_key(curName)):
                        #cur.execute("INSERT INTO headers (child_of,description, id_section) VALUES ("+str(self.id_crf)+",'"+curName+"',"+`ifrm+1`+")")
                        #cur.execute("SELECT last_insert_rowid()")
                        #self.hashHeaders[curName]=cur.fetchone()[0]
                        longCRC=UniqueLongCRC(str(self.id_crf)+","+curName,cur)
                        cur.execute("INSERT INTO headers (id_header,child_of,description, id_section, date_time, status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'"+curName+"',"+`ifrm+1`+",'0000-00-00 00:00:00',0)")
                        #cur.execute("SELECT last_insert_rowid()")
                        self.arrHeaders[longCRC]=curName
                        self.hashHeaders[curName]=longCRC
                        self.arrHeadersEvents[curName]=None
                        self.arrHeadersDefaults[curName]=None
                        self.arrHeadersTypedesc[curName]=None
                        self.arrHeadersSubtypedesc[curName]=None
                        self.arrHeadersValidations[curName]=None
                        self.arrHeadersIdSection[curName]=ifrm+1
                        self.arrHeadersPos[curName]=0
                    else:
                        ct[0]=ct[0]+1
                        #riallineamento headers sulla base dell'xrc
                        if not (tmpHeadersAligned.has_key(curName)):
                            tmpHeadersAligned[curName]=True
                            if (tmpHeadersIdSection.has_key(curName) and tmpHeadersPos.has_key(curName)):
                                if `tmpHeadersIdSection[curName]`!=`ifrm+1` or `tmpHeadersPos[curName]`!=`ct[0]`:
                                    #print curName  #Debug - verifica di quali campi sono stati riallineati
                                    cur.execute("UPDATE headers SET id_section="+`ifrm+1`+", pos="+`ct[0]`+ " WHERE child_of="+str(self.id_crf)+" AND description='"+curName+"'")
                                    self.arrHeadersIdSection[curName]=ifrm+1
                                    self.arrHeadersPos[curName]=ct[0]
                    if list[i].GetClassName() in ("wxComboBox","wxChoice"):
                        if not self.arrHeadersComboPreload.has_key(curName):
                            self.arrHeadersComboPreload[curName]=list[i].GetItems()
                            #print curName+": "+str(list[i].GetItems())

    def OnButtonLogoutButton(self, event):
        event.Skip
        self.DoLogout()
        
    def DoLogout(self):
        self.lstFound.Clear()
        self.txtUsername.SetValue("")
        self.txtPassword.SetValue("")
        #self.txtCryptoKey.SetValue("")
        self.frameLogin.SetPosition(wx.Point(0,0))
        self.frameSearch.SetPosition(wx.Point(5000,0))
        self.frameRecordset.SetPosition(wx.Point(5000,0))
        self.EnableDisableButtons(1)
        #chiusura parametrica attuale frame datamanager
        cur.execute("SELECT count(*) FROM headers WHERE child_of=0")
        row = cur.fetchone()
        if row!=None:
            if row[0]>=2:
                self.parent.OnDataEntry(None)
                self.Close()

    def OnSearchEnterPressed(self, event):
        if event.GetKeyCode() in (wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER):
            self.OnButtonSearchButton(event)

    def OnButtonSearchButton(self, event):
        event.Skip()
        self.DoSearch()

    def GetStrLocalDatetime(self,strUtcDatetime):
        par=strUtcDatetime
        if len(par)>19: par=par[:19]    #prendo la prima data/ora della stringa, non l'ultima
        #UtcDateTime=datetime.datetime(*time.strptime(strUtcDatetime, "%Y-%m-%d %H:%M:%S")[0:6])
        #now=datetime.datetime.now()
        #utcnow=datetime.datetime.utcnow()
        #diff=now-utcnow
        #LocalDatetime=UtcDateTime+diff
        #return LocalDatetime.strftime("%Y-%m-%d %H:%M:%S")
        
        #da local a utc:
        #time.strftime("%Y-%m-%d %H:%M:%S",time.gmtime(time.mktime(time.strptime("2008-09-17 14:04:00","%Y-%m-%d %H:%M:%S"))))
        import time,calendar
        return time.strftime("%Y-%m-%d %H:%M:%S",time.localtime(calendar.timegm(time.strptime(par,"%Y-%m-%d %H:%M:%S"))))
        
    def DoSearch(self):
        wx.BeginBusyCursor()
        self.lstFound.Clear()

        ##contents_dictionary
        #dictionary={}
        #cur.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary")
        #for row in cur:
            #dictionary[str(row[0])+"-"+str(row[1])]=row[2]
        ##populate table

        #pre-filter - begin
        con_m = sqlite.connect(":memory:")
        cur_m = con_m.cursor()

        #q1="SELECT r.id_row AS id_row"
        #q2=" FROM (select id_row from rows where rows.status=0) as r"

        list=self.PanelSearch.GetChildren()
        strlstfields=""
        lstfields=[]
        fldNames={}
        dValues={}
        cValues={}
        ct=0
        firstOrderBy=0
        for i in range(len(list)):
            if self.IsInputField(list[i]):
                if list[i].GetName()!="heavybase_memo1":
                    if strlstfields!="": strlstfields=strlstfields+","
                    strlstfields=strlstfields+list[i].GetName()
                    lstfields.append(list[i].GetName())
                    tmp=self.GetInputFieldValue(list[i])
                    if tmp:
                        if firstOrderBy=="": firstOrderBy=i
                        dValues[ct]=tmp
                        cValues[ct]=base64.b64encode(rijndael.EncryptData(self.digestkey,dValues[ct]))
                        fldNames[ct]=list[i].GetName()
                        ct=ct+1
        self.parent.lastSearch_lstfields=lstfields
        nSearchElm=len(lstfields)

        nSearchValues=len(cValues)
        doSearch=True
        if nSearchValues==0:
            if GetSetting(self.parent,'pool_mode')!="1":
                # if wx.MessageBox("No valid search criteria entered.\nDo you wish to list all of the records?", "Search", wx.YES_NO) == wx.NO:
                if wx.MessageBox(TT("Do you wish to list all of the records?"), TT("Search"), wx.ICON_QUESTION | wx.YES_NO) == wx.NO:
                    doSearch=False
        if doSearch:
            try: cur_m.execute("DROP TABLE search")
            except: pass
            
            #Comandi amministratore
            candoit=False
            permission=GetSetting(self.parent,"sysadmin_groups")
            if permission!="":
                s1=","+permission+","
                s2=","+str(self.id_group)+","
                if s2 in s1: candoit=True
            else:
                id_group_admin=GetSetting(self.parent,"id_group_admin")
                if str(self.id_group)==str(id_group_admin):
                    candoit=True
                        
            reqstatus=0
            adminmode=False
            if candoit: 
                crfTypeIdx=self.crfTypeChoice.GetSelection()
                if crfTypeIdx==1: adminmode=True
                elif crfTypeIdx==2: 
                    reqstatus=-1
                    adminmode=True
                elif crfTypeIdx==3: 
                    reqstatus=1
                    adminmode=True
            
            while strlstfields.find("#")>=0:
                strlstfields = strlstfields[:strlstfields.find("#")] + "__" + strlstfields[strlstfields.find("#")+1:]
            square(con,con_m,"search",lstfields,self.id_group,self.id_crf,keyfields="",adminmode=adminmode,reqstatus=reqstatus)
            sql="SELECT id_row,id_user,id_instance,date_time,"+strlstfields+" FROM search"
            cur_m.execute(sql)
            cols = [i[0] for i in cur_m.description]; hc={}; i=0
            for col in cols: hc[col]=i; i=i+1
            rowlst=[]
            self.arrFound={}
            elmSize={}
            cc=0
            for row in cur_m:
                dofilter=False
                for i in range(nSearchValues):
                    colname=fldNames[i]
                    while colname.find("#")>=0:
                        colname = colname[:colname.find("#")] + "__" + colname[colname.find("#")+1:]
                    #if self.chkTolerance.GetValue()==True and row[hc[fldNames[i]]] is not None:
                    if self.chkTolerance.GetValue()==True and row[hc[colname]] is not None:
                        if not self.parent.trans.has_key(row[i]):
                            #dData=HB_DecryptOne(self.digestkey,row[hc[fldNames[i]]],"latin-1")
                            dData=HB_DecryptOne(self.digestkey,row[hc[colname]],"latin-1")
                            #self.parent.trans[row[hc[fldNames[i]]]]=dData
                            self.parent.trans[row[hc[colname]]]=dData
                        else:
                            #dData=self.parent.trans[row[hc[fldNames[i]]]]
                            dData=self.parent.trans[row[hc[colname]]]
                        normData=dData.strip().lower()
                        normPattern=dValues[i].strip().lower()
                        if levenshtein_distance(normData,normPattern) > 2 and normData.find(normPattern) < 0: dofilter=True
                    else:
                        normData=""
                        #if row[hc[fldNames[i]]] is not None: normData=row[hc[fldNames[i]]]
                        if row[hc[colname]] is not None: normData=row[hc[colname]]
                        normPattern=cValues[i]
                        if normData!=normPattern: dofilter=True
                if not dofilter:
                    #print row
                    elmlst=[]
                    dDatalst=[]
                    cc=cc+1
                    for i in range(len(row)):
                        elm=row[i]
                        if i>3:
                            colname=lstfields[i-4]
                            if colname.find("#")>=0: colname=colname[:colname.find("#")]
                            if elm!=None:
                                if not self.parent.trans.has_key(row[i]):
                                    dData=HB_DecryptOne(self.digestkey,row[i],"latin-1")
                                    self.parent.trans[row[i]]=dData
                                else:
                                    dData=self.parent.trans[row[i]]
                                elm=dData
                                #typedesc=self.arrHeadersTypedesc[cols[i]]
                                typedesc=self.arrHeadersTypedesc[colname]
                                if typedesc==None: typedesc=""
                                #if cols[i]==self.keyname or typedesc in ("integer","float") or "number" in typedesc:
                                if colname==self.keyname or typedesc in ("integer","float") or "number" in typedesc:
                                    elm=elm.strip().replace(",",".")
                                    dec=elm.find(".")
                                    if dec>=0:
                                        elm=("0000000000"+elm[:dec])[-10:] + (elm[dec:]+"0000000000")[:11]
                                    else:
                                        elm=("0000000000"+elm)[-10:]
                                if typedesc=="date":
                                    if len(elm)==10:
                                        elm=elm[6:10]+elm[3:5]+elm[0:2]
                                curElmSize=len(dData)
                                if elmSize.has_key(cols[i]):
                                    if elmSize[cols[i]]<curElmSize: elmSize[cols[i]]=curElmSize
                                else:
                                    elmSize[cols[i]]=curElmSize
                            else:
                                dData=""
                                if not elmSize.has_key(cols[i]): elmSize[cols[i]]=0
                            dDatalst.append(dData)
                        elmlst.append(elm)
                    elmlst.append(dDatalst)
                    rowlst.append(elmlst)
            orderBy=4+firstOrderBy
            rowlst.sort(key=lambda x:(x[orderBy:]))
            self.parent.lastSearch=[]
            for i in range(len(rowlst)):
                currow=""
                for i1 in range(nSearchElm):
                    currow=currow+rowlst[i][nSearchElm+4][i1].ljust(elmSize[cols[i1+4]]+1)
                currow=currow+"["+self.GetStrLocalDatetime(rowlst[i][3])+"]"
                self.parent.lastSearch.append(rowlst[i][nSearchElm+4])
                self.lstFound.Append(currow)
                self.arrFound[i+1]=str(rowlst[i][0])+";"+str(rowlst[i][1])+";"+str(rowlst[i][2])+";"+self.GetStrLocalDatetime(rowlst[i][3])
            cur_m.close()
            con_m.close()
            if GetSetting(self.parent,'pool_mode')=="1":
                if len(self.arrFound)==1:
                    self.id_row=long(self.arrFound[1].split(";")[0])
                    self.id_user=long(self.arrFound[1].split(";")[1])
                    self.id_rowinstance=long(self.arrFound[1].split(";")[2])
                    self.OpenCurRecord()
                elif len(self.arrFound)==0:
                    self.DoAddNew()
        wx.EndBusyCursor()

    def FindDupes(self):
        con_m = sqlite.connect(":memory:")
        cur_m = con_m.cursor()
        list=self.PanelSearch.GetChildren()
        strlstfields=""
        lstfields=[]
        fldNames={}
        dValues={}
        cValues={}
        ct=0
        firstOrderBy=0
        emptyvalue=base64.b64encode(rijndael.EncryptData(self.digestkey,""))
        strPresetFields=GetSetting(self.parent,'find_dupes_fields')
        presetFields=()
        if strPresetFields!="":
            presetFields=strPresetFields.split(",")
            if "group_shortcode" not in presetFields: presetFields.append("group_shortcode")
            if self.keyname not in presetFields: presetFields.append(self.keyname)
        for i in range(len(list)):
            if self.IsInputField(list[i]):
                if list[i].GetName()!="heavybase_memo1" and (len(presetFields)==0 or list[i].GetName() in presetFields):
                    fldname=list[i].GetName()
                    if strlstfields!="": strlstfields=strlstfields+","
                    strlstfields=strlstfields+fldname
                    lstfields.append(fldname)
                    tmp=""
                    if self.contents.has_key(fldname):
                        tmp=self.contents[fldname]
                    if tmp:
                        if firstOrderBy=="": firstOrderBy=i
                        dValues[ct]=tmp
                        cValues[ct]=base64.b64encode(rijndael.EncryptData(self.digestkey,dValues[ct]))
                        fldNames[ct]=fldname
                        ct=ct+1
        for fldname in ("group_shortcode",self.keyname):
            if fldname not in lstfields:
                if strlstfields!="": strlstfields=strlstfields+","
                strlstfields=strlstfields+fldname
                lstfields.append(fldname)
                tmp=""
                if self.contents.has_key(fldname):
                    tmp=self.contents[fldname]
                if tmp:
                    if firstOrderBy=="": firstOrderBy=i
                    dValues[ct]=tmp
                    cValues[ct]=base64.b64encode(rijndael.EncryptData(self.digestkey,dValues[ct]))
                    fldNames[ct]=fldname
                    ct=ct+1
        nSearchElm=len(lstfields)
        nSearchValues=len(cValues)
        doSearch=True
        if nSearchValues==0: 
            return None,None
        else:
            try: cur_m.execute("DROP TABLE search")
            except: pass
            square(con,con_m,"search",lstfields,self.id_group,self.id_crf,keyfields="")
            while strlstfields.find("#")>=0:
                strlstfields = strlstfields[:strlstfields.find("#")] + "__" + strlstfields[strlstfields.find("#")+1:]
            sql="SELECT id_row,id_user,id_instance,date_time,"+strlstfields+" FROM search"
            cur_m.execute(sql)
            cols = [i[0] for i in cur_m.description]; hc={}; i=0
            for col in cols: hc[col]=i; i=i+1
            rowlst=[]
            self.arrFound={}
            elmSize={}
            cc=0
            for row in cur_m:
                atleast1=False
                dofilter=False
                for i in range(nSearchValues):
                    if fldNames[i] not in ("group_shortcode", self.keyname):
                        normData=""
                        if row[hc[fldNames[i].replace("#","__")]] is not None: normData=row[hc[fldNames[i].replace("#","__")]]
                        normPattern=cValues[i]
                        if normData!=emptyvalue and normPattern!=emptyvalue and normData!="" and normPattern!="":
                            atleast1=True
                            if normData!=normPattern: dofilter=True
                if atleast1 and not dofilter:
                    if (self.contents.has_key("group_shortcode") and hc.has_key("group_shortcode")) or (self.contents.has_key(self.keyname) and hc.has_key(self.keyname)):
                        dofilter=True
                        for i in range(nSearchValues):
                            if fldNames[i] in ("group_shortcode", self.keyname):
                                normData=""
                                if row[hc[fldNames[i]]] is not None: normData=row[hc[fldNames[i]]]
                                normPattern=cValues[i]
                                if normData!=normPattern: dofilter=False
                if atleast1 and not dofilter:
                    elmlst=[]
                    dDatalst=[]
                    cc=cc+1
                    for i in range(len(row)):
                        elm=row[i]
                        if i>3:
                            if elm!=None:
                                if not self.parent.trans.has_key(row[i]):
                                    dData=HB_DecryptOne(self.digestkey,row[i],"latin-1")
                                    self.parent.trans[row[i]]=dData
                                else:
                                    dData=self.parent.trans[row[i]]
                                elm=dData
                                typedesc=self.arrHeadersTypedesc[cols[i]]
                                if typedesc==None: typedesc=""
                                if cols[i]==self.keyname or typedesc in ("integer","float") or "number" in typedesc:
                                    elm=elm.strip().replace(",",".")
                                    dec=elm.find(".")
                                    if dec>=0:
                                        elm=("0000000000"+elm[:dec])[-10:] + (elm[dec:]+"0000000000")[:11]
                                    else:
                                        elm=("0000000000"+elm)[-10:]
                                if typedesc=="date":
                                    if len(elm)==10:
                                        elm=elm[6:10]+elm[3:5]+elm[0:2]
                                curElmSize=len(dData)
                                if elmSize.has_key(cols[i]):
                                    if elmSize[cols[i]]<curElmSize: elmSize[cols[i]]=curElmSize
                                else:
                                    elmSize[cols[i]]=curElmSize
                            else:
                                dData=""
                                if not elmSize.has_key(cols[i]): elmSize[cols[i]]=0
                            dDatalst.append(dData)
                        elmlst.append(elm)
                    elmlst.append(dDatalst)
                    rowlst.append(elmlst)
            orderBy=4+firstOrderBy
            rowlst.sort(key=lambda x:(x[orderBy:]))
            lstFound=[]
            for i in range(len(rowlst)):
                currow=""
                for i1 in range(nSearchElm):
                    currow=currow+rowlst[i][nSearchElm+4][i1].ljust(elmSize[cols[i1+4]]+1)
                currow=currow+"["+self.GetStrLocalDatetime(rowlst[i][3])+"]"
                lstFound.append(currow)
            cur_m.close()
            con_m.close()
            return fldNames,lstFound
        
    def OnButtonExportSearchButton(self, event):
        event.Skip()
        if len(self.parent.lastSearch)>0:
            wx.BeginBusyCursor()
            try: w=xlwt.Workbook()
            except: w=Workbook()
            ws=w.add_sheet('search')
            for x in range(len(self.parent.lastSearch_lstfields)):
                ws.write(0,x,self.parent.lastSearch_lstfields[x])
            for y in range(len(self.parent.lastSearch)):
                for x in range(len(self.parent.lastSearch[y])):
                    ws.write(y+1,x,self.parent.lastSearch[y][x])
            reportname="lastSearch.xls"
            w.save(REPORTS_PATH + reportname)
            wx.EndBusyCursor()
            if not assoc_open(REPORTS_PATH+reportname):
                wx.MessageBox("File '"+reportname+"' created.", "Search", wx.ICON_INFORMATION | wx.OK, self)

    def OnButtonExportButton(self, event):
        wx.BeginBusyCursor()
        import time
        time.sleep(1)
        now = datetime.datetime.utcnow()
        ts=now.strftime("%Y%m%d%H%M%S")
        dbexport="export_"+str(self.id_logged_user)+"_"+ts+".db"
        shutil.copy(DATABASE, REPORTS_PATH+dbexport)
        con2 = sqlite.connect(REPORTS_PATH+dbexport, isolation_level=None)
        cur2 = con2.cursor()
        #cur2.execute("UPDATE contents SET data='' WHERE id_header IN (SELECT id_header FROM headers WHERE stype='P')")
        cur2.execute("DELETE FROM contents_index WHERE id_header IN (SELECT id_header FROM headers WHERE stype='P')")
        cur2.execute("DELETE FROM contents_dictionary WHERE id_header IN (SELECT id_header FROM headers WHERE stype='P')")
        cur2.close()
        con2.close()
        wx.EndBusyCursor()
#        wx.MessageBox("File '"+os.path.abspath(os.path.dirname(dbexport))+os.path.sep+dbexport+"' created.", "Export", wx.ICON_INFORMATION | wx.OK, self)
#        wx.MessageBox("File '"+dbexport+"' created.", "Export", wx.ICON_INFORMATION | wx.OK, self)

        #autoupdates - begin
        if wx.MessageBox("File '"+dbexport+"' created.\nDo you wish to send it directly?\n(internet connection required)", "Export", wx.YES_NO) == wx.YES:
            #init
            serverurl = GetSetting(self.parent,'heavybase_server')
            projectname = GetSetting(self.parent,'project_name')
            currelease = GetSetting(self.parent,'heavybase_release')
            username = self.txtUsername.GetValue()
            import hashlib
            password = hashlib.md5(self.txtUsername.GetValue()).hexdigest()
            #begin
            import os
            if GetSetting(self.parent,"Network_Connection_Mode")!="use MSIE" or os.name!='nt':
                if GetSetting(self.parent,"Network_Connection_Mode")!="use http":
                    import xmlrpclib
                    customRouting=CustomRouter(self.parent.arrSettings,serverurl)
                    customRouting.DoOpen()
                    print "connecting to "+customRouting.serverurl
                    server = xmlrpclib.ServerProxy("http://"+customRouting.serverurl, transport = customRouting.customTransport)
                else:
                    customRouting=CustomRouter(self.parent.arrSettings,serverurl)
                    customRouting.GetSettings()
                    #serverurl, useproxy, proxy_host, proxy_port, proxy_username, proxy_password
                    server = HTTPconn(serverurl, customRouting.useproxy, customRouting.Network_Proxy_Host, customRouting.Network_Proxy_Port, 
                                      customRouting.Network_Proxy_Username, customRouting.Network_Proxy_Password)
            else:
                server = IEconn.getInstance(serverurl)
            try:
                # upload datafile
                import zlib
                filename=".."+os.path.sep+dbexport
                infile = open(filename)
                contents=infile.read()
                infile.close()
                encoded = base64.b64encode(zlib.compress(contents))
                res = server.uploadFile(projectname,username,password,dbexport,encoded)
                cur.execute("INSERT INTO exports (id_user,timestamp,filename,synced) VALUES ("+str(self.id_logged_user)+",'"+now.strftime("%d/%m/%Y, %H:%M:%S")+"','"+dbexport+"','automatic, succeeded')")
                wx.MessageBox(res, "Update succes", wx.ICON_INFORMATION | wx.OK, self)
            except:
                cur.execute("INSERT INTO exports (id_user,timestamp,filename,synced) VALUES ("+str(self.id_logged_user)+",'"+now.strftime("%d/%m/%Y, %H:%M:%S")+"','"+dbexport+"','automatic, failed. reverting to manual')")
                wx.MessageBox("Error sending datafile", "Update failed", wx.ICON_ERROR | wx.OK, self)
                wx.MessageBox("Please send via mail as attachment the file\n'"+dbexport+"'\nto the Coordinating Centre.", "Export", wx.ICON_INFORMATION | wx.OK, self)

            if GetSetting(self.parent,"Network_Connection_Mode")!="use MSIE" or os.name!='nt':
                customRouting.DoClose()
        else:
            cur.execute("INSERT INTO exports (id_user,timestamp,filename,synced) VALUES ("+str(self.id_logged_user)+",'"+now.strftime("%d/%m/%Y, %H:%M:%S")+"','"+dbexport+"','manual')")
            wx.MessageBox("Please send via mail as attachment the file\n'"+dbexport+"'\nto the Coordinating Centre.", "Export", wx.ICON_INFORMATION | wx.OK, self)

        #autoupdates - end

    def EnableCrfField(self,pagename,fieldname,value,pattern=""):
        if wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,pagename),fieldname)!=None:
            okfield=False
            if pattern in ("","M","S"): 
                okfield=True
            else:
                #if "," in pattern:
                lstfields=pattern.split(",")
                for fld in lstfields:
                    if fieldname==fld:
                        okfield=True
                        break
                    else:
                        import re
                        fldre=re.compile(pattern)
                        if fldre.match(fieldname): 
                            okfield=True
                            break
            if okfield: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,pagename),fieldname).Enable(value)

    def MonitoringCheck(self,page):
        wx.BeginBusyCursor()
        #if DEBUG_MODE: print "MonitoringCheck"
        id_group_admin=GetSetting(self.parent,"id_group_admin")
        id_group_monitor=GetSetting(self.parent,"id_group_monitor")
        id_group_auditor=GetSetting(self.parent,"id_group_auditor")
        id_group_sysadmin=GetSetting(self.parent,"sysadmin_groups")
        if str(self.id_group)==str(id_group_admin) or str(self.id_group) in str(id_group_sysadmin).split(","):
            # Admin mode
            if self.HeadersMonBySect.has_key(page):
                for mon_var in self.HeadersMonBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],mon_var,True)
                    for usr_var in self.HeadersUsrBySect[page]:
                        if self.GetInputFieldValue(wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),mon_var)):
                            self.EnableCrfField(self.FormName[page-1],usr_var,False,self.HeadersMonPattern[mon_var])
                        else:
                            self.EnableCrfField(self.FormName[page-1],usr_var,True,self.HeadersMonPattern[mon_var])
            if self.HeadersAdmBySect.has_key(page):
                for adm_var in self.HeadersAdmBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],adm_var,True)
        elif str(self.id_group)==str(id_group_monitor):
            # Monitor mode
            if self.HeadersUsrBySect.has_key(page):
                for var in self.HeadersUsrBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],var,False)
            if self.HeadersMonBySect.has_key(page):
                for var in self.HeadersMonBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],var,True)
            if self.HeadersAdmBySect.has_key(page):
                for adm_var in self.HeadersAdmBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],adm_var,True)
        elif str(self.id_group)==str(id_group_auditor):
            # Auditor mode
            if self.HeadersUsrBySect.has_key(page):
                for var in self.HeadersUsrBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],var,False)
            if self.HeadersMonBySect.has_key(page):
                for var in self.HeadersMonBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],var,False)
            if self.HeadersAdmBySect.has_key(page):
                for adm_var in self.HeadersAdmBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],adm_var,False)
        else:
            ## User mode
            if self.HeadersMonBySect.has_key(page):
                for mon_var in self.HeadersMonBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],mon_var,False)
                    for usr_var in self.HeadersUsrBySect[page]:
                        if self.GetInputFieldValue(wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),mon_var)):
                            self.EnableCrfField(self.FormName[page-1],usr_var,False,self.HeadersMonPattern[mon_var])
                        else:
                            self.EnableCrfField(self.FormName[page-1],usr_var,True,self.HeadersMonPattern[mon_var])
            if self.HeadersAdmBySect.has_key(page):
                for adm_var in self.HeadersAdmBySect[page]:
                    self.EnableCrfField(self.FormName[page-1],adm_var,False)
        if str(self.id_group)==str(id_group_admin):
            try: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),'rnd_trialcode').SetEditable(True)
            except: pass
            try: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),'rnd_datetime').SetEditable(True)
            except: pass
            try: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),'rnd_arm').SetEditable(True)
            except: pass
        else:
            try: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),'rnd_trialcode').SetEditable(False)
            except: pass
            try: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),'rnd_datetime').SetEditable(False)
            except: pass
            try: wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]),'rnd_arm').SetEditable(False)
            except: pass
        list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[page-1]).GetChildren()
        self.ColorizeFields(page-1,list,self.contents)
        wx.EndBusyCursor()


    def OnButtonAddnewButton(self, event):
        event.Skip()
        self.DoAddNew()
        
    def DoAddNew(self):
        wx.BeginBusyCursor()
        self.dont_trigger_field_change=True
        self.preloadCombo={}
        self.preloadComboMatch={}
        self.preloadComboFulfilled={}
        
        #Riempimento campi con inizializzazione chiave
        self.id_row=0
        self.id_rowinstance=self.parent.id_instance
        self.id_user=self.id_logged_user
        self.rap=self.uap

        #salvataggio dati caricati per il saverecord
        self.curcycle=1
        self.contents={}
        self.ifrm=1

        self.cur_group_shortcode=""
        
        maxval=0
        if self.keymode in ("private","public"):
            #contents_dictionary
            dictionary={}
            cur.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary")
            for row in cur:
                dictionary[str(row[0])+"-"+str(row[1])]=row[2]
                
            if self.keymode == "private":
                # qui non c'e' il filtraggio rows.status=0 perche' devo intercettare tutte le schede anche costruite dall'utente e poi modificate eventualmente da un altro (amministratore o altri), quindi devo considerare anche quelle sostituite (e per politica anche quelle cancellate xche non so distinguere tra quelle cancellate dall'utente e quelle cancellate da altri)
                query="SELECT id_dictionary FROM users,rows,contents_index WHERE users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND users.id_user="+str(self.id_user)+" AND rows.id_header="+str(self.id_crf)
            else:
                query="SELECT id_dictionary FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND rows.id_header="+str(self.id_crf)
            maxval=0
            cur.execute(query)
            for row in cur:
                cData=""
                if dictionary.has_key(str(self.hashHeaders[self.keyname])+"-"+str(row[0])):
                    cData=dictionary[str(self.hashHeaders[self.keyname])+"-"+str(row[0])]
                if not self.parent.trans.has_key(cData):
                    dData=HB_DecryptOne(self.digestkey,cData,"latin-1")
                    self.parent.trans[cData]=dData
                else:
                    dData=self.parent.trans[cData]
                try: 
                    if maxval<int(dData): maxval=int(dData)
                except: pass
        else:
            #keymode=Groups
            lstfields=[]
            lstfields.append("group_shortcode")
            lstfields.append(self.keyname)
            strlstfields="group_shortcode"+","+self.keyname
            con_m = sqlite.connect(":memory:")
            cur_m = con_m.cursor()
            square(con,con_m,"search",lstfields,self.id_group,self.id_crf)
            enc_group_shortcode=HB_EncryptOne(self.digestkey,self.group_shortcode)
            sql="SELECT id_row,id_user,id_instance,date_time,"+strlstfields+" FROM search WHERE group_shortcode='"+enc_group_shortcode+"'"
            cur_m.execute(sql)
            cols = [i[0] for i in cur_m.description]; hc={}; i=0
            for col in cols: hc[col]=i; i=i+1
            for row in cur_m:
                if not self.parent.trans.has_key(row[hc[self.keyname]]):
                    dData=HB_DecryptOne(self.digestkey,row[hc[self.keyname]],"latin-1")
                    self.parent.trans[row[hc[self.keyname]]]=dData
                else:
                    dData=self.parent.trans[row[hc[self.keyname]]]
                curval=0
                try: curval=int(dData)
                except: pass
                if curval>maxval: maxval=curval
            
        keyvalue=str(maxval+1)
        
        #keyvalue padding - begin
        try:
            if self.arrHeadersTypedesc[self.keyname][:6]=="format":
                if keyvalue!="":
                    paddingformat=self.arrHeadersTypedesc[self.keyname][self.arrHeadersTypedesc[self.keyname].find("=")+1:]
                    keyvalue=(paddingformat+str(keyvalue))[-len(paddingformat):]
        except: pass
        #keyvalue padding - end
        
        if DEBUG_MODE: print "new keyvalue="+self.group_shortcode+"-"+keyvalue
        self.curkeyvalue=keyvalue
        #assegnazione chiave
        self.contents[self.keyname]=keyvalue
        self.samename={}
        for ifrm in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
            self.PopulateFieldListDefault(ifrm,list,keyvalue,self.contents)
        #Attivazione di tutti i campi calcolati
        for ifrm in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
            self.BindEventsFieldList(ifrm,list)
        #Attivazione pagina 1
        self.frameCrf.Scroll(0,0)
        wx.xrc.XRCCTRL(self.frameRecordset,'choice').SetSelection(0)
        for i in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            if wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetSelection()==i:
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(10, 10))
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(True)
                self.frameCrf.SetVirtualSize(wx.Size(SCREEN_X-50,self.FormHeight[i]/10))
                self.frameCrf.SetScrollbars(10,10,1,self.FormHeight[i]/10)
            else:
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(False)
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(5000, 0))
        if self.ActivateCyclicity:
            if self.Cyclicity[1]==0: 
                wx.xrc.XRCCTRL(self.frameRecordset,'cycles').Show(False)
                try: wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(False)
                except: pass
            else: 
                try: 
                    wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(True)
                    wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Scroll(0,0)
                except: pass
                wx.xrc.XRCCTRL(self.frameRecordset,'cycles').Show(True)
                if DEBUG_MODE: print "objCycles showed"
        self.ifrm=0
        self.MonitoringCheck(self.ifrm+1)
        if not DEBUG_MODE:
            try: exec(self.OnFormActivate[self.ifrm+1])
            except: pass
        else:
            print "Executing OnActivate script"
            exec(self.OnFormActivate[self.ifrm+1])
        self.loadedcontents={}  #in questo modo salvo anche in caso di pulsante Save con tutti default
        
        #exec(self.OnFormActivate[self.ifrm+1])
        #Visibilita' buttonDelete, buttonSave e buttonCommit
        if GetSetting(self.parent,'hide_forbidden_buttons')=="1":
            checkrap=False
            if len(self.rap.split(";"))<3: checkrap=True
            elif str(self.id_group) in self.rap.split(";")[2].split(","): checkrap="True"
            if checkrap:
                self.buttonDelete.Show(True)
            else:
                self.buttonDelete.Show(False)
            checkrap=False
            if len(self.rap.split(";"))<3: checkrap=True
            elif str(self.id_group) in self.rap.split(";")[1].split(","): checkrap="True"
            if checkrap:
                self.buttonSave.Show(True)
                self.buttonCommit.Show(True)
            else:
                self.buttonSave.Show(False)
                self.buttonCommit.Show(False)
        #Apertura Scheda
        self.frameLogin.SetPosition(wx.Point(5000,0))
        self.frameSearch.SetPosition(wx.Point(5000,0))
        self.frameRecordset.SetPosition(wx.Point(0,0))
        self.EnableDisableButtons(3)

        self.dont_trigger_field_change=False
        self.lastfocusvariable=""
        self.lastfocusvariableobj=None
        wx.EndBusyCursor()

    def OnChoice(self, event):
        if not self.CrfPageEnabled[wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetSelection()]:
            event.Skip(False)
            wx.xrc.XRCCTRL(self.frameRecordset,'choice').SetSelection(0)
            wx.MessageBox(TT("This Page has been disabled"), TT("Invalid Page"), wx.ICON_ERROR | wx.OK, self)
            return
            
        event.Skip()

        if self.ActivateCyclicity:
            if self.Cyclicity[self.ifrm+1]>0: 
                self.SaveCycle(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1)
            else:
                ifrm=self.ifrm
                list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
                self.SaveFieldList(list,0)
        
        self.frameCrf.Scroll(0,0)
        for i in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            if wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetSelection()==i:
                self.ifrm=i
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(10, 10))
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(True)
                self.frameCrf.SetVirtualSize(wx.Size(SCREEN_X-50,self.FormHeight[i]/10))
                self.frameCrf.SetScrollbars(10,10,1,self.FormHeight[i]/10)
                if self.ActivateCyclicity:
                    if self.Cyclicity[i+1]==0: 
                        wx.xrc.XRCCTRL(self.frameRecordset,'cycles').Show(False)
                        try: wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(False)
                        except: pass
                        if DEBUG_MODE: print "objCycles hidden"
                    else:
                        try: 
                            wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(True)
                            wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Scroll(0,0)
                        except: pass
                        objCycles=wx.xrc.XRCCTRL(self.frameRecordset,'cycles') 
                        objCycles.Show(True)
                        if DEBUG_MODE: print "objCycles showed"
                        showallcycles=True
                        if GetSetting(self.parent,"show_all_cycles")=="0": showallcycles=False
                        firstemptycycle=self.FindFirstEmptyCycle()
                        self.curcycle=max(1,(firstemptycycle-1))
                        if DEBUG_MODE: print "curcycle="+str(self.curcycle)
                        wx.xrc.XRCCTRL(self.frameRecordset,'cycles').SetSelection(self.curcycle-1)
                        for t in range(100):
                            if t<self.Cyclicity[i+1] and (t<firstemptycycle or showallcycles):
                                try: objCycles.ShowItem(t,True)
                                except: pass
                            else:
                                try: objCycles.ShowItem(t,False)
                                except: pass
                        self.LoadCycle(self.curcycle)
                self.MonitoringCheck(self.ifrm+1)
                if not DEBUG_MODE:
                    try: exec(self.OnFormActivate[self.ifrm+1])
                    except: pass
                else:
                    print "Executing OnActivate script"
                    exec(self.OnFormActivate[self.ifrm+1])
                #exec(self.OnFormActivate[self.ifrm+1])
            else:
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(False)
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(5000, 0))
        self.lastfocusvariable=""
        self.lastfocusvariableobj=None
        self.frameCrf.Layout()
        self.OnWindowResize(None)

#    def OnScrollUpFrmCycles(self,event):
#        wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Scroll(0,0)
#        #wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Update()
#        print "scrollup"
#    def OnScrollDnFrmCycles(self,event):
#        wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Scroll(0,100)
#        #wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Update()
#        print "scrolldown"

    def OnChangeCycle(self,event):
        self.frameCrf.Scroll(0,0)
        objCycles = event.GetEventObject()
        idx = objCycles.GetSelection()+1
        if idx!=self.curcycle:
            wx.BeginBusyCursor()
            self.SaveCycle(self.curcycle)
            self.curcycle=idx
            self.LoadCycle(self.curcycle)
            wx.EndBusyCursor()
        self.MonitoringCheck(self.ifrm+1)
        if not DEBUG_MODE:
            try: exec(self.OnFormActivate[self.ifrm+1])
            except: pass
        else:
            print "Executing OnActivate script"
            exec(self.OnFormActivate[self.ifrm+1])
        self.lastfocusvariable=""
        self.lastfocusvariableobj=None
        self.OnWindowResize(None)
      
    def SaveRecord(self):
        wx.BeginBusyCursor()
        import time
#        while len(self.parent.p2p_defUpdate)>0:
#            print "SaveRecord sleeping"
#            time.sleep(0.5)
            
        self.parent.syncroTaskPause=True
        
        now = datetime.datetime.utcnow()
        ts=now.strftime("%d/%m/%Y, %H:%M:%S")
        ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
        #if self.ActivateCyclicity:
        #    if self.Cyclicity[self.ifrm+1]>0: self.SaveCycle(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1)
        
        # la chiamata che segue e' ancora obbligatoria sebbene ridondante perche' in alcuni plugin c'e' la chiamata al saverecord senza la precedente presaverecord
        self.PreSaveRecord()
        
        #dummy save test - begin
        dummysave=True
        if self.anyrecordchange:
            dummysave=False
            self.anyrecordchange=False
        else:
            checkedkeys={}
            for keycont in self.contents:
                if self.loadedcontents.has_key(keycont):
                    if self.contents[keycont]!=self.loadedcontents[keycont]:
                        #print "updated "+keycont+": '"+self.loadedcontents[keycont]+"' -> '"+self.contents[keycont]+"'"
                        dummysave=False
                else:
                    dummysave=False
                    #print "added "+keycont+": '"+self.contents[keycont]+"'"
                if not dummysave: break
                checkedkeys[keycont]=True
            if dummysave:
                for keycont in self.loadedcontents:
                    if not checkedkeys.has_key(keycont):
                        dummysave=False
                        #print "removed "+keycont+": '"+self.loadedcontents[keycont]+"'"
        #dummy save test - end
        
        if not dummysave:
            self.parent.syncroTaskPause=True
            import time
            while self.parent.syncroTaskUpdating:
                time.sleep(1)
                if DEBUG_MODE: print "waiting synchro standby..."
            cur.execute("BEGIN TRANSACTION")
                    
            id_oldrow=0
            if self.id_row<>0:
                id_oldrow=self.id_row
            tmprap=self.rap
            if not (tmprap): tmprap=self.uap
            if not (tmprap): tmprap=""
            #new id_row - begin
            cur.execute("SELECT max(id_row) FROM rows")
            row = cur.fetchone()
            if row==None:
                self.id_row=1
            else:
                if row[0]==None: self.id_row=1 
                else: self.id_row=row[0]+1
            #new id_row - end

            cur.execute("INSERT INTO rows(id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(self.id_row)+","+str(self.id_logged_user)+","+str(self.id_crf)+",'"+tmprap+"','"+ts_db+"',0,0,0,"+str(self.parent.id_instance)+")")
            self.SaveInDatabase()
            cur.execute("UPDATE rows SET status="+str(self.id_row)+",status_user="+str(self.id_logged_user)+",status_instance="+str(self.parent.id_instance)+" WHERE id_row="+str(id_oldrow)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance)+" AND id_header="+str(self.id_crf))
            self.id_rowinstance=self.parent.id_instance
            self.id_user=self.id_logged_user

            retry_commit=True
            while retry_commit:
                try:
                    if DEBUG_MODE: print "trying commit..."
                    cur.execute('COMMIT TRANSACTION')
                    retry_commit=False
                except: 
                    outfile = open("quit_synchro.cmd", 'w')
                    outfile.write("\n")
                    outfile.close()
                    time.sleep(1)
            self.parent.syncroTaskPause=False

            self.parent.syncrotablesupd["rows"]=True
            
            self.loadedcontents={}
            for keycont in self.contents:
                self.loadedcontents[keycont]=self.contents[keycont] #in questo caso salvo i soli record realmente modificati
            
            if not self.parent.quickreportRunning:
                if DEBUG_MODE: print "starting quickreport"
                self.QuickReporterTask = QuickReporter(self.parent,self,self.digestkey)
                self.QuickReporterTask.start()
            else:
                if DEBUG_MODE: print "quickreport already running"

        self.parent.syncroTaskPause=False
        self.lstFound.Clear()
        wx.EndBusyCursor()
        wx.MessageBox("Record saved\nid transaction="+str(self.id_logged_user)+"-"+str(self.id_row)+"-"+str(self.parent.id_instance), "Info", wx.ICON_INFORMATION | wx.OK, self)

    def PopulateComboAfterSave(self,ifrm,list):
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                self.PopulateComboAfterSave(ifrm,sublist)
            else:
                if self.IsInputField(list[i]):
                    curName=list[i].GetName()
                    if list[i].GetClassName() in ("wxComboBox"):
                        self.preloadComboFulfilled[curName]=True
                        #oldValue=self.GetInputFieldValue(list[i])
                        self.FulfillComboList(list[i],curName)
                        #self.SetInputFieldValue(list[i],oldValue)

    def OnButtonCloneButton(self, event):
        event.Skip()
        #if wx.MessageBox("Are you sure you wish to\n clone this record (all the pages)\n and open the clone?", "Confirmation request", wx.YES_NO) == wx.YES:
        if wx.MessageBox(TT("Are you sure you wish to clone this record (all the pages) and open the clone?"), TT("Record Clone"), wx.YES_NO) == wx.YES:
            self.PreSaveRecord()
            self.SaveRecord()
            now = datetime.datetime.utcnow()
            ts=now.strftime("%d/%m/%Y, %H:%M:%S")
            ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
            id_oldrow=0
            if self.id_row<>0:
                id_oldrow=self.id_row

                tmprap=self.rap
                if not (tmprap): tmprap=self.uap
                if not (tmprap): tmprap=""
                #new id_row - begin
                cur.execute("SELECT max(id_row) FROM rows")
                row = cur.fetchone()
                if row==None:
                    self.id_row=1
                else:
                    if row[0]==None: self.id_row=1 
                    else: self.id_row=row[0]+1
                #new id_row - end
                cur.execute("INSERT INTO rows(id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(self.id_row)+","+str(self.id_logged_user)+","+str(self.id_crf)+",'"+tmprap+"','"+ts_db+"',0,0,0,"+str(self.parent.id_instance)+")")
                if GetSetting(self.parent,"clone_cyclic_variables")!="0":
                    cur.execute("INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) SELECT "+str(self.id_row)+","+str(self.id_logged_user)+",id_header,id_dictionary,"+str(self.parent.id_instance)+",id_cycle FROM contents_index WHERE id_row="+str(id_oldrow)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance))
                else:
                    cur.execute("INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) SELECT "+str(self.id_row)+","+str(self.id_logged_user)+",id_header,id_dictionary,"+str(self.parent.id_instance)+",id_cycle FROM contents_index WHERE id_row="+str(id_oldrow)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance)+" AND id_cycle=0")
                    
                self.id_user=self.id_logged_user
                self.id_rowinstance=self.parent.id_instance
                self.parent.syncrotablesupd["rows"]=True
                self.OpenCurRecord()

                txt=GetSetting(self.parent,"on_clone")
                if txt!="":
                    exec(txt)
                    #except: pass

    def OnButtonSaveButton(self, event):
        event.Skip()
        checkrap=False
        if len(self.rap.split(";"))<2: checkrap=True
        elif str(self.id_group) in self.rap.split(";")[1].split(","): checkrap="True"
        if checkrap:
            if wx.MessageBox(TT("Are you sure you wish to save this record (all the pages)?"), TT("Save record"), wx.YES_NO) == wx.YES:
                self.PreSaveRecord()
                self.lstFound.Clear()
                cur.execute("SELECT validation FROM headers WHERE id_header="+str(self.id_crf))
                row = cur.fetchone()
                rule=None
                if row!=None: 
                    if row[0]!=None:
                        if len(row[0])>0: rule=row[0]
                outcome=True
                outcome_message=""
                if rule!=None:
                    #self.PreSaveRecord()
                    rule=rule.replace("wx.xrc.XRCCTRL","XRCCTRL")
                    rule=rule.replace("XRCCTRL","wx.xrc.XRCCTRL")
                    if not DEBUG_MODE:
                        try: exec(rule)
                        except: pass
                    else:
                        print "Executing record validation script"
                        exec(rule)
                if not outcome:
                    wx.MessageBox(outcome_message, "Invalid data", wx.ICON_ERROR | wx.OK, self)
                else:
                    self.outcome=True
                    warnings={}
                    if not DEBUG_MODE:
                        try: 
                            import analysis
                            warnings=analysis.GetWarnings(self.contents)
                        except: pass
                    else:
                        print "Executing Warnings validation"
                        found=False
                        try:
                            import analysis
                            found=True
                        except ImportError: pass
                        if found: 
                            warnings=analysis.GetWarnings(self.contents)
                            if DEBUG_MODE: print "GetWarnings executed."
                    #Check dupes
                    lstNames,lstValues=self.FindDupes()
                    if lstNames!=None:
                        if len(lstValues)>0:
                            nwarn=len(warnings)+1
                            for elm in lstValues:
                                warnings[nwarn]=['warning','record',TT('possible duplicate'),elm]
                                nwarn=nwarn+1
                                
                    id_sections={}
                    for keycont in self.contents:
                        dcurValue=self.contents[keycont]
                        curName=keycont
                        cycle="0"
                        if keycont.find("#")>=0: 
                            curName=keycont[:keycont.find("#")]
                            cycle=keycont[keycont.find("#")+1:]
                        if curName in self.arrHeadersTypedesc:
                            if self.arrHeadersTypedesc[curName]=="date":
                                if not valid_date_or_null(self.contents[keycont]):
                                    if len(id_sections)==0:
                                        id_section=0
                                        cur.execute("SELECT label FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
                                        for row in cur:
                                            id_section+=1
                                            id_sections[id_section]=row[0]
                                    cur.execute("select id_section from headers where description='"+curName+"'")
                                    row = cur.fetchone()
                                    if row!=None:
                                        if row[0]!=None:
                                            pagetitle=id_sections[row[0]]
                                            if cycle!="0": pagetitle+=" (#"+cycle+")"
                                            warnings[len(warnings)+1]=['warning',pagetitle,curName,TT('incorrect date')]
                        
                    if len(warnings)>0:
                        #self.PreSaveRecord()
                        self.outcome=False
                        top = WarningsFrame(self,"Warnings",warnings)
                        top.ShowModal()
                        top.Destroy()
                        
                    if self.outcome:
                        self.SaveRecord()
                        
                        #combo begin
                        wx.BeginBusyCursor()
                        self.preloadCombo={}
                        self.preloadComboMatch={}
                        for ifrm in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
                            list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
                            self.PopulateComboAfterSave(ifrm,list)
                        self.preloadComboFulfilled={}
                        wx.EndBusyCursor()
                        #combo end

                        self.lstFound.Clear()
        else:
            wx.MessageBox("Update denied", "Authorization failed", wx.ICON_ERROR | wx.OK, self)

    def OnButtonCommitButton(self, event):
        event.Skip()
        checkrap=False
        if len(self.rap.split(";"))<2: checkrap=True
        elif str(self.id_group) in self.rap.split(";")[1].split(","): checkrap="True"
        if checkrap:
            if wx.MessageBox(TT("Are you sure you wish to save this record (all the pages) and exit?"), TT("Save record and exit"), wx.YES_NO) == wx.YES:
                self.PreSaveRecord()
                self.lstFound.Clear()
                cur.execute("SELECT validation FROM headers WHERE id_header="+str(self.id_crf))
                row = cur.fetchone()
                rule=None
                if row[0]!=None: rule=row[0]
                if rule!=None:
                    if len(rule)==0: rule=None
                outcome=True
                outcome_message=""
                if rule!=None:
                    #self.PreSaveRecord()
                    rule=rule.replace("wx.xrc.XRCCTRL","XRCCTRL")
                    rule=rule.replace("XRCCTRL","wx.xrc.XRCCTRL")
                    if not DEBUG_MODE:
                        try: exec(rule)
                        except: pass
                    else:
                        print "Executing record validation script"
                        exec(rule)
                if not outcome:
                    wx.MessageBox(outcome_message, "Invalid data", wx.ICON_ERROR | wx.OK, self)
                else:
                    self.outcome=True
                    warnings={}
                    if not DEBUG_MODE:
                        try: 
                            import analysis
                            warnings=analysis.GetWarnings(self.contents)
                        except: pass
                    else:
                        print "Executing Warnings validation"
                        found=False
                        try:
                            import analysis
                            found=True
                        except ImportError: pass
                        if found: 
                            warnings=analysis.GetWarnings(self.contents)
                        
                    #Check dupes
                    lstNames,lstValues=self.FindDupes()
                    if lstNames!=None:
                        if len(lstValues)>0:
                            nwarn=len(warnings)+1
                            for elm in lstValues:
                                warnings[nwarn]=['warning','record',TT('possible duplicate'),elm]
                                nwarn=nwarn+1
                        
                    id_sections={}
                    for keycont in self.contents:
                        dcurValue=self.contents[keycont]
                        curName=keycont
                        cycle="0"
                        if keycont.find("#")>=0: 
                            curName=keycont[:keycont.find("#")]
                            cycle=keycont[keycont.find("#")+1:]
                        if curName in self.arrHeadersTypedesc:
                            if self.arrHeadersTypedesc[curName]=="date":
                                if not valid_date_or_null(self.contents[keycont]):
                                    if len(id_sections)==0:
                                        id_section=0
                                        cur.execute("SELECT label FROM eforms WHERE id_header="+str(self.id_crf)+" AND status=0 AND pos>0 AND (id_group="+str(self.id_group)+" OR (id_group=0 AND pos NOT IN (SELECT pos FROM eforms WHERE id_group="+str(self.id_group)+"))) ORDER BY POS,id_eform")
                                        for row in cur:
                                            id_section+=1
                                            id_sections[id_section]=row[0]
                                    cur.execute("select id_section from headers where description='"+curName+"'")
                                    row = cur.fetchone()
                                    if row!=None:
                                        if row[0]!=None:
                                            pagetitle=id_sections[row[0]]
                                            if cycle!="0": pagetitle+=" (#"+cycle+")"
                                            warnings[len(warnings)+1]=['warning',pagetitle,curName,TT('incorrect date')]
                                    
                    if len(warnings)>0:
                        #self.PreSaveRecord()
                        self.outcome=False
                        top = WarningsFrame(self,"Warnings",warnings)
                        top.ShowModal()
                        top.Destroy()
                        
                    if self.outcome:
                        self.SaveRecord()
                        
                        if GetSetting(self.parent,'pool_mode')!="1" or len(self.arrFound)>1:
                            self.lstFound.Clear()
                            self.frameLogin.SetPosition(wx.Point(5000,0))
                            self.frameSearch.SetPosition(wx.Point(0,0))
                            self.frameRecordset.SetPosition(wx.Point(5000,0))
                            self.EnableDisableButtons(2)
                        else:
                            self.DoLogout()
        else:
            wx.MessageBox("Update denied", "Authorization failed", wx.ICON_ERROR | wx.OK, self)

    def OnButtonCancelButton(self, event):
        event.Skip()
        if wx.MessageBox(TT("Are you sure you wish to quit without saving this record (all the pages)?"), TT("Quit record"), wx.YES_NO) == wx.YES:
            if GetSetting(self.parent,'pool_mode')!="1" or len(self.arrFound)>1:
                self.frameLogin.SetPosition(wx.Point(5000,0))
                self.frameSearch.SetPosition(wx.Point(0,0))
                self.frameRecordset.SetPosition(wx.Point(5000,0))
                self.EnableDisableButtons(2)
            else:
                self.DoLogout()
                
    def OnButtonDeleteButton(self, event):
        event.Skip()
        checkrap=False
        if len(self.rap.split(";"))<3: checkrap=True
        elif str(self.id_group) in self.rap.split(";")[2].split(","): checkrap="True"
        if checkrap:
            if wx.MessageBox(TT("Are you sure you wish to delete this record (all the pages)?"), TT("Delete record"), wx.YES_NO) == wx.YES:
                now = datetime.datetime.utcnow()
                ts=now.strftime("%d/%m/%Y, %H:%M:%S")
                ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
                id_oldrow=0
                if self.id_row<>0:
                    id_oldrow=self.id_row
                tmprap=self.rap
                if not (tmprap): tmprap=self.uap
                if not (tmprap): tmprap=""
                #new id_row - begin
                cur.execute("SELECT max(id_row) FROM rows")
                row = cur.fetchone()
                if row==None:
                    self.id_row=1
                else:
                    if row[0]==None: self.id_row=1 
                    else: self.id_row=row[0]+1
                #new id_row - end
                cur.execute("INSERT INTO rows(id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(self.id_row)+","+str(self.id_logged_user)+","+str(self.id_crf)+",'"+tmprap+"','"+ts_db+"',-1,0,0,"+str(self.parent.id_instance)+")")
                #cur.execute("SELECT last_insert_rowid()")
                #self.id_row=cur.fetchone()[0]
                cur.execute("UPDATE rows SET status="+str(self.id_row)+",status_user="+str(self.id_logged_user)+",status_instance="+str(self.parent.id_instance)+" WHERE id_row="+str(id_oldrow)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance)+" AND id_header="+str(self.id_crf))
                #cur.execute("UPDATE rows SET status="+str(self.id_row)+",status_user="+str(self.id_logged_user)+",status_instance="+str(self.parent.id_instance)+",date_time=date_time||' -> '||'"+ts_db+"' WHERE id_row="+str(id_oldrow)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance))
                wx.MessageBox("Record deleted\n"+"id transaction="+str(self.id_logged_user)+"-"+str(self.id_row)+"-"+str(self.parent.id_instance), "Info", wx.ICON_INFORMATION | wx.OK, self)

                if not self.parent.quickreportRunning:
                    self.QuickReporterTask = QuickReporter(self.parent,self,self.digestkey)
                    self.QuickReporterTask.start()

                self.parent.syncrotablesupd["rows"]=True
                
                self.lstFound.Clear()
                self.frameLogin.SetPosition(wx.Point(5000,0))
                self.frameSearch.SetPosition(wx.Point(0,0))
                self.frameRecordset.SetPosition(wx.Point(5000,0))
                self.EnableDisableButtons(2)
        else:
            wx.MessageBox(TT("Delete denied"), TT("Authorization failed"), wx.ICON_ERROR | wx.OK, self)

    def IsInputField(self,elm):
        ret=False
        try:
            if elm.GetClassName() in ("wxTextCtrl","wxSpinCtrl","wxDatePickerCtrl","wxCalendarCtrl","wxChoice","wxRadioButton","wxCheckBox","wxComboBox","wxRadioBox","wxPyComboCtrl"):
                ret=True
        except: pass
        return ret

    def IsInputFieldEditable(self,elm):
        ret=False
        if self.IsInputField(elm):
            if elm.GetClassName() in ("wxTextCtrl"):
                ret=elm.IsEditable()
            else:
                ret=True
        return ret

    def GetInputFieldChangeEventId(self,elm):
        ret=""
        if elm.GetClassName() in ("wxTextCtrl","wxSpinCtrl","wxDatePickerCtrl","wxCalendarCtrl","wxComboBox","wxPyComboCtrl"):
            ret=wx.EVT_TEXT
        if elm.GetClassName() in ("wxCheckBox"):
            ret=wx.EVT_CHECKBOX
        if elm.GetClassName() in ("wxRadioButton"):
            ret=wx.EVT_RADIOBUTTON
        if elm.GetClassName() in ("wxChoice","wxRadioBox"):
            ret=wx.EVT_CHOICE
        return ret

    def IsButton(self,elm):
        if elm.GetClassName() in ("wxButton"):
            return True
        else:
            return False

    def IsDate(self,testValue):
        ret=False
        try:
            import time
            a=time.strptime(testValue, "%d/%m/%Y")
            ret=True
        except: pass
        return ret
        
    def GetPageFieldValue(self,pagename,fieldname):
        return self.GetInputFieldValue(wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,pagename),fieldname)).strip()
        
    def GetInputFieldValue(self,elm):
        if not self.IsInputField(elm): return ""
        #print elm.GetName()+" ("+elm.GetClassName()+")"
        curValue=""
        if elm.GetName() not in self.arrHeadersTypedesc: self.arrHeadersTypedesc[elm.GetName()]=None
        if elm.GetClassName() in ("wxTextCtrl","wxSpinCtrl","wxComboBox","wxPyComboCtrl"):
            if elm.GetValue():
                curValue=elm.GetValue()
                if curValue=="  /  /    ": curValue=""
                if curValue=="  /    ": curValue=""
                if self.arrHeadersTypedesc[elm.GetName()]=="date":
                    if len(curValue)!=10: curValue=""
                try: curValue=str(curValue).strip()
                except: pass
                #if self.arrHeadersTypedesc[elm.GetName()]!=None:
                    #if self.arrHeadersTypedesc[elm.GetName()][:4]=="mask":
                        #if DEBUG_MODE: print "(Get): "+elm.GetName()+"=\""+curValue+"\""

        if elm.GetClassName() in ("wxDatePickerCtrl","wxCalendarCtrl"):
            try:
                if elm.GetValue():
                    curValue=elm.GetValue().Format("%d/%m/%Y")
            except:
                pass
        if elm.GetClassName() in ("wxRadioBox"):
            curValue=elm.GetStringSelection()
        if elm.GetClassName() in ("wxChoice"):
            if not self.multidict:
                if elm.GetStringSelection():
                    curValue=elm.GetStringSelection()
            else:
                tmpValue=""
                if elm.GetStringSelection():
                    tmpValue=elm.GetStringSelection()
                    for opt in self.arrHeadersComboPreload[elm.GetName()]:
                        tmpOpt=""
                        try: tmpOpt=opt.split("|")[self.FormIdxDict[self.ifrm]]
                        except: tmpOpt=opt
                        if tmpValue==tmpOpt:
                            curValue=opt
                            break
        if elm.GetClassName() in ("wxRadioButton"):
            if elm.GetValue():
                curValue=elm.GetLabel()
        if elm.GetClassName() in ("wxCheckBox"):
            if elm.Is3State():
                if elm.Get3StateValue()==wx.CHK_CHECKED:
                    curValue="1"
            else:
                if elm.IsChecked():
                    curValue="1"
        #print elm.GetName()+" = "+curValue
        return curValue

    def GetSqlValue(self,value):
        sqlValue=""
        if value==None: value=""
        for i in range(len(value)):
            if value[i]=="'":
                sqlValue+="''"
            else:
                sqlValue+=value[i]
        return sqlValue

    def GetCorrectedDate(self,value):
        value=value.lower().strip()
        value=value.replace("/  /","")
        value=value.replace("00/00/0000","")
        if value=="": return ""
        value=value.replace("gennaio","01/")
        value=value.replace("febbraio","02/")
        value=value.replace("marzo","03/")
        value=value.replace("aprile","04/")
        value=value.replace("maggio","05")
        value=value.replace("giugno","06/")
        value=value.replace("luglio","07/")
        value=value.replace("agosto","08/")
        value=value.replace("settembre","09/")
        value=value.replace("ottobre","10/")
        value=value.replace("novembre","11/")
        value=value.replace("dicembre","12/")
        value=value.replace("anno","")
        value=value.replace("nel","")
        
        value=value.replace(" ","")
        value=value.replace("-","/")
        value=value.replace("../","0/")
        value=value.replace(".","/")
        value=value.replace("\\","/")
        value=value.replace("&","/")
        value=value.replace("//","/")
        if len(value)>0:
            if value[-1:]=="/": value=value[:-1]
            if value[:1]=="/": value=value[1:]
        lst=value.split("/")
        if len(lst)==1:
            if len(value)<=4:
                val=0
                try: val=int(value)
                except: pass
                if val>=1900 and val<=2020: 
                    value="00/00/"+value
                    lst=value.split("/")
            elif len(value)==6:
                gg=0
                mm=0
                aaaa=0
                try:
                    gg=int(value[0:2])
                    mm=int(value[2:4])
                    aaaa=int(value[4:6])
                except: pass
                if gg>=1 and gg<=31 and mm>=1 and mm<=12:
                    value=value[0:2]+"/"+value[2:4]+"/"+value[4:6]
                    lst=value.split("/")
            elif len(value)==8:
                gg=0
                mm=0
                aaaa=0
                try:
                    gg=int(value[0:2])
                    mm=int(value[2:4])
                    aaaa=int(value[4:8])
                except: pass
                if gg>=1 and gg<=31 and mm>=1 and mm<=12 and aaaa>=1900 and aaaa<=2020:
                    value=value[0:2]+"/"+value[2:4]+"/"+value[4:8]
                    lst=value.split("/")
            elif len(value)==10:
                if value[2:3]=="7": value=value[0:2]+"/"+value[3:10]
                if value[5:6]=="7": value=value[0:5]+"/"+value[6:10]
                lst=value.split("/")
        if len(lst)==2: 
            if len(value)==9:
                if len(lst[0])==4:
                    gg=0
                    mm=0
                    aaaa=0
                    try:
                        gg=int(value[0:2])
                        mm=int(value[2:4])
                        aaaa=int(value[5:9])
                    except: pass
                    if gg>=1 and gg<=31 and mm>=1 and mm<=12 and aaaa>=1900 and aaaa<=2020:
                        value=value[0:2]+"/"+value[2:4]+"/"+value[5:9]
                        lst=value.split("/")
                elif len(lst[1])==6:
                    gg=0
                    mm=0
                    aaaa=0
                    try:
                        gg=int(value[0:2])
                        mm=int(value[3:5])
                        aaaa=int(value[5:9])
                    except: pass
                    if gg>=1 and gg<=31 and mm>=1 and mm<=12 and aaaa>=1900 and aaaa<=2020:
                        value=value[0:2]+"/"+value[3:5]+"/"+value[5:9]
                        lst=value.split("/")
            elif len(value)==10:
                if value[2:3]=="7": value=value[0:2]+"/"+value[3:10]
                if value[5:6]=="7": value=value[0:5]+"/"+value[6:10]
                lst=value.split("/")
            else:
                value="00/"+value
                lst=value.split("/")
        if len(lst)==3:
            gg=lst[0]
            if len(gg)>2: gg=gg[:2]
            gg=("00"+gg)[-2:]
            val=0
            try: val=int(gg)
            except: pass
            #if val<1 or val>31: gg="15"
            mm=lst[1]
            if len(mm)>2: mm=mm[:2]
            mm=("00"+mm)[-2:]
            val=0
            try: val=int(mm)
            except: pass
            #if val<1 or val>12: mm="06"
            aaaa="0000"
            if len(lst[2])==4: 
                aaaa=lst[2]
            elif len(lst[2])==3:
                if lst[2]!="0": aaaa="1"+lst[2]
                else: aaaa="2"+lst[2]
            elif len(lst[2])==2:
                if lst[2]>="10": aaaa="19"+lst[2]
                else: aaaa="20"+lst[2]
            value=gg+"/"+mm+"/"+aaaa
        else:
            value=""
        return value
        
    def SetPageFieldVisibility(self,pagename,fieldname,showmode):
        wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,pagename),fieldname).Show(showmode)
        
    def SetPageFieldLabel(self,pagename,fieldname,value):
        wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,pagename),fieldname).SetLabel(str(value))
        
    def SetPageFieldValue(self,pagename,fieldname,value):
        self.SetInputFieldValue(wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,pagename),fieldname),str(value))
        
    def SetInputFieldValue(self,elm,value,iPg=-1):
        if not self.IsInputField(elm): return
        if value==None: value=""
        if value=="  /  /    ": value=""
        if value=="00/00/0000": value=""
        value=value.strip()
        #if DEBUG_MODE and value!="": print "Writing "+elm.GetName()+"="+value
        if elm.GetClassName() in ("wxTextCtrl","wxComboBox"):
            if self.arrHeadersTypedesc.has_key(elm.GetName()):
                if not self.arrHeadersTypedesc[elm.GetName()] is None:
                    #print elm.GetName(),self.arrHeadersTypedesc[elm.GetName()],self.arrHeadersTypedesc[elm.GetName()]
                    if self.arrHeadersTypedesc[elm.GetName()] in ("date","mask=##/##/####"):
                        value=self.GetCorrectedDate(value)
                    elif self.arrHeadersTypedesc[elm.GetName()]=="integer":
                        if value=="":
                            value=None
                        else:
                            try: value=int(value)
                            except: pass
                    elif self.arrHeadersTypedesc[elm.GetName()]=="float":
                        if value=="":
                            value=None
                        else:
                            try: value=float(value)
                            except: pass
                    elif self.arrHeadersTypedesc[elm.GetName()][:6]=="number":
                        if value=="":
                            value=None
                        else:
                            mask=self.arrHeadersTypedesc[elm.GetName()][self.arrHeadersTypedesc[elm.GetName()].find("=")+1:]
                            intpart=int(mask[:mask.find(".")])
                            decpart=int(mask[mask.find(".")+1:])
                            if decpart==0:
                                try: value=int(value)
                                except: pass
                            else:
                                try: value=float(value)
                                except: pass
                    elif self.arrHeadersTypedesc[elm.GetName()][:4]=="mask":
                        pass
                        #if value=="": 
                        #    if elm._masklength==7: value="  /    "
                        #else:
                        paddingformat=self.arrHeadersTypedesc[elm.GetName()][self.arrHeadersTypedesc[elm.GetName()].find("=")+1:]
                        if paddingformat.find("{")>=0:
                            for t in range(9):
                                paddingformat=paddingformat.replace("#{"+str(t+1)+"}","#".ljust(t+1,"#"))
                        for splitter in [".",":","-"]:
                            if splitter in paddingformat and value.count(splitter)==1:
                                if value==splitter: value=""
                                if value!="":
                                    if "," not in paddingformat: value=value.replace(",",".")
                                    padpos=paddingformat.find(splitter)
                                    valpos=value.find(splitter)
                                    value=value[:valpos].zfill(padpos)+value[valpos:].ljust(elm._masklength-padpos-1,'0')
                    elif self.arrHeadersTypedesc[elm.GetName()][:6]=="format":
                        if value!="":
                            paddingformat=self.arrHeadersTypedesc[elm.GetName()][self.arrHeadersTypedesc[elm.GetName()].find("=")+1:]
                            value=(paddingformat+str(value))[-len(paddingformat):]
                            
                    if self.arrHeadersTypedesc[elm.GetName()][:4]!="mask":
                        try: elm.SetValue(unicode(value,"latin-1"))
                        except:
                            try: 
                                elm.SetValue(value)
                            except:
                                if DEBUG_MODE: print "Error writing value \""+value+"\" to field \""+elm.GetName()
                    else:   #masked
                        try: 
                            elm.SetValue(value)
                        except:
                            if DEBUG_MODE: print "Error writing value \""+value+"\" to field \""+elm.GetName()+"\" with mask \""+self.arrHeadersTypedesc[elm.GetName()]+"\""
                        #if value!="": 
                            #try: elm.SetValue(value)
                            #except: pass
                        #else: elm.ClearValue()
                else: elm.SetValue(value)
            else: elm.SetValue(value)
        if elm.GetClassName() in ("wxSpinCtrl"):
            val=0
            try: val=int(value)
            except: val=elm.GetMin()
            try: elm.SetValue(val)
            except: pass
        if elm.GetClassName() in ("wxPyComboCtrl"):
            elm.Value=self.GetCorrectedDate(value)
            elm.convert_to_wx_date()
        if elm.GetClassName() in ("wxDatePickerCtrl","wxCalendarCtrl"):
            if value:
                try:
                    day=value[0:2]
                    while day[0]=="0": day=day[1:]
                    month=value[3:5]
                    while month[0]=="0": month=month[1:]
                    year=value[6:10]
                    while year[0]=="0": year=year[1:]
                    elm.SetValue(wx.DateTimeFromDMY(int(day), int(month)-1, int(year)))
                except:
                    pass
            else:
                #elm.SetValue(wx.DateTime.Today())
                blankdt = wx.DateTime()
                elm.SetValue(blankdt)
        if elm.GetClassName() in ("wxRadioBox"):
            elm.SetStringSelection(value)
        if elm.GetClassName() in ("wxChoice"):
            if value!="":
                #print "trying to set "+elm.GetName()+" = "+value+" (current ontology: "+str(self.FormIdxDict[self.ifrm])+")"
                if not self.multidict:
                    elm.SetStringSelection(value)
                else:
                    for opt in self.arrHeadersComboPreload[elm.GetName()]:
                        if value==opt:
                            ifrm=self.ifrm
                            if iPg!=-1: ifrm=iPg
                            try: elm.SetStringSelection(opt.split("|")[self.FormIdxDict[ifrm]])
                            except: elm.SetStringSelection(value)
                            break
                        
                    #for opt in self.arrHeadersComboPreload[elm.GetName()]:
                        #tmpOpt=""
                        #try: tmpOpt=opt.split("|")[self.FormIdxDict[self.ifrm]]
                        #except: tmpOpt=opt
                        #print value,tmpOpt,value==tmpOpt
                        #if value==tmpOpt:
                            #ifrm=self.ifrm
                            #if iPg!=-1: ifrm=iPg
                            #print "new ontology: "+self.FormIdxDict[ifrm]+" --- value="+opt.split("|")[self.FormIdxDict[ifrm]]
                            #try: elm.SetStringSelection(opt.split("|")[self.FormIdxDict[ifrm]])
                            #except: elm.SetStringSelection(value)
                            #break
            else:
                elm.SetSelection(0)
                #try: elm.SetStringSelection("")
                #except: elm.SetSelection(0) 

        if elm.GetClassName() in ("wxRadioButton"):
            if value:
                if elm.GetLabel()==value:
                    elm.SetValue(True)
            else:
                elm.SetValue(False)
        if elm.GetClassName() in ("wxCheckBox"):
            if value:
                import string
                if string.lower(value)=="true" or value=="1":
                    elm.SetValue(True)
                else:
                    elm.SetValue(False)
            else:
                elm.SetValue(False)

    def OnButtonAutoMergeDupes(self,event):
        event.Skip()
        if wx.MessageBox(TT("Are you sure you wish to continue?"), TT("Dupes auto-merge"), wx.YES_NO) != wx.YES:
            return
        wx.BeginBusyCursor()
        self.DoAutoMergeDupes()
        self.lstFound.Clear()
        wx.EndBusyCursor()
        wx.MessageBox(TT("Done fixing errors"), TT("Dupes auto-merge"), wx.ICON_INFORMATION | wx.OK, self)

    def DoAutoMergeDupes(self):
        if GetSetting(self.parent,'automergedupes')=="0": return
    
        self.parent.syncroTaskPause=True
        import time
        while self.parent.syncroTaskUpdating:
            time.sleep(1)
            if DEBUG_MODE: print "waiting synchro standby..."
        
        cur.execute("BEGIN TRANSACTION")
        
        # duplicate headers descriptions
        corrections=True
        while corrections:
            corrections=False
            cur.execute("select min(id_header) as minIdHeader, max(id_header) as maxIdHeader, headers.description from headers, (select description from (select description,count(*) as quanti from headers group by description) as s1 where quanti>1) as s2 where headers.description=s2.description group by headers.description")
            for row in cur:
                corrections=True
                #print "update contents_index set id_header="+str(row[0])+" where id_header="+str(row[1])
                curw.execute("update contents_index set id_header="+str(row[0])+" where id_header="+str(row[1]))
                curw.execute("delete from headers where id_header="+str(row[1]))
        #duplicate rows records
        cur.execute("select max(dup) as maxdup from (select id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance, count(*) as dup from rows group by id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance)")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>1:
                try: cur.execute('DROP TABLE rows_nodup')
                except: pass
                cur.execute('CREATE TABLE rows_nodup (id_row INTEGER, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER, rap TEXT, hpath TEXT, status_user INTEGER, id_locking INTEGER, id_instance INTEGER, status_instance INTEGER)')
                cur.execute('INSERT INTO rows_nodup (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT DISTINCT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows')
                cur.execute('DELETE FROM rows')
                cur.execute('INSERT INTO rows (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows_nodup')
                cur.execute('DROP TABLE rows_nodup')
                print "dupes from rows removed."
        #duplicate contents_index records
        cur.execute("select max(dup) as maxdup from (select id_row, id_header, id_user, id_dictionary, id_instance, id_cycle, count(*) as dup from contents_index group by id_row, id_header, id_user, id_dictionary, id_instance, id_cycle)")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>1:
                try: cur.execute('DROP TABLE contents_index_nodup')
                except: pass
                cur.execute('CREATE TABLE contents_index_nodup (id_row INTEGER, id_user INTEGER, id_header INTEGER, id_dictionary INTEGER, id_instance INTEGER, id_cycle INTEGER)')
                cur.execute('INSERT INTO contents_index_nodup (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT DISTINCT id_row, id_header, id_user, id_dictionary, id_instance, id_cycle FROM contents_index')
                cur.execute('DELETE FROM contents_index')
                cur.execute('INSERT INTO contents_index (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT id_row, id_header, id_user, id_dictionary, id_instance, id_cycle FROM contents_index_nodup')
                cur.execute('DROP TABLE contents_index_nodup')
                print "dupes from contents_index removed."
                
        #perfect duplicate contents_dictionary records
        cur.execute("select max(dup) as maxdup from (select id_header, id_dictionary, data, count(*) as dup from contents_dictionary group by id_header, id_dictionary, data)")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>1:
                try: cur.execute('DROP TABLE contents_dictionary_nodup')
                except: pass
                cur.execute('CREATE TABLE contents_dictionary_nodup (id_header INTEGER, id_dictionary INTEGER, data BLOB)')
                cur.execute('INSERT INTO contents_dictionary_nodup (id_header, id_dictionary, data) SELECT DISTINCT id_header, id_dictionary, data FROM contents_dictionary')
                cur.execute('DELETE FROM contents_dictionary')
                cur.execute('INSERT INTO contents_dictionary (id_header, id_dictionary, data) SELECT id_header, id_dictionary, data FROM contents_dictionary_nodup')
                cur.execute('DROP TABLE contents_dictionary_nodup')
                print "dupes from contents_dictionary removed."
                
        #not perfect duplicate contents_dictionary records (same indexes, different value, unknown problem)
        cur.execute("select max(dup) as maxdup from (select id_header, id_dictionary, count(*) as dup from contents_dictionary group by id_header, id_dictionary)")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>1:
                try: cur.execute('DROP TABLE contents_dictionary_nodup')
                except: pass
                cur.execute('CREATE TABLE contents_dictionary_nodup (id_header INTEGER, id_dictionary INTEGER, data BLOB)')
                chk={}
                cur.execute('SELECT id_header,id_dictionary,data FROM contents_dictionary')
                for row in cur:
                    if not chk.has_key(str(row[0])+"-"+str(row[1])):
                        curw.execute('INSERT INTO contents_dictionary_nodup (id_header, id_dictionary, data) VALUES ('+str(row[0])+','+str(row[1])+',\''+row[2]+'\')')
                        chk[str(row[0])+"-"+str(row[1])]=True
                cur.execute('DELETE FROM contents_dictionary')
                cur.execute('INSERT INTO contents_dictionary (id_header, id_dictionary, data) SELECT id_header, id_dictionary, data FROM contents_dictionary_nodup')
                cur.execute('DROP TABLE contents_dictionary_nodup')
                print "inconsistent dupes from contents_dictionary removed."
        #rows with status>=0 without contents_index (unknown problem, only lonely rows with status<0 are allowed)
        cur.execute("select count(*) from rows left join contents_index on rows.id_row=contents_index.id_row and rows.id_user=contents_index.id_user and rows.id_instance=contents_index.id_instance where rows.status>=0 and contents_index.id_row is null")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>0:
                cur.execute('CREATE TABLE rows_notalone (id_row INTEGER, id_header INTEGER, status INTEGER, date_time TEXT, id_user INTEGER, rap TEXT, hpath TEXT, status_user INTEGER, id_locking INTEGER, id_instance INTEGER, status_instance INTEGER)')
                cur.execute('INSERT INTO rows_notalone (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows WHERE status<0')
                cur.execute('INSERT INTO rows_notalone (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT DISTINCT rows.id_row, rows.id_header, rows.status, rows.date_time, rows.id_user, rows.rap, rows.hpath, rows.status_user, rows.id_locking, rows.id_instance, rows.status_instance FROM rows inner join contents_index on rows.id_row=contents_index.id_row and rows.id_user=contents_index.id_user and rows.id_instance=contents_index.id_instance')
                cur.execute('DELETE FROM rows')
                cur.execute('INSERT INTO rows (id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance) SELECT id_row, id_header, status, date_time, id_user, rap, hpath, status_user, id_locking, id_instance, status_instance FROM rows_notalone')
                cur.execute('DROP TABLE rows_notalone')
                print "table rows fixed."
        #contents_index without rows (previous syncronizations interrupted)
        cur.execute("select count(*) from contents_index left join rows on contents_index.id_row=rows.id_row and contents_index.id_user=rows.id_user and contents_index.id_instance=rows.id_instance where rows.id_row is null")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>0:
                cur.execute('CREATE TABLE contents_index_notalone (id_row INTEGER, id_user INTEGER, id_header INTEGER, id_dictionary INTEGER, id_instance INTEGER, id_cycle INTEGER)')
                cur.execute('INSERT INTO contents_index_notalone (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT DISTINCT contents_index.id_row, contents_index.id_header, contents_index.id_user, contents_index.id_dictionary, contents_index.id_instance, contents_index.id_cycle FROM contents_index inner join rows on contents_index.id_row=rows.id_row and contents_index.id_user=rows.id_user and contents_index.id_instance=rows.id_instance')
                cur.execute('DELETE FROM contents_index')
                cur.execute('INSERT INTO contents_index (id_row, id_header, id_user, id_dictionary, id_instance, id_cycle) SELECT id_row, id_header, id_user, id_dictionary, id_instance, id_cycle FROM contents_index_notalone')
                cur.execute('DROP TABLE contents_index_notalone')
                print "table contents_index fixed."
        #contents_dictionary without contents_index (previous syncronizations interrupted)
        cur.execute("select count(*) from contents_dictionary left join contents_index on contents_dictionary.id_header=contents_index.id_header and contents_dictionary.id_dictionary=contents_index.id_dictionary where contents_index.id_header is null")
        row = cur.fetchone()
        if row!=None: 
            if row[0]>0:
                cur.execute('CREATE TABLE contents_dictionary_notalone (id_header INTEGER, id_dictionary INTEGER, data BLOB)')
                cur.execute('INSERT INTO contents_dictionary_notalone (id_header,id_dictionary,data) SELECT DISTINCT contents_dictionary.id_header,contents_dictionary.id_dictionary,contents_dictionary.data from contents_dictionary inner join contents_index on contents_dictionary.id_header=contents_index.id_header and contents_dictionary.id_dictionary=contents_index.id_dictionary')
                cur.execute('DELETE FROM contents_dictionary')
                cur.execute('INSERT INTO contents_dictionary (id_header,id_dictionary,data) SELECT id_header,id_dictionary,data FROM contents_dictionary_notalone');
                cur.execute('DROP TABLE contents_dictionary_notalone')
                print "table contents_dictionary fixed."

        print "Dupes checking started."
        automerge_group_shortcode=GetSetting(self.parent,'automerge_group_shortcode')
        if automerge_group_shortcode=="": automerge_group_shortcode="group_shortcode"
        automerge_keyname=GetSetting(self.parent,'automerge_keyname')
        if automerge_keyname=="": automerge_keyname=self.keyname
        query=""
        if self.keymode=="private":
            query="SELECT users.id_user,id_dictionary,count(*) FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[automerge_keyname])+" AND users.id_user="+str(self.id_user)+" GROUP BY users.id_user,id_dictionary"
        elif self.keymode=="group":
            query="SELECT c1.id_dictionary,c2.id_dictionary,count(*) FROM users,rows,contents_index as c1, contents_index as c2 WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=c1.id_row AND rows.id_user=c1.id_user AND rows.id_instance=c1.id_instance AND rows.id_row=c2.id_row AND rows.id_user=c2.id_user AND rows.id_instance=c2.id_instance AND c1.id_header="+str(self.hashHeaders[automerge_group_shortcode])+" AND c2.id_header="+str(self.hashHeaders[automerge_keyname])+" GROUP BY c1.id_dictionary,c2.id_dictionary"
        else:
            query="SELECT 0,id_dictionary,count(*) FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[automerge_keyname])+" GROUP BY id_dictionary"
        repl=[]
        cur.execute(query)
        for row in cur:
            if row[2]>1: repl.append(str(row[0])+"-"+str(row[1]))
        if len(repl)==0:
            print "Dupes checking: no errors."
        else:
            print "Dupes checking: fixing errors..."
            cur.execute("SELECT max(id_row) FROM ROWS")
            row=cur.fetchone()
            max_id_row=row[0]
            #id_instance=9999999999
            id_instance=self.parent.id_instance
            for pat in repl:
                lastdatetime=""
                records=[]
                query=""
                if self.keymode=="private":
                    #query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND users.id_user="+str(self.id_user)+" AND contents_index.id_dictionary="+pat.split("-")[1]
                    query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[automerge_keyname])+" AND users.id_user="+str(self.id_user)+" AND contents_index.id_dictionary="+pat.split("-")[1]
                elif self.keymode=="group":
                    #query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index as c1,contents_index as c2 WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=c1.id_row AND rows.id_user=c1.id_user AND rows.id_instance=c1.id_instance AND rows.id_row=c2.id_row AND rows.id_user=c2.id_user AND rows.id_instance=c2.id_instance AND c1.id_header="+str(self.hashHeaders["group_shortcode"])+" AND c2.id_header="+str(self.hashHeaders[self.keyname])+" AND c1.id_dictionary="+pat.split("-")[0]+" AND c2.id_dictionary="+pat.split("-")[1]
                    query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM users,rows,contents_index as c1,contents_index as c2 WHERE rows.status=0 AND users.id_user=rows.id_user AND rows.id_row=c1.id_row AND rows.id_user=c1.id_user AND rows.id_instance=c1.id_instance AND rows.id_row=c2.id_row AND rows.id_user=c2.id_user AND rows.id_instance=c2.id_instance AND c1.id_header="+str(self.hashHeaders[automerge_group_shortcode])+" AND c2.id_header="+str(self.hashHeaders[automerge_keyname])+" AND c1.id_dictionary="+pat.split("-")[0]+" AND c2.id_dictionary="+pat.split("-")[1]
                else:   #public
                    #query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[self.keyname])+" AND contents_index.id_dictionary="+pat.split("-")[1]
                    query="SELECT DISTINCT rows.id_row,rows.id_user,rows.id_instance,rows.date_time,rows.rap FROM rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_header="+str(self.hashHeaders[automerge_keyname])+" AND contents_index.id_dictionary="+pat.split("-")[1]
                cur.execute(query)
                for row in cur:
                    record={}
                    record["id_row"]=row[0]
                    record["id_user"]=row[1]
                    record["id_instance"]=row[2]
                    record["date_time"]=row[3]
                    if record["date_time"]>lastdatetime or lastdatetime=="": 
                        lastdatetime=record["date_time"]
                    record["rap"]=row[4]
                    records.append(record)
                id_user=0
                rap=""
                fields={}
                fields_ts={}
                lastdatetime=""
                for item in records:
                    cur.execute("SELECT DISTINCT id_header,id_dictionary,id_cycle FROM contents_index WHERE id_row="+str(item["id_row"])+" AND id_user="+str(item["id_user"])+" AND id_instance="+str(item["id_instance"]))
                    for row in cur:
                        curkey=str(row[0])+"-"+str(row[2])
                        dopick=False
                        if not fields.has_key(curkey):
                            dopick=True
                        elif item["date_time"]>fields_ts[curkey]:
                            dopick=True
                        if dopick:
                            fields[curkey]=row[1]
                            fields_ts[curkey]=item["date_time"]
                            id_user=item["id_user"]
                            rap=item["rap"]
                import datetime
                now = datetime.datetime.utcnow()
                ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
                if lastdatetime=="": lastdatetime=ts_db

                max_id_row=max_id_row+1
                #qy="INSERT INTO rows (id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(max_id_row)+","+str(id_user)+",1,'"+rap+"','"+ts_db+"',0,0,0,"+id_instance+")"
                qy="INSERT INTO rows (id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(max_id_row)+","+str(id_user)+",1,'"+rap+"','"+lastdatetime+"',0,0,0,"+id_instance+")"
                #print qy
                cur.execute(qy)
                for elm in fields.keys():
                    id_header=elm.split("-")[0]
                    id_cycle=elm.split("-")[1]
                    #qy="INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) VALUES ("+str(max_id_row)+","+str(id_user)+","+str(id_header)+","+str(fields[id_header])+","+id_instance+",0)"
                    qy="INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) VALUES ("+str(max_id_row)+","+str(id_user)+","+id_header+","+str(fields[elm])+","+id_instance+","+id_cycle+")"
                    #print qy
                    cur.execute(qy)
                for item in records:
                    #qy="UPDATE rows SET status="+str(max_id_row)+",status_user="+str(id_user)+",status_instance="+str(id_instance)+" WHERE id_row="+str(item["id_row"])+" AND id_user="+str(item["id_user"])+" AND id_instance="+str(item["id_instance"])
                    qy="UPDATE rows SET status="+str(max_id_row)+",status_user="+str(id_user)+",status_instance="+id_instance+" WHERE id_row="+str(item["id_row"])+" AND id_user="+str(item["id_user"])+" AND id_instance="+str(item["id_instance"])
                    #print qy
                    cur.execute(qy)
            print "Dupes checking: done fixing errors."
            
        cur.execute('COMMIT TRANSACTION')
        self.parent.syncroTaskPause=False
        
        self.parent.syncrotablesupd["headers"]=True
        self.parent.syncrotablesupd["rows"]=True

        
    def OnButtonShowDiff(self,event):
        event.Skip()
            
        # wx.MessageBox("Fuction disabled", "Show differences between records", wx.ICON_INFORMATION | wx.OK, self)
        # return
        if len(self.lstFound.GetSelections())!=2:
            wx.MessageBox("This function requires exactly 2 records selected.", "Show differences between records", wx.ICON_ERROR | wx.OK, self)
        else:
            id_row1=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[0])
            id_user1=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[1])
            id_rowinstance1=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[2])
            datetime1=self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[3]
            id_row2=long(self.arrFound[self.lstFound.GetSelections()[1]+1].split(";")[0])
            id_user2=long(self.arrFound[self.lstFound.GetSelections()[1]+1].split(";")[1])
            id_rowinstance2=long(self.arrFound[self.lstFound.GetSelections()[1]+1].split(";")[2])
            datetime2=self.arrFound[self.lstFound.GetSelections()[1]+1].split(";")[3]
            
            q1="SELECT rap FROM rows WHERE id_row="+str(id_row1)+" AND id_user="+str(id_user1)+" AND id_instance="+str(id_rowinstance1)+" AND id_header="+str(self.id_crf)
            cur.execute(q1)
            row = cur.fetchone()
            rap1 = row[0]
            q2="SELECT rap FROM rows WHERE id_row="+str(id_row2)+" AND id_user="+str(id_user2)+" AND id_instance="+str(id_rowinstance2)+" AND id_header="+str(self.id_crf)
            cur.execute(q2)
            row = cur.fetchone()
            rap2 = row[0]
            #
            tmp_rap=[]
            for trunk in rap1.split(";"):
                arr=trunk.split(",")
                arr.sort()
                tmp_rap.append(",".join(arr))
            norm_rap1=";".join(tmp_rap)
            #
            tmp_rap=[]
            for trunk in rap2.split(";"):
                arr=trunk.split(",")
                arr.sort()
                tmp_rap.append(",".join(arr))
            norm_rap2=";".join(tmp_rap)
            #
            if norm_rap1!=norm_rap2:
                if DEBUG_MODE:
                    print "incompatible access permissions"
                    print q1
                    print rap1
                    print q2
                    print rap2
                wx.MessageBox("The two selected record have incompatible access permissions.", "Show differences between records", wx.ICON_ERROR | wx.OK, self)
            else:
                wx.BeginBusyCursor()
                
                form1={}
                cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(id_row1)+" AND contents.id_user="+str(id_user1)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(id_rowinstance1))
                for row in cur:
                    #if not newer.has_key(row[0]):
                    keycont=row[0]
                    if row[3]!=0: keycont=keycont+"#"+str(row[3])
                    form1[keycont]=row[1]
                form2={}
                cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(id_row2)+" AND contents.id_user="+str(id_user2)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(id_rowinstance2))
                for row in cur:
                    #if not newer.has_key(row[0]):
                    keycont=row[0]
                    if row[3]!=0: keycont=keycont+"#"+str(row[3])
                    form2[keycont]=row[1]
                    
                wx.EndBusyCursor()
                
                dlg = ShowDiffDlg(self,id_row1,id_user1,id_rowinstance1,datetime1,form1,id_row2,id_user2,id_rowinstance2,datetime2,form2,rap1)
                dlg.ShowModal()
                dlg.Destroy()  

    def OnLstFoundClicked(self,event):
        event.Skip()
        candoit=False
        permission=GetSetting(self.parent,"sysadmin_groups")
        if permission!="":
            s1=","+permission+","
            s2=","+str(self.id_group)+","
            if s2 in s1: candoit=True
        else:
            id_group_admin=GetSetting(self.parent,"id_group_admin")
            if str(self.id_group)==str(id_group_admin):
                candoit=True
        if candoit:
            log_id_row=-1
            log_id_user=-1
            log_id_rowinstance=-1
            try: log_id_row=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[0])
            except: pass
            try: log_id_user=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[1])
            except: pass
            try: log_id_rowinstance=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[2])
            except: pass
            self.parent.SetStatusText("key: "+str(log_id_row)+","+str(log_id_user)+","+str(log_id_rowinstance),2)

    def OnLstFoundDoubleClicked(self,event):
        event.Skip()
        self.id_row=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[0])
        self.id_user=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[1])
        self.id_rowinstance=long(self.arrFound[self.lstFound.GetSelections()[0]+1].split(";")[2])
        self.OpenCurRecord()

    def OpenCurRecord(self):
        print "opening "+str(self.id_row)+"-"+str(self.id_user)+"-"+str(self.id_rowinstance)
        wx.BeginBusyCursor()
        
        self.dont_trigger_field_change=True
        self.preloadCombo={}
        self.preloadComboMatch={}
        self.preloadComboFulfilled={}

        self.rap=""
        cur.execute("SELECT rap FROM rows WHERE id_row="+str(self.id_row)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance))
        #print "SELECT rap FROM rows WHERE id_row="+str(self.id_row)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance)
        row = cur.fetchone()
        if row!=None: self.rap=row[0]

        self.cur_group_shortcode=""
        #trans={}
        contents={}
        self.curkeyvalue=""
        idKey=self.hashHeaders[self.keyname]
        
        if self.mode_shared_variables:
            #shared variables - begin
            curKeyIdDic=0
            cur.execute("SELECT id_dictionary FROM contents_index WHERE id_row="+str(self.id_row)+" AND id_user="+str(self.id_user)+" AND id_header="+str(idKey)+" AND id_instance="+str(self.id_rowinstance))
            row = cur.fetchone()
            if row!=None: curKeyIdDic=row[0]
            if self.keymode=="private":
                query="SELECT id_row,id_user FROM rows WHERE id_row IN (SELECT MAX(rows.id_row) FROM rows,contents_index WHERE rows.id_header="+str(self.id_crf)+" AND rows.status=0 AND contents_index.id_header="+str(idKey)+" AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_dictionary="+str(curKeyIdDic)+" AND rows.id_user="+str(self.id_user)+")"
            elif self.keymode=="group":
                query="SELECT id_row,id_user FROM rows WHERE id_row IN (SELECT MAX(rows.id_row) FROM users,rows,contents_index WHERE rows.id_header="+str(self.id_crf)+" AND rows.status=0 AND contents_index.id_header="+str(idKey)+" AND users.id_user=rows.id_user AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_dictionary="+str(curKeyIdDic)+" AND users.id_group="+str(self.id_group)+")"
            else:
                query="SELECT id_row,id_user FROM rows WHERE id_row IN (SELECT MAX(rows.id_row) FROM rows,contents_index WHERE rows.id_header="+str(self.id_crf)+" AND rows.status=0 AND contents_index.id_header="+str(idKey)+" AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND contents_index.id_dictionary="+str(curKeyIdDic)+")"
            cur.execute(query)        
            row = cur.fetchone()
            if row!=None:
                id_sharedrow=row[0]
                id_shareduser=row[1]
                #cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(id_sharedrow)+" AND contents.id_user="+str(id_shareduser)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(self.id_rowinstance)+" AND deftype='S' ORDER BY contents.id_row, contents.id_user, contents.id_instance, contents.id_header, contents.id_cycle")
                qy="SELECT contents.id_header,data,id_cycle FROM headers,contents WHERE contents.id_row="+str(id_sharedrow)+" AND contents.id_user="+str(id_shareduser)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(self.id_rowinstance)+" AND deftype='S'"
                if self.mode_multi_database:
                    qy="SELECT contents.id_header,data,id_cycle FROM rows,headers,contents WHERE rows.id_header="+str(self.id_crf)+" AND rows.id_row="+str(id_sharedrow)+" AND rows.id_user="+str(id_shareduser)+" AND rows.id_instance="+str(self.id_rowinstance)+" AND rows.id_row=contents.id_row AND rows.id_user=contents.id_user AND rows.id_instance=contents.id_instance AND headers.id_header=contents.id_header AND deftype='S'"
                cur.execute(qy)
                for row in cur:
                    # Eventuale reload arrHeaders - begin
                    if row[0] not in self.arrHeaders:
                        cur2 = con.cursor()
                        cur2.execute("SELECT id_header,description,onchange,defaultvalue,validation,id_section,pos,typedesc,subtypedesc,label FROM headers WHERE child_of="+str(self.id_crf))
                        for row2 in cur2:
                            self.arrHeaders[row2[0]]=row2[1]
                            self.hashHeaders[row2[1]]=row2[0]
                    # Eventuale reload arrHeaders - end
                    keycont=self.arrHeaders[row[0]]
                    if row[2]!=0: keycont=keycont+"#"+str(row[2])
                    if not contents.has_key(keycont):
                        if not self.parent.trans.has_key(row[1]):
                            dData=HB_DecryptOne(self.digestkey,row[1],"latin-1")
                            self.parent.trans[row[1]]=dData
                        else:
                            dData=self.parent.trans[row[1]]
                        contents[keycont]=dData
                        if row[0]==idKey and self.curkeyvalue=="": self.curkeyvalue=dData
            #shared variables - end
        #cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(self.id_row)+" AND contents.id_user="+str(self.id_user)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(self.id_rowinstance)+" ORDER BY contents.id_row, contents.id_user, contents.id_instance, contents.id_header, contents.id_cycle")
        #cur.execute("SELECT description,data,contents.id_header,contents.id_cycle FROM headers,contents WHERE contents.id_row="+str(self.id_row)+" AND contents.id_user="+str(self.id_user)+" AND headers.id_header=contents.id_header AND contents.id_instance="+str(self.id_rowinstance))
        qy="SELECT id_header,data,id_cycle FROM contents WHERE id_row="+str(self.id_row)+" AND id_user="+str(self.id_user)+" AND id_instance="+str(self.id_rowinstance)
        if self.mode_multi_database:
            qy="SELECT contents.id_header,data,id_cycle FROM rows,contents WHERE rows.id_header="+str(self.id_crf)+" AND rows.id_row="+str(self.id_row)+" AND rows.id_user="+str(self.id_user)+" AND rows.id_instance="+str(self.id_rowinstance)+" AND rows.id_row=contents.id_row AND rows.id_user=contents.id_user AND rows.id_instance=contents.id_instance"
        cur.execute(qy)
        for row in cur:
            # Eventuale reload arrHeaders - begin
            if row[0] not in self.arrHeaders:
                cur2 = con.cursor()
                cur2.execute("SELECT id_header,description,onchange,defaultvalue,validation,id_section,pos,typedesc,subtypedesc,label FROM headers WHERE child_of="+str(self.id_crf))
                for row2 in cur2:
                    self.arrHeaders[row2[0]]=row2[1]
                    self.hashHeaders[row2[1]]=row2[0]
            # Eventuale reload arrHeaders - end
            keycont=self.arrHeaders[row[0]]
            if row[2]!=0: keycont=keycont+"#"+str(row[2])
            if not contents.has_key(keycont):
                if not self.parent.trans.has_key(row[1]):
                    dData=HB_DecryptOne(self.digestkey,row[1],"latin-1")
                    self.parent.trans[row[1]]=dData
                else:
                    dData=self.parent.trans[row[1]]
                contents[keycont]=dData
                if row[0]==idKey and self.curkeyvalue=="": self.curkeyvalue=dData
        #Riempimento campi
        self.ifrm=0
        self.samename={}
        for ifrm in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
            self.PopulateFieldList(ifrm,list,contents)
        #Attivazione di tutti i campi calcolati
        for ifrm in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
            self.BindEventsFieldList(ifrm,list)
        #Attivazione pagina 1
        self.frameCrf.Scroll(0,0)
        wx.xrc.XRCCTRL(self.frameRecordset,'choice').SetSelection(0)
        for i in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
            if wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetSelection()==i:
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(10, 10))
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(True)
                self.frameCrf.SetVirtualSize(wx.Size(SCREEN_X-50,self.FormHeight[i]/10))
                self.frameCrf.SetScrollbars(10,10,1,self.FormHeight[i]/10)
            else:
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).Show(False)
                wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]).SetPosition(wx.Point(5000, 0))
                
        self.curcycle=1
        self.contents=contents

        if self.ActivateCyclicity:
            objCycles=wx.xrc.XRCCTRL(self.frameRecordset,'cycles')                
            if self.Cyclicity[1]==0: 
                objCycles.Show(False)
                if DEBUG_MODE: print "objCycles hidden"
                try: wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(False)
                except: pass
            else: 
                try: 
                    wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Show(True)
                    wx.xrc.XRCCTRL(self.frameRecordset,'frmCycles').Scroll(0,0)
                except: pass
                objCycles=wx.xrc.XRCCTRL(self.frameRecordset,'cycles') 
                objCycles.Show(True)
                if DEBUG_MODE: print "objCycles showed"
                firstemptycycle=self.FindFirstEmptyCycle()
                self.curcycle=max(1,(firstemptycycle-1))
                wx.xrc.XRCCTRL(self.frameRecordset,'cycles').SetSelection(self.curcycle-1)
                if DEBUG_MODE: print "curcycle="+str(self.curcycle)
                for t in range(100):
                    if t<self.Cyclicity[i+1] and t<firstemptycycle:
                        objCycles.ShowItem(t,True)
                    else:
                        try: objCycles.ShowItem(t,False)
                        except: pass
                self.LoadCycle(self.curcycle)
        #salvataggio dati caricati per il saverecord
        if self.ActivateCyclicity:
            wx.xrc.XRCCTRL(self.frameRecordset,'cycles').SetSelection(0)
        self.MonitoringCheck(self.ifrm+1)
        if not DEBUG_MODE:
            try: exec(self.OnFormActivate[self.ifrm+1])
            except: pass
        else:
            print "Executing OnActivate script"
            exec(self.OnFormActivate[self.ifrm+1])
        
        self.loadedcontents={}
        for keycont in self.contents:
            self.loadedcontents[keycont]=self.contents[keycont] #in questo caso salvo i soli record realmente modificati
        #exec(self.OnFormActivate[self.ifrm+1])
        if GetSetting(self.parent,'hide_forbidden_buttons')=="1":
            #Visibilita' buttonDelete, buttonSave e buttonCommit
            checkrap=False
            if len(self.rap.split(";"))<3: checkrap=True
            elif str(self.id_group) in self.rap.split(";")[2].split(","): checkrap="True"
            if checkrap:
                self.buttonDelete.Show(True)
            else:
                self.buttonDelete.Show(False)
            checkrap=False
            if len(self.rap.split(";"))<3: checkrap=True
            elif str(self.id_group) in self.rap.split(";")[1].split(","): checkrap="True"
            if checkrap:
                self.buttonSave.Show(True)
                self.buttonCommit.Show(True)
            else:
                self.buttonSave.Show(False)
                self.buttonCommit.Show(False)
        #Apertura Scheda
        self.frameLogin.SetPosition(wx.Point(5000,0))
        self.frameSearch.SetPosition(wx.Point(5000,0))
        self.frameRecordset.SetPosition(wx.Point(0,0))
        self.EnableDisableButtons(3)
        
        self.dont_trigger_field_change=False
        self.lastfocusvariable=""
        self.lastfocusvariableobj=None
        wx.EndBusyCursor()

    def IsCyclicFormEmpty(self,list,contents,cycle):
        notemptyform=False
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                notemptyform=self.IsCyclicFormEmpty(sublist,contents,cycle)
                if notemptyform: break
            else:
                if self.IsInputField(list[i]):
                    curName=list[i].GetName()
                    if curName not in ("group_shortcode", self.keyname):
                        isReadOnly=False
                        if not list[i].IsEnabled(): isReadOnly=True
                        try:
                            if not list[i].IsEditable():  isReadOnly=True
                        except: pass
                        keycont=curName
                        if not contents.has_key(curName): keycont=keycont+"#"+`cycle`
                        hasValue=False
                        if contents.has_key(keycont):
                            if contents[keycont]!="": hasValue=True
                        #print list[i].GetName()+": "+str(hasValue)+" ("+str(isReadOnly)+")"
                        if not isReadOnly and hasValue: 
                            notemptyform=True
                            break
        return notemptyform
        
    def FindFirstEmptyCycle(self):
        firstemptycycle=100
        ifrm=self.ifrm
        list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
        for t in range(100):
            notemptyform=self.IsCyclicFormEmpty(list,self.contents,t+1)
            if not notemptyform:
                firstemptycycle=t+1
                break
        if firstemptycycle>=99: firstemptycycle=1
        return firstemptycycle
    
    def LoadCycle(self,cycle):
        self.dont_trigger_field_change=True
        ifrm=self.ifrm
        list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
        self.LoadCyclicFieldList(list,self.contents,cycle)
        self.dont_trigger_field_change=False
    
    def LoadCyclicFieldList(self,list,contents,cycle):
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                self.LoadCyclicFieldList(sublist,contents,cycle)
            else:
                if self.IsInputField(list[i]):
                    curName=list[i].GetName()
                    keycont=curName
                    if not contents.has_key(curName): keycont=keycont+"#"+`cycle`
                    if contents.has_key(keycont):
                        self.SetInputFieldValue(list[i],contents[keycont])
                        #print "loading cycle "+str(cycle)+": "+curName+" = "+contents[keycont]
                    else:
                        self.SetInputFieldValue(list[i],"")
                        #print "loading cycle "+str(cycle)+": "+curName+" = ''"
                    #colorazione campi - inizio
                    if self.IsInputFieldEditable(list[i]):
                        confidence=keycont+"__metadata_confidence"
                        if keycont+"__metadata_confidence" in contents:
                            if int(contents[confidence])==1: 
                                list[i].SetBackgroundColour(wx.Colour(255, 0, 0))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 0, 0), validBackgroundColour=wx.Colour(255, 0, 0), invalidBackgroundColour=wx.Colour(255, 0, 0))
                                except: pass
                            if int(contents[confidence])==2: 
                                list[i].SetBackgroundColour(wx.Colour(0, 255, 0))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(0, 255, 0), validBackgroundColour=wx.Colour(0, 255, 0), invalidBackgroundColour=wx.Colour(0, 255, 0))
                                except: pass
                            if int(contents[confidence])==3: 
                                list[i].SetBackgroundColour(wx.Colour(255, 255, 0))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 0), validBackgroundColour=wx.Colour(255, 255, 0), invalidBackgroundColour=wx.Colour(255, 255, 0))
                                except: pass
                        else:
                            list[i].SetBackgroundColour(wx.Colour(255, 255, 255))
                            try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 255), validBackgroundColour=wx.Colour(255, 255, 255), invalidBackgroundColour=wx.Colour(255, 255, 255))
                            except: pass
                        list[i].Refresh()
                    #colorazione campi - fine

    def SaveCycle(self,cycle):
        ifrm=self.ifrm
        list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[ifrm]).GetChildren()
        self.SaveFieldList(list,cycle)

    def SaveFieldList(self,list,cycle):
        for ielm in range(len(list)):
            if list[ielm].GetClassName() in ("wxPanel"):
                sublist=list[ielm].GetChildren()
                self.SaveFieldList(sublist,cycle)
            else:
                if self.IsInputField(list[ielm]):
                    curName=list[ielm].GetName()
                    dcurValue=self.GetInputFieldValue(list[ielm])
                    #print "saving cycle "+str(cycle)+": "+curName+" = "+dcurValue
                    if type(dcurValue).__name__=="int": dcurValue=str(dcurValue)
                    if dcurValue=="":
                        if cycle==0:
                            if self.contents.has_key(curName):
                                del self.contents[curName]
                        else:
                            if self.contents.has_key(curName+"#"+`cycle`):
                                del self.contents[curName+"#"+`cycle`]
                            elif self.contents.has_key(curName):
                                del self.contents[curName]
                    else:
                        if cycle==0:
                            self.contents[curName]=dcurValue
                        else:
                            if self.contents.has_key(curName): 
                                self.contents[curName]=dcurValue
                            else: 
                                self.contents[curName+"#"+`cycle`]=dcurValue

    def SaveInDatabase(self):
        self.parent.syncroTaskWaitSave=True
        revtrans={}
        checkContents={}
        queue=[]
        for keycont in self.contents:
            dcurValue=self.contents[keycont]
            if (self.parent.arrSettings.has_key(keycont)):
                if (self.parent.arrSettings[keycont]!=dcurValue):
                    cur.execute("UPDATE settings SET setting_value='"+self.GetSqlValue(dcurValue)+"' WHERE setting_key='"+keycont+"'")
                    self.parent.arrSettings[keycont]=dcurValue
            else:
                if (keycont and dcurValue):
                    curName=keycont
                    cycle="0"
                    if keycont.find("#")>=0: 
                        curName=keycont[:keycont.find("#")]
                        cycle=keycont[keycont.find("#")+1:]
                    if not (checkContents.has_key(keycont)):
                        if revtrans.has_key(dcurValue):
                            curValue=revtrans[dcurValue]
                        else:
                            if dcurValue in self.parent.trans.values():
                                #reverse lookup
                                curValue=[key for key,value in self.parent.trans.items() if value==dcurValue ][0]
                            else:
                                curValue=HB_EncryptOne(self.digestkey,dcurValue)
                            revtrans[dcurValue]=curValue
                        #aggiunta eventuali headers x metadati apparsi - begin
                        if not (self.hashHeaders.has_key(curName)):
                            longCRC=UniqueLongCRC(str(self.id_crf)+","+curName,cur)
                            #cur.execute("INSERT INTO headers (id_header,child_of,description, id_section, date_time, status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'"+curName+"',"+`self.ifrm+1`+",'0000-00-00 00:00:00',0)")
                            cur.execute("INSERT INTO headers (id_header,child_of,description, date_time, status) VALUES ("+str(longCRC)+","+str(self.id_crf)+",'"+curName+"','0000-00-00 00:00:00',0)")
                            self.hashHeaders[curName]=longCRC
                            self.arrHeadersEvents[curName]=None
                            self.arrHeadersDefaults[curName]=None
                            self.arrHeadersTypedesc[curName]=None
                            self.arrHeadersSubtypedesc[curName]=None
                            self.arrHeadersValidations[curName]=None
                            self.parent.syncrotablesupd["headers"]=True
                        #aggiunta eventuali headers x metadati apparsi - end
                        id_dictionary=0
                        cur.execute("SELECT id_header,id_dictionary,data FROM contents_dictionary WHERE id_header="+str(self.hashHeaders[curName])+" AND data='"+self.GetSqlValue(curValue)+"'")
                        row = cur.fetchone()
                        if row!=None:
                            id_dictionary=row[1]
                        if id_dictionary==0:
                            if self.parent.max_idHeadDict.has_key(str(self.hashHeaders[curName])):
                                #cur.execute("SELECT MAX(id_dictionary) FROM contents_dictionary WHERE id_header="+`self.hashHeaders[curName]`)
                                #row = cur.fetchone()
                                #if row!=None:
                                    #id_dictionary=row[0]
                                #if id_dictionary==None: id_dictionary=0
                                #id_dictionary=id_dictionary+1
                                self.parent.max_idHeadDict[str(self.hashHeaders[curName])]=self.parent.max_idHeadDict[str(self.hashHeaders[curName])]+1
                            else:
                                self.parent.max_idHeadDict[str(self.hashHeaders[curName])]=1
                            # Correttore di sicurezza per il caso multiutente sullo stesso DB - Inizio
                            cur.execute("SELECT MAX(id_dictionary) FROM contents_dictionary WHERE id_header="+str(self.hashHeaders[curName]))
                            row = cur.fetchone()
                            if row!=None:
                                if row[0]>(self.parent.max_idHeadDict[str(self.hashHeaders[curName])]-1):
                                    self.parent.max_idHeadDict[str(self.hashHeaders[curName])]=row[0]+1
                                    print "corrected max_idHeadDict for field "+curName
                            # Correttore di sicurezza per il caso multiutente sullo stesso DB - Fine
                            id_dictionary=self.parent.max_idHeadDict[str(self.hashHeaders[curName])]
                            sql="INSERT INTO contents_dictionary (id_header,id_dictionary,data) VALUES ("+str(self.hashHeaders[curName])+","+str(id_dictionary)+",'"+self.GetSqlValue(curValue)+"')"
                            #print sql
                            cur.execute(sql)
                        #sql="INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) VALUES ("+str(self.id_row)+","+str(self.id_logged_user)+","+`self.hashHeaders[curName]`+","+`id_dictionary`+","+str(self.parent.id_instance)+","+cycle+")"
                        if len(queue)>=500:
                            cur.execute(''.join(queue))
                            queue=[]
                        if len(queue)==0: queue.append("INSERT INTO contents_index (id_row,id_user,id_header,id_dictionary,id_instance,id_cycle) SELECT "+str(self.id_row)+","+str(self.id_logged_user)+","+str(self.hashHeaders[curName])+","+str(id_dictionary)+","+str(self.parent.id_instance)+","+cycle)
                        else: queue.append(" UNION SELECT "+str(self.id_row)+","+str(self.id_logged_user)+","+str(self.hashHeaders[curName])+","+str(id_dictionary)+","+str(self.parent.id_instance)+","+cycle)
                        #cur.execute(sql)
                        checkContents[curName]=True;
        if len(queue)>0: cur.execute(''.join(queue))
        self.parent.syncroTaskWaitSave=False

    def BindEventsFieldList(self,ifrm,list):
        for ielm in range(len(list)):
            if list[ielm].GetClassName() in ("wxPanel"):
                sublist=list[ielm].GetChildren()
                self.BindEventsFieldList(ifrm,sublist)
            else:
                if self.IsInputField(list[ielm]):
                    curName=list[ielm].GetName()
                    if (self.arrHeadersEvents.has_key(curName)):
                        if (self.arrHeadersEvents[curName]!=None):
                            if list[ielm].GetClassName() not in ("wxCheckBox","wxButton"):
                                if not DEBUG_MODE:
                                    try: exec(self.arrHeadersEvents[curName])
                                    except: pass
                                else:
                                    print "Executing script for field '"+curName+"'"
                                    exec(self.arrHeadersEvents[curName])
                    list[ielm].Bind(self.GetInputFieldChangeEventId(list[ielm]), self.OnInputFieldChange)
#                    list[ielm].Bind(wx.EVT_ENTER_WINDOW, self.OnInputFieldMetadata)
#                    list[ielm].Bind(wx.EVT_MIDDLE_UP, self.OnInputFieldMetadata)
                    if list[ielm].GetClassName() in ("wxTextCtrl","wxComboBox"):
                        list[ielm].Bind(wx.EVT_KEY_UP, self.onTextBoxKeyUp)
                    if list[ielm].GetClassName() in ("wxComboBox","wxChoice"):
                        list[ielm].Bind(wx.EVT_MOUSEWHEEL, self.OnComboMouseWheel)
                        if list[ielm].GetClassName() in ("wxComboBox"):
                            if sys.platform != 'darwin':
                                list[ielm].Bind(wx.EVT_SET_FOCUS, self.FulfillCombo)    #questa funzione poi chiama la OnInputFieldFocus che serve x tutti
                            else:
                                list[ielm].Bind(wx.EVT_ENTER_WINDOW, self.FulfillCombo)    #questa funzione poi chiama la OnInputFieldFocus che serve x tutti
                        #if list[ielm].GetClassName() in ("wxChoice"):
                            #if self.multidict:
                                #if sys.platform != 'darwin':
                                    #list[ielm].Bind(wx.EVT_SET_FOCUS, self.FulfillChoice)    #questa funzione poi chiama la OnInputFieldFocus che serve x tutti
                                #else:
                                    #list[ielm].Bind(wx.EVT_ENTER_WINDOW, self.FulfillChoice)    #questa funzione poi chiama la OnInputFieldFocus che serve x tutti
                    if list[ielm].GetClassName() in ("wxComboBox","wxChoice","wxCheckBox"):
                        list[ielm].Bind(wx.EVT_CHAR, self.onChoiceKeyPress)
                    #Popup menu - Begin
                    list[ielm].Bind(wx.EVT_MIDDLE_DOWN, self.onPopupMenu)
                    #Popup menu - End
                if self.IsButton(list[ielm]):
                    curName=list[ielm].GetName()
                    if (self.arrHeadersEvents.has_key(curName)):
                        if (self.arrHeadersEvents[curName]!=None):
                            list[ielm].Bind(wx.EVT_BUTTON, self.OnInputButtonClick)
                if self.IsInputField(list[ielm]) or self.IsButton(list[ielm]):
                    if list[ielm].GetClassName() not in ("wxComboBox"):         #nel caso combo questa funzione e' chiamata a valle del FulfillCombo
                        list[ielm].Bind(wx.EVT_SET_FOCUS, self.OnInputFieldFocus)
                    list[ielm].Bind(wx.EVT_ENTER_WINDOW, self.OnInputFieldSetTooltip)
                    list[ielm].Bind(wx.EVT_LEAVE_WINDOW, self.OnInputFieldDelTooltip)

    def FulfillChoice(self,evt):
        evt.Skip()
        evtobj=evt.GetEventObject()
        evtname=evtobj.GetName()
        if (evtname!=None):
            if not self.preloadComboFulfilled.has_key(evtname):
                if self.arrHeadersComboPreload.has_key(evtname):
                    wx.BeginBusyCursor()
                    oldValue=self.GetInputFieldValue(evtobj)
                    elmlst=[]
                    for elm in self.arrHeadersComboPreload[evtname]:
                        elmval=elm
                        try: elmval=elm.split("|")[self.FormIdxDict[self.ifrm]]
                        except: pass
                        elmlst.append(elmval)
                    evtobj.SetItems(elmlst)
                    self.SetInputFieldValue(evtobj,oldValue)
                    wx.EndBusyCursor()
                self.preloadComboFulfilled[evtname]=True
        self.OnInputFieldFocus(evt)
        
    def FulfillCombo(self,evt):
        evt.Skip()
        evtobj=evt.GetEventObject()
        evtname=evtobj.GetName()
        if (evtname!=None):
            if not self.preloadComboFulfilled.has_key(evtname) or GetSetting(self.parent,"combo_learn")=="0":
                wx.BeginBusyCursor()
                oldValue=self.GetInputFieldValue(evtobj)

                self.preloadComboFulfilled[evtname]=True
                elmlst={}
                try: elmlst=self.preloadCombo[self.preloadComboMatch[evtname]]
                except: pass
                #print evtname+": "+str(elmlst)
                #for ct in range(len(elmlst)): evtobj.Append(elmlst[ct])
                evtobj.SetItems(elmlst)

                self.SetInputFieldValue(evtobj,oldValue)
                wx.EndBusyCursor()
        self.OnInputFieldFocus(evt)

    def FulfillComboList(self,obj,fieldname):
        #if self.preloadComboMatch.has_key(fieldname): return
        if GetSetting(self.parent,"combo_learn")=="0":
            if self.arrHeadersComboPreload.has_key(fieldname):
                elmlst=self.arrHeadersComboPreload[fieldname]
                self.preloadCombo[fieldname]=elmlst
                self.preloadComboMatch[fieldname]=fieldname
            return
        self.DoFulfillComboList(obj,fieldname)

    def DoFulfillComboList(self,obj,fieldname):
        lstfields=[]
        if fieldname.find("__")<0:
#            lstfields.append(fieldname)
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND pos is not null AND description='"+fieldname+"' GROUP BY description, id_cycle")
            for row in cur: lstfields.append(row[0])
        else:
#            cur.execute("SELECT description FROM headers WHERE description like '"+fieldname.split("__")[0]+"__%'")
            basename=fieldname.split("__")[0]
            cur.execute("SELECT description||case when max(id_cycle)>0 then '#'||id_cycle else '' end FROM headers,rows,contents_index WHERE rows.status=0 AND rows.id_row=contents_index.id_row AND rows.id_user=contents_index.id_user AND rows.id_instance=contents_index.id_instance AND headers.id_header=contents_index.id_header and child_of="+str(self.id_crf)+" AND pos is not null AND description like '"+basename+"__%' GROUP BY description, id_cycle")
            for row in cur: lstfields.append(row[0])
        if len(lstfields)==0: lstfields.append(fieldname)

        if len(lstfields)>1: lstfields.sort()
        import string
        strlstfields=string.join(lstfields,",")

        lstfields2=[]
        for elm in lstfields:
            lstfields2.append(elm.replace("#","__"))

        strlstfields2=""
        strlstfields2=string.join(lstfields2,",")
        
        elmlst=[]
        if self.preloadCombo.has_key(strlstfields):
            elmlst=self.preloadCombo[strlstfields]
        else:
            if self.arrHeadersComboPreload.has_key(fieldname):
                elmlst=self.arrHeadersComboPreload[fieldname]
            if len(lstfields)>0:
                elmkeys={}
                con_m = sqlite.connect(":memory:")
                cur_m = con_m.cursor()
                square(con,con_m,"search",lstfields,self.id_group,self.id_crf)
                
                sql="SELECT DISTINCT "+strlstfields2+" FROM search"
                #print sql
                cur_m.execute(sql)
                cols = [i[0] for i in cur_m.description]; hc={}; i=0
                for col in cols: hc[col]=i; i=i+1
                for row in cur_m:
                    for col in cols:
                        if row[hc[col]]!=None:
                            if not elmkeys.has_key(row[hc[col]]):
                                elmkeys[row[hc[col]]]=True
                                if not self.parent.trans.has_key(row[hc[col]]):
                                    dData=HB_DecryptOne(self.digestkey,row[hc[col]],"latin-1")
                                    self.parent.trans[row[hc[col]]]=dData
                                else:
                                    dData=self.parent.trans[row[hc[col]]]
                                if dData not in elmlst:
                                    elmlst.append(dData)
                cur_m.close()
                con_m.close()

        if len(elmlst)>1:
            #elmlst.sort(key=str(str.lower))
            elmlst.sort()

        self.preloadCombo[strlstfields]=elmlst
        self.preloadComboMatch[fieldname]=strlstfields
        #obj.Clear()
        
        #print fieldname+" ("+strlstfields+"): "+str(elmlst)
        # for ct in range(len(elmlst)): obj.Append(elmlst[ct])

    def ColorizeFields(self,ifrm,list,contents):
        #metadata_used=False
        #for key in contents:
            #if "metadata" in key:
                #metadata_used=True
                #break
        #if not metadata_used: return
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                self.ColorizeFields(ifrm,sublist,contents)
            else:
                if self.IsInputField(list[i]):
                    curName=list[i].GetName()
                    keycont=curName
                    if self.ActivateCyclicity:
                        if self.Cyclicity[ifrm+1]>0 and (not contents.has_key(curName)): keycont=curName+"#1"
                    #colorazione campi - inizio
                    if self.IsInputFieldEditable(list[i]):
                        confidence=keycont+"__metadata_confidence"
                        if keycont+"__metadata_confidence" in contents:
                            if int(contents[confidence])==1: 
                                list[i].SetBackgroundColour(wx.Colour(255, 0, 0))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 0, 0), validBackgroundColour=wx.Colour(255, 0, 0), invalidBackgroundColour=wx.Colour(255, 0, 0))
                                except: pass
                            if int(contents[confidence])==2: 
                                list[i].SetBackgroundColour(wx.Colour(0, 255, 0))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(0, 255, 0), validBackgroundColour=wx.Colour(0, 255, 0), invalidBackgroundColour=wx.Colour(0, 255, 0))
                                except: pass
                            if int(contents[confidence])==3: 
                                list[i].SetBackgroundColour(wx.Colour(255, 255, 0))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 0), validBackgroundColour=wx.Colour(255, 255, 0), invalidBackgroundColour=wx.Colour(255, 255, 0))
                                except: pass
                        else:
                            if list[i].GetBackgroundColour()!=wx.Colour(255, 255, 255):
                                list[i].SetBackgroundColour(wx.Colour(255, 255, 255))
                                try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 255), validBackgroundColour=wx.Colour(255, 255, 255), invalidBackgroundColour=wx.Colour(255, 255, 255))
                                except: pass
                        list[i].Refresh()
                    #colorazione campi - fine
                    
    def PopulateFieldListDefault(self,ifrm,list,keyvalue,contents):
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                self.PopulateFieldListDefault(ifrm,sublist,keyvalue,contents)
            else:
                if self.IsInputField(list[i]):
                    curName=list[i].GetName()
                    curValue=""
                    if not self.samename.has_key(curName): self.samename[curName]=[]
                    self.samename[curName].append(ifrm)
                    if curName==self.keyname:
                        curValue=keyvalue
                        self.SetInputFieldValue(list[i],curValue)
                    else:
                        #Stored or Defaults values
                        if (self.parent.arrSettings.has_key(curName)):
                            curValue=self.parent.arrSettings[curName]
                            self.SetInputFieldValue(list[i],curValue)
                        else:
                            if curName=="group_shortcode":
                                curValue=self.group_shortcode
                                self.cur_group_shortcode=curValue
                                self.SetInputFieldValue(list[i],curValue)
                            else:
                                #Preloading ComboBox da Dictionary
                                if list[i].GetClassName() in ("wxComboBox"):
                                    self.FulfillComboList(list[i],curName)
                                if self.multidict:
                                    if list[i].GetClassName() in ("wxChoice"):
                                        #print curName+" ontology: "+str(ifrm)
                                        elmlst=[]
                                        for elm in self.arrHeadersComboPreload[curName]:
                                            elmval=elm
                                            try: 
                                                elmval=elm.split("|")[self.FormIdxDict[ifrm]]
                                                #print elmval
                                            except:
                                                #print "ERROR ON "+elm
                                                pass
                                            elmlst.append(elmval)
                                        list[i].SetItems(elmlst)
                                curValue=""
                                if (self.arrHeadersDefaults.has_key(curName)):
                                    if (self.arrHeadersDefaults[curName]!=None):
                                        curValue=self.arrHeadersDefaults[curName]
                                self.SetInputFieldValue(list[i],curValue)
                    keycont=curName
                    if self.ActivateCyclicity:
                        if self.Cyclicity[ifrm+1]>0 and (not contents.has_key(curName)): keycont=curName+"#1"
                    contents[keycont]=curValue
                    #Sblocca campi chiave - amministratore - Inizio
                    if curName in ("group_shortcode", self.keyname):
                        candoit=False
                        permission=GetSetting(self.parent,"sysadmin_groups")
                        if permission!="":
                            s1=","+permission+","
                            s2=","+str(self.id_group)+","
                            if s2 in s1: candoit=True
                        else:
                            id_group_admin=GetSetting(self.parent,"id_group_admin")
                            if str(self.id_group)==str(id_group_admin):
                                candoit=True
                        #if candoit:
                        #    list[i].style=0
                        #else:
                        #    list[i].style=wx.TE_READONLY
                        #list[i].Refresh()
                        list[i].SetEditable(candoit)
                    #Sblocca campi chiave - amministratore - Fine

    def PopulateFieldList(self,ifrm,list,contents):
        for i in range(len(list)):
            if list[i].GetClassName() in ("wxPanel"):
                sublist=list[i].GetChildren()
                self.PopulateFieldList(ifrm,sublist,contents)
            else:
                if self.IsInputField(list[i]):
                    curName=list[i].GetName()
                    if curName=="group_shortcode": 
                        if contents.has_key(curName): self.cur_group_shortcode=contents[curName]
                    if not self.samename.has_key(curName): self.samename[curName]=[]
                    self.samename[curName].append(ifrm)
                    keycont=""
                    if (self.parent.arrSettings.has_key(curName)):
                        self.SetInputFieldValue(list[i],self.parent.arrSettings[curName])
                        keycont=curName
                    else:
                        keycont=curName
                        if self.ActivateCyclicity:
                            if self.Cyclicity[ifrm+1]>0 and (not contents.has_key(curName)): keycont=curName+"#1"
                        #Preloading ComboBox da Dictionary
                        if list[i].GetClassName() in ("wxComboBox"):
                            self.FulfillComboList(list[i],curName)
                        if self.multidict:
                            if list[i].GetClassName() in ("wxChoice"):
                                if DEBUG_MODE: print curName+" ontology: "+str(self.FormIdxDict[ifrm])
                                elmlst=[]
                                for elm in self.arrHeadersComboPreload[curName]:
                                    elmval=elm
                                    try: elmval=elm.split("|")[self.FormIdxDict[ifrm]]
                                    except: pass
                                    elmlst.append(elmval)
                                list[i].SetItems(elmlst)
                        if contents.has_key(keycont):
                            self.SetInputFieldValue(list[i],contents[keycont])
                        else:
                            self.SetInputFieldValue(list[i],"")
                        variablename=keycont
                    ##colorazione campi - inizio
                    #if self.IsInputFieldEditable(list[i]):
                        #confidence=keycont+"__metadata_confidence"
                        #if keycont+"__metadata_confidence" in contents:
                            #if int(contents[confidence])==1: 
                                #list[i].SetBackgroundColour(wx.Colour(255, 0, 0))
                                #try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 0, 0), validBackgroundColour=wx.Colour(255, 0, 0), invalidBackgroundColour=wx.Colour(255, 0, 0))
                                #except: pass
                            #if int(contents[confidence])==2: 
                                #list[i].SetBackgroundColour(wx.Colour(0, 255, 0))
                                #try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(0, 255, 0), validBackgroundColour=wx.Colour(0, 255, 0), invalidBackgroundColour=wx.Colour(0, 255, 0))
                                #except: pass
                            #if int(contents[confidence])==3: 
                                #list[i].SetBackgroundColour(wx.Colour(255, 255, 0))
                                #try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 0), validBackgroundColour=wx.Colour(255, 255, 0), invalidBackgroundColour=wx.Colour(255, 255, 0))
                                #except: pass
                        #else:
                            #list[i].SetBackgroundColour(wx.Colour(255, 255, 255))
                            #try: list[i].SetCtrlParameters(emptyBackgroundColour=wx.Colour(255, 255, 255), validBackgroundColour=wx.Colour(255, 255, 255), invalidBackgroundColour=wx.Colour(255, 255, 255))
                            #except: pass
                        #list[i].Refresh()
                    ##colorazione campi - fine
                    #Sblocca campi chiave - amministratore - Inizio
                    if curName in ("group_shortcode", self.keyname):
                        candoit=False
                        permission=GetSetting(self.parent,"sysadmin_groups")
                        if permission!="":
                            s1=","+permission+","
                            s2=","+str(self.id_group)+","
                            if s2 in s1: candoit=True
                        else:
                            id_group_admin=GetSetting(self.parent,"id_group_admin")
                            if str(self.id_group)==str(id_group_admin):
                                candoit=True
                        #if candoit:
                        #    list[i].style=0
                        #else:
                        #    list[i].style=wx.TE_READONLY
                        #list[i].Refresh()
                        list[i].SetEditable(candoit)
                    #Sblocca campi chiave - amministratore - Fine
    
    def OnInputFieldChange(self,evt):
        if self.dont_trigger_field_change: return
        self.dont_trigger_field_change=True
        elmName=evt.GetEventObject().GetName()
        eventhandler=False
        if self.arrHeadersEvents.has_key(elmName):
            if self.arrHeadersEvents[elmName]!=None: eventhandler=True
        if eventhandler:
            customEventFunction=self.arrHeadersEvents[elmName]
            #print "evento:\n"+str(customEventFunction)
            if customEventFunction!=None:
                if not DEBUG_MODE:
                    try: exec(customEventFunction)
                    except: pass
                else:
                    print "Executing script for field '"+elmName+"'"
                    exec(customEventFunction)
                #print customEventFunction
                #try: exec(customEventFunction)
                #except: pass
                #exec(customEventFunction)
        else:
            if self.HeadersMonPattern.has_key(elmName):
                deftype=self.HeadersMonPattern[elmName]
                if deftype!="" and deftype!="S" and deftype!="A":
                    self.MonitoringCheck(self.ifrm+1)
                
        if self.samename.has_key(elmName):
            if len(self.samename[elmName]) > 1:
                #print elmName
                iPg=0
                newValue=""
                for i in range(wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetCount()):
                    if wx.xrc.XRCCTRL(self.frameRecordset,'choice').GetSelection()==i:
                        iPg=i
                        newValue=self.GetInputFieldValue(wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[i]),elmName))
                for i in range(len(self.samename[elmName])):
                    iPg2=self.samename[elmName][i]
                    if iPg2!=iPg:
                        try:
                            dupElm=wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[iPg2]),elmName)
                        except:
                            print "field name error: "+elmName
                        curValue=self.GetInputFieldValue(dupElm)
                        #print "pag."+str(iPg2)+": curValue="+curValue+"; newValue="+newValue
                        if curValue!=newValue:
                            #if dupElm.GetClassName() in ("wxChoice"):
                                #print curName+" ontology: "+str(iPg2)
                                #elmlst=[]
                                #for elm in self.arrHeadersComboPreload[curName]:
                                    #elmval=elm
                                    #try: elmval=elm.split("|")[self.FormIdxDict[self.ifrm]]
                                    #except: pass
                                    #elmlst.append(elmval)
                                #dupElm.SetItems(elmlst)
                            self.SetInputFieldValue(dupElm,newValue,iPg2)
                            #print elmName,newValue
        self.dont_trigger_field_change=False

    def OnInputButtonClick(self,evt):
        eventFunction=self.arrHeadersEvents[evt.GetEventObject().GetName()]
        if (self.arrHeadersEvents[evt.GetEventObject().GetName()]!=None):
            #wx.MessageBox(evt.GetString()+"\n"+eventFunction, "Info", wx.ICON_INFORMATION | wx.OK, self)
            #try:
            #    exec(eventFunction)
            #except:
            #    pass
            
            #exec(eventFunction)
            if not DEBUG_MODE:
                try: exec(eventFunction)
                except: pass
            else:
                print "Executing script for field '"+evt.GetEventObject().GetName()+"'"
                exec(eventFunction)

    def FindNextField(self,list,startfieldname,shiftDown):
        return self.FindNextField2(list,startfieldname,shiftDown,"",False)
    def FindNextField2(self,list,startfieldname,shiftDown,lastfieldname,foundstartfield):
        ret=""
        for ielm in range(len(list)):
            if list[ielm].GetClassName() in ("wxPanel"):
                sublist=list[ielm].GetChildren()
                ret=self.FindNextField2(sublist,startfieldname,shiftDown,lastfieldname,foundstartfield)
                if ret!="": break
            else:
                if self.IsInputField(list[ielm]) or self.IsButton(list[ielm]):
                    if not foundstartfield:
                        fieldname=list[ielm].GetName()
                        if fieldname==startfieldname:
                            if shiftDown:
                                ret=lastfieldname
                                break
                            else:
                                foundstartfield=True
                        else:
                            lastfieldname=fieldname
                    else:
                        ret=list[ielm].GetName()
                        break
        return ret
    
    def onChoiceKeyPress(self, event):
        event.Skip()
        keycode = event.GetKeyCode()
        controlDown = event.CmdDown()
        altDown = event.AltDown()
        shiftDown = event.ShiftDown()
        if keycode==wx.WXK_TAB:
            #if sys.platform == 'win32':
            if os.name=='nt':
                list=wx.xrc.XRCCTRL(self.frameCrf,self.FormName[self.ifrm]).GetChildren()
                nextfieldname=self.FindNextField(list,event.GetEventObject().GetName(),shiftDown)
                #print "nextfieldname="+nextfieldname
                if nextfieldname:
                    wx.xrc.XRCCTRL(wx.xrc.XRCCTRL(self.frameCrf,self.FormName[self.ifrm]),nextfieldname).SetFocus()
                #self.Navigate()
        
    def onTextBoxKeyUp(self, event):
        event.Skip()
        keycode = event.GetKeyCode()
        controlDown = event.CmdDown()
        altDown = event.AltDown()
        shiftDown = event.ShiftDown()
        #print str(keycode)+"-"+str(controlDown)+"-"+str(altDown)+"-"+str(shiftDown)
        if (keycode==wx.WXK_BACK or keycode==wx.WXK_DELETE) and (shiftDown or controlDown or altDown):
            obj = event.GetEventObject()
            self.SetInputFieldValue(obj,None)
        
    def OnComboMouseWheel(self, event):
        event.SetEventObject(self.parent)
        self.parent.GetEventHandler().ProcessEvent(event)

    def OnInputFieldFocus(self, event):
        event.Skip()
        tab_scroll=GetSetting(self.parent,"tab_scroll")
        if tab_scroll!="0":
            if event.GetEventObject().GetClassName()!="wxButton":
                fieldpos_in_page=event.GetEventObject().GetPosition().y-(self.frameCrf.GetScrollPos(wx.VERTICAL)*self.frameCrf.GetScrollPixelsPerUnit()[1])
                clientsize=self.frameCrf.GetClientSize().y
                #print str(event.GetEventObject().GetPosition().y),str(fieldpos_in_page),str(clientsize),str(self.frameCrf.GetScrollPos(wx.VERTICAL))
                if fieldpos_in_page>(clientsize-50):
                    self.frameCrf.Scroll(0,self.frameCrf.GetScrollPos(wx.VERTICAL)+clientsize/3/self.frameCrf.GetScrollPixelsPerUnit()[1])
                elif fieldpos_in_page<30 and self.frameCrf.GetScrollPos(wx.VERTICAL)>0:
                    self.frameCrf.Scroll(0,self.frameCrf.GetScrollPos(wx.VERTICAL)-clientsize/3/self.frameCrf.GetScrollPixelsPerUnit()[1])
        try: 
            self.lastfocusvariable=event.GetEventObject().GetName()
            self.lastfocusvariableobj=event.GetEventObject()
        except: pass
            
    def OnInputFieldSetTooltip(self, event):
        event.Skip()
        obj=event.GetEventObject()
        objname=obj.GetName()
        if objname in self.arrHeadersLabels:
            self.parent.SetStatusText(objname+": "+self.arrHeadersLabels[objname],2)
    def OnInputFieldDelTooltip(self, event):
        event.Skip()
        self.parent.SetStatusText("",2)
        
    def onPopupMenu(self, event):
        o = event.GetEventObject()
        sx,sy       = self.frameCrf.ScreenToClient(o.GetPositionTuple())
        dx,dy       = self.frameCrf.ScreenToClient(wx.GetMousePosition())
        try: 
            self.lastfocusvariable=event.GetEventObject().GetName()
            self.lastfocusvariableobj=event.GetEventObject()
        except: pass
        
        #self.frameCrf.PopupMenu(menu, o.GetPosition())
        
        ##self.frameCrf.PopupMenu(menu, wx.Point(self.frameCrf.ClientToScreen(o.GetPosition().x,self.frameCrf.ClientToScreen(o.GetPosition().y-100))))
        yunit, xunit = self.frameCrf.GetScrollPixelsPerUnit()
        yoff = self.frameCrf.GetScrollPos(wx.VERTICAL) * yunit
        #yoff = self.frameCrf.GetScrollPos(wx.VERTICAL) * yunit
        #print str(yunit), str(yoff), str(o.GetPosition().y)
        menu = MyPopupMenu(o.GetName(), self, o, event)
        #self.frameCrf.PopupMenu(menu, wx.Point(o.GetPosition().x, o.GetPosition().y+30-yoff))
        #self.frameCrf.PopupMenu(menu, o.ScreenToClient(wx.GetMousePosition()))
        self.frameCrf.PopupMenu(menu, self.frameCrf.ScreenToClient(wx.GetMousePosition()))
        
        menu.Destroy()
        
    def GetReportInVarCombination(seft,digits,bases,idx):
        # Costruzione massimi numeri raggiungibili con "i" cifre
        mux=1
        tops={}
        tops[1]=1
        for i in range(2,digits+1):
            mux=mux*bases[i-1]
            tops[i]=mux
        # Costruzione combinazione
        comb={}
        part=0
        i1=digits;
        while i1>=1:
            c=0
            for i2 in range(0,bases[i1]):
                if part+tops[i1]*i2<=idx:
                    c=i2
            part+=tops[i1]*c
            comb[i1]=c+1
            i1-=1
        return comb

    def OnReportChoice(self,event):
        event.Skip
        wx.BeginBusyCursor()
        selection=self.reportChoice.GetSelection()
        id_profile=0
        if selection>0:
            id_profile=self.arrReports[selection]
        else:
            id_profile=0

        if id_profile==0:
            return

        cur.execute("SELECT ylabel,xlabel,outcome FROM reports_profiles WHERE id_profile="+`id_profile`)
        row=cur.fetchone()
        ylabel=row[0]
        xlabel=row[1]
        outcome=row[2]

        #Report type (supported types: 'survival', 'export')
        reportType="survival"
        try:
            cur.execute("SELECT report_type FROM reports_profiles WHERE id_profile="+`id_profile`)
            row=cur.fetchone()
            reportType=row[0]
        except:
            pass
        if (reportType==""): reportType="export"

        ccind=0
        indName={}
        hashName={}
        indCount={}
        cur.execute("SELECT DISTINCT variable_name FROM reports_details WHERE id_profile="+`id_profile`+" AND inout='in'")
        for row in cur:
            ccind=ccind+1
            indName[ccind]=row[0]
            hashName[row[0]]=ccind
            indCount[ccind]=0
        ccin=0
        inValue={}
        inLabel={}
        ccout=0
        outName={}
        outValue={}
        outLabel={}
        lstfields=[]
        cur.execute("SELECT inout, variable_name, variable_value, variable_label FROM reports_details WHERE id_profile="+`id_profile`)
        for row in cur:
            if row[0]=="out":
                ccout=ccout+1
                outName[ccout]=row[1]
                outValue[ccout]=row[2]
                outLabel[ccout]=row[3]
                lstfields.append(row[1])
            if row[0]=="in":
                ccin=ccin+1
                indCount[hashName[row[1]]]=indCount[hashName[row[1]]]+1
                inValue[str(hashName[row[1]])+","+str(indCount[hashName[row[1]]])]=row[2]
                inLabel[str(hashName[row[1]])+","+str(indCount[hashName[row[1]]])]=row[3]

        report=""

        #survival - begin
        if (reportType=="survival"):
            if ccind==0:
                for i1 in range(1,ccout+1):
                    if i1>1:
                        report+="UNION"+"\n"
                    report+="SELECT '"+outLabel[i1]+"' AS "+ylabel+", "+outcome+" AS "+xlabel+"\n"
                    report+="FROM"+"\n"
                    report+="(SELECT id_row FROM rows WHERE status=0) AS r"+"\n"
                    report+="LEFT JOIN (SELECT id_row, data FROM headers INNER JOIN contents ON headers.id_header=contents.id_header AND headers.description='"+outName[i1]+"') AS c1 ON r.id_row=c1.id_row"+"\n"
                    report+="WHERE c1.data IN ("+outValue[i1]+")"+"\n"
            else:
                mux=1
                for i in range(1,ccind+1):
                    mux=mux*indCount[i]

                #intestazione - inizio
                report+="SELECT col1."+ylabel+" AS "+ylabel
                for i in range(mux):
                    comb=self.GetReportInVarCombination(ccind,indCount,i);

                    report+=", col"+`i+1`+"."+xlabel+" AS "
                    for i1 in range(1,ccind+1):
                        if i1>1:
                            report+="_"
                        report+=indName[i1]+"_"+inLabel[`i1`+","+`comb[i1]`]
                report+="\nFROM\n"
                #intestazione - fine
                for i in range(mux):
                    comb=self.GetReportInVarCombination(ccind,indCount,i);
                    if i>0:
                        report+="INNER JOIN\n"
                    report+="(\n"
                    #common - begin
                    for i1 in range(1,ccout+1):
                        if i1>1:
                            report+="UNION"+"\n"
                        report+="SELECT '"+outLabel[i1]+"' AS "+ylabel+", "+outcome+" AS "+xlabel+"\n"
                        report+="FROM"+"\n"
                        report+="(SELECT id_row FROM rows WHERE status=0) AS r"+"\n"
                        report+="LEFT JOIN (SELECT id_row, data FROM headers INNER JOIN contents ON headers.id_header=contents.id_header AND headers.description='"+outName[i1]+"') AS c1 ON r.id_row=c1.id_row"+"\n"
                        for i2 in range(1,ccind+1):
                            report+="LEFT JOIN (SELECT id_row, data FROM headers INNER JOIN contents ON headers.id_header=contents.id_header AND headers.description='"+indName[i2]+"') AS c"+`i2+1`+" ON r.id_row=c"+`i2+1`+".id_row"+"\n"
                        report+="WHERE c1.data IN ("+outValue[i1]+")"+"\n"
                        for i2 in range(1,ccind+1):
                            report+=" AND c"+`i2+1`+".data IN ("+inValue[`i2`+","+`comb[i2]`]+")"
                        report+="\n"
                    #common - end
                    report+=") AS col"+`i+1`+"\n"
                for i2 in range(1,mux):
                    if i2==1:
                        report+="ON "
                    else:
                        report+=" AND "
                    report+="col1."+ylabel+"=col"+`i2+1`+"."+ylabel
                report+=";\n"
        #survival - end
        #export - begin
        if (reportType=="export"):
            try: cur.execute("DROP TABLE export")
            except: pass
            square(con,con,"export",lstfields,self.id_group,self.id_crf)
            report="SELECT * FROM export"

        cur.execute(report)

        #trans={}
        if self.report_format=="text":
            reportname="report_"+`id_profile`+".txt"
            theFile = open(REPORTS_PATH+reportname, 'w')
            currow=""
            for fieldDesc in cur.description:
                if currow!="":
                    currow+="\t"
                currow+=str(fieldDesc[0])
            theFile.write(currow + '\n')
            for row in cur:
                currow=""
                for i in range(len(row)):
                    if currow!="":
                        currow+="\t"
                    dData=row[i]
                    if dData==None:
                        dData=""
                    elif i>2:
                        if not self.parent.trans.has_key(row[i]):
                            try:
                                dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(row[i])),"latin-1")
                            except:
                                dData=row[i]
#                                try: print "-"+row[i]+"-"
#                                except: pass
                            self.parent.trans[row[i]]=dData
                        else:
                            dData=self.parent.trans[row[i]]
                    currow+=dData
                theFile.write(currow + '\n')
            theFile.close
        elif self.report_format=="excel":
            reportname="report_"+`id_profile`+".xls"
            try: w=xlwt.Workbook()
            except: w=Workbook()
            ws=w.add_sheet('data')
            ccol=0
            for fieldDesc in cur.description:
                ws.write(0,ccol,fieldDesc[0])
                ccol+=1
            crow=1
            for row in cur:
                ccol=0
                for i in range(len(row)):
                    dData=row[i]
                    if dData==None:
                        dData=""
                    elif i>2:
                        if not self.parent.trans.has_key(row[i]):
                            try:
                                dData=unicode(rijndael.DecryptData(self.digestkey,base64.b64decode(row[i])),"latin-1")
                            except:
                                dData=row[i]
#                                try: print "-"+row[i]+"-"
#                                except: pass
                            self.parent.trans[row[i]]=dData
                        else:
                            dData=self.parent.trans[row[i]]
                    try:
                        ws.write(crow,ccol,dData)
                    except:
                        ws.write(crow,ccol,str(dData))
                    ccol+=1
                crow+=1
            w.save(REPORTS_PATH+reportname)
#        cur.close()
#        wx.MessageBox("File '"+os.path.abspath(os.path.dirname(reportname))+os.path.sep+reportname+"' created.", "Report", wx.ICON_INFORMATION | wx.OK, self)
        self.reportChoice.SetStringSelection("")
        wx.EndBusyCursor()

        #wx.MessageBox("File '"+reportname+"' created.", "Report", wx.ICON_INFORMATION | wx.OK, self)
        if not assoc_open(REPORTS_PATH+reportname):
            wx.MessageBox("File '"+reportname+"' created.", "Report", wx.ICON_INFORMATION | wx.OK, self)

    def Attach_Add(self):
        
        from os.path import expanduser
        home = expanduser("~")
        
        filename=""
        dlg = wx.FileDialog(
            self, message="Choose a file",
            defaultDir=home, 
            defaultFile="",
            wildcard="All files (*.*)|*.*",
            style=wx.OPEN
            )    
        ret = dlg.ShowModal()    
        if ret == wx.ID_OK:
            wx.BeginBusyCursor()

            #attach_dir=DATABASE_PATH+"attach"+os.path.sep+str(self.group_shortcode)+os.path.sep+str(self.curkeyvalue)
            attach_dir=DATABASE_PATH+"attach"+os.path.sep+self.contents["group_shortcode"]+os.path.sep+str(self.curkeyvalue)
            secondkeyname=GetSetting(self.parent,'secondkeyname')
            if secondkeyname!="":
                secondkeyvalue="1"
                if self.contents.has_key(secondkeyname): secondkeyvalue=self.contents[secondkeyname]
                attach_dir=attach_dir+os.path.sep+secondkeyvalue
            attach_dir=attach_dir+os.path.sep+self.FormName[self.ifrm]
            if self.ActivateCyclicity:
                if self.Cyclicity[self.ifrm+1]>0: attach_dir=attach_dir+os.path.sep+str(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1)
            try: os.makedirs(attach_dir)
            except: pass
            path = dlg.GetPath()
            filename=dlg.GetFilename()
            f_in = open(path, 'rb')
            content_in = f_in.read()
            f_in.close()

            #from gzip import GzipFile
            #from StringIO import StringIO
            #sio = StringIO()
            #gzf = GzipFile(fileobj=sio, mode='wb')
            #gzf.write(content_in)
            #gzf.close()
            zipstring=HB_ZipString(content_in)
            
            #content_out=rijndael.EncryptData(self.digestkey,sio.getvalue())
            
            # crypt algo begin
            #content_out=HB_EncryptOne(self.digestkey,zipstring)
            #outputfilename=attach_dir+os.path.sep+filename+".aes"
            content_out=xor_crypt_string(zipstring, self.digestkey)
            outputfilename=attach_dir+os.path.sep+filename+".hb"
            # crypt algo end
            
            f_out = open(outputfilename, 'wb')
            f_out.write(content_out)
            f_out.close()
            fsp = FileSplitter()
            fsp.split(outputfilename,0)
            os.unlink(outputfilename)

            wx.EndBusyCursor()
        dlg.Destroy()
        return filename

    def Attach_View(self,filename):
        outcome="error"
        #attach_dir=DATABASE_PATH+"attach"+os.path.sep+str(self.group_shortcode)+os.path.sep+str(self.curkeyvalue)
        attach_dir=DATABASE_PATH+"attach"+os.path.sep+self.contents["group_shortcode"]+os.path.sep+str(self.curkeyvalue)
        secondkeyname=GetSetting(self.parent,'secondkeyname')
        if secondkeyname!="":
            secondkeyvalue="1"
            if self.contents.has_key(secondkeyname): secondkeyvalue=self.contents[secondkeyname]
            attach_dir=attach_dir+os.path.sep+secondkeyvalue
        attach_dir=attach_dir+os.path.sep+self.FormName[self.ifrm]
        if self.ActivateCyclicity:
            if self.Cyclicity[self.ifrm+1]>0: attach_dir=attach_dir+os.path.sep+str(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1)

        # crypt algo begin
        #inputfilename=attach_dir+os.path.sep+filename+".aes"
        inputfilename=attach_dir+os.path.sep+filename+".hb"
        if DEBUG_MODE: print "Retrieving "+inputfilename
        # crypt algo end
        fsp = FileSplitter()
        if not fsp.combine(inputfilename):
            wx.MessageBox("File "+filename+" "+TT("not available"), TT("File extraction cancelled"), wx.ICON_ERROR | wx.OK, self)
            try: wx.EndBusyCursor()
            except: pass
            return outcome
            
        import tempfile
        dlgpath = os.path.normpath(tempfile.gettempdir() + os.path.sep + filename)
        if(os.path.exists(dlgpath)):
            os.unlink(dlgpath)
                         
        #from os.path import expanduser
        #home = expanduser("~")
        #dlg = wx.FileDialog(
            #self, message="Save file as ...", 
            #defaultDir=home, 
            #defaultFile=filename, 
            #wildcard="All files (*.*)|*.*",
            #style=wx.SAVE
            #)
        #ret = dlg.ShowModal()    
        #if ret == wx.ID_OK:
        if True:
            wx.BeginBusyCursor()

            #path = dlg.GetPath()
            path=dlgpath

            f_in = open(inputfilename, 'rb')
            content_in = f_in.read()
            f_in.close()
            os.unlink(inputfilename)
            #content_out=rijndael.DecryptData(self.digestkey,content_in)
            
            # crypt algo begin
            #content_out=HB_DecryptOne(self.digestkey,content_in,"")
            content_out=xor_crypt_string(content_in,self.digestkey)
            # crypt algo end

            from gzip import GzipFile
            from StringIO import StringIO
            sio = StringIO(content_out)
            gz = GzipFile(fileobj=sio, mode='rb')
            content_out2=gz.read()

            f_out = open(path, 'wb')
            #f_out.write(content_out)
            f_out.write(content_out2)
            f_out.close()
            outcome="ok"
            if wx.MessageBox(TT("Do you want to open the extracted file?"), TT("File extract"), wx.YES_NO) == wx.YES:
                if not assoc_open(path):
                    if DEBUG_MODE: print "file extracted but impossible to open"
                    wx.MessageBox("File "+path+" " + TT("extracted but impossible to open"), TT("File extract"), wx.ICON_ERROR | wx.OK, self)
            else:
                from os.path import expanduser
                home = expanduser("~")
                dlg = wx.FileDialog(
                    self, message="Save file as ...", 
                    defaultDir=home, 
                    defaultFile=filename, 
                    wildcard="All files (*.*)|*.*",
                    style=wx.SAVE
                    )
                ret = dlg.ShowModal()    
                if ret == wx.ID_OK:
                    path = dlg.GetPath()
                    f_out = open(path, 'wb')
                    f_out.write(content_out2)
                    f_out.close()
                    wx.MessageBox("File "+path+" " + TT("created."), TT("File extract"), wx.ICON_INFORMATION | wx.OK, self)
                dlg.Destroy()

            wx.EndBusyCursor()
        else:
            outcome="cancelled"
        return outcome

    def Attach_Del(self,filename):    
        outcome="error"
        #attach_dir=DATABASE_PATH+"attach"+os.path.sep+str(self.group_shortcode)+os.path.sep+str(self.curkeyvalue)
        attach_dir=DATABASE_PATH+"attach"+os.path.sep+self.contents["group_shortcode"]+os.path.sep+str(self.curkeyvalue)
        secondkeyname=GetSetting(self.parent,'secondkeyname')
        if secondkeyname!="":
            secondkeyvalue="1"
            if self.contents.has_key(secondkeyname): secondkeyvalue=self.contents[secondkeyname]
            attach_dir=attach_dir+os.path.sep+secondkeyvalue
        attach_dir=attach_dir+os.path.sep+self.FormName[self.ifrm]
        if self.ActivateCyclicity:
            if self.Cyclicity[self.ifrm+1]>0: attach_dir=attach_dir+os.path.sep+str(wx.xrc.XRCCTRL(self.frameRecordset,'cycles').GetSelection()+1)
        # if not os.path.exists(attach_dir+os.path.sep+filename):
        #     wx.MessageBox("File "+filename+" not available", "File delete cancelled", wx.ICON_ERROR | wx.OK, self)
        #     return outcome

        # crypt algo begin
        #inputfilename=attach_dir+os.path.sep+filename+".aes"
        inputfilename=attach_dir+os.path.sep+filename+".hb"
        # crypt algo end
        
        fsp = FileSplitter()
        if not fsp.remove(inputfilename):
             wx.MessageBox("File "+filename+" "+TT("not available"), TT("File delete cancelled"), wx.ICON_ERROR | wx.OK, self)
             return outcome
        else:
            try: os.unlink(inputfilename)
            except: pass
            outcome="ok"
        return outcome


class SettingsFrame(wx.aui.AuiMDIChildFrame):
    def __init__(self, parent, childTitle):
        wx.aui.AuiMDIChildFrame.__init__(self, parent, -1, title=childTitle)
        self.parent=parent
        self.Bind(wx.EVT_SIZE, self.OnSize)

        self.settingsPanel = SettingsPanel(self)
        self.settingsPanel.SetSize(self.GetClientSize())

    def OnSize(self, evt):
        self.settingsPanel.SetSize(self.GetClientSize())

class SettingsPanel(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent, -1)
        self.parent=parent
        self.Bind(wx.EVT_SIZE, self.OnSize)

        self.settingsWindow = wx.ScrolledWindow(parent=self, pos=wx.Point(0, 0), style=wx.VSCROLL)
#        self.grid = SimpleGrid(self.settingsWindow,6,5)
        self.grid = SimpleGrid(self.settingsWindow,6,5)
        self.InitGrid()

        self.buttonOk = wx.Button(label='Ok', parent=self, size=wx.Size(100, 30), style=0)
        self.buttonOk.Bind(wx.EVT_BUTTON, self.OnButtonOk)
        self.buttonCancel = wx.Button(label=TT('Cancel'), parent=self, size=wx.Size(100, 30), style=0)
        self.buttonCancel.Bind(wx.EVT_BUTTON, self.OnButtonCancel)


    def InitGrid(self):
        #Headings - Begin
        self.grid.SetRowLabelSize(50)
        self.grid.SetDefaultColSize(120)

        self.grid.SetColLabelValue(0, "Section")
        self.grid.SetColLabelValue(1, "Subsection")
        self.grid.SetColLabelValue(2, "Parameter")
        self.grid.SetColLabelValue(3, "Default")
        self.grid.SetColLabelValue(4, "Value")
        attr = gridlib.GridCellAttr()
        attr.SetTextColour(wx.BLACK)
        attr.SetBackgroundColour(wx.LIGHT_GREY)
        attr.SetFont(wx.Font(10, wx.SWISS, wx.NORMAL, wx.BOLD))
        attr.SetReadOnly(True)
        self.grid.SetColAttr(0, attr)
        self.grid.SetColAttr(1, attr)
        self.grid.SetColAttr(2, attr)
        self.grid.SetColAttr(3, attr)

        self.grid.SetCellValue(0, 0, "Network")
        self.grid.SetCellValue(0, 1, "Proxy")
        self.grid.SetCellValue(0, 2, "Setting")
        self.grid.SetCellValue(0, 3, "automatic")
        self.grid.SetCellValue(0, 4, GetSetting(self.parent.parent,"Network_Proxy_Setting"))
        self.grid.SetCellEditor(0, 4, gridlib.GridCellChoiceEditor(["automatic","no proxy","manual proxy","NTLM proxy"], False))

        self.grid.SetCellValue(1, 0, "Network")
        self.grid.SetCellValue(1, 1, "Proxy")
        self.grid.SetCellValue(1, 2, "Host")
        self.grid.SetCellValue(1, 3, "")
        self.grid.SetCellValue(1, 4, GetSetting(self.parent.parent,"Network_Proxy_Host"))

        self.grid.SetCellValue(2, 0, "Network")
        self.grid.SetCellValue(2, 1, "Proxy")
        self.grid.SetCellValue(2, 2, "Port")
        self.grid.SetCellValue(2, 3, "")
        self.grid.SetCellValue(2, 4, GetSetting(self.parent.parent,"Network_Proxy_Port"))

        self.grid.SetCellValue(3, 0, "Network")
        self.grid.SetCellValue(3, 1, "Proxy")
        self.grid.SetCellValue(3, 2, "Username")
        self.grid.SetCellValue(3, 3, "")
        self.grid.SetCellValue(3, 4, GetSetting(self.parent.parent,"Network_Proxy_Username"))

        self.grid.SetCellValue(4, 0, "Network")
        self.grid.SetCellValue(4, 1, "Proxy")
        self.grid.SetCellValue(4, 2, "Password")
        self.grid.SetCellValue(4, 3, "")
        self.grid.SetCellValue(4, 4, GetSetting(self.parent.parent,"Network_Proxy_Password"))
        
        self.grid.SetCellValue(5, 0, "Network")
        self.grid.SetCellValue(5, 1, "Connection")
        self.grid.SetCellValue(5, 2, "Mode")
        self.grid.SetCellValue(5, 3, "automatic")
        self.grid.SetCellValue(5, 4, GetSetting(self.parent.parent,"Network_Connection_Mode"))
        self.grid.SetCellEditor(5, 4, gridlib.GridCellChoiceEditor(["automatic","use xmlrpc","use http","use MSIE"], False))

    def OnButtonOk(self, event):
        event.Skip()
        self.grid.SetGridCursor(0,0)
        SaveCustomSetting("Network_Proxy_Setting",self.grid.GetCellValue(0,4))
        SaveCustomSetting("Network_Proxy_Host",self.grid.GetCellValue(1,4))
        SaveCustomSetting("Network_Proxy_Port",self.grid.GetCellValue(2,4))
        SaveCustomSetting("Network_Proxy_Username",self.grid.GetCellValue(3,4))
        SaveCustomSetting("Network_Proxy_Password",self.grid.GetCellValue(4,4))
        SaveCustomSetting("Network_Connection_Mode",self.grid.GetCellValue(5,4))
        self.parent.parent.reloadSettings()
        
        if GetSetting(self.parent.parent,"Network_Proxy_Setting")=="NTLM proxy":
            try:
                ntlmaps_server.terminate()
                ntlmaps_server.wait()
                print "NTLMAPS_SERVER stopped."
                SetNTLMAPSconfig()
                if os.name=='nt':
                    ntlmaps_server = subprocess.Popen(["pythonw","ntlmaps.py","-c","ntlmaps.cfg"])
                else:
                    ntlmaps_server = subprocess.Popen(["python","ntlmaps.py","-c","ntlmaps.cfg"])
                print "NTLMAPS_SERVER restarted."
            except:
                print "ERROR restarting NTLMAPS_SERVER."
        
        self.parent.parent.syncroTaskPause=False
        self.parent.Close()

    def OnButtonCancel(self, event):
        event.Skip()
        self.parent.parent.syncroTaskPause=False
        self.parent.Close()

    def OnSize(self, evt):
        self.settingsWindow.SetSize(wx.Size(self.GetClientSize().x,self.GetClientSize().y-50))
        self.grid.SetSize(self.GetClientSize())
        self.settingsWindow.SetScrollbars(10,10,1,50)
        self.buttonOk.SetPosition(wx.Point(self.GetClientSize().x/2-150,self.GetClientSize().y-40))
        self.buttonCancel.SetPosition(wx.Point(self.GetClientSize().x/2+50,self.GetClientSize().y-40))

class SimpleGrid(gridlib.Grid): ##, mixins.GridAutoEditMixin):
    def __init__(self, parent,rows,columns):
        gridlib.Grid.__init__(self, parent, -1)
        ##mixins.GridAutoEditMixin.__init__(self)
        self.moveTo = None

        self.Bind(wx.EVT_IDLE, self.OnIdle)

        #self.CreateGrid(25, 25)#, gridlib.Grid.SelectRows)
        self.CreateGrid(rows, columns)#, gridlib.Grid.SelectRows)
        ##self.EnableEditing(False)

        # test all the events
        self.Bind(gridlib.EVT_GRID_CELL_LEFT_CLICK, self.OnCellLeftClick)
        self.Bind(gridlib.EVT_GRID_CELL_RIGHT_CLICK, self.OnCellRightClick)
        self.Bind(gridlib.EVT_GRID_CELL_LEFT_DCLICK, self.OnCellLeftDClick)
        self.Bind(gridlib.EVT_GRID_CELL_RIGHT_DCLICK, self.OnCellRightDClick)

        self.Bind(gridlib.EVT_GRID_LABEL_LEFT_CLICK, self.OnLabelLeftClick)
        self.Bind(gridlib.EVT_GRID_LABEL_RIGHT_CLICK, self.OnLabelRightClick)
        self.Bind(gridlib.EVT_GRID_LABEL_LEFT_DCLICK, self.OnLabelLeftDClick)
        self.Bind(gridlib.EVT_GRID_LABEL_RIGHT_DCLICK, self.OnLabelRightDClick)

        self.Bind(gridlib.EVT_GRID_ROW_SIZE, self.OnRowSize)
        self.Bind(gridlib.EVT_GRID_COL_SIZE, self.OnColSize)

        self.Bind(gridlib.EVT_GRID_RANGE_SELECT, self.OnRangeSelect)
        self.Bind(gridlib.EVT_GRID_CELL_CHANGE, self.OnCellChange)
        self.Bind(gridlib.EVT_GRID_SELECT_CELL, self.OnSelectCell)

        self.Bind(gridlib.EVT_GRID_EDITOR_SHOWN, self.OnEditorShown)
        self.Bind(gridlib.EVT_GRID_EDITOR_HIDDEN, self.OnEditorHidden)
        self.Bind(gridlib.EVT_GRID_EDITOR_CREATED, self.OnEditorCreated)


    def OnCellLeftClick(self, evt):
#        self.log.write("OnCellLeftClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnCellRightClick(self, evt):
#        self.log.write("OnCellRightClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnCellLeftDClick(self, evt):
#        self.log.write("OnCellLeftDClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnCellRightDClick(self, evt):
#        self.log.write("OnCellRightDClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelLeftClick(self, evt):
#        self.log.write("OnLabelLeftClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelRightClick(self, evt):
#        self.log.write("OnLabelRightClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelLeftDClick(self, evt):
#        self.log.write("OnLabelLeftDClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnLabelRightDClick(self, evt):
#        self.log.write("OnLabelRightDClick: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()

    def OnRowSize(self, evt):
#        self.log.write("OnRowSize: row %d, %s\n" %
#                       (evt.GetRowOrCol(), evt.GetPosition()))
        evt.Skip()

    def OnColSize(self, evt):
#        self.log.write("OnColSize: col %d, %s\n" %
#                       (evt.GetRowOrCol(), evt.GetPosition()))
        evt.Skip()

    def OnRangeSelect(self, evt):
        if evt.Selecting():
#            self.log.write("OnRangeSelect: top-left %s, bottom-right %s\n" %
#                           (evt.GetTopLeftCoords(), evt.GetBottomRightCoords()))
            pass
        evt.Skip()


    def OnCellChange(self, evt):
#        self.log.write("OnCellChange: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))

        # Show how to stay in a cell that has bad data.  We can't just
        # call SetGridCursor here since we are nested inside one so it
        # won't have any effect.  Instead, set coordinates to move to in
        # idle time.
        value = self.GetCellValue(evt.GetRow(), evt.GetCol())

        if value == 'no good':
            self.moveTo = evt.GetRow(), evt.GetCol()


    def OnIdle(self, evt):
        if self.moveTo != None:
            self.SetGridCursor(self.moveTo[0], self.moveTo[1])
            self.moveTo = None

        evt.Skip()


    def OnSelectCell(self, evt):
#        self.log.write("OnSelectCell: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))

        # Another way to stay in a cell that has a bad value...
        row = self.GetGridCursorRow()
        col = self.GetGridCursorCol()

        if self.IsCellEditControlEnabled():
            self.HideCellEditControl()
            self.DisableCellEditControl()

        value = self.GetCellValue(row, col)

        if value == 'no good 2':
            return  # cancels the cell selection

        evt.Skip()


    def OnEditorShown(self, evt):
        if evt.GetRow() == 6 and evt.GetCol() == 3 and \
           wx.MessageBox("Are you sure you wish to edit this cell?",
                        "Checking", wx.YES_NO) == wx.NO:
            evt.Veto()
            return

#        self.log.write("OnEditorShown: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()


    def OnEditorHidden(self, evt):
        if evt.GetRow() == 6 and evt.GetCol() == 3 and \
           wx.MessageBox("Are you sure you wish to  finish editing this cell?",
                        "Checking", wx.YES_NO) == wx.NO:
            evt.Veto()
            return

#        self.log.write("OnEditorHidden: (%d,%d) %s\n" %
#                       (evt.GetRow(), evt.GetCol(), evt.GetPosition()))
        evt.Skip()


    def OnEditorCreated(self, evt):
 #       self.log.write("OnEditorCreated: (%d, %d) %s\n" %
 #                      (evt.GetRow(), evt.GetCol(), evt.GetControl()))
        pass

class DataExtractionPanel ( wx.Panel ):
    
    def __init__( self, parent ):
        wx.Panel.__init__  ( self, parent, id = wx.ID_ANY, pos = wx.DefaultPosition, size = wx.Size( 900,450 ), style = wx.TAB_TRAVERSAL )
        
        fgSizer1 = wx.FlexGridSizer( 2, 1, 0, 0 )
        fgSizer1.AddGrowableCol( 0 )
        fgSizer1.AddGrowableRow( 1 )
        fgSizer1.SetFlexibleDirection( wx.BOTH )
        fgSizer1.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
        
        self.m_staticText2 = wx.StaticText( self, wx.ID_ANY, TT("DATA EXTRACTION"), wx.DefaultPosition, wx.DefaultSize, wx.ALIGN_CENTRE )
        self.m_staticText2.Wrap( -1 )
        self.m_staticText2.SetFont( wx.Font( 10, 74, 90, 92, False, "Sans" ) )
        
        fgSizer1.Add( self.m_staticText2, 0, wx.ALL|wx.ALIGN_CENTER_HORIZONTAL, 5 )
        
        fgSizer4 = wx.FlexGridSizer( 2, 2, 0, 0 )
        fgSizer4.AddGrowableCol( 0 )
        fgSizer4.AddGrowableCol( 1 )
        fgSizer4.AddGrowableRow( 0 )
        fgSizer4.SetFlexibleDirection( wx.BOTH )
        fgSizer4.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
        
        fgSizer2 = wx.FlexGridSizer( 2, 1, 0, 0 )
        fgSizer2.AddGrowableCol( 0 )
        fgSizer2.AddGrowableRow( 2 )
        fgSizer2.SetFlexibleDirection( wx.BOTH )
        fgSizer2.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
        
        self.m_staticText1 = wx.StaticText( self, wx.ID_ANY, TT("AVAILABLE FIELDS"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText1.Wrap( -1 )
        fgSizer2.Add( self.m_staticText1, 0, wx.ALL, 5 )
        
        #optFieldOrderChoices = [ u"alpha", u"pos" ]
        #optFieldOrderChoices = [ TT("alpha"), TT("pos"), TT("alpha, all vars"), TT("pos, all vars") ]
        #self.optFieldOrder = wx.RadioBox( self, wx.ID_ANY, TT("order by"), wx.DefaultPosition, wx.DefaultSize, optFieldOrderChoices, 2, wx.RA_SPECIFY_COLS )
        #self.optFieldOrder.SetSelection( 0 )
        optFieldOrderChoices = [ TT("alphabetic, non empty variables only"), TT("by position, non empty variables only"), TT("alphabetic, all representative vars."), TT("by position, all representative vars."), TT("alphabetic, all variables"), TT("by position, all variables") ]
        self.optFieldOrder = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, optFieldOrderChoices, 0)
        self.optFieldOrder.SetSelection( 0 )
        fgSizer2.Add( self.optFieldOrder, 0, wx.ALL|wx.EXPAND, 5 )
        
        select_fieldsChoices = []
        if sys.platform == 'darwin':
            self.select_fields = wx.ListBox( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( -1,-1 ), select_fieldsChoices, wx.LB_EXTENDED )
        else:
            self.select_fields = wx.CheckListBox( self, wx.ID_ANY, wx.DefaultPosition, wx.Size( -1,-1 ), select_fieldsChoices)
        fgSizer2.Add( self.select_fields, 0, wx.ALL|wx.EXPAND, 5 )
        
        fgSizer4.Add( fgSizer2, 1, wx.EXPAND, 5 )
        
        fgSizer5 = wx.FlexGridSizer( 4, 1, 0, 0 )
        fgSizer5.AddGrowableCol( 0 )
        fgSizer5.AddGrowableRow( 3 )
        fgSizer5.SetFlexibleDirection( wx.BOTH )
        fgSizer5.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
        
        self.m_staticText4 = wx.StaticText( self, wx.ID_ANY, TT("QUERY BUILDER"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText4.Wrap( -1 )
        fgSizer5.Add( self.m_staticText4, 0, wx.ALL, 5 )
        
        fgSizer6 = wx.FlexGridSizer( 6, 7, 0, 0 )
        fgSizer6.AddGrowableCol( 1 )
        fgSizer6.AddGrowableCol( 4 )
        fgSizer6.SetFlexibleDirection( wx.BOTH )
        fgSizer6.SetNonFlexibleGrowMode( wx.FLEX_GROWMODE_SPECIFIED )
        
        
        fgSizer6.AddSpacer( ( 0, 0), 1, wx.EXPAND, 2 )
        
        self.m_staticText91 = wx.StaticText( self, wx.ID_ANY, TT("field")+" 1", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText91.Wrap( -1 )
        fgSizer6.Add( self.m_staticText91, 0, wx.ALL, 2 )
        
        
        self.m_staticText10 = wx.StaticText( self, wx.ID_ANY, TT("condition"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText10.Wrap( -1 )
        fgSizer6.Add( self.m_staticText10, 0, wx.ALL, 2 )
        
        self.m_staticText6 = wx.StaticText( self, wx.ID_ANY, TT("type")+" elm2", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText6.Wrap( -1 )
        fgSizer6.Add( self.m_staticText6, 0, wx.ALL, 2 )
        
        self.m_staticText7 = wx.StaticText( self, wx.ID_ANY, TT("value or field")+" 2", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText7.Wrap( -1 )
        fgSizer6.Add( self.m_staticText7, 0, wx.ALL, 2 )
        
        
        fgSizer6.AddSpacer( ( 0, 0), 1, wx.EXPAND, 2 )
        
        self.m_staticText8 = wx.StaticText( self, wx.ID_ANY, u"and/or", wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText8.Wrap( -1 )
        fgSizer6.Add( self.m_staticText8, 0, wx.ALL, 2 )
        
        filter_pre1Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre1 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre1Choices, 0 )
        self.filter_pre1.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre1, 0, wx.ALL, 2 )
        
        filter_field1Choices = []
        self.filter_field1 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field1Choices, 0 )
        fgSizer6.Add( self.filter_field1, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond1Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond1 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond1Choices, 0 )
        self.filter_cond1.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond1, 0, wx.ALL, 2 )
        
        filter_type1Choices = [ u"value", u"field" ]
        self.filter_type1 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type1Choices, 0 )
        self.filter_type1.SetSelection( 0 )
        fgSizer6.Add( self.filter_type1, 0, wx.ALL, 2 )
        
        filter_value1Choices = []
        self.filter_value1 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value1Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value1, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post1Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post1 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post1Choices, 0 )
        self.filter_post1.SetSelection( 0 )
        fgSizer6.Add( self.filter_post1, 0, wx.ALL, 2 )
        
        filter_andor1Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor1 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor1Choices, 0 )
        self.filter_andor1.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor1, 0, wx.ALL, 2 )
        
        filter_pre2Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre2 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre2Choices, 0 )
        self.filter_pre2.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre2, 0, wx.ALL, 2 )
        
        filter_field2Choices = []
        self.filter_field2 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field2Choices, 0 )
        fgSizer6.Add( self.filter_field2, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond2Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond2 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond2Choices, 0 )
        self.filter_cond2.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond2, 0, wx.ALL, 2 )
        
        filter_type2Choices = [ u"value", u"field" ]
        self.filter_type2 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type2Choices, 0 )
        self.filter_type2.SetSelection( 0 )
        fgSizer6.Add( self.filter_type2, 0, wx.ALL, 2 )
        
        filter_value2Choices = []
        self.filter_value2 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value2Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value2, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post2Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post2 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post2Choices, 0 )
        self.filter_post2.SetSelection( 0 )
        fgSizer6.Add( self.filter_post2, 0, wx.ALL, 2 )
        
        filter_andor2Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor2 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor2Choices, 0 )
        self.filter_andor2.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor2, 0, wx.ALL, 2 )
        
        filter_pre3Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre3 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre3Choices, 0 )
        self.filter_pre3.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre3, 0, wx.ALL, 2 )
        
        filter_field3Choices = []
        self.filter_field3 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field3Choices, 0 )
        fgSizer6.Add( self.filter_field3, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond3Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond3 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond3Choices, 0 )
        self.filter_cond3.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond3, 0, wx.ALL, 2 )
        
        filter_type3Choices = [ u"value", u"field" ]
        self.filter_type3 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type3Choices, 0 )
        self.filter_type3.SetSelection( 0 )
        fgSizer6.Add( self.filter_type3, 0, wx.ALL, 2 )
        
        filter_value3Choices = []
        self.filter_value3 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value3Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value3, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post3Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post3 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post3Choices, 0 )
        self.filter_post3.SetSelection( 0 )
        fgSizer6.Add( self.filter_post3, 0, wx.ALL, 2 )
        
        filter_andor3Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor3 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor3Choices, 0 )
        self.filter_andor3.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor3, 0, wx.ALL, 2 )
        
        filter_pre4Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre4 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre4Choices, 0 )
        self.filter_pre4.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre4, 0, wx.ALL, 2 )
        
        filter_field4Choices = []
        self.filter_field4 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field4Choices, 0 )
        fgSizer6.Add( self.filter_field4, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond4Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond4 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond4Choices, 0 )
        self.filter_cond4.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond4, 0, wx.ALL, 2 )
        
        filter_type4Choices = [ u"value", u"field" ]
        self.filter_type4 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type4Choices, 0 )
        self.filter_type4.SetSelection( 0 )
        fgSizer6.Add( self.filter_type4, 0, wx.ALL, 2 )
        
        filter_value4Choices = []
        self.filter_value4 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value4Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value4, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post4Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post4 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post4Choices, 0 )
        self.filter_post4.SetSelection( 0 )
        fgSizer6.Add( self.filter_post4, 0, wx.ALL, 2 )
        
        filter_andor4Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor4 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor4Choices, 0 )
        self.filter_andor4.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor4, 0, wx.ALL, 2 )
        
        filter_pre5Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre5 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre5Choices, 0 )
        self.filter_pre5.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre5, 0, wx.ALL, 2 )
        
        filter_field5Choices = []
        self.filter_field5 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field5Choices, 0 )
        fgSizer6.Add( self.filter_field5, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond5Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond5 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond5Choices, 0 )
        self.filter_cond5.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond5, 0, wx.ALL, 2 )
        
        filter_type5Choices = [ u"value", u"field" ]
        self.filter_type5 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type5Choices, 0 )
        self.filter_type5.SetSelection( 0 )
        fgSizer6.Add( self.filter_type5, 0, wx.ALL, 2 )
        
        filter_value5Choices = []
        self.filter_value5 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value5Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value5, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post5Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post5 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post5Choices, 0 )
        self.filter_post5.SetSelection( 0 )
        fgSizer6.Add( self.filter_post5, 0, wx.ALL, 2 )
        
        filter_andor5Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor5 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor5Choices, 0 )
        self.filter_andor5.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor5, 0, wx.ALL, 2 )
        
        filter_pre6Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre6 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre6Choices, 0 )
        self.filter_pre6.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre6, 0, wx.ALL, 2 )
        
        filter_field6Choices = []
        self.filter_field6 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field6Choices, 0 )
        fgSizer6.Add( self.filter_field6, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond6Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond6 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond6Choices, 0 )
        self.filter_cond6.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond6, 0, wx.ALL, 2 )
        
        filter_type6Choices = [ u"value", u"field" ]
        self.filter_type6 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type6Choices, 0 )
        self.filter_type6.SetSelection( 0 )
        fgSizer6.Add( self.filter_type6, 0, wx.ALL, 2 )
        
        filter_value6Choices = []
        self.filter_value6 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value6Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value6, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post6Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post6 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post6Choices, 0 )
        self.filter_post6.SetSelection( 0 )
        fgSizer6.Add( self.filter_post6, 0, wx.ALL, 2 )
        
        filter_andor6Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor6 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor6Choices, 0 )
        self.filter_andor6.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor6, 0, wx.ALL, 2 )
        
        filter_pre7Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre7 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre7Choices, 0 )
        self.filter_pre7.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre7, 0, wx.ALL, 2 )
        
        filter_field7Choices = []
        self.filter_field7 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field7Choices, 0 )
        fgSizer6.Add( self.filter_field7, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond7Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond7 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond7Choices, 0 )
        self.filter_cond7.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond7, 0, wx.ALL, 2 )
        
        filter_type7Choices = [ u"value", u"field" ]
        self.filter_type7 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type7Choices, 0 )
        self.filter_type7.SetSelection( 0 )
        fgSizer6.Add( self.filter_type7, 0, wx.ALL, 2 )
        
        filter_value7Choices = []
        self.filter_value7 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value7Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value7, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post7Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post7 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post7Choices, 0 )
        self.filter_post7.SetSelection( 0 )
        fgSizer6.Add( self.filter_post7, 0, wx.ALL, 2 )
        
        filter_andor7Choices = [ wx.EmptyString, u"and", u"or" ]
        self.filter_andor7 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_andor7Choices, 0 )
        self.filter_andor7.SetSelection( 0 )
        fgSizer6.Add( self.filter_andor7, 0, wx.ALL, 2 )
        
        filter_pre8Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_pre8 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_pre8Choices, 0 )
        self.filter_pre8.SetSelection( 0 )
        fgSizer6.Add( self.filter_pre8, 0, wx.ALL, 2 )
        
        filter_field8Choices = []
        self.filter_field8 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_field8Choices, 0 )
        fgSizer6.Add( self.filter_field8, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_cond8Choices = [ u"=", u"<=", u">=", u"<", u">", u"<>", u"in", u"not in", u"contains" ]
        self.filter_cond8 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_cond8Choices, 0 )
        self.filter_cond8.SetSelection( 0 )
        fgSizer6.Add( self.filter_cond8, 0, wx.ALL, 2 )
        
        filter_type8Choices = [ u"value", u"field" ]
        self.filter_type8 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_type8Choices, 0 )
        self.filter_type8.SetSelection( 0 )
        fgSizer6.Add( self.filter_type8, 0, wx.ALL, 2 )
        
        filter_value8Choices = []
        self.filter_value8 = wx.ComboBox( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,-1 ), filter_value8Choices, wx.CB_SORT )
        fgSizer6.Add( self.filter_value8, 0, wx.ALL|wx.EXPAND, 2 )
        
        filter_post8Choices = [ wx.EmptyString, u"(", u")" ]
        self.filter_post8 = wx.Choice( self, wx.ID_ANY, wx.DefaultPosition, wx.DefaultSize, filter_post8Choices, 0 )
        self.filter_post8.SetSelection( 0 )
        fgSizer6.Add( self.filter_post8, 0, wx.ALL, 2 )
        
        
        fgSizer6.AddSpacer( ( 0, 0), 1, wx.EXPAND, 2 )
        
        fgSizer5.Add( fgSizer6, 1, wx.EXPAND, 5 )
        
        self.m_staticText9 = wx.StaticText( self, wx.ID_ANY, TT("FREE TEXT"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.m_staticText9.Wrap( -1 )
        fgSizer5.Add( self.m_staticText9, 0, wx.ALL, 2 )
        
        self.filter_freetext = wx.TextCtrl( self, wx.ID_ANY, wx.EmptyString, wx.DefaultPosition, wx.Size( -1,50 ), wx.HSCROLL|wx.TE_MULTILINE )
        fgSizer5.Add( self.filter_freetext, 0, wx.ALL|wx.EXPAND, 2 )
        
        fgSizer4.Add( fgSizer5, 1, wx.EXPAND, 5 )
        
        gSizer2 = wx.GridSizer( 2, 2, 0, 0 )
        
        self.cmdMarkAll = wx.Button( self, wx.ID_ANY, TT("mark all"), wx.DefaultPosition, wx.Size( 110,25 ), 0 )
        gSizer2.Add( self.cmdMarkAll, 0, wx.ALIGN_BOTTOM|wx.ALIGN_LEFT, 2 )
        
        self.cmdUnmarkAll = wx.Button( self, wx.ID_ANY, TT("unmark all"), wx.DefaultPosition, wx.Size( 110,25 ), 0 )
        gSizer2.Add( self.cmdUnmarkAll, 0, wx.ALIGN_BOTTOM|wx.ALIGN_RIGHT, 2 )
        
        fgSizer4.Add( gSizer2, 1, wx.EXPAND, 5 )
        
        gSizer1 = wx.GridSizer( 1, 2, 0, 0 )
        
        self.chkDistinct = wx.CheckBox( self, wx.ID_ANY, TT("use DISTINCT clause"), wx.DefaultPosition, wx.DefaultSize, 0 )
        self.chkDistinct.SetValue(True) 
        gSizer1.Add( self.chkDistinct, 0, wx.ALIGN_BOTTOM|wx.ALIGN_LEFT, 5 )
        
        self.cmdrun = wx.Button( self, wx.ID_ANY, TT("Run Query"), wx.DefaultPosition, wx.Size( 130,25 ), 0 )
        gSizer1.Add( self.cmdrun, 0, wx.ALIGN_BOTTOM|wx.ALIGN_RIGHT, 5 )
        
        fgSizer4.Add( gSizer1, 1, wx.EXPAND, 0 )
        
        fgSizer1.Add( fgSizer4, 1, wx.EXPAND, 5 )
        
        self.SetSizer( fgSizer1 )
        self.Layout()
    
    def __del__( self ):
        pass

#STATISTICS - INIZIO
class StatisticsFrame(wx.aui.AuiMDIChildFrame):
    def __init__(self, parent, childTitle):
        wx.aui.AuiMDIChildFrame.__init__(self, parent, -1, title=childTitle)
        self.parent=parent
        self.Bind(wx.EVT_SIZE, self.OnSize)

        self.statisticsPanel = StatisticsPanel(self)
        self.statisticsPanel.SetSize(self.GetClientSize())

    def OnSize(self, evt):
        self.statisticsPanel.SetSize(self.GetClientSize())

class StatisticsPanel(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent, -1)
        self.parent=parent
        self.Bind(wx.EVT_SIZE, self.OnSize)

        self.statisticsWindow = wx.ScrolledWindow(parent=self, pos=wx.Point(0, 0), style=wx.VSCROLL)
        self.buttonStartAnalysis = wx.Button(label='Start Analysis', parent=self.statisticsWindow, size=wx.Size(100, 30), style=0)
        self.buttonStartAnalysis.Bind(wx.EVT_BUTTON, self.OnButtonStartAnalysis)
        self.buttonOk = wx.Button(label='Close', parent=self, size=wx.Size(100, 30), style=0)
        self.buttonOk.Bind(wx.EVT_BUTTON, self.OnButtonOk)

    def OpenDataDir(self):
        dialog = wx.DirDialog(None, 'Choose a directory with data to be imported:', style = wx.DD_DEFAULT_STYLE)
        result = dialog.ShowModal()
        if (result == wx.ID_OK):
            path=dialog.GetPath()
            now = datetime.datetime.utcnow()
            ts=now.strftime("%Y%m%d%H%M%S")
            dbmergefullname=path+os.path.sep+"dbmerge_"+ts+".db"
            con = sqlite.connect(dbmergefullname, isolation_level=None)
            #con = sqlite.connect(":memory:")
            cur = con.cursor()
            initMergeDb(cur)
            #IMPORT
            for entry in custom_listdir(path):
                if entry.ljust(6)[0:6]=="export" and entry.rjust(3)[-3:]==".db": 
                    dbimportfullname=path+os.path.sep+entry
                    print "Importing "+dbimportfullname
                    con2 = sqlite.connect(dbimportfullname, isolation_level=None)
                    cur2 = con2.cursor()
                    #HEADERS
                    cur2.execute("SELECT id_header,child_of,description FROM headers")
                    for row in cur2:
                        cur.execute("SELECT * FROM headers WHERE id_header="+str(row[0])+" AND child_of="+str(row[1])+" AND description='"+row[2]+"'")
                        rowtmp = cur.fetchone()
                        if rowtmp==None:
                            cur.execute("INSERT INTO headers (id_header,child_of,description) VALUES ("+str(row[0])+","+str(row[1])+",'"+row[2]+"')")
                    # Cleaning of ROWS and CONTENTS
                    cur2.execute("SELECT DISTINCT id_user FROM rows WHERE status=0")
                    for row in cur2:
                        cur.execute("DELETE FROM rows WHERE id_user="+str(row[0]))
                        cur.execute("DELETE FROM contents WHERE id_user="+str(row[0]))
                    #ROWS
                    cur2.execute("SELECT id_row,id_header,status,date_time,id_user,rap FROM rows WHERE status=0")
                    for row in cur2:
                        cur.execute("INSERT INTO rows (id_row,id_header,status,date_time,id_user,rap) VALUES ("+str(row[0])+","+str(row[1])+","+str(row[2])+",'"+row[3]+"',"+str(row[4])+",'"+row[5]+"')")
                    #CONTENTS
                    i=0
                    cur2.execute("SELECT contents.id_row,contents.id_header,data,contents.id_user FROM rows INNER JOIN contents ON rows.id_row=contents.id_row AND rows.id_user=contents.id_user WHERE rows.status=0")
                    for row in cur2:
                        i=i+1
                        if ((i%1000)==0): print `i`,
                        if ((i%30000)==0): print ""
                        cur.execute("INSERT INTO contents (id_row,id_header,data,id_user) VALUES ("+str(row[0])+","+str(row[1])+",'"+GetSqlValue(row[2])+"',"+str(row[3])+")")
                    cur2.close()
                    con2.close()
                    print "Done."
            #ANALYSIS
            #1 - Prevalenza
            print "Creating View 1..."
            #square(con,"test1",["key","data_visita","malattia_base"])
            square(con,"step1",["key","data_visita","data_nascita","data_inizio_nad","malattia_base",
                                "indicazione_principale_nad","accesso_ne","accesso_np","tipo_cv","tipo_miscela_ne",
                                "composizione_miscela_ne","tipo_miscela_np","produzione_miscela_np","lipidi_np","glutamina_np",
                                "taurina_np","vitamine_np","oligoelementi_np","infusioni_x_settimana","fabbisogno_calorico_BEE",
                                "modalita_infusione","pompa_infusione","fornitura_domicilio_materiale","assistenza_infermieristica","gestione_domiciliare_nad",
                                "bmi","percentile_peso","percentile_altezza","stato_idratazione","stato_civile",
                                "indice_karnofsky","attivita_lavorativa_pz","capacita_lavorativa_pz"])
            print "Opening View 1..."
    #        cur.execute("SELECT count(*) AS tot, malattia_base FROM test1 GROUP BY malattia_base")
            cur.execute("SELECT id_user,count(distinct key) FROM step1 GROUP BY id_user")
            for row in cur:
                print row
            
            cur.close()
            con.close()
        else:
            wx.MessageBox("Cancelled.", "Report", wx.ICON_INFORMATION | wx.OK)
        dialog.Destroy()
    def OnSize(self, evt):
        self.statisticsWindow.SetSize(wx.Size(self.GetClientSize().x,self.GetClientSize().y-50))
        self.buttonOk.SetPosition(wx.Point(self.GetClientSize().x/2-50,self.GetClientSize().y-40))

    def OnButtonOk(self, event):
        event.Skip()
        self.parent.Close()

    def OnButtonStartAnalysis(self, event):
        event.Skip()
        self.OpenDataDir()
#STATISTICS - FINE

#CROSS PLATFORM OPEN - BEGIN
'''Utilities for opening files or URLs in the registered default application
and for sending e-mail using the user's preferred composer.

'''

import os
import sys
import webbrowser
import subprocess

_controllers = {}
_open = None


class BaseController(object):
    '''Base class for open program controllers.'''

    def __init__(self, name):
        self.name = name

    def open(self, filename):
        raise NotImplementedError


class Controller(BaseController):
    '''Controller for a generic open program.'''

    def __init__(self, *args):
        super(Controller, self).__init__(os.path.basename(args[0]))
        self.args = list(args)

    def _invoke(self, cmdline):
        if sys.platform[:3] == 'win':
            closefds = False
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        else:
            closefds = True
            startupinfo = None

        if (os.environ.get('DISPLAY') or sys.platform[:3] == 'win' or
                                                    sys.platform == 'darwin'):
            inout = file(os.devnull, 'r+')
        else:
            # for TTY programs, we need stdin/out
            inout = None

        # if possible, put the child precess in separate process group,
        # so keyboard interrupts don't affect child precess as well as
        # Python
        setsid = getattr(os, 'setsid', None)
        if not setsid:
            setsid = getattr(os, 'setpgrp', None)

        pipe = subprocess.Popen(cmdline, stdin=inout, stdout=inout,
                                stderr=inout, close_fds=closefds,
                                preexec_fn=setsid, startupinfo=startupinfo)

        # It is assumed that this kind of tools (gnome-open, kfmclient,
        # exo-open, xdg-open and open for OSX) immediately exit after lauching
        # the specific application
        returncode = pipe.wait()
        if hasattr(self, 'fixreturncode'):
            returncode = self.fixreturncode(returncode)
        return not returncode

    def open(self, filename):
        if isinstance(filename, basestring):
            cmdline = self.args + [filename]
        else:
            # assume it is a sequence
            cmdline = self.args + filename
        try:
            return self._invoke(cmdline)
        except OSError:
            return False


# Platform support for Windows
if sys.platform[:3] == 'win':

    class Start(BaseController):
        '''Controller for the win32 start progam through os.startfile.'''

        def open(self, filename):
            try:
                os.startfile(filename)
            except WindowsError:
                # [Error 22] No application is associated with the specified
                # file for this operation: '<URL>'
                return False
            else:
                return True

    _controllers['windows-default'] = Start('start')
    _open = _controllers['windows-default'].open


# Platform support for MacOS
elif sys.platform == 'darwin':
    _controllers['open']= Controller('open')
    _open = _controllers['open'].open


# Platform support for Unix
else:

    import commands

    # @WARNING: use the private API of the webbrowser module
    from webbrowser import _iscommand

    class KfmClient(Controller):
        '''Controller for the KDE kfmclient program.'''

        def __init__(self, kfmclient='kfmclient'):
            super(KfmClient, self).__init__(kfmclient, 'exec')
            self.kde_version = self.detect_kde_version()

        def detect_kde_version(self):
            kde_version = None
            try:
                info = commands.getoutput('kde-config --version')

                for line in info.splitlines():
                    if line.startswith('KDE'):
                        kde_version = line.split(':')[-1].strip()
                        break
            except (OSError, RuntimeError):
                pass

            return kde_version

        def fixreturncode(self, returncode):
            if returncode is not None and self.kde_version > '3.5.4':
                return returncode
            else:
                return os.EX_OK

    def detect_desktop_environment():
        '''Checks for known desktop environments

        Return the desktop environments name, lowercase (kde, gnome, xfce)
        or "generic"

        '''

        desktop_environment = 'generic'

        if os.environ.get('KDE_FULL_SESSION') == 'true':
            desktop_environment = 'kde'
        elif os.environ.get('GNOME_DESKTOP_SESSION_ID'):
            desktop_environment = 'gnome'
        else:
            try:
                info = commands.getoutput('xprop -root _DT_SAVE_MODE')
                if ' = "xfce4"' in info:
                    desktop_environment = 'xfce'
            except (OSError, RuntimeError):
                pass

        return desktop_environment


    def register_X_controllers():
        if _iscommand('kfmclient'):
            _controllers['kde-open'] = KfmClient()

        for command in ('gnome-open', 'exo-open', 'xdg-open'):
            if _iscommand(command):
                _controllers[command] = Controller(command)

    def get():
        controllers_map = {
            'gnome': 'gnome-open',
            'kde': 'kde-open',
            'xfce': 'exo-open',
        }

        desktop_environment = detect_desktop_environment()

        try:
            controller_name = controllers_map[desktop_environment]
            return _controllers[controller_name].open

        except KeyError:
            if _controllers.has_key('xdg-open'):
                return _controllers['xdg-open'].open
            else:
                return webbrowser.open


    if os.environ.get("DISPLAY"):
        register_X_controllers()
    _open = get()


def assoc_open(filename):
    print "opening "+filename
    import time
    time.sleep(3)
    '''Open a file or an URL in the registered default application.'''
    if len(filename)>3:
        # if filename[-3:].lower() in ("rtf","txt","csv","abw"):
        if filename[-3:].lower() in ("abw"):
            if sys.platform[:3] == 'win':
                try: return subprocess.Popen(["AbiWord"+os.path.sep+"bin"+os.path.sep+"AbiWord.exe",filename])
                except: return _open(filename)
            elif sys.platform == 'darwin':
                #if filename[-3:].lower()=="abw":
                    ##try: return subprocess.Popen(["Portable AbiWord.app/Contents/Resources/app/AbiWord.app/Contents/MacOS/AbiWord",filename])
                    #try: return subprocess.Popen(["open","-a","Portable\ AbiWord.app",filename])
                    #except: return _open(filename)
                #else:
                    #return _open(filename)
                try: return subprocess.Popen(["open","-a","AbiWord.app",filename])
                except: return _open(filename)
            else:
                try: return subprocess.Popen(["abiword",filename])
                except: return _open(filename)
        elif filename[-3:].lower() in ("sps"):
            if sys.platform[:3] == 'win':
                try: return subprocess.Popen(["PSPP"+os.path.sep+"bin"+os.path.sep+"psppire.exe","-q",filename])
                except: return _open(filename)
            elif sys.platform == 'darwin':
                #if filename[-3:].lower()=="abw":
                    ##try: return subprocess.Popen(["Portable AbiWord.app/Contents/Resources/app/AbiWord.app/Contents/MacOS/AbiWord",filename])
                    #try: return subprocess.Popen(["open","-a","Portable\ AbiWord.app",filename])
                    #except: return _open(filename)
                #else:
                    #return _open(filename)
                try: return subprocess.Popen(["open","-a","PSPP.app","-q",filename])
                except: return _open(filename)
            else:
                try: return subprocess.Popen(["psppire","-q",filename])
                except: return _open(filename)
        elif filename[-3:].lower() in ("xls","xlsx"):
            if sys.platform[:3] == 'win':
                try: return subprocess.Popen(["Gnumeric"+os.path.sep+"bin"+os.path.sep+"gnumeric.exe","--no-splash",filename])
                except: return _open(filename)
            else:
                try: return subprocess.Popen(["gnumeric","--no-splash",filename])
                except: return _open(filename)
        else:
            return _open(filename)
    else:
        wx.MessageBox("Incorrect file name", "Open file", wx.ICON_ERROR | wx.OK, None)
#CROSS PLATFORM OPEN - END

#FILE SPLITTER - BEGIN
class FileSplitter:

    def split(self,filename,numchunks):
        # Split the file and save chunks to separate files

        try:
            f = open(filename, 'rb')
        except (OSError, IOError), e:
            raise FileSplitterException, str(e)

        bname = filename
        # Get the file size
        fsize = os.path.getsize(filename)
        
        if numchunks==0:
            maxchunksize=100000
            import math
            numchunks=int(math.ceil(float(fsize)/float(maxchunksize)))

        # Get size of each chunk
        chunksize = int(float(fsize)/float(numchunks))

        chunksz = chunksize
        total_bytes = 0

        for x in range(numchunks):
            chunkfilename = bname + ('-%03d' % (numchunks)) + ('-%03d' % (x+1))

            # if reading the last section, calculate correct
            # chunk size.
            if x == numchunks - 1:
                chunksz = fsize - total_bytes

            try:
                data = f.read(chunksz)
                total_bytes += len(data)
                chunkf = file(chunkfilename, 'wb')
                chunkf.write(data)
                chunkf.close()
            except (OSError, IOError), e:
                print e
                continue
            except EOFError, e:
                print e
                break

    def sort_index(self, f1, f2):

        index1 = f1.rfind('-')
        index2 = f2.rfind('-')
        
        if index1 != -1 and index2 != -1:
            i1 = int(f1[index1:len(f1)])
            i2 = int(f2[index2:len(f2)])
            return i2 - i1
        
    def check_chunks(self,chunkfiles):
        tmp=""
        try: tmp=chunkfiles[len(chunkfiles)-1].split("-")
        except: return False
        if tmp[len(tmp)-1]!=tmp[len(tmp)-2]: return False
        ct=0
        for f in chunkfiles:
            ct=ct+1
            tmp=f.split("-")
            if ct!=int(tmp[len(tmp)-1]): return False
        return True

    def remove(self,filename):
        import re
        bname = (os.path.split(filename))[1]
        bname2 = bname
        filepath = (os.path.split(filename))[0]
        if filepath=="": filepath="."
        
        for a, b in zip(['+', '.', '[', ']','$', '(', ')'],
                        ['\+','\.','\[','\]','\$', '\(', '\)']):
            bname2 = bname2.replace(a, b)
            
        chunkre = re.compile(bname2 + '-' + '[0-9]+')
        
        found=False
        for f in os.listdir(filepath):
            if chunkre.match(f):
                found=True
                os.unlink(filepath+os.path.sep+f)
        return found

    def combine(self,filename):
        # Combine existing chunks to recreate the file.
        ret=True
        
        import re
        
        bname = (os.path.split(filename))[1]
        bname2 = bname
        filepath = (os.path.split(filename))[0]
        if filepath=="": filepath="."
        
        try: test=os.listdir(filepath)
        except: return False
        
        # bugfix: if file contains characters like +,.,[]
        # properly escape them, otherwise re will fail to match.
        for a, b in zip(['+', '.', '[', ']','$', '(', ')'],
                        ['\+','\.','\[','\]','\$', '\(', '\)']):
            bname2 = bname2.replace(a, b)
            
        chunkre = re.compile(bname2 + '-' + '[0-9]+')
        
        chunkfiles = []
        #for f in os.listdir("."):
        for f in os.listdir(filepath):
            if chunkre.match(f):
                chunkfiles.append(f)

        chunkfiles.sort(self.sort_index)

        if self.check_chunks(chunkfiles):
            try:
                #cmbf = open(bname, 'ab')
                cmbf = open(filepath+os.path.sep+bname, 'ab')
                for f in chunkfiles:

                    try:
                        #data = open(f, 'rb').read()
                        #print "Opening "+filepath+os.path.sep+f
                        data = open(filepath+os.path.sep+f, 'rb').read()
                        cmbf.write(data)
                    except (OSError, IOError, EOFError), e:
                        ret=False
                        print e
                        continue
                cmbf.close()

            except (OSError, IOError, EOFError), e:
                ret=False
                raise FileSplitterException, str(e)
        else:
            ret=False
        return ret
#FILE SPLITTER - END

# Date Control - Begin
import wx.combo
import wx.calendar
from datetime import date as dt

class DateCtrl(wx.combo.ComboCtrl):
    INPUT_FORMAT = 0
    DISPLAY_FORMAT = 1

#    def __init__(self, parent, size, pos, input_format, display_format, title, default_to_today, allow_null):
    def __init__(self, name, parent, size, pos, input_format, display_format, title, default_to_today, allow_null):
#        wx.combo.ComboCtrl.__init__(self, parent, size=size, pos=pos)
        wx.combo.ComboCtrl.__init__(self, parent, size=size, pos=pos, name=name)

        self.input_format = input_format
        self.display_format = display_format
        self.title = title
        self.default_to_today = default_to_today
        self.allow_null = allow_null

        self.TextCtrl.Bind(wx.EVT_SET_FOCUS, self.on_got_focus)
        self.TextCtrl.Bind(wx.EVT_CHAR, self.on_char)
        self.Bind(wx.EVT_ENTER_WINDOW, self.on_mouse_enter)
        self.Bind(wx.EVT_LEAVE_WINDOW, self.on_mouse_leave)
        self.TextCtrl.Bind(wx.EVT_KILL_FOCUS, self.on_lost_focus)

        self.nav = False  # force navigation after selecting date
        self.is_valid = True  # unlike IsValid(), a blank date can be valid
        self.current_format = self.DISPLAY_FORMAT
        self.date = wx.DateTime()
        self.setup_button()  # create a custom button for popup
        (self.blank_string, self.yr_pos, self.mth_pos, self.day_pos,
            self.literal_pos) = self.setup_input_format()

        # set up button coords for mouse hit-test
        self.b_x1 = self.TextRect[2] - 2
        self.b_y1 = self.TextRect[1] - 1
        self.b_x2 = self.b_x1 + self.ButtonSize[0] + 3
        self.b_y2 = self.b_y1 + self.ButtonSize[1] + 1
        self.on_button = False

        self.timer = wx.Timer(self)
        self.Bind(wx.EVT_TIMER, self.show_tooltip)

    def on_lost_focus(self, evt):
        self.convert_to_wx_date()
        
    def on_mouse_enter(self, evt):
        if self.b_x1 <= evt.X <= self.b_x2:
            if self.b_y1 <= evt.Y <= self.b_y2:
                self.on_button = True
                self.timer.Start(500, oneShot=True)
        evt.Skip()

    def on_mouse_leave(self, evt):
        if self.on_button:
            self.on_button = False
            self.timer.Stop()
        evt.Skip()

    def show_tooltip(self, evt):
        abs_x, abs_y = self.ScreenPosition
        rect = wx.Rect(abs_x+self.b_x1, abs_y+self.b_y1,
            self.b_x2-self.b_x1+1, self.b_y2-self.b_y1+1)
        tip = wx.TipWindow(self, 'Show calendar\n(F4 or space)')
        # tip will be destroyed when mouse leaves this rect
        tip.SetBoundingRect(rect)

    def setup_button(self):  # copied directly from demo
        # make a custom bitmap showing "..."
        bw, bh = 14, 16
        bmp = wx.EmptyBitmap(bw, bh)
        dc = wx.MemoryDC(bmp)

        # clear to a specific background colour
        bgcolor = wx.Colour(255, 254, 255)
        dc.SetBackground(wx.Brush(bgcolor))
        dc.Clear()

        # draw the label onto the bitmap
        label = u'\u2026'  # unicode ellipsis
        font = wx.SystemSettings.GetFont(wx.SYS_DEFAULT_GUI_FONT)
        font.SetWeight(wx.FONTWEIGHT_BOLD)
        dc.SetFont(font)
        tw, th = dc.GetTextExtent(label)
        dc.DrawText(label, (bw-tw)/2, (bw-tw)/2)
        del dc

        # now apply a mask using the bgcolor
        bmp.SetMaskColour(bgcolor)

        # and tell the ComboCtrl to use it
        self.SetButtonBitmaps(bmp, True)

    def setup_input_format(self):
        """
        Modify the defined input format to a string where each character
        represents one character of the input string.
        Generate and return a blank string to fill in the control.
        Return positions within the string of yr, mth, day and literals.
        """
        format = self.input_format
        blank_string = format

        yr_pos = format.find('%y')
        if yr_pos > -1:
            blank_string = blank_string[:yr_pos]+'  '+blank_string[yr_pos+2:]
            yr_pos = (yr_pos, yr_pos+2)
        else:
            yr_pos = format.find('%Y')
            if yr_pos > -1:
                blank_string = blank_string[:yr_pos]+'    '+blank_string[yr_pos+2:]
                format = format[:yr_pos+2]+'YY'+format[yr_pos+2:]
                yr_pos = (yr_pos, yr_pos+4)

        mth_pos = format.find('%m')
        if mth_pos > -1:
            blank_string = blank_string[:mth_pos]+'  '+blank_string[mth_pos+2:]
            mth_pos = (mth_pos, mth_pos+2)

        day_pos = format.find('%d')
        if day_pos > -1:
            blank_string = blank_string[:day_pos]+'  '+blank_string[day_pos+2:]
            day_pos = (day_pos, day_pos+2)

        literal_pos = [i for (i, ch) in enumerate(blank_string)
            if blank_string[i] == format[i]]

        return blank_string, yr_pos, mth_pos, day_pos, literal_pos

    # Overridden from ComboCtrl, called when the combo button is clicked
    def OnButtonClick(self):
        self.SetFocus()  # in case we do not have focus
        dlg = CalendarDlg(self)
        dlg.CentreOnScreen()
        if dlg.ShowModal() == wx.ID_OK:
            self.date = dlg.cal.Date
            self.Value = self.date.Format(self.display_format)
            self.current_format = self.DISPLAY_FORMAT
            self.nav = True  # force navigation to next control
        dlg.Destroy()

    # Overridden from ComboCtrl to avoid assert since there is no ComboPopup
    def DoSetPopupControl(self, popup):
        pass

    def on_got_focus(self, evt):
        if self.nav:  # user has made a selection, so move on
            self.nav = False
            wx.CallAfter(self.Navigate)
        else:
            text_ctrl = self.TextCtrl
            if not self.is_valid:  # re-focus after error
                pass  # leave Value alone
            elif self.date.IsValid():
                text_ctrl.Value = self.date.Format(self.input_format)
            elif self.default_to_today:
                self.date = wx.DateTime.Today()
                text_ctrl.Value = self.date.Format(self.input_format)
            else:
                text_ctrl.Value = self.blank_string
            self.current_format = self.INPUT_FORMAT
            text_ctrl.InsertionPoint = 0
            text_ctrl.SetSelection(-1, -1)
            text_ctrl.pos = 0
        evt.Skip()

    def convert_to_wx_date(self):  # conversion and validation method
        self.is_valid = True

        value = self.Value
        if value in (self.blank_string, ''):
            if self.default_to_today:
                self.date = wx.DateTime.Today()
                self.Value = self.date.Format(self.display_format)
            elif self.allow_null:
                self.date = wx.DateTime()
                self.Value = ''
            else:
                wx.CallAfter(self.display_error, 'Date is required')
            return

        if self.current_format == self.DISPLAY_FORMAT:  # no validation reqd
            self.TextCtrl.SetSelection(0, 0)
            return

        today = dt.today()

        if self.yr_pos == -1:  # 'yr' not an element of input_format
            year = today.year
        else:
            year = value[self.yr_pos[0]:self.yr_pos[1]].strip()
            if year == '':
                year = today.year
            elif len(year) == 2:
                # assume year is in range (today-90) to (today+10)
                year = int(year) + int(today.year/100)*100
                if year - today.year > 10:
                    year -= 100
                elif year - today.year < -90:
                    year += 100
            else:
                year = int(year)

        if self.mth_pos == -1:  # 'mth' not an element of input_format
            month = today.month
        else:
            month = value[self.mth_pos[0]:self.mth_pos[1]].strip()
            if month == '':
                month = today.month
            else:
                month = int(month)

        if self.day_pos == -1:  # 'day' not an element of input_format
            day = today.day
        else:
            day = value[self.day_pos[0]:self.day_pos[1]].strip()
            if day == '':
                day = today.day
            else:
                day = int(day)

        try:
            date = dt(year, month, day)  # validate using python datetime
#        except ValueError as error:  # gives a meaningful error message
#            wx.CallAfter(self.display_error, error.args[0])
        except:  # gives a meaningful error message
            wx.MessageDialog(self, "datetime error", self.title, wx.OK | wx.ICON_INFORMATION)
        else:  # date is valid
            self.date = wx.DateTimeFromDMY(day, month-1, year)
            self.Value = self.date.Format(self.display_format)
            self.current_format = self.DISPLAY_FORMAT
            print self.Value

    def display_error(self, errmsg):
        self.is_valid = False
        self.SetFocus()
        dlg = wx.MessageDialog(self, errmsg,
            self.title, wx.OK | wx.ICON_INFORMATION)
        dlg.ShowModal()
        dlg.Destroy()

    def on_char(self, evt):
        text_ctrl = self.TextCtrl
        code = evt.KeyCode
        if code in (wx.WXK_SPACE, wx.WXK_F4) and not evt.AltDown():
            self.OnButtonClick()
            return
        max = len(self.blank_string)
        if code in (wx.WXK_LEFT, wx.WXK_RIGHT, wx.WXK_HOME, wx.WXK_END):
            if text_ctrl.Selection == (0, max):
                text_ctrl.SetSelection(0, 0)
            if code == wx.WXK_LEFT:
                if text_ctrl.pos > 0:
                    text_ctrl.pos -= 1
                    while text_ctrl.pos in self.literal_pos:
                        text_ctrl.pos -= 1
            elif code == wx.WXK_RIGHT:
                if text_ctrl.pos < max:
                    text_ctrl.pos += 1
                    while text_ctrl.pos in self.literal_pos:
                        text_ctrl.pos += 1
            elif code == wx.WXK_HOME:
                text_ctrl.pos = 0
            elif code == wx.WXK_END:
                text_ctrl.pos = max
            text_ctrl.InsertionPoint = text_ctrl.pos
            return
        if code in (wx.WXK_BACK, wx.WXK_DELETE):
            if text_ctrl.Selection == (0, max):
                text_ctrl.Value = self.blank_string
                text_ctrl.SetSelection(0, 0)
            if code == wx.WXK_BACK:
                if text_ctrl.pos == 0:
                    return
                text_ctrl.pos -= 1
                while text_ctrl.pos in self.literal_pos:
                    text_ctrl.pos -= 1
            elif code == wx.WXK_DELETE:
                if text_ctrl.pos == max:
                    return
            curr_val = text_ctrl.Value
            text_ctrl.Value = curr_val[:text_ctrl.pos]+' '+curr_val[text_ctrl.pos+1:]
            text_ctrl.InsertionPoint = text_ctrl.pos
            return
        if code in (wx.WXK_TAB, wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER) or code > 255:
            evt.Skip()
            return
        if text_ctrl.pos == max:
            wx.Bell()
            return
        ch = chr(code)
        if ch not in ('0123456789'):
            wx.Bell()
            return
        if text_ctrl.Selection == (0, max):
            curr_val = self.blank_string
        else:
            curr_val = text_ctrl.Value
        text_ctrl.Value = curr_val[:text_ctrl.pos]+ch+curr_val[text_ctrl.pos+1:]
        text_ctrl.pos += 1
        while text_ctrl.pos in self.literal_pos:
            text_ctrl.pos += 1
        text_ctrl.InsertionPoint = text_ctrl.pos

class CalendarDlg(wx.Dialog):
    def __init__(self, parent):

        wx.Dialog.__init__(self, parent, title=parent.title)
        panel = wx.Panel(self, -1)

        sizer = wx.BoxSizer(wx.VERTICAL)
        panel.SetSizer(sizer)

        cal = wx.calendar.CalendarCtrl(panel, date=parent.date)

        #if sys.platform != 'win32':
        if os.name!='nt':
            # gtk truncates the year - this fixes it
            w, h = cal.Size
            cal.Size = (w+25, h)
            cal.MinSize = cal.Size

        sizer.Add(cal, 0)

        button_sizer = wx.BoxSizer(wx.HORIZONTAL)
        button_sizer.Add((0, 0), 1)
        btn_ok = wx.Button(panel, wx.ID_OK)
        btn_ok.SetDefault()
        button_sizer.Add(btn_ok, 0, wx.ALL, 2)
        button_sizer.Add((0, 0), 1)
        btn_can = wx.Button(panel, wx.ID_CANCEL)
        button_sizer.Add(btn_can, 0, wx.ALL, 2)
        button_sizer.Add((0, 0), 1)
        sizer.Add(button_sizer, 1, wx.EXPAND | wx.ALL, 10)
        sizer.Fit(panel)
        self.ClientSize = panel.Size

        cal.Bind(wx.EVT_KEY_DOWN, self.on_key_down)
        cal.SetFocus()
        self.cal = cal

    def on_key_down(self, evt):
        code = evt.KeyCode
        if code == wx.WXK_TAB:
            self.cal.Navigate()
        elif code in (wx.WXK_RETURN, wx.WXK_NUMPAD_ENTER):
            self.EndModal(wx.ID_OK)
        elif code == wx.WXK_ESCAPE:
            self.EndModal(wx.ID_CANCEL)
        else:
            evt.Skip()
# Date Control - End

class AccessRightsDlg(wx.Dialog):
    def __init__(self, parent):
        wx.Dialog.__init__(self, parent, title="Modify Access Rights for: "+parent.cur_group_shortcode+"-"+parent.curkeyvalue, size=wx.Size(940,430))

        self.parent=parent
        self.hashPosById={}
        self.hashIdByPos={}
        self.hashDescByPos={}
        pos=0
        groups=[]
        #cur.execute("SELECT id_group,shortcode,description AS grp FROM groups order by shortcode,description")
        cur.execute("SELECT id_group,shortcode,description AS grp FROM groups order by description,shortcode")
        for row in cur:
            groupdesc=str(row[1])
            if row[2]!=None: groupdesc=groupdesc+"__"+row[2]
            groups.append(groupdesc)
            self.hashPosById[str(row[0])]=pos
            self.hashIdByPos[pos]=str(row[0])
            self.hashDescByPos[pos]=groupdesc
            pos=pos+1

        wx.StaticText(label=TT('can read'), parent=self, pos=wx.Point(10, 10), size=wx.Size(300, 25), style=0)
        wx.StaticText(label=TT('can modify'), parent=self, pos=wx.Point(320, 10), size=wx.Size(300, 25), style=0)
        wx.StaticText(label=TT('can delete'), parent=self, pos=wx.Point(630, 10), size=wx.Size(300, 25), style=0)

        self.lb_r = wx.CheckListBox(parent=self, pos=wx.Point(10, 35), size=wx.Size(300, 300), choices=groups)
        self.lb_w = wx.CheckListBox(parent=self, pos=wx.Point(320, 35), size=wx.Size(300, 300), choices=groups)
        self.lb_x = wx.CheckListBox(parent=self, pos=wx.Point(630, 35), size=wx.Size(300, 300), choices=groups)

        btnOk = wx.Button(label='OK', parent=self, pos=wx.Point(340,350), size=wx.Size(100,25))
        btnCancel = wx.Button(label=TT('Cancel'), parent=self, pos=wx.Point(500,350), size=wx.Size(100,25))
        btnOk.Bind(wx.EVT_BUTTON, self.OnConfirmModifyRights)
        btnCancel.Bind(wx.EVT_BUTTON, self.OnCancelModifyRights)

        #Init
        rap=self.parent.rap
        if rap:
            part_rap=rap.split(";")
            if len(part_rap)>=1:
                lst=[]
                for elm in part_rap[0].split(","):
                    lst.append(self.hashPosById[elm])
                self.lb_r.SetChecked(lst)
            if len(part_rap)>=2:
                lst=[]
                for elm in part_rap[1].split(","):
                    lst.append(self.hashPosById[elm])
                self.lb_w.SetChecked(lst)
            if len(part_rap)>=3:
                lst=[]
                for elm in part_rap[2].split(","):
                    lst.append(self.hashPosById[elm])
                self.lb_x.SetChecked(lst)
            
        
    def OnConfirmModifyRights(self,event):
        warnings=""
        for i in range(self.lb_x.GetCount()):
            if self.lb_x.IsChecked(i) and not self.lb_w.IsChecked(i):
                warnings=warnings+"group "+self.hashDescByPos[i].split("__")[0]+" would be allowed to delete but not update\n"
            if self.lb_w.IsChecked(i) and not self.lb_r.IsChecked(i):
                warnings=warnings+"group "+self.hashDescByPos[i].split("__")[0]+" would be allowed to update but not read\n"
#            print self.hashIdByPos[i]+"---"+str(self.parent.id_group)
            if self.hashIdByPos[i]==str(self.parent.id_group):
                if (not self.lb_r.IsChecked(i)):
                    warnings=warnings+"administrators would not be allowed to read\n"
                if (not self.lb_w.IsChecked(i)):
                    warnings=warnings+"administrators would not be allowed to update\n"
                if (not self.lb_x.IsChecked(i)):
                    warnings=warnings+"administrators would not be allowed to delete\n"
        if warnings!="":
            wx.MessageBox(warnings, "Error", wx.ICON_ERROR | wx.OK, None)
        else:
            canread=""
            for i in range(self.lb_r.GetCount()):
                if self.lb_r.IsChecked(i):
                    if canread!="": canread=canread+","
                    canread=canread+self.hashIdByPos[i]
            canmodify=""
            for i in range(self.lb_w.GetCount()):
                if self.lb_w.IsChecked(i):
                    if canmodify!="": canmodify=canmodify+","
                    canmodify=canmodify+self.hashIdByPos[i]
            candelete=""
            for i in range(self.lb_x.GetCount()):
                if self.lb_x.IsChecked(i):
                    if candelete!="": candelete=candelete+","
                    candelete=candelete+self.hashIdByPos[i]
            newrap=canread+";"+canmodify+";"+candelete
            print "oldrap:"+self.parent.rap
            print "newrap:"+newrap
            self.parent.anyrecordchange=True
            self.parent.rap=newrap
            
            self.Destroy()

    def OnCancelModifyRights(self,event):
        self.Destroy()
        
class ShowDiffDlg(wx.Dialog):
    def __init__(self, parent, id_row1,id_user1,id_rowinstance1,datetime1,form1,id_row2,id_user2,id_rowinstance2,datetime2,form2,rap):
        wx.Dialog.__init__(self, parent, title="Diff & Merge", size=wx.Size(940,430))

        self.parent=parent
        self.id_row1=id_row1
        self.id_user1=id_user1
        self.id_rowinstance1=id_rowinstance1
        self.datetime1=datetime1
        self.form1=form1
        self.id_row2=id_row2
        self.id_user2=id_user2
        self.id_rowinstance2=id_rowinstance2
        self.datetime2=datetime2
        self.form2=form2
        self.rap=rap
        
        group_shortcode1=""
        if self.form1.has_key("group_shortcode"):
            dData=""
            if not self.parent.parent.trans.has_key(self.form1["group_shortcode"]):
                dData=HB_DecryptOne(self.parent.digestkey,self.form1["group_shortcode"],"latin-1")
                self.parent.parent.trans[self.form1["group_shortcode"]]=dData
            else:
                dData=self.parent.parent.trans[self.form1["group_shortcode"]]
            group_shortcode1=dData
        key1=""
        if self.form1.has_key(self.parent.keyname):
            dData=""
            if not self.parent.parent.trans.has_key(self.form1[self.parent.keyname]):
                dData=HB_DecryptOne(self.parent.digestkey,self.form1[self.parent.keyname],"latin-1")
                self.parent.parent.trans[self.form1[self.parent.keyname]]=dData
            else:
                dData=self.parent.parent.trans[self.form1[self.parent.keyname]]
            key1=dData

        group_shortcode2=""
        if self.form2.has_key("group_shortcode"):
            dData=""
            if not self.parent.parent.trans.has_key(self.form2["group_shortcode"]):
                dData=HB_DecryptOne(self.parent.digestkey,self.form2["group_shortcode"],"latin-1")
                self.parent.parent.trans[self.form2["group_shortcode"]]=dData
            else:
                dData=self.parent.parent.trans[self.form2["group_shortcode"]]
            group_shortcode2=dData
        key2=""
        if self.form2.has_key(self.parent.keyname):
            dData=""
            if not self.parent.parent.trans.has_key(self.form2[self.parent.keyname]):
                dData=HB_DecryptOne(self.parent.digestkey,self.form2[self.parent.keyname],"latin-1")
                self.parent.parent.trans[self.form2[self.parent.keyname]]=dData
            else:
                dData=self.parent.parent.trans[self.form2[self.parent.keyname]]
            key2=dData
        
        label1=group_shortcode1+"-"+key1+" ["+self.datetime1+"]"
        label2=group_shortcode2+"-"+key2+" ["+self.datetime2+"]"
        
        only1=[]
        for fldname in self.form1:
            if not self.form2.has_key(fldname):
                dData=""
                if not self.parent.parent.trans.has_key(self.form1[fldname]):
                    dData=HB_DecryptOne(self.parent.digestkey,self.form1[fldname],"latin-1")
                    self.parent.parent.trans[self.form1[fldname]]=dData
                else:
                    dData=self.parent.parent.trans[self.form1[fldname]]
                only1.append(fldname+" = "+dData)
        only2=[]
        for fldname in self.form2:
            if not self.form1.has_key(fldname):
                dData=""
                if not self.parent.parent.trans.has_key(self.form2[fldname]):
                    dData=HB_DecryptOne(self.parent.digestkey,self.form2[fldname],"latin-1")
                    self.parent.parent.trans[self.form2[fldname]]=dData
                else:
                    dData=self.parent.parent.trans[self.form2[fldname]]
                only2.append(fldname+" = "+dData)
        both=[]
        for fldname in self.form1:
            if self.form2.has_key(fldname):
                if self.form1[fldname]!=self.form2[fldname]:
                    dData=""
                    if not self.parent.parent.trans.has_key(self.form1[fldname]):
                        dData=HB_DecryptOne(self.parent.digestkey,self.form1[fldname],"latin-1")
                        self.parent.parent.trans[self.form1[fldname]]=dData
                    else:
                        dData=self.parent.parent.trans[self.form1[fldname]]
                    both.append("< "+fldname+" = "+dData)
                    dData=""
                    if not self.parent.parent.trans.has_key(self.form2[fldname]):
                        dData=HB_DecryptOne(self.parent.digestkey,self.form2[fldname],"latin-1")
                        self.parent.parent.trans[self.form2[fldname]]=dData
                    else:
                        dData=self.parent.parent.trans[self.form2[fldname]]
                    both.append("> "+fldname+" = "+dData)
        
        wx.StaticText(label="only in "+label1, parent=self, pos=wx.Point(10, 10), size=wx.Size(300, 25), style=0)
        wx.StaticText(label='both present but different', parent=self, pos=wx.Point(320, 10), size=wx.Size(300, 25), style=0)
        wx.StaticText(label="only in "+label2, parent=self, pos=wx.Point(630, 10), size=wx.Size(300, 25), style=0)

        if sys.platform == 'darwin':
            self.only1 = wx.ListBox(parent=self, pos=wx.Point(10, 35), size=wx.Size(300, 300), choices=only1, style=wx.LB_EXTENDED)
            self.both = wx.ListBox(parent=self, pos=wx.Point(320, 35), size=wx.Size(300, 300), choices=both, style=wx.LB_EXTENDED)
            self.only2 = wx.ListBox(parent=self, pos=wx.Point(630, 35), size=wx.Size(300, 300), choices=only2, style=wx.LB_EXTENDED)
        else:
            self.only1 = wx.CheckListBox(parent=self, pos=wx.Point(10, 35), size=wx.Size(300, 300), choices=only1)
            self.both = wx.CheckListBox(parent=self, pos=wx.Point(320, 35), size=wx.Size(300, 300), choices=both)
            self.only2 = wx.CheckListBox(parent=self, pos=wx.Point(630, 35), size=wx.Size(300, 300), choices=only2)
                    
        obj=self.only1
        for i in range(obj.GetCount()):
            if sys.platform == 'darwin':
                obj.SetSelection(i, True)
            else:
                obj.Check(i, True)
        obj=self.only2
        for i in range(obj.GetCount()):
            if sys.platform == 'darwin':
                obj.SetSelection(i, True)
            else:
                obj.Check(i, True)
                
        resto=1
        if self.datetime1>=self.datetime2: resto=0
        ct=0
        obj=self.both
        for i in range(obj.GetCount()):
            if ct%2==resto:
                if sys.platform == 'darwin':
                    obj.SetSelection(i, True)
                else:
                    obj.Check(i, True)
            ct=ct+1
            
        btnOk = wx.Button(label='Merge', parent=self, pos=wx.Point(340,350), size=wx.Size(100,25))
        btnCancel = wx.Button(label=TT('Cancel'), parent=self, pos=wx.Point(500,350), size=wx.Size(100,25))
        btnOk.Bind(wx.EVT_BUTTON, self.OnConfirmShowDiff)
        btnCancel.Bind(wx.EVT_BUTTON, self.OnCancelShowDiff)
        
    def DoDecrypt(self,word):
        dData=""
        if not self.parent.parent.trans.has_key(word):
            dData=HB_DecryptOne(self.parent.digestkey,word,"latin-1")
            self.parent.parent.trans[word]=dData
        else:
            dData=self.parent.parent.trans[word]
        return dData
        
    def OnConfirmShowDiff(self,event):
        if wx.MessageBox("Are you sure you wish to merge these two records keeping the selected fields?", "Diff & Merge", wx.YES_NO) == wx.YES:
            validation=True
            checklist={}
            if sys.platform == 'darwin':
                for i in range(len(self.both.GetSelections())):
                    fldName = self.both.GetSelections()[i].split(" = ")[0][2:]
                    if checklist.has_key(fldName):
                        wx.MessageBox("The field "+fldName+" has been selected two times.", "Diff & Merge", wx.ICON_ERROR | wx.OK, self)
                        validation=False
                        break
                    else:
                        checklist[fldName]=True
            else:
                for i in range(len(self.both.GetChecked())):
                    fldName = self.both.GetCheckedStrings()[i].split(" = ")[0][2:]
                    if checklist.has_key(fldName):
                        wx.MessageBox("The field "+fldName+" has been selected two times.", "Diff & Merge", wx.ICON_ERROR | wx.OK, self)
                        validation=False
                        break
                    else:
                        checklist[fldName]=True
            if validation:
                wx.BeginBusyCursor()
                self.parent.syncroTaskPause=True
                
                self.parent.contents={}
                
                for fldname in self.form1:
                    if self.form2.has_key(fldname):
                        if self.form1[fldname]==self.form2[fldname]:
                            self.parent.contents[fldname]=self.DoDecrypt(self.form1[fldname])
                                
                if sys.platform == 'darwin':
                    for i in range(len(self.only1.GetSelections())):
                        fldName = self.only1.GetSelections()[i].split(" = ")[0]
                        self.parent.contents[fldName]=self.DoDecrypt(self.form1[fldName])
                    for i in range(len(self.only2.GetSelections())):
                        fldName = self.only2.GetSelections()[i].split(" = ")[0]
                        self.parent.contents[fldName]=self.DoDecrypt(self.form2[fldName])
                    for i in range(len(self.both.GetSelections())):
                        fldName = self.both.GetSelections()[i].split(" = ")[0][2:]
                        if self.both.GetSelections()[i].split(" = ")[0][:2]=="< ":
                            self.parent.contents[fldName]=self.DoDecrypt(self.form1[fldName])
                        else:
                            self.parent.contents[fldName]=self.DoDecrypt(self.form2[fldName])
                else:
                    for i in range(len(self.only1.GetChecked())):
                        fldName = self.only1.GetCheckedStrings()[i].split(" = ")[0]
                        self.parent.contents[fldName]=self.DoDecrypt(self.form1[fldName])
                    for i in range(len(self.only2.GetChecked())):
                        fldName = self.only2.GetCheckedStrings()[i].split(" = ")[0]
                        self.parent.contents[fldName]=self.DoDecrypt(self.form2[fldName])
                    for i in range(len(self.both.GetChecked())):
                        fldName = self.both.GetCheckedStrings()[i].split(" = ")[0][2:]
                        if self.both.GetCheckedStrings()[i].split(" = ")[0][:2]=="< ":
                            self.parent.contents[fldName]=self.DoDecrypt(self.form1[fldName])
                        else:
                            self.parent.contents[fldName]=self.DoDecrypt(self.form2[fldName])
                            
                now = datetime.datetime.utcnow()
                ts=now.strftime("%d/%m/%Y, %H:%M:%S")
                ts_db=now.strftime("%Y-%m-%d %H:%M:%S")
                
                #new id_row - begin
                cur.execute("SELECT max(id_row) FROM rows")
                row = cur.fetchone()
                self.parent.id_row=row[0]+1
                #new id_row - end
                
                self.parent.parent.syncroTaskPause=True
                import time
                while self.parent.parent.syncroTaskRunning:
                    time.sleep(1)
                    if DEBUG_MODE: print "waiting synchro standby..."
                cur.execute("BEGIN TRANSACTION")
                
                cur.execute("INSERT INTO rows(id_row,id_user,id_header,rap,date_time,status,status_user,status_instance,id_instance) VALUES ("+str(self.parent.id_row)+","+str(self.parent.id_logged_user)+","+str(self.parent.id_crf)+",'"+self.rap+"','"+ts_db+"',0,0,0,"+str(self.parent.parent.id_instance)+")")
                
                self.parent.SaveInDatabase()
                
                cur.execute("UPDATE rows SET status="+str(self.parent.id_row)+",status_user="+str(self.parent.id_logged_user)+",status_instance="+str(self.parent.parent.id_instance)+" WHERE id_row="+str(self.id_row1)+" AND id_user="+str(self.id_user1)+" AND id_instance="+str(self.id_rowinstance1)+" AND id_header="+str(self.parent.id_crf))
                cur.execute("UPDATE rows SET status="+str(self.parent.id_row)+",status_user="+str(self.parent.id_logged_user)+",status_instance="+str(self.parent.parent.id_instance)+" WHERE id_row="+str(self.id_row2)+" AND id_user="+str(self.id_user2)+" AND id_instance="+str(self.id_rowinstance2)+" AND id_header="+str(self.parent.id_crf))

                cur.execute("COMMIT TRANSACTION")
                self.parent.parent.syncroTaskPause=False

                self.parent.parent.syncrotablesupd["rows"]=True
                self.parent.lstFound.Clear()
                
                wx.EndBusyCursor()
            
                self.Destroy()
        else:
            self.Destroy()
    def OnCancelShowDiff(self,event):
        self.Destroy()


class WinSysTrayIcon(object):
    '''TODO'''
    QUIT = 'QUIT'
    SPECIAL_ACTIONS = [QUIT]
    
    FIRST_ID = 1023
    
    def __init__(self,
                 parent,
                 icon,
                 hover_text,
                 menu_options,
                 on_quit=None,
                 default_menu_index=None,
                 window_class_name=None):
        
        self.parent=parent
        self.icon = icon
        self.hover_text = hover_text
        self.on_quit = on_quit
        
        #menu_options = menu_options + (('Exit', None, self.QUIT),)
        self._next_action_id = self.FIRST_ID
        self.menu_actions_by_id = set()
        self.menu_options = self._add_ids_to_menu_options(list(menu_options))
        self.menu_actions_by_id = dict(self.menu_actions_by_id)
        del self._next_action_id
        
        
        self.default_menu_index = (default_menu_index or 0)
        self.window_class_name = window_class_name or "SysTrayIconPy"
        
        message_map = {win32gui.RegisterWindowMessage("TaskbarCreated"): self.restart,
                       win32con.WM_DESTROY: self.destroy,
                       win32con.WM_COMMAND: self.command,
                       win32con.WM_USER+20 : self.notify,}
        # Register the Window class.
        window_class = win32gui.WNDCLASS()
        hinst = window_class.hInstance = win32gui.GetModuleHandle(None)
        window_class.lpszClassName = self.window_class_name
        window_class.style = win32con.CS_VREDRAW | win32con.CS_HREDRAW;
        window_class.hCursor = win32gui.LoadCursor(0, win32con.IDC_ARROW)
        window_class.hbrBackground = win32con.COLOR_WINDOW
        window_class.lpfnWndProc = message_map # could also specify a wndproc.
        classAtom = win32gui.RegisterClass(window_class)
        # Create the Window.
        style = win32con.WS_OVERLAPPED | win32con.WS_SYSMENU
        self.hwnd = win32gui.CreateWindow(classAtom,
                                          self.window_class_name,
                                          style,
                                          0,
                                          0,
                                          win32con.CW_USEDEFAULT,
                                          win32con.CW_USEDEFAULT,
                                          0,
                                          0,
                                          hinst,
                                          None)
        self.parent.taskbarIconHwnd=self.hwnd
        win32gui.UpdateWindow(self.hwnd)
        self.notify_id = None
        self.refresh_icon()
        
        win32gui.PumpMessages()

    def _add_ids_to_menu_options(self, menu_options):
        result = []
        for menu_option in menu_options:
            option_text, option_icon, option_action = menu_option
            if callable(option_action) or option_action in self.SPECIAL_ACTIONS:
                self.menu_actions_by_id.add((self._next_action_id, option_action))
                result.append(menu_option + (self._next_action_id,))
            elif non_string_iterable(option_action):
                result.append((option_text,
                               option_icon,
                               self._add_ids_to_menu_options(option_action),
                               self._next_action_id))
            else:
                print 'Unknown item', option_text, option_icon, option_action
            self._next_action_id += 1
        return result
        
    def refresh_icon(self):
        # Try and find a custom icon
        hinst = win32gui.GetModuleHandle(None)
        if os.path.isfile(self.icon):
            icon_flags = win32con.LR_LOADFROMFILE | win32con.LR_DEFAULTSIZE
            hicon = win32gui.LoadImage(hinst,
                                       self.icon,
                                       win32con.IMAGE_ICON,
                                       0,
                                       0,
                                       icon_flags)
        else:
            print "Can't find icon file - using default."
            hicon = win32gui.LoadIcon(0, win32con.IDI_APPLICATION)

        if self.notify_id: message = win32gui.NIM_MODIFY
        else: message = win32gui.NIM_ADD
        self.notify_id = (self.hwnd,
                          0,
                          win32gui.NIF_ICON | win32gui.NIF_MESSAGE | win32gui.NIF_TIP,
                          win32con.WM_USER+20,
                          hicon,
                          self.hover_text)
        icon_ok=False
        while not icon_ok:
            try:
                win32gui.Shell_NotifyIcon(message, self.notify_id)
                icon_ok=True
            except:
                #win32gui.PumpWaitingMessages()
                time.sleep(0.5)
            
    def restart(self, hwnd, msg, wparam, lparam):
        self.refresh_icon()

    def destroy(self, hwnd, msg, wparam, lparam):
        if self.on_quit: self.on_quit(self)
        nid = (self.hwnd, 0)
        win32gui.Shell_NotifyIcon(win32gui.NIM_DELETE, nid)
        win32gui.PostQuitMessage(0) # Terminate the app.

    def notify(self, hwnd, msg, wparam, lparam):
        if lparam==win32con.WM_LBUTTONDBLCLK:
            self.execute_menu_option(self.default_menu_index + self.FIRST_ID)
        elif lparam==win32con.WM_RBUTTONUP:
            self.show_menu()
        elif lparam==win32con.WM_LBUTTONUP:
            self.execute_menu_option(self.default_menu_index + self.FIRST_ID)
            #pass
        return True
        
    def show_menu(self):
        menu = win32gui.CreatePopupMenu()
        self.create_menu(menu, self.menu_options)
        #win32gui.SetMenuDefaultItem(menu, 1000, 0)
        
        pos = win32gui.GetCursorPos()
        # See http://msdn.microsoft.com/library/default.asp?url=/library/en-us/winui/menus_0hdi.asp
        win32gui.SetForegroundWindow(self.hwnd)
        win32gui.TrackPopupMenu(menu,
                                win32con.TPM_LEFTALIGN,
                                pos[0],
                                pos[1],
                                0,
                                self.hwnd,
                                None)
        win32gui.PostMessage(self.hwnd, win32con.WM_NULL, 0, 0)
    
    def create_menu(self, menu, menu_options):
        for option_text, option_icon, option_action, option_id in menu_options[::-1]:
            if option_icon:
                option_icon = self.prep_menu_icon(option_icon)
            
            if option_id in self.menu_actions_by_id:                
                item, extras = win32gui_struct.PackMENUITEMINFO(text=option_text,
                                                                hbmpItem=option_icon,
                                                                wID=option_id)
                win32gui.InsertMenuItem(menu, 0, 1, item)
            else:
                submenu = win32gui.CreatePopupMenu()
                self.create_menu(submenu, option_action)
                item, extras = win32gui_struct.PackMENUITEMINFO(text=option_text,
                                                                hbmpItem=option_icon,
                                                                hSubMenu=submenu)
                win32gui.InsertMenuItem(menu, 0, 1, item)

    def prep_menu_icon(self, icon):
        # First load the icon.
        ico_x = win32api.GetSystemMetrics(win32con.SM_CXSMICON)
        ico_y = win32api.GetSystemMetrics(win32con.SM_CYSMICON)
        hicon = win32gui.LoadImage(0, icon, win32con.IMAGE_ICON, ico_x, ico_y, win32con.LR_LOADFROMFILE)

        hdcBitmap = win32gui.CreateCompatibleDC(0)
        hdcScreen = win32gui.GetDC(0)
        hbm = win32gui.CreateCompatibleBitmap(hdcScreen, ico_x, ico_y)
        hbmOld = win32gui.SelectObject(hdcBitmap, hbm)
        # Fill the background.
        brush = win32gui.GetSysColorBrush(win32con.COLOR_MENU)
        win32gui.FillRect(hdcBitmap, (0, 0, 16, 16), brush)
        # unclear if brush needs to be feed.  Best clue I can find is:
        # "GetSysColorBrush returns a cached brush instead of allocating a new
        # one." - implies no DeleteObject
        # draw the icon
        win32gui.DrawIconEx(hdcBitmap, 0, 0, hicon, ico_x, ico_y, 0, 0, win32con.DI_NORMAL)
        win32gui.SelectObject(hdcBitmap, hbmOld)
        win32gui.DeleteDC(hdcBitmap)
        
        return hbm

    def command(self, hwnd, msg, wparam, lparam):
        id = win32gui.LOWORD(wparam)
        self.execute_menu_option(id)
        
    def execute_menu_option(self, id):
        menu_action = self.menu_actions_by_id[id]      
        if menu_action == self.QUIT:
            try: win32gui.DestroyWindow(self.hwnd)
            except: pass
        else:
            menu_action(self)

class HeavyBase(wx.App):
    def OnInit(self):
        # Ensure Single Instance - Begin
        import sys, os, errno, tempfile
        import binascii
        curdir=os.getcwd()
        # self.lockfile = os.path.normpath(tempfile.gettempdir() + '/' + os.path.basename(__file__) + "_" + hex(abs(binascii.crc32(os.path.abspath(__file__)))) + '.lock')
        self.lockfile = os.path.normpath(tempfile.gettempdir() + '/' + os.path.basename(__file__) + "_" + hex(abs(binascii.crc32(curdir))) + '.lock')
        #if sys.platform == 'win32':
        if os.name=='nt':
            try:
                # file already exists, we try to remove (in case previous execution was interrupted)
                if(os.path.exists(self.lockfile)):
                    os.unlink(self.lockfile)
                self.fd =  os.open(self.lockfile, os.O_CREAT|os.O_EXCL|os.O_RDWR)
            except OSError, e:
                if e.errno == 13:
                    #wx.MessageBox("Another instance is already running.", "HEAVyBASE", wx.ICON_ERROR | wx.OK, None)
                    f=open('restore.cmd','w')
                    f.write('\n')
                    f.close()
                    sys.exit(-1)
                print e.errno
                raise
        else: # non Windows
            import fcntl, sys
            self.fp = open(self.lockfile, 'w')
            try:
                fcntl.lockf(self.fp, fcntl.LOCK_EX | fcntl.LOCK_NB)
            except IOError:
                wx.MessageBox("Another instance is already running.", "HEAVyBASE", wx.ICON_ERROR | wx.OK, None)
                sys.exit(-1)
        # Ensure Single Instance - End
        
        # Authorized Host - Begin
        authorization_check=False
        authorized_host=False
        import uuid
        cur_host=str(uuid.getnode())
        cur.execute("select setting_value from settings where setting_key='authorized_host'")
        for row in cur:
            authorization_check=True
            if str(row[0])==cur_host:
                authorized_host=True
                break
        if authorization_check and (not authorized_host):
            wx.MessageBox("Authorizaton denied.", "HEAVyBASE", wx.ICON_ERROR | wx.OK, None)
            sys.exit(-1)
        # Authorized Host - End
        
        log(PROGNAME+" Started",mode="w")
        frame = wx.MiniFrame(None, -1, TT("Loading HEAVyBASE"), size=wx.Size(340,80))
        panel = wx.Panel(frame)
        lbl=wx.StaticText(label=TT('Please wait'), parent=panel, style=wx.ALIGN_CENTRE_HORIZONTAL|wx.EXPAND)
        lbl.SetFont(wx.Font( 24, wx.SWISS, wx.NORMAL, wx.BOLD))
        sizer = wx.BoxSizer(wx.VERTICAL)
        sizer.Add(lbl, 0, wx.ALL, 10)
        panel.SetSizer(sizer)
        panel.Layout()
        frame.Show(1)
        frame.Center(direction=wx.BOTH)
        self.SetTopWindow(frame)
        wx.Yield()
        if os.name=='nt':
            import win32gui
            win32gui.PumpWaitingMessages()
        else:
            import time
            time.sleep(0.5)
        frame.Update()

        # Standard output and Standard error redirection to a file for production use with Windows - Begin
        #if sys.platform == 'win32' and ("pythonw" in sys.executable) and not DEBUG_MODE:
        if os.name=='nt' and ("pythonw" in sys.executable) and not DEBUG_MODE:
            print "redirecting standard output and standard error to 'lastlog.txt'"
            LOGFILE = open('lastlog.txt','w')
            sys.stdout = LOGFILE
            sys.stderr = LOGFILE
        # End
        
        import datetime
        now = datetime.datetime.utcnow()
        ts=now.strftime("%Y%m%d%H%M%S")
        print "\n"+"Start timestamp: "+ts

        self.main = create(None)
        self.main.Show()
        self.SetTopWindow(frame)
        frame.SetPosition(wx.Point(5000,0))
        frame.Destroy()
        self.SetTopWindow(self.main)
        # Taskbar - Begin
        if os.name=='nt' and (IsRunningFromRemovable()==False):
            import win32api
            import win32con
            import win32gui_struct
            try:
                import winxpgui as win32gui
            except ImportError:
                import win32gui
            title=LoadCustomSetting("project_title")
            if title!="": title=title+" - "
            title=title+PROGNAME+" "+RELEASE
            #hover_text = "HEAVyBASE"
            hover_text = title
            def show(sysTrayIcon): self.main.Show()
            def hide(sysTrayIcon): 
                self.main.Hide()
            def quitApp(sysTrayIcon): 
                self.main.syncroTaskPause=True
                if self.main.DoCloseHeavyBase():
                    self.main.Destroy()
                    import win32gui
                    try: win32gui.DestroyWindow(sysTrayIcon.hwnd)        #fallisce se l'interfaccia e' gia' stata chiusa dal distruttore dei 10 secondi
                    except: pass
                else:
                    self.main.syncroTaskPause=False
            def bye(sysTrayIcon): print "bye."
            
            menu_options = ((TT('Activate'), None, show),(TT('Hide'), None, hide),(TT('Exit'), None, quitApp))
            WinSysTrayIcon(self.main,"ship.ico", hover_text, menu_options, on_quit=bye, default_menu_index=0)
        # Taskbar - End
        
        return True

    def __del__(self):
        # Ensure Single Instance - Begin
        #if sys.platform == 'win32':
        if os.name=='nt':
            if hasattr(self, 'fd'):
                os.close(self.fd)
                os.unlink(self.lockfile)
        # Ensure Single Instance - End

        # Standard output and Standard error redirection to a file for production use with Windows - Begin
        #if sys.platform == 'win32' and ("pythonw" in sys.executable) and not DEBUG_MODE:
        if os.name=='nt' and ("pythonw" in sys.executable) and not DEBUG_MODE:
            sys.stdout=sys.__stdout__
            sys.stderr=sys.__stderr__
            try: LOGFILE.close()
            except: pass
        # End
            
def Get_User_ShellFolders():
    # Routine to grab all the Windows Shell Folder locations from the registry.  If successful, returns dictionary
    # of shell folder locations indexed on Windows keyword for each; otherwise, returns an empty dictionary.
    import _winreg
    return_dict = {}
 
    # First open the registry hive
    try:
        Hive = _winreg.ConnectRegistry(None, _winreg.HKEY_CURRENT_USER)
    except WindowsError:
        print "Can't connect to registry hive HKEY_CURRENT_USER."
        return return_dict
 
    # Then open the registry key where Windows stores the Shell Folder locations
    try:
        Key = _winreg.OpenKey(Hive, "Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
    except WindowsError:
        print "Can't open registry key Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders."
        _winreg.CloseKey(Hive)
        return return_dict
 
    # Nothing failed above, so enumerate through all the Shell Folder values and return in a dictionary
    # This relies on error at end of 
    try:
        #i = 0
        #while 1:
        for i in range(0, _winreg.QueryInfoKey(Key)[1]):
            name, value, val_type = _winreg.EnumValue(Key, i)
            return_dict[name] = value
            i += 1
        _winreg.CloseKey(Key)                           # Only use with for loop
        _winreg.CloseKey(Hive)                          # Only use with for loop
        return return_dict                              # Only use with for loop
    except WindowsError:
        # In case of failure before read completed, don't return partial results
        _winreg.CloseKey(Key)
        _winreg.CloseKey(Hive)
        return {}
    
def main():
    #Interface profile inference
    interface_profile=LoadCustomSetting("interface_profile")
    if interface_profile=="":
        lbl_salva=LoadCustomSetting("lbl_salva")
        if lbl_salva=="salva":
            SaveCustomSetting("interface_profile","italiano")
    #Ricerca package
    filename=sys.argv[0]
    fullpathfilename=os.path.abspath(filename)
    fullpath=fullpathfilename[:fullpathfilename.rfind(os.path.sep)]
    refpath=os.path.abspath(fullpath+os.path.sep+".."+os.path.sep+".."+os.path.sep)
    #Spostamento automatico package vicino alla cartella
    appfoldername=fullpath.split(os.path.sep)[-2]
    package_name=appfoldername[:-len("_database")]
    if sys.platform[:3] == 'win':
        package_name+=".exe"
    elif sys.platform == 'darwin':
        package_name+=".command"
    else:   #Linux
        package_name+=".sh"
    if not os.path.exists(refpath+os.path.sep+package_name) and not os.path.exists(refpath+appfoldername+os.path.sep+package_name):
        if sys.platform[:3] == 'win':
            localsettings=Get_User_ShellFolders()["Cache"]
            try: 
                shutil.move('"'+localsettings+os.path.sep+package_name+'"', refpath)
            except: pass
    #Spostamento su desktop su richiesta
    chfilename=fullpath+os.path.sep+"id_install.txt"
    if os.path.exists(chfilename):
        ckfile = open(chfilename,"r")
        contents=ckfile.read()
        ckfile.close()
        if len(contents)<5:
            import random
            random.seed()
            id_install=str(int(random.random()*10000000000))[0:10]
            ckfile = open(chfilename,"w")
            ckfile.write(id_install+"\n")
            ckfile.close()
            #
            desktoppath=os.path.join(os.path.expanduser("~"), "Desktop")
            if refpath!=desktoppath:
                app = wx.App(redirect=False)
                dlg = wx.MessageDialog(None,TT("Can the application be moved on the Desktop?"),TT("Application relocation"),wx.OK|wx.CANCEL|wx.ICON_QUESTION)
                result = dlg.ShowModal()
                dlg.Destroy()
                if result == wx.ID_OK:
                    try: os.unlink(desktoppath+os.path.sep+package_name)
                    except: pass
                    error=True
                    try: 
                        shutil.move(refpath+os.path.sep+package_name, desktoppath+os.path.sep)
                        error=False
                    except: pass
                    if error:
                        wx.MessageBox(TT("The application cannot be moved.\nPlease don't click 'EXECUTE' directly from the Browser, click on 'SAVE' instead."), TT("Application relocation"), wx.ICON_EXCLAMATION | wx.OK, None)
                        shutil.rmtree(refpath+os.path.sep+appfoldername,ignore_errors=True)
                    else:
                        import subprocess
                        if sys.platform[:3] != 'win':
                            shutil.move(refpath+os.path.sep+appfoldername, desktoppath+os.path.sep)
                            os.chdir(desktoppath+os.path.sep+appfoldername+os.path.sep+"bin"+os.path.sep)
                            newheavybase = subprocess.Popen(["python",PROGNAME+".py"])
                        else:
                            os.chdir(desktoppath)
                            newheavybase = subprocess.Popen([package_name])
                            shutil.rmtree(refpath+os.path.sep+appfoldername,ignore_errors=True)
                    app=None
                    del app
                    job=job_server_local.submit(AsyncKillAll,(0,pp_fnames), (log,), ())
                    sys.exit()
                    return
                app=None
                del app
    else:
        import random
        random.seed()
        id_install=str(int(random.random()*10000000000))[0:10]
        ckfile = open(chfilename,"w")
        ckfile.write(id_install+"\n")
        ckfile.close()
    
    # cartella in un'altra cartella
    parentdir=".."+os.path.sep+".."+os.path.sep+"bin"+os.path.sep
    if os.path.abspath(parentdir)!=os.path.abspath("."):
        if os.path.isfile(parentdir+DATABASE_NAME) and os.path.isfile(parentdir+PROFILE_NAME) and os.path.isfile(parentdir+PROGNAME+".py"):
            os.chdir(parentdir)
            import subprocess
            if os.name=='nt':
                if not DEBUG_MODE:
                    newheavybase = subprocess.Popen(["pythonw",PROGNAME+".py"])
                else:
                    newheavybase = subprocess.Popen(["python",PROGNAME+".py"])
            else:
                newheavybase = subprocess.Popen(["python",PROGNAME+".py"])
            sys.exit()
            return
        
    par_filename="outlog.txt"
    par_redirect=True
    try: os.unlink(par_filename)
    except: pass
    if DEBUG_MODE: 
        par_filename=None
        par_redirect=False
        #if sys.platform != 'win32':
            #par_redirect=False
        #else:
            #pass
    application = HeavyBase(redirect=par_redirect, filename=par_filename)
    #application = HeavyBase(0)
    application.MainLoop()

if __name__ == '__main__':
    main()
